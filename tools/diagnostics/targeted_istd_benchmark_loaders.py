"""Input loaders for the targeted ISTD benchmark diagnostic."""

from __future__ import annotations

import json
import re
from collections.abc import Mapping
from pathlib import Path

from openpyxl import load_workbook

from tools.diagnostics.diagnostic_io import (
    bool_value,
    read_tsv_required,
)
from tools.diagnostics.diagnostic_io import (
    optional_float as _float_value,
)
from tools.diagnostics.diagnostic_io import (
    require_fields as _require_fields,
)
from tools.diagnostics.diagnostic_io import (
    required_float as _required_float,
)
from tools.diagnostics.diagnostic_io import (
    required_indexes as _required_indexes,
)
from tools.diagnostics.diagnostic_io import (
    split_semicolon_labels as _split_semicolon_labels,
)
from tools.diagnostics.diagnostic_io import (
    text_value as _text,
)
from tools.diagnostics.targeted_istd_benchmark_models import (
    AlignmentCell,
    AlignmentFeature,
    AlignmentMatrixData,
    TargetDefinition,
    TargetedPoint,
    TargetedReliabilityPoint,
)


def read_target_definitions(path: Path) -> tuple[TargetDefinition, ...]:
    if not path.exists():
        raise FileNotFoundError(str(path))
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook["Targets"]
        rows = sheet.iter_rows(values_only=True)
        header = next(rows)
        cols = _required_indexes(
            header,
            (
                "Label",
                "Role",
                "m/z",
                "RT min",
                "RT max",
                "ppm tol",
                "NL (Da)",
                "Expected product m/z",
            ),
            "Targets",
        )
        targets: list[TargetDefinition] = []
        for row in rows:
            role = _text(row[cols["Role"]])
            if role != "ISTD":
                continue
            label = _text(row[cols["Label"]])
            if not label:
                continue
            targets.append(
                TargetDefinition(
                    label=label,
                    role=role,
                    mz=_required_float(row[cols["m/z"]], "m/z", label),
                    rt_min=_required_float(row[cols["RT min"]], "RT min", label),
                    rt_max=_required_float(row[cols["RT max"]], "RT max", label),
                    ppm_tol=_required_float(row[cols["ppm tol"]], "ppm tol", label),
                    neutral_loss_da=_required_float(
                        row[cols["NL (Da)"]],
                        "NL (Da)",
                        label,
                    ),
                    product_mz=_required_float(
                        row[cols["Expected product m/z"]],
                        "Expected product m/z",
                        label,
                    ),
                )
            )
        return tuple(targets)
    finally:
        workbook.close()


def read_targeted_points(path: Path) -> dict[str, tuple[TargetedPoint, ...]]:
    if not path.exists():
        raise FileNotFoundError(str(path))
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook["XIC Results"]
        rows = sheet.iter_rows(values_only=True)
        header = next(rows)
        cols = _required_indexes(
            header,
            (
                "SampleName",
                "Target",
                "Role",
                "RT",
                "Area",
                "NL",
                "Confidence",
                "Reason",
            ),
            "XIC Results",
        )
        current_sample = ""
        grouped: dict[str, list[TargetedPoint]] = {}
        for row in rows:
            raw_sample = row[cols["SampleName"]]
            if raw_sample not in (None, ""):
                current_sample = _normalize_sample_id(_text(raw_sample))
            if not current_sample:
                continue
            label = _text(row[cols["Target"]])
            role = _text(row[cols["Role"]])
            if not label or role != "ISTD":
                continue
            grouped.setdefault(label, []).append(
                TargetedPoint(
                    sample_stem=current_sample,
                    target_label=label,
                    role=role,
                    rt=_float_value(row[cols["RT"]]),
                    area=_float_value(row[cols["Area"]]),
                    nl=_text(row[cols["NL"]]),
                    confidence=_text(row[cols["Confidence"]]),
                    reason=_text(row[cols["Reason"]]),
                )
            )
        return {label: tuple(points) for label, points in grouped.items()}
    finally:
        workbook.close()


def read_targeted_reliability_points(
    path: Path,
) -> dict[tuple[str, str], TargetedReliabilityPoint]:
    if not path.exists():
        raise FileNotFoundError(str(path))
    payload = json.loads(path.read_text(encoding="utf-8"))
    rows = payload.get("rows")
    if not isinstance(rows, list):
        raise ValueError(f"{path} is missing rows list")
    points: dict[tuple[str, str], TargetedReliabilityPoint] = {}
    for index, row in enumerate(rows, start=1):
        if not isinstance(row, dict):
            raise ValueError(f"{path} rows[{index}] must be an object")
        sample = _normalize_sample_id(
            _required_text_value(path, index, row, "sample_name")
        )
        target = _required_text_value(path, index, row, "target_label")
        state = _required_text_value(path, index, row, "reliability_state")
        points[(sample, target)] = TargetedReliabilityPoint(
            sample_stem=sample,
            target_label=target,
            reliability_state=state,
            risk_reasons=_risk_reasons_value(row.get("risk_reasons")),
        )
    return points


def read_alignment_review(path: Path) -> tuple[AlignmentFeature, ...]:
    rows = list(read_tsv_required(path, ()))
    _require_fields(
        rows,
        (
            "feature_family_id",
            "neutral_loss_tag",
            "family_center_mz",
            "family_center_rt",
            "family_product_mz",
            "family_observed_neutral_loss_da",
            "include_in_primary_matrix",
        ),
        path,
    )
    return tuple(
        AlignmentFeature(
            feature_family_id=row["feature_family_id"],
            neutral_loss_tag=row["neutral_loss_tag"],
            family_center_mz=_required_float(
                row.get("family_center_mz"),
                "family_center_mz",
                row["feature_family_id"],
            ),
            family_center_rt=_required_float(
                row.get("family_center_rt"),
                "family_center_rt",
                row["feature_family_id"],
            ),
            family_product_mz=_required_float(
                row.get("family_product_mz"),
                "family_product_mz",
                row["feature_family_id"],
            ),
            family_observed_neutral_loss_da=_required_float(
                row.get("family_observed_neutral_loss_da"),
                "family_observed_neutral_loss_da",
                row["feature_family_id"],
            ),
            include_in_primary_matrix=_is_primary_review_row(row),
        )
        for row in rows
    )


def read_alignment_matrix(path: Path) -> AlignmentMatrixData:
    rows = list(read_tsv_required(path, ()))
    if rows and "feature_family_id" not in rows[0]:
        return _read_clean_alignment_matrix(path, rows)
    return _read_legacy_alignment_matrix(path, rows)


def _read_legacy_alignment_matrix(
    path: Path,
    rows: list[dict[str, str]],
) -> AlignmentMatrixData:
    _require_fields(rows, ("feature_family_id",), path)
    metadata_columns = {
        "feature_family_id",
        "neutral_loss_tag",
        "family_center_mz",
        "family_center_rt",
        "family_product_mz",
        "family_observed_neutral_loss_da",
    }
    fieldnames = set(rows[0])
    sample_columns = sorted(fieldnames - metadata_columns)
    matrix: dict[str, dict[str, float]] = {}
    normalized_samples = frozenset(
        _normalize_sample_id(sample) for sample in sample_columns
    )
    for row in rows:
        family_id = row["feature_family_id"]
        values: dict[str, float] = {}
        for sample in sample_columns:
            area = _float_value(row.get(sample))
            if area is not None and area > 0:
                values[_normalize_sample_id(sample)] = area
        matrix[family_id] = values
    return AlignmentMatrixData(
        areas_by_family=matrix,
        sample_stems=normalized_samples,
    )


def _read_clean_alignment_matrix(
    path: Path,
    rows: list[dict[str, str]],
) -> AlignmentMatrixData:
    _require_fields(rows, ("Mz", "RT"), path)
    identity_path = path.with_name("alignment_matrix_identity.tsv")
    identity_rows = list(read_tsv_required(identity_path, ()))
    _require_fields(
        identity_rows,
        ("matrix_row_index", "source_feature_family_ids"),
        identity_path,
    )
    identity_by_index = _matrix_identity_by_index(identity_path, identity_rows)
    metadata_columns = {"Mz", "RT"}
    fieldnames = set(rows[0])
    sample_columns = sorted(fieldnames - metadata_columns)
    matrix: dict[str, dict[str, float]] = {}
    normalized_samples = frozenset(
        _normalize_sample_id(sample) for sample in sample_columns
    )
    for row_index, row in enumerate(rows, start=1):
        family_ids = identity_by_index.get(row_index)
        if family_ids is None:
            raise ValueError(
                f"{identity_path} is missing matrix_row_index {row_index}"
            )
        values: dict[str, float] = {}
        for sample in sample_columns:
            area = _float_value(row.get(sample))
            if area is not None and area > 0:
                values[_normalize_sample_id(sample)] = area
        for family_id in family_ids:
            matrix[family_id] = dict(values)
    return AlignmentMatrixData(
        areas_by_family=matrix,
        sample_stems=normalized_samples,
    )


def _matrix_identity_by_index(
    path: Path,
    rows: list[dict[str, str]],
) -> dict[int, tuple[str, ...]]:
    identity_by_index: dict[int, tuple[str, ...]] = {}
    for row_number, row in enumerate(rows, start=2):
        index_text = _text(row.get("matrix_row_index"))
        try:
            matrix_row_index = int(index_text)
        except ValueError as exc:
            raise ValueError(
                f"{path} row {row_number} has invalid matrix_row_index: "
                f"{index_text!r}"
            ) from exc
        if matrix_row_index in identity_by_index:
            raise ValueError(
                f"{path} has duplicate matrix_row_index {matrix_row_index}"
            )
        family_ids = tuple(
            _split_semicolon_labels(row.get("source_feature_family_ids"))
        )
        if not family_ids:
            raise ValueError(
                f"{path} row {row_number} is missing source_feature_family_ids"
            )
        identity_by_index[matrix_row_index] = family_ids
    return identity_by_index


def _is_primary_review_row(row: Mapping[str, str]) -> bool:
    if not _is_trueish(row.get("include_in_primary_matrix")):
        return False
    identity_decision = (row.get("identity_decision") or "").strip()
    if identity_decision and identity_decision != "production_family":
        return False
    return True


def read_alignment_cells(path: Path) -> dict[tuple[str, str], AlignmentCell]:
    rows = list(read_tsv_required(path, ()))
    _require_fields(
        rows,
        ("feature_family_id", "sample_stem", "status", "area", "apex_rt"),
        path,
    )
    cells: dict[tuple[str, str], AlignmentCell] = {}
    for row in rows:
        sample = _normalize_sample_id(row["sample_stem"])
        cell = AlignmentCell(
            feature_family_id=row["feature_family_id"],
            sample_stem=sample,
            status=row.get("status", ""),
            area=_float_value(row.get("area")),
            apex_rt=_float_value(row.get("apex_rt")),
        )
        cells[(cell.feature_family_id, sample)] = cell
    return cells

def _required_text_value(
    path: Path,
    row_number: int,
    row: Mapping[str, object],
    field: str,
) -> str:
    value = _text(row.get(field))
    if not value:
        raise ValueError(f"{path} rows[{row_number}] is missing {field}")
    return value


def _risk_reasons_value(value: object) -> tuple[str, ...]:
    if isinstance(value, list):
        return tuple(_text(item) for item in value if _text(item))
    return tuple(part for part in _text(value).split(";") if part)



def _normalize_sample_id(sample_id: str) -> str:
    value = sample_id.strip()
    return re.sub(
        r"(^|_)QC_(\d+)$",
        lambda match: f"{match.group(1)}QC{match.group(2)}",
        value,
    )


def _is_trueish(value: str | None) -> bool:
    return bool_value(value) is True
