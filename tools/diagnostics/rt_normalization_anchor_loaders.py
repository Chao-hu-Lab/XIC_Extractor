"""Input loaders for the RT normalization anchor diagnostic."""

from __future__ import annotations

import re
from collections.abc import Mapping
from pathlib import Path

from openpyxl import load_workbook

from tools.diagnostics.diagnostic_io import (
    bool_value,
    read_tsv_required,
)
from tools.diagnostics.diagnostic_io import (
    optional_float as _float_value,
)
from tools.diagnostics.diagnostic_io import (
    require_fields as _require_fields,
)
from tools.diagnostics.diagnostic_io import (
    required_float as _required_float,
)
from tools.diagnostics.diagnostic_io import (
    required_indexes as _required_indexes,
)
from tools.diagnostics.diagnostic_io import (
    text_value as _text,
)
from tools.diagnostics.rt_normalization_anchor_models import (
    AlignmentCell,
    AlignmentFeature,
    AnchorDefinition,
)
from xic_extractor.alignment.rt_normalization import AnchorPoint
from xic_extractor.injection_rolling import read_injection_order


def _read_anchor_definitions(
    path: Path,
    *,
    active_neutral_loss_da: float,
    active_neutral_loss_tolerance_da: float,
) -> dict[str, AnchorDefinition]:
    if not path.exists():
        raise FileNotFoundError(str(path))
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook["Targets"]
        rows = sheet.iter_rows(values_only=True)
        header = next(rows)
        cols = _required_indexes(
            header,
            ("Label", "Role", "RT min", "RT max", "NL (Da)"),
            "Targets",
        )
        anchors: dict[str, AnchorDefinition] = {}
        for row in rows:
            label = _text(row[cols["Label"]])
            role = _text(row[cols["Role"]])
            if not label or role != "ISTD":
                continue
            neutral_loss = _required_float(row[cols["NL (Da)"]], "NL (Da)", label)
            if abs(neutral_loss - active_neutral_loss_da) > (
                active_neutral_loss_tolerance_da
            ):
                continue
            rt_min = _required_float(row[cols["RT min"]], "RT min", label)
            rt_max = _required_float(row[cols["RT max"]], "RT max", label)
            anchors[label] = AnchorDefinition(
                label=label,
                role=role,
                neutral_loss_da=neutral_loss,
                reference_rt_min=(rt_min + rt_max) / 2.0,
            )
        return anchors
    finally:
        workbook.close()


def _read_anchor_points(
    path: Path,
    anchors: Mapping[str, AnchorDefinition],
) -> tuple[AnchorPoint, ...]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook["XIC Results"]
        rows = sheet.iter_rows(values_only=True)
        header = next(rows)
        cols = _required_indexes(
            header,
            ("SampleName", "Target", "Role", "RT"),
            "XIC Results",
        )
        current_sample = ""
        points: list[AnchorPoint] = []
        for row in rows:
            raw_sample = row[cols["SampleName"]]
            if raw_sample not in (None, ""):
                current_sample = _normalize_sample_id(_text(raw_sample))
            if not current_sample:
                continue
            label = _text(row[cols["Target"]])
            role = _text(row[cols["Role"]])
            if role != "ISTD" or label not in anchors:
                continue
            rt = _float_value(row[cols["RT"]])
            if rt is None:
                continue
            points.append(
                AnchorPoint(
                    sample_stem=current_sample,
                    target_label=label,
                    observed_rt_min=rt,
                    reference_rt_min=anchors[label].reference_rt_min,
                )
            )
        return tuple(points)
    finally:
        workbook.close()


def _read_optional_injection_order(
    sample_info: Path | None,
    *,
    reference_source: str,
) -> dict[str, int] | None:
    if not reference_source.startswith("injection-"):
        return None
    if sample_info is None:
        raise ValueError(f"sample_info is required for {reference_source}")
    return read_injection_order(sample_info)


def _read_alignment_review(path: Path) -> dict[str, AlignmentFeature]:
    rows = list(read_tsv_required(path, ()))
    _require_fields(rows, ("feature_family_id",), path)
    features: dict[str, AlignmentFeature] = {}
    for row in rows:
        family_id = row["feature_family_id"]
        features[family_id] = AlignmentFeature(
            feature_family_id=family_id,
            include_in_primary_matrix=_is_trueish(
                row.get("include_in_primary_matrix"),
            ),
            family_center_mz=_float_value(row.get("family_center_mz")),
            family_center_rt=_float_value(row.get("family_center_rt")),
        )
    return features


def _read_alignment_cells(path: Path) -> tuple[AlignmentCell, ...]:
    rows = list(read_tsv_required(path, ()))
    _require_fields(
        rows,
        ("feature_family_id", "sample_stem", "apex_rt"),
        path,
    )
    cells: list[AlignmentCell] = []
    for row in rows:
        apex_rt = _float_value(row.get("apex_rt"))
        if apex_rt is None:
            continue
        cells.append(
            AlignmentCell(
                feature_family_id=row["feature_family_id"],
                sample_stem=_normalize_sample_id(row["sample_stem"]),
                apex_rt=apex_rt,
            )
        )
    return tuple(cells)


def _normalize_sample_id(sample_id: str) -> str:
    value = sample_id.strip()
    return re.sub(
        r"(^|_)QC_(\d+)$",
        lambda match: f"{match.group(1)}QC{match.group(2)}",
        value,
    )


def _is_trueish(value: str | None) -> bool:
    return bool_value(value) is True
