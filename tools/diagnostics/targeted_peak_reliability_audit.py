"""Targeted peak reliability audit for benchmark eligibility."""

from __future__ import annotations

import argparse
import csv
import json
import math
import sys
from collections import Counter, defaultdict
from collections.abc import Iterator, Mapping, Sequence
from dataclasses import asdict, dataclass
from pathlib import Path
from statistics import median
from typing import Any, Literal, Protocol, cast

from openpyxl import load_workbook

from xic_extractor.evidence_semantics import (
    EvidenceSignalSet,
    classify_evidence_consistency,
)

ReliabilityState = Literal[
    "benchmark_eligible",
    "targeted_review_positive",
    "targeted_review",
    "targeted_negative",
]

ROWS_COLUMNS = (
    "sample_name",
    "target_label",
    "role",
    "rt",
    "area",
    "confidence",
    "nl",
    "prior_rt",
    "prior_source",
    "total_severity",
    "quality_flags",
    "reliability_state",
    "risk_reasons",
    "known_exception",
    "target_area_median",
    "area_to_target_median_ratio",
    "weak_area_threshold_ratio",
)

SUMMARY_COLUMNS = (
    "target_label",
    "role",
    "benchmark_eligible_count",
    "targeted_review_positive_count",
    "targeted_review_count",
    "targeted_negative_count",
    "top_risk_reasons",
    "known_exception",
)

_PEAK_CANDIDATE_COLUMNS = (
    "sample_name",
    "target_label",
    "proposal_sources",
    "selected",
    "raw_score",
    "support_labels",
    "concern_labels",
    "quality_flags",
    "ms2_present",
    "nl_match",
)

_WEAK_AREA_THRESHOLD_RATIO = 0.05


@dataclass(frozen=True)
class TargetedReliabilityOutputs:
    summary_tsv: Path
    rows_tsv: Path
    json_path: Path
    markdown_path: Path


@dataclass(frozen=True)
class TargetedReliabilityRow:
    sample_name: str
    target_label: str
    role: str
    rt: float | None
    area: float | None
    confidence: str
    nl: str
    prior_rt: float | None
    prior_source: str
    total_severity: int | None
    quality_flags: str
    reliability_state: ReliabilityState
    risk_reasons: tuple[str, ...]
    known_exception: str
    target_area_median: float | None = None
    area_to_target_median_ratio: float | None = None
    weak_area_threshold_ratio: float | None = None


@dataclass(frozen=True)
class TargetedReliabilitySummary:
    target_label: str
    role: str
    benchmark_eligible_count: int
    targeted_review_positive_count: int
    targeted_review_count: int
    targeted_negative_count: int
    top_risk_reasons: str
    known_exception: str


@dataclass(frozen=True)
class TargetedReliabilityResult:
    rows: tuple[TargetedReliabilityRow, ...]
    summaries: tuple[TargetedReliabilitySummary, ...]

    @property
    def benchmark_eligible_count(self) -> int:
        return sum(row.reliability_state == "benchmark_eligible" for row in self.rows)

    @property
    def targeted_review_count(self) -> int:
        return sum(row.reliability_state == "targeted_review" for row in self.rows)

    @property
    def targeted_review_positive_count(self) -> int:
        return sum(
            row.reliability_state == "targeted_review_positive"
            for row in self.rows
        )

    @property
    def targeted_negative_count(self) -> int:
        return sum(row.reliability_state == "targeted_negative" for row in self.rows)


@dataclass(frozen=True)
class _TargetedInputRow:
    sample_name: str
    target_label: str
    role: str
    rt: float | None
    area: float | None
    confidence: str
    nl: str
    reason: str


@dataclass(frozen=True)
class _ScoreBreakdown:
    prior_rt: float | None
    prior_source: str
    total_severity: int | None
    quality_flags: str


@dataclass(frozen=True)
class _CandidateEvidence:
    support_labels: tuple[str, ...]
    concern_labels: tuple[str, ...]
    proposal_sources: tuple[str, ...]
    quality_flags: tuple[str, ...]
    ms2_present: bool | None
    nl_match: bool | None
    raw_score: float | None
    diagnostic_product_absence_reason: str = ""


@dataclass(frozen=True)
class _AreaContext:
    target_area_median: float
    area_to_target_median_ratio: float
    weak_area_threshold_ratio: float = _WEAK_AREA_THRESHOLD_RATIO

    @property
    def weak_area(self) -> bool:
        return self.area_to_target_median_ratio < self.weak_area_threshold_ratio


class _WorksheetLike(Protocol):
    def iter_rows(
        self,
        *,
        values_only: bool = False,
    ) -> Iterator[tuple[object, ...]]: ...


def run_targeted_peak_reliability_audit(
    *,
    targeted_workbook: Path,
    output_dir: Path,
    known_target_exceptions: Sequence[str] = (),
    peak_candidates_tsv: Path | None = None,
) -> tuple[TargetedReliabilityOutputs, TargetedReliabilityResult]:
    rows, score_by_key = _load_targeted_workbook(targeted_workbook)
    candidate_by_key = _load_selected_candidate_evidence(peak_candidates_tsv)
    known = _parse_known_exceptions(known_target_exceptions)
    area_context = _area_context_by_key(rows)
    reliability_rows = tuple(
        _classify_row(
            row,
            score=score_by_key.get((row.sample_name, row.target_label)),
            candidate_evidence=candidate_by_key.get(
                (row.sample_name, row.target_label)
            ),
            known_exception=known.get(row.target_label, ""),
            area_context=area_context.get((row.sample_name, row.target_label)),
        )
        for row in rows
    )
    result = TargetedReliabilityResult(
        rows=reliability_rows,
        summaries=_summarize_rows(reliability_rows),
    )
    output_dir.mkdir(parents=True, exist_ok=True)
    outputs = TargetedReliabilityOutputs(
        summary_tsv=output_dir / "targeted_peak_reliability_summary.tsv",
        rows_tsv=output_dir / "targeted_peak_reliability_rows.tsv",
        json_path=output_dir / "targeted_peak_reliability.json",
        markdown_path=output_dir / "targeted_peak_reliability.md",
    )
    _write_outputs(outputs, result, known_target_exceptions=known)
    return outputs, result


def main(argv: Sequence[str] | None = None) -> int:
    args = _parse_args(argv)
    try:
        outputs, result = run_targeted_peak_reliability_audit(
            targeted_workbook=args.targeted_workbook.resolve(),
            output_dir=args.output_dir.resolve(),
            known_target_exceptions=tuple(args.known_target_exception),
            peak_candidates_tsv=(
                args.peak_candidates_tsv.resolve()
                if args.peak_candidates_tsv is not None
                else None
            ),
        )
    except (FileNotFoundError, KeyError, ValueError) as exc:
        print(str(exc), file=sys.stderr)
        return 2
    print(f"Summary TSV: {outputs.summary_tsv}")
    print(f"Rows TSV: {outputs.rows_tsv}")
    print(f"Reliability JSON: {outputs.json_path}")
    print(f"Reliability report: {outputs.markdown_path}")
    return 0 if result.targeted_review_count == 0 else 0


def _parse_args(argv: Sequence[str] | None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Audit targeted peak reliability for benchmark eligibility.",
    )
    parser.add_argument("--targeted-workbook", type=Path, required=True)
    parser.add_argument("--output-dir", type=Path, required=True)
    parser.add_argument(
        "--peak-candidates-tsv",
        type=Path,
        help=(
            "Optional selected peak candidate evidence TSV. When provided, "
            "selected candidate consistency can mark NL_FAIL rows as "
            "targeted_review_positive without making them benchmark eligible."
        ),
    )
    parser.add_argument(
        "--known-target-exception",
        action="append",
        default=[],
        help="Known targeted-side exception in TARGET:FAILURE_MODE form.",
    )
    return parser.parse_args(argv)


def _load_targeted_workbook(
    path: Path,
) -> tuple[tuple[_TargetedInputRow, ...], dict[tuple[str, str], _ScoreBreakdown]]:
    if not path.exists():
        raise FileNotFoundError(str(path))
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        rows = _read_xic_results(workbook["XIC Results"])
        score_by_key = (
            _read_score_breakdown(workbook["Score Breakdown"])
            if "Score Breakdown" in workbook.sheetnames
            else {}
        )
        return rows, score_by_key
    finally:
        workbook.close()


def _read_xic_results(sheet: _WorksheetLike) -> tuple[_TargetedInputRow, ...]:
    iterator = sheet.iter_rows(values_only=True)
    header = next(iterator)
    cols = _required_indexes(
        header,
        (
            "SampleName",
            "Target",
            "Role",
            "RT",
            "Area",
            "NL",
            "Confidence",
        ),
        "XIC Results",
    )
    header_indexes = {
        str(value).strip(): index for index, value in enumerate(header) if value
    }
    reason_index = header_indexes.get("Reason")
    current_sample = ""
    rows: list[_TargetedInputRow] = []
    for row in iterator:
        raw_sample = row[cols["SampleName"]]
        if raw_sample not in (None, ""):
            current_sample = _text(raw_sample)
        if not current_sample:
            continue
        target_label = _text(row[cols["Target"]])
        if not target_label:
            continue
        rows.append(
            _TargetedInputRow(
                sample_name=current_sample,
                target_label=target_label,
                role=_text(row[cols["Role"]]),
                rt=_float_value(row[cols["RT"]]),
                area=_float_value(row[cols["Area"]]),
                confidence=_text(row[cols["Confidence"]]).upper(),
                nl=_text(row[cols["NL"]]).upper(),
                reason=_text(row[reason_index]) if reason_index is not None else "",
            )
        )
    return tuple(rows)


def _read_score_breakdown(
    sheet: _WorksheetLike,
) -> dict[tuple[str, str], _ScoreBreakdown]:
    iterator = sheet.iter_rows(values_only=True)
    header = next(iterator)
    cols = _required_indexes(
        header,
        (
            "SampleName",
            "Target",
            "Prior RT",
            "Prior Source",
            "Total Severity",
            "Quality Flags",
        ),
        "Score Breakdown",
    )
    scores: dict[tuple[str, str], _ScoreBreakdown] = {}
    for row in iterator:
        sample = _text(row[cols["SampleName"]])
        target = _text(row[cols["Target"]])
        if not sample or not target:
            continue
        scores[(sample, target)] = _ScoreBreakdown(
            prior_rt=_float_value(row[cols["Prior RT"]]),
            prior_source=_text(row[cols["Prior Source"]]),
            total_severity=_int_value(row[cols["Total Severity"]]),
            quality_flags=_text(row[cols["Quality Flags"]]),
        )
    return scores


def _classify_row(
    row: _TargetedInputRow,
    *,
    score: _ScoreBreakdown | None,
    candidate_evidence: _CandidateEvidence | None,
    known_exception: str,
    area_context: _AreaContext | None,
) -> TargetedReliabilityRow:
    if row.rt is None or row.area is None or row.area <= 0:
        return TargetedReliabilityRow(
            sample_name=row.sample_name,
            target_label=row.target_label,
            role=row.role,
            rt=row.rt,
            area=row.area,
            confidence=row.confidence,
            nl=row.nl,
            prior_rt=score.prior_rt if score is not None else None,
            prior_source=score.prior_source if score is not None else "",
            total_severity=score.total_severity if score is not None else None,
            quality_flags=score.quality_flags if score is not None else "",
            reliability_state="targeted_negative",
            risk_reasons=("no_usable_peak",),
            known_exception=known_exception,
            target_area_median=(
                area_context.target_area_median if area_context is not None else None
            ),
            area_to_target_median_ratio=(
                area_context.area_to_target_median_ratio
                if area_context is not None
                else None
            ),
            weak_area_threshold_ratio=(
                area_context.weak_area_threshold_ratio
                if area_context is not None
                else None
            ),
        )

    risk_reasons: list[str] = []
    if row.confidence == "VERY_LOW" or (
        row.confidence == "LOW" and not _is_accepted_low_istd(row)
    ):
        risk_reasons.append("low_confidence")
    if _is_nl_fail(row.nl):
        if _is_plausible_nl_dropout(row, score, candidate_evidence):
            risk_reasons.append("plausible_nl_dropout")
        else:
            risk_reasons.append("hard_nl_conflict")
    elif _is_no_ms2(row.nl):
        risk_reasons.append("no_ms2")
    if score is None:
        risk_reasons.append("score_breakdown_unavailable")
    elif score.quality_flags:
        risk_reasons.append("quality_flags")
    if area_context is not None and area_context.weak_area:
        risk_reasons.append("weak_area_rank")
    if _is_nl_fail(row.nl) and candidate_evidence is not None:
        risk_reasons.extend(_candidate_product_context_reasons(candidate_evidence))

    blocking_reasons = tuple(
        reason for reason in risk_reasons if reason != "score_breakdown_unavailable"
    )
    if not blocking_reasons:
        state: ReliabilityState = "benchmark_eligible"
    elif _is_review_positive(blocking_reasons):
        state = "targeted_review_positive"
    else:
        state = "targeted_review"
    return TargetedReliabilityRow(
        sample_name=row.sample_name,
        target_label=row.target_label,
        role=row.role,
        rt=row.rt,
        area=row.area,
        confidence=row.confidence,
        nl=row.nl,
        prior_rt=score.prior_rt if score is not None else None,
        prior_source=score.prior_source if score is not None else "",
        total_severity=score.total_severity if score is not None else None,
        quality_flags=score.quality_flags if score is not None else "",
        reliability_state=state,
        risk_reasons=tuple(dict.fromkeys(risk_reasons)),
        known_exception=known_exception,
        target_area_median=(
            area_context.target_area_median if area_context is not None else None
        ),
        area_to_target_median_ratio=(
            area_context.area_to_target_median_ratio
            if area_context is not None
            else None
        ),
        weak_area_threshold_ratio=(
            area_context.weak_area_threshold_ratio
            if area_context is not None
            else None
        ),
    )


def _area_context_by_key(
    rows: Sequence[_TargetedInputRow],
) -> dict[tuple[str, str], _AreaContext]:
    by_target: dict[str, list[_TargetedInputRow]] = defaultdict(list)
    for row in rows:
        if row.area is not None and row.area > 0:
            by_target[row.target_label].append(row)
    context: dict[tuple[str, str], _AreaContext] = {}
    for target_rows in by_target.values():
        if len(target_rows) < 3:
            continue
        target_median = median(float(row.area) for row in target_rows if row.area)
        if target_median <= 0:
            continue
        for row in target_rows:
            if row.area is None:
                continue
            context[(row.sample_name, row.target_label)] = _AreaContext(
                target_area_median=float(target_median),
                area_to_target_median_ratio=float(row.area) / float(target_median),
            )
    return context


def _summarize_rows(
    rows: Sequence[TargetedReliabilityRow],
) -> tuple[TargetedReliabilitySummary, ...]:
    grouped: dict[str, list[TargetedReliabilityRow]] = defaultdict(list)
    for row in rows:
        grouped[row.target_label].append(row)
    summaries: list[TargetedReliabilitySummary] = []
    for target_label in sorted(grouped):
        target_rows = grouped[target_label]
        reasons = Counter(
            reason for row in target_rows for reason in row.risk_reasons
        )
        top_reasons = ";".join(
            reason for reason, _count in reasons.most_common(5)
        )
        known_exception = next(
            (row.known_exception for row in target_rows if row.known_exception),
            "",
        )
        summaries.append(
            TargetedReliabilitySummary(
                target_label=target_label,
                role=target_rows[0].role,
                benchmark_eligible_count=sum(
                    row.reliability_state == "benchmark_eligible"
                    for row in target_rows
                ),
                targeted_review_positive_count=sum(
                    row.reliability_state == "targeted_review_positive"
                    for row in target_rows
                ),
                targeted_review_count=sum(
                    row.reliability_state == "targeted_review"
                    for row in target_rows
                ),
                targeted_negative_count=sum(
                    row.reliability_state == "targeted_negative"
                    for row in target_rows
                ),
                top_risk_reasons=top_reasons,
                known_exception=known_exception,
            )
        )
    return tuple(summaries)


def _write_outputs(
    outputs: TargetedReliabilityOutputs,
    result: TargetedReliabilityResult,
    *,
    known_target_exceptions: Mapping[str, str],
) -> None:
    _write_tsv(outputs.rows_tsv, ROWS_COLUMNS, _row_dicts(result.rows))
    _write_tsv(
        outputs.summary_tsv,
        SUMMARY_COLUMNS,
        _summary_dicts(result.summaries),
    )
    outputs.json_path.write_text(
        json.dumps(
            _json_payload(result, known_target_exceptions=known_target_exceptions),
            indent=2,
            sort_keys=True,
        ),
        encoding="utf-8",
    )
    _write_markdown(outputs.markdown_path, result)


def _row_dicts(
    rows: Sequence[TargetedReliabilityRow],
) -> list[dict[str, object]]:
    return [
        {
            **asdict(row),
            "risk_reasons": ";".join(row.risk_reasons),
        }
        for row in rows
    ]


def _summary_dicts(
    summaries: Sequence[TargetedReliabilitySummary],
) -> list[dict[str, object]]:
    return [asdict(summary) for summary in summaries]


def _json_payload(
    result: TargetedReliabilityResult,
    *,
    known_target_exceptions: Mapping[str, str],
) -> dict[str, object]:
    return {
        "overall_status": (
            "WARN"
            if result.targeted_review_count
            or result.targeted_review_positive_count
            else "PASS"
        ),
        "summary": {
            "benchmark_eligible_count": result.benchmark_eligible_count,
            "targeted_review_positive_count": (
                result.targeted_review_positive_count
            ),
            "targeted_review_count": result.targeted_review_count,
            "targeted_negative_count": result.targeted_negative_count,
        },
        "known_target_exceptions": dict(known_target_exceptions),
        "rows": _row_dicts(result.rows),
        "summaries": _summary_dicts(result.summaries),
    }


def _write_markdown(path: Path, result: TargetedReliabilityResult) -> None:
    overall_status = (
        "WARN"
        if result.targeted_review_count or result.targeted_review_positive_count
        else "PASS"
    )
    lines = [
        "# Targeted Peak Reliability Audit",
        "",
        f"Overall status: {overall_status}",
        "",
        (
            "| Target | Eligible | Review-positive | Review | Negative | "
            "Known exception | Top risk reasons |"
        ),
        "|---|---:|---:|---:|---:|---|---|",
    ]
    for summary in result.summaries:
        lines.append(
            "| "
            f"{summary.target_label} | "
            f"{summary.benchmark_eligible_count} | "
            f"{summary.targeted_review_positive_count} | "
            f"{summary.targeted_review_count} | "
            f"{summary.targeted_negative_count} | "
            f"{summary.known_exception} | "
            f"{summary.top_risk_reasons} |"
        )
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def _write_tsv(
    path: Path,
    fieldnames: Sequence[str],
    rows: Sequence[Mapping[str, object]],
) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, delimiter="\t")
        writer.writeheader()
        for row in rows:
            writer.writerow(
                {key: _format_value(row.get(key, "")) for key in fieldnames}
            )


def _parse_known_exceptions(values: Sequence[str]) -> dict[str, str]:
    known: dict[str, str] = {}
    for value in values:
        if ":" not in value:
            raise ValueError(
                "--known-target-exception must use TARGET:FAILURE_MODE form"
            )
        target, mode = value.split(":", 1)
        target = target.strip()
        mode = mode.strip()
        if not target or not mode:
            raise ValueError(
                "--known-target-exception must use TARGET:FAILURE_MODE form"
            )
        known[target] = mode
    return known


def _is_nl_fail(value: str) -> bool:
    token = value.strip().upper()
    return token == "NL_FAIL" or "NL_FAIL" in token or token.startswith("✗")


def _is_no_ms2(value: str) -> bool:
    token = value.strip().upper().replace(" ", "_")
    return token == "NO_MS2" or "NO_MS2" in token


def _is_review_positive(blocking_reasons: Sequence[str]) -> bool:
    if "plausible_nl_dropout" not in blocking_reasons:
        return False
    hard_reasons = {
        "hard_nl_conflict",
        "no_ms2",
        "quality_flags",
        "weak_area_rank",
    }
    return not any(reason in hard_reasons for reason in blocking_reasons)


def _is_plausible_nl_dropout(
    row: _TargetedInputRow,
    score: _ScoreBreakdown | None,
    candidate_evidence: _CandidateEvidence | None,
) -> bool:
    if candidate_evidence is not None:
        return "plausible_nl_dropout" in classify_evidence_consistency(
            _candidate_evidence_signal_set(candidate_evidence)
        )
    return "plausible_nl_dropout" in classify_evidence_consistency(
        _evidence_signal_set(row, score)
    )


def _candidate_evidence_signal_set(
    evidence: _CandidateEvidence,
) -> EvidenceSignalSet:
    return EvidenceSignalSet(
        support_labels=evidence.support_labels,
        concern_labels=evidence.concern_labels,
        proposal_sources=evidence.proposal_sources,
        quality_flags=evidence.quality_flags,
        ms2_present=evidence.ms2_present,
        nl_match=evidence.nl_match,
        raw_score=evidence.raw_score,
    )


def _candidate_product_context_reasons(
    evidence: _CandidateEvidence,
) -> tuple[str, ...]:
    if not evidence.diagnostic_product_absence_reason:
        return ()
    return (evidence.diagnostic_product_absence_reason,)


def _evidence_signal_set(
    row: _TargetedInputRow,
    score: _ScoreBreakdown | None,
) -> EvidenceSignalSet:
    reason = row.reason.upper()
    support_labels: list[str] = []
    concern_labels: list[str] = []
    proposal_sources: list[str] = []
    if "LOCAL S/N STRONG" in reason:
        support_labels.append("local_sn_strong")
    if "TRACE CLEAN" in reason:
        support_labels.append("trace_clean")
    if "SHAPE CLEAN" in reason:
        support_labels.append("shape_clean")
    if "CENTWAVE_CWT" in reason:
        proposal_sources.append("centwave_cwt")
    if _is_nl_fail(row.nl):
        concern_labels.append("nl_fail")
    if _is_no_ms2(row.nl) or "NO MS2" in reason or "NO_MS2" in reason:
        concern_labels.append("no_ms2")
    concern_labels.extend(_reason_hard_concern_labels(reason))
    return EvidenceSignalSet(
        support_labels=tuple(dict.fromkeys(support_labels)),
        concern_labels=tuple(dict.fromkeys(concern_labels)),
        proposal_sources=tuple(dict.fromkeys(proposal_sources)),
        quality_flags=(
            tuple(_split_semicolon_labels(score.quality_flags))
            if score is not None
            else ()
        ),
        ms2_present=False
        if "no_ms2" in concern_labels
        else True
        if _is_nl_fail(row.nl)
        else None,
        nl_match=False if _is_nl_fail(row.nl) else None,
    )


def _reason_hard_concern_labels(reason: str) -> list[str]:
    labels: list[str] = []
    for token, label in (
        ("HARD QUALITY FLAG", "hard_quality_flag"),
        ("WEAK CANDIDATE", "hard_quality_flag"),
        ("LOW SCAN SUPPORT", "low_scan_support"),
        ("LOW_TRACE_CONTINUITY", "low_trace_continuity"),
        ("LOW TRACE CONTINUITY", "low_trace_continuity"),
        ("POOR EDGE RECOVERY", "poor_edge_recovery"),
        ("EDGE CLIPPED", "poor_edge_recovery"),
        ("RT CENTRALITY POOR", "rt_centrality_poor"),
        ("RT_CENTRALITY_POOR", "rt_centrality_poor"),
        ("SHAPE POOR", "shape_poor"),
        ("SHAPE_POOR", "shape_poor"),
    ):
        if token in reason:
            labels.append(label)
    return labels


def _split_semicolon_labels(value: str) -> list[str]:
    return [part.strip() for part in value.split(";") if part.strip()]


def _load_selected_candidate_evidence(
    path: Path | None,
) -> dict[tuple[str, str], _CandidateEvidence]:
    if path is None:
        return {}
    rows = _read_required_tsv(path, _PEAK_CANDIDATE_COLUMNS)
    selected_by_key: dict[tuple[str, str], list[_CandidateEvidence]] = defaultdict(
        list
    )
    for row in rows:
        if _bool_value(row["selected"]) is not True:
            continue
        key = (row["sample_name"], row["target_label"])
        selected_by_key[key].append(
            _CandidateEvidence(
                support_labels=tuple(_split_semicolon_labels(row["support_labels"])),
                concern_labels=tuple(_split_semicolon_labels(row["concern_labels"])),
                proposal_sources=tuple(
                    _split_semicolon_labels(row["proposal_sources"])
                ),
                quality_flags=tuple(_split_semicolon_labels(row["quality_flags"])),
                ms2_present=_bool_value(row["ms2_present"]),
                nl_match=_bool_value(row["nl_match"]),
                raw_score=_float_value(row["raw_score"]),
                diagnostic_product_absence_reason=row.get(
                    "diagnostic_product_absence_reason",
                    "",
                ).strip(),
            )
        )
    return {
        key: values[0]
        for key, values in selected_by_key.items()
        if len(values) == 1
    }


def _read_required_tsv(
    path: Path,
    required: Sequence[str],
) -> tuple[dict[str, str], ...]:
    if not path.exists():
        raise FileNotFoundError(str(path))
    with path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle, delimiter="\t")
        fieldnames = tuple(reader.fieldnames or ())
        missing = [column for column in required if column not in fieldnames]
        if missing:
            raise ValueError(
                f"{path}: missing required columns: {', '.join(missing)}"
            )
        return tuple(dict(row) for row in reader)


def _bool_value(value: object) -> bool | None:
    token = _text(value).strip().upper()
    if token in {"TRUE", "T", "YES", "Y", "1"}:
        return True
    if token in {"FALSE", "F", "NO", "N", "0"}:
        return False
    return None


def _is_accepted_low_istd(row: _TargetedInputRow) -> bool:
    reason = row.reason.upper()
    if any(
        token in reason
        for token in ("HARD QUALITY FLAG", "WEAK CANDIDATE", "NL FAIL", "NO MS2")
    ):
        return False
    return (
        row.role.strip().upper() == "ISTD"
        and row.confidence == "LOW"
        and not _is_nl_fail(row.nl)
        and not _is_no_ms2(row.nl)
        and "DECISION: ACCEPTED" in reason
    )


def _required_indexes(
    header: Sequence[object],
    required: Sequence[str],
    sheet_name: str,
) -> dict[str, int]:
    indexes = {str(value).strip(): index for index, value in enumerate(header) if value}
    missing = [field for field in required if field not in indexes]
    if missing:
        raise ValueError(f"{sheet_name} is missing required columns: {missing}")
    return {field: indexes[field] for field in required}


def _float_value(value: object) -> float | None:
    try:
        numeric = float(cast(Any, value))
    except (TypeError, ValueError):
        return None
    return numeric if math.isfinite(numeric) else None


def _int_value(value: object) -> int | None:
    numeric = _float_value(value)
    if numeric is None:
        return None
    return int(numeric)


def _text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _format_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if math.isnan(value):
            return ""
        return f"{value:.6g}"
    return str(value)


if __name__ == "__main__":
    raise SystemExit(main())
