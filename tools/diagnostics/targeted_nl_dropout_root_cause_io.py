"""Input loaders for targeted NL dropout root-cause audit."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from tools.diagnostics.diagnostic_io import (
    bool_value as _bool_value,
    optional_float as _optional_float,
    optional_int as _optional_int,
    read_tsv_required as _read_required_tsv,
    required_indexes,
    split_semicolon_labels as _split_labels,
    text_value as _text,
)
from tools.diagnostics.targeted_nl_dropout_root_cause_models import (
    _CANDIDATE_COLUMNS,
    _RELIABILITY_COLUMNS,
    CandidateRow,
    ReliabilityRow,
)


def _read_reliability_rows(path: Path) -> tuple[ReliabilityRow, ...]:
    rows = _read_required_tsv(path, _RELIABILITY_COLUMNS)
    return tuple(
        ReliabilityRow(
            sample_name=row["sample_name"],
            target_label=row["target_label"],
            role=row["role"],
            reliability_state=row["reliability_state"],
            risk_reasons=tuple(_split_labels(row["risk_reasons"])),
        )
        for row in rows
    )


def _read_candidate_rows(path: Path) -> tuple[CandidateRow, ...]:
    rows = _read_required_tsv(path, _CANDIDATE_COLUMNS)
    return tuple(
        CandidateRow(
            sample_name=row["sample_name"],
            target_label=row["target_label"],
            resolver_mode=row["resolver_mode"],
            candidate_id=row["candidate_id"],
            proposal_sources=tuple(_split_labels(row["proposal_sources"])),
            rt_apex_min=_optional_float(row["rt_apex_min"]),
            selected=_bool_value(row["selected"]) is True,
            confidence=row["confidence"],
            raw_score=_optional_float(row["raw_score"]),
            support_labels=tuple(_split_labels(row["support_labels"])),
            concern_labels=tuple(_split_labels(row["concern_labels"])),
            quality_flags=tuple(_split_labels(row["quality_flags"])),
            ms2_present=_bool_value(row["ms2_present"]),
            nl_match=_bool_value(row["nl_match"]),
            nl_status=row["nl_status"],
            best_loss_ppm=_optional_float(row["best_loss_ppm"]),
            best_ms2_scan_rt_min=_optional_float(row["best_ms2_scan_rt_min"]),
            apex_ms2_delta_min=_optional_float(row["apex_ms2_delta_min"]),
            best_product_base_ratio=_optional_float(row["best_product_base_ratio"]),
            trigger_scan_count=_optional_int(row["trigger_scan_count"]),
            strict_nl_scan_count=_optional_int(row["strict_nl_scan_count"]),
            ms2_alignment_source=row["ms2_alignment_source"],
            diagnostic_product_absence_reason=row.get(
                "diagnostic_product_absence_reason",
                "",
            ),
            nearest_product_loss_ppm=_optional_float(
                row.get("nearest_product_loss_ppm", "")
            ),
            nearest_product_base_ratio=_optional_float(
                row.get("nearest_product_base_ratio", "")
            ),
            nearest_product_mz=_optional_float(row.get("nearest_product_mz", "")),
        )
        for row in rows
    )


def _read_target_mz(path: Path) -> dict[str, float]:
    if not path.exists():
        raise FileNotFoundError(str(path))
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook["Targets"]
        rows = sheet.iter_rows(values_only=True)
        header = next(rows)
        indexes = required_indexes(header, ("Label", "m/z"), "Targets")
        values: dict[str, float] = {}
        for row in rows:
            label = _text(row[indexes["Label"]])
            mz = _optional_float(_text(row[indexes["m/z"]]))
            if label and mz is not None:
                values[label] = mz
        return values
    finally:
        workbook.close()
