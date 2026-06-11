from __future__ import annotations

from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

from tools.diagnostics.diagnostic_io import read_tsv_required
from tools.diagnostics.targeted_gt_alignment_audit_models import (
    AuditConfig,
    TargetGroundTruth,
)
from tools.diagnostics.targeted_gt_alignment_audit_utils import (
    _to_float,
    _unescape_excel_formula,
)


def _load_target_ground_truth(config: AuditConfig) -> list[TargetGroundTruth]:
    rows = _target_workbook_rows(config.target_workbook)
    _propagate_sample_context(rows)
    rows_by_target_role = _rows_by_target_role_map(
        rows,
        (
            (config.target_label, "Analyte"),
            (config.istd_label, "ISTD"),
        ),
    )
    analyte_rows = rows_by_target_role[(config.target_label, "Analyte")]
    istd_rows = rows_by_target_role[(config.istd_label, "ISTD")]
    targets: list[TargetGroundTruth] = []
    for sample in sorted(analyte_rows):
        analyte = analyte_rows[sample]
        istd = istd_rows.get(sample)
        target_rt = _to_float(analyte.get("RT"))
        istd_rt = _to_float(istd.get("RT")) if istd else None
        rt_delta = (
            (target_rt - istd_rt) * 60.0
            if target_rt is not None and istd_rt is not None
            else None
        )
        targets.append(
            TargetGroundTruth(
                sample_stem=sample,
                group=str(analyte.get("Group") or ""),
                target_mz=config.target_mz,
                target_rt_min=target_rt,
                target_peak_start_min=_to_float(analyte.get("PeakStart")),
                target_peak_end_min=_to_float(analyte.get("PeakEnd")),
                target_peak_width_min=_to_float(analyte.get("PeakWidth")),
                target_area=_to_float(analyte.get("Area")),
                target_confidence=str(analyte.get("Confidence") or ""),
                target_nl_ok=str(analyte.get("NL") or ""),
                target_reason=str(analyte.get("Reason") or ""),
                istd_rt_min=istd_rt,
                istd_rt_delta_sec=rt_delta,
            ),
        )
    return targets


def _target_workbook_rows(path: Path) -> list[dict[str, object]]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    worksheet = workbook["XIC Results"]
    iterator = worksheet.iter_rows(values_only=True)
    headers = [str(value or "") for value in next(iterator)]
    rows: list[dict[str, object]] = []
    for row in iterator:
        record = {
            header: row[index] if index < len(row) else None
            for index, header in enumerate(headers)
        }
        if record.get("SampleName") is None and record.get("Target") is None:
            continue
        rows.append(record)
    return rows


def _propagate_sample_context(rows: list[dict[str, object]]) -> None:
    last_sample: object = None
    last_group: object = None
    for row in rows:
        if row.get("SampleName"):
            last_sample = row["SampleName"]
            last_group = row.get("Group")
        else:
            row["SampleName"] = last_sample
            row["Group"] = last_group


def _rows_by_target_role(
    rows: list[dict[str, object]],
    target: str,
    role: str,
) -> dict[str, dict[str, object]]:
    return _rows_by_target_role_map(rows, ((target, role),))[(target, role)]


def _rows_by_target_role_map(
    rows: list[dict[str, object]],
    target_roles: tuple[tuple[str, str], ...],
) -> dict[tuple[str, str], dict[str, dict[str, object]]]:
    selected: dict[tuple[str, str], dict[str, dict[str, object]]] = {
        target_role: {} for target_role in target_roles
    }
    for row in rows:
        target = row.get("Target")
        role = row.get("Role")
        if not isinstance(target, str) or not isinstance(role, str):
            continue
        bucket = selected.get((target, role))
        if bucket is None:
            continue
        sample = row.get("SampleName")
        if not isinstance(sample, str) or not sample:
            raise ValueError(f"Missing sample for {target}/{role}")
        bucket[sample] = row
    return selected


def _load_tsv(path: Path) -> list[dict[str, str]]:
    return [
        {key: _unescape_excel_formula(value) for key, value in row.items()}
        for row in read_tsv_required(path, ())
    ]


def _filter_review_by_mz(
    rows: list[dict[str, str]],
    config: AuditConfig,
) -> list[dict[str, str]]:
    low = config.target_mz * (1.0 - config.ppm * 1e-6)
    high = config.target_mz * (1.0 + config.ppm * 1e-6)
    return [
        row
        for row in rows
        if (mz := _to_float(row.get("family_center_mz"))) is not None
        and low <= mz <= high
    ]


def _cells_by_sample_in_review_range(
    cell_rows: list[dict[str, str]],
    review_index: dict[str, dict[str, str]],
) -> dict[str, tuple[dict[str, str], ...]]:
    grouped: dict[str, list[dict[str, str]]] = defaultdict(list)
    for cell in cell_rows:
        review_row = review_index.get(cell.get("feature_family_id", ""))
        if review_row is None:
            continue
        enriched_cell = dict(cell)
        if "include_in_primary_matrix" in review_row:
            enriched_cell["_review_include_in_primary_matrix"] = review_row.get(
                "include_in_primary_matrix",
                "",
            )
        if "accepted_cell_count" in review_row:
            enriched_cell["_review_accepted_cell_count"] = review_row.get(
                "accepted_cell_count",
                "",
            )
        if "identity_decision" in review_row:
            enriched_cell["_review_identity_decision"] = review_row.get(
                "identity_decision",
                "",
            )
        grouped[cell["sample_stem"]].append(enriched_cell)
    return {sample: tuple(rows) for sample, rows in grouped.items()}
