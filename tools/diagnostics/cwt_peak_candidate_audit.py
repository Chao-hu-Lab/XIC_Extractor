from __future__ import annotations

import argparse
import csv
import json
import sys
from collections.abc import Sequence
from dataclasses import asdict, dataclass
from pathlib import Path

from openpyxl import load_workbook

_CWT_SOURCE = "centwave_cwt"
_DEFAULT_NEAR_RT_WINDOW_MIN = 0.08

_REQUIRED_COLUMNS = (
    "sample_name",
    "target_label",
    "resolver_mode",
    "candidate_id",
    "proposal_sources",
    "rt_apex_min",
    "selected",
    "confidence",
    "raw_score",
    "reason",
)

_SUMMARY_COLUMNS = (
    "candidate_row_count",
    "candidate_group_count",
    "cwt_row_count",
    "cwt_only_row_count",
    "selected_cwt_agreed_group_count",
    "selected_cwt_nearby_group_count",
    "selected_cwt_far_alternative_group_count",
    "selected_without_cwt_group_count",
    "cwt_selected_support_group_count",
    "cwt_far_unconfirmed_group_count",
    "cwt_far_chemically_plausible_group_count",
)

_GROUP_COLUMNS = (
    "group_id",
    "sample_name",
    "target_label",
    "target_mz",
    "resolver_mode",
    "cwt_agreement_class",
    "cwt_conditioned_class",
    "candidate_count",
    "cwt_row_count",
    "cwt_only_row_count",
    "selected_candidate_id",
    "selected_rt_apex_min",
    "selected_proposal_sources",
    "selected_ms2_present",
    "selected_nl_match",
    "selected_ms2_trace_strength",
    "nearest_cwt_candidate_id",
    "nearest_cwt_rt_apex_min",
    "nearest_cwt_delta_min",
    "nearest_cwt_ms2_present",
    "nearest_cwt_nl_match",
    "nearest_cwt_ms2_trace_strength",
    "selected_confidence",
    "selected_raw_score",
    "selected_reason",
)

_CWT_ONLY_COLUMNS = (
    "group_id",
    "sample_name",
    "target_label",
    "target_mz",
    "resolver_mode",
    "candidate_id",
    "rt_apex_min",
    "confidence",
    "raw_score",
    "reason",
)


@dataclass(frozen=True)
class CwtCandidateRow:
    sample_name: str
    target_label: str
    resolver_mode: str
    candidate_id: str
    proposal_sources: str
    rt_apex_min: float
    selected: bool
    confidence: str
    raw_score: str
    reason: str
    ms2_present: str
    nl_match: str
    ms2_trace_strength: str

    @property
    def group_id(self) -> str:
        return "|".join((self.sample_name, self.target_label, self.resolver_mode))

    @property
    def source_set(self) -> frozenset[str]:
        return frozenset(
            source.strip()
            for source in self.proposal_sources.split(";")
            if source.strip()
        )

    @property
    def has_cwt(self) -> bool:
        return _CWT_SOURCE in self.source_set

    @property
    def cwt_only(self) -> bool:
        return self.source_set == frozenset({_CWT_SOURCE})


@dataclass(frozen=True)
class CwtGroupAuditRow:
    group_id: str
    sample_name: str
    target_label: str
    target_mz: float | None
    resolver_mode: str
    cwt_agreement_class: str
    cwt_conditioned_class: str
    candidate_count: int
    cwt_row_count: int
    cwt_only_row_count: int
    selected_candidate_id: str
    selected_rt_apex_min: float | None
    selected_proposal_sources: str
    selected_ms2_present: str
    selected_nl_match: str
    selected_ms2_trace_strength: str
    nearest_cwt_candidate_id: str
    nearest_cwt_rt_apex_min: float | None
    nearest_cwt_delta_min: float | None
    nearest_cwt_ms2_present: str
    nearest_cwt_nl_match: str
    nearest_cwt_ms2_trace_strength: str
    selected_confidence: str
    selected_raw_score: str
    selected_reason: str


@dataclass(frozen=True)
class CwtOnlyAuditRow:
    group_id: str
    sample_name: str
    target_label: str
    target_mz: float | None
    resolver_mode: str
    candidate_id: str
    rt_apex_min: float
    confidence: str
    raw_score: str
    reason: str


def main(argv: Sequence[str] | None = None) -> int:
    args = _parse_args(argv)
    try:
        rows = _read_peak_candidates(args.peak_candidates_tsv)
        target_mz_by_label = (
            _read_target_mz(args.targeted_workbook)
            if args.targeted_workbook is not None
            else {}
        )
        groups = _audit_groups(
            rows,
            target_mz_by_label=target_mz_by_label,
            near_rt_window_min=args.near_rt_window_min,
        )
        cwt_only_rows = _cwt_only_rows(rows, target_mz_by_label=target_mz_by_label)
        payload = {
            "summary": _summary(rows, groups, cwt_only_rows),
            "groups": [asdict(row) for row in groups],
            "cwt_only_rows": [asdict(row) for row in cwt_only_rows],
        }
        _write_outputs(args.output_dir, payload, groups, cwt_only_rows)
    except (OSError, ValueError) as exc:
        print(str(exc), file=sys.stderr)
        return 2
    print(f"CWT audit JSON: {args.output_dir / 'cwt_peak_candidate_audit.json'}")
    return 0


def _parse_args(argv: Sequence[str] | None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Audit CWT peak candidate agreement from peak_candidates.tsv.",
    )
    parser.add_argument("--peak-candidates-tsv", type=Path, required=True)
    parser.add_argument(
        "--targeted-workbook",
        type=Path,
        help="Optional XIC workbook with Targets sheet used to enrich target_mz.",
    )
    parser.add_argument("--output-dir", type=Path, required=True)
    parser.add_argument(
        "--near-rt-window-min",
        type=float,
        default=_DEFAULT_NEAR_RT_WINDOW_MIN,
        help=(
            "Classify selected candidates as selected_cwt_nearby when the "
            "nearest CWT proposal is within this RT window."
        ),
    )
    return parser.parse_args(argv)


def _read_peak_candidates(path: Path) -> tuple[CwtCandidateRow, ...]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle, delimiter="\t")
        fieldnames = tuple(reader.fieldnames or ())
        missing = [column for column in _REQUIRED_COLUMNS if column not in fieldnames]
        if missing:
            raise ValueError(f"{path}: missing required columns: {', '.join(missing)}")
        return tuple(
            _row_from_dict(path, index, row)
            for index, row in enumerate(reader, 2)
        )


def _read_target_mz(path: Path) -> dict[str, float]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if "Targets" not in workbook.sheetnames:
            raise ValueError(f"{path}: missing required sheet: Targets")
        rows = workbook["Targets"].iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            raise ValueError(f"{path}: Targets sheet is empty")
        indexes = _required_indexes(header, ("Label", "m/z"), "Targets")
        target_mz: dict[str, float] = {}
        for row_number, row in enumerate(rows, 2):
            label = _text(row[indexes["Label"]])
            if not label:
                continue
            target_mz[label] = _float_value(
                path,
                row_number,
                "m/z",
                _text(row[indexes["m/z"]]),
            )
        return target_mz
    finally:
        workbook.close()


def _required_indexes(
    header: object,
    required: tuple[str, ...],
    sheet_name: str,
) -> dict[str, int]:
    if not isinstance(header, tuple):
        raise ValueError(f"{sheet_name}: header row is invalid")
    indexes = {_text(value): index for index, value in enumerate(header)}
    missing = [column for column in required if column not in indexes]
    if missing:
        raise ValueError(
            f"{sheet_name}: missing required columns: {', '.join(missing)}"
        )
    return {column: indexes[column] for column in required}


def _row_from_dict(
    path: Path,
    row_number: int,
    row: dict[str, str],
) -> CwtCandidateRow:
    return CwtCandidateRow(
        sample_name=row["sample_name"],
        target_label=row["target_label"],
        resolver_mode=row["resolver_mode"],
        candidate_id=row["candidate_id"],
        proposal_sources=row["proposal_sources"],
        rt_apex_min=_float_value(path, row_number, "rt_apex_min", row["rt_apex_min"]),
        selected=row["selected"].strip().upper() == "TRUE",
        confidence=row["confidence"],
        raw_score=row["raw_score"],
        reason=row["reason"],
        ms2_present=row.get("ms2_present", ""),
        nl_match=row.get("nl_match", ""),
        ms2_trace_strength=row.get("ms2_trace_strength", ""),
    )


def _audit_groups(
    rows: tuple[CwtCandidateRow, ...],
    *,
    target_mz_by_label: dict[str, float],
    near_rt_window_min: float,
) -> tuple[CwtGroupAuditRow, ...]:
    grouped: dict[str, list[CwtCandidateRow]] = {}
    for row in rows:
        grouped.setdefault(row.group_id, []).append(row)
    return tuple(
        _audit_group(
            group_rows,
            target_mz_by_label=target_mz_by_label,
            near_rt_window_min=near_rt_window_min,
        )
        for _, group_rows in sorted(grouped.items(), key=lambda item: item[0])
    )


def _audit_group(
    rows: list[CwtCandidateRow],
    *,
    target_mz_by_label: dict[str, float],
    near_rt_window_min: float,
) -> CwtGroupAuditRow:
    first = rows[0]
    selected = next((row for row in rows if row.selected), None)
    cwt_rows = [row for row in rows if row.has_cwt]
    cwt_only_count = sum(row.cwt_only for row in rows)
    nearest = _nearest_cwt(selected, cwt_rows)
    nearest_delta = (
        abs(selected.rt_apex_min - nearest.rt_apex_min)
        if selected is not None and nearest is not None
        else None
    )
    agreement_class = _agreement_class(
        selected,
        cwt_rows,
        nearest_delta_min=nearest_delta,
        near_rt_window_min=near_rt_window_min,
    )
    return CwtGroupAuditRow(
        group_id=first.group_id,
        sample_name=first.sample_name,
        target_label=first.target_label,
        target_mz=target_mz_by_label.get(first.target_label),
        resolver_mode=first.resolver_mode,
        cwt_agreement_class=agreement_class,
        cwt_conditioned_class=_conditioned_class(agreement_class, nearest),
        candidate_count=len(rows),
        cwt_row_count=len(cwt_rows),
        cwt_only_row_count=cwt_only_count,
        selected_candidate_id=selected.candidate_id if selected else "",
        selected_rt_apex_min=selected.rt_apex_min if selected else None,
        selected_proposal_sources=selected.proposal_sources if selected else "",
        selected_ms2_present=selected.ms2_present if selected else "",
        selected_nl_match=selected.nl_match if selected else "",
        selected_ms2_trace_strength=selected.ms2_trace_strength if selected else "",
        nearest_cwt_candidate_id=nearest.candidate_id if nearest else "",
        nearest_cwt_rt_apex_min=nearest.rt_apex_min if nearest else None,
        nearest_cwt_delta_min=nearest_delta,
        nearest_cwt_ms2_present=nearest.ms2_present if nearest else "",
        nearest_cwt_nl_match=nearest.nl_match if nearest else "",
        nearest_cwt_ms2_trace_strength=nearest.ms2_trace_strength if nearest else "",
        selected_confidence=selected.confidence if selected else "",
        selected_raw_score=selected.raw_score if selected else "",
        selected_reason=selected.reason if selected else "",
    )


def _agreement_class(
    selected: CwtCandidateRow | None,
    cwt_rows: list[CwtCandidateRow],
    *,
    nearest_delta_min: float | None,
    near_rt_window_min: float,
) -> str:
    if selected is None:
        return "no_selected_candidate"
    if selected.has_cwt:
        return "selected_cwt_agreed"
    if cwt_rows:
        if nearest_delta_min is not None and nearest_delta_min <= near_rt_window_min:
            return "selected_cwt_nearby"
        return "selected_cwt_far_alternative"
    return "selected_without_cwt"


def _conditioned_class(
    agreement_class: str,
    nearest_cwt: CwtCandidateRow | None,
) -> str:
    if agreement_class in {"selected_cwt_agreed", "selected_cwt_nearby"}:
        return "cwt_selected_support"
    if agreement_class == "selected_cwt_far_alternative":
        if nearest_cwt is not None and _chemically_plausible(nearest_cwt):
            return "cwt_far_chemically_plausible"
        return "cwt_far_unconfirmed"
    if agreement_class == "selected_without_cwt":
        return "no_cwt_proposal"
    return agreement_class


def _chemically_plausible(row: CwtCandidateRow) -> bool:
    return (
        row.nl_match.strip().upper() == "TRUE"
        and row.ms2_trace_strength.strip().lower() in {"moderate", "strong"}
    )


def _nearest_cwt(
    selected: CwtCandidateRow | None,
    cwt_rows: list[CwtCandidateRow],
) -> CwtCandidateRow | None:
    if selected is None or not cwt_rows:
        return None
    return min(cwt_rows, key=lambda row: abs(selected.rt_apex_min - row.rt_apex_min))


def _cwt_only_rows(
    rows: tuple[CwtCandidateRow, ...],
    *,
    target_mz_by_label: dict[str, float],
) -> tuple[CwtOnlyAuditRow, ...]:
    return tuple(
        CwtOnlyAuditRow(
            group_id=row.group_id,
            sample_name=row.sample_name,
            target_label=row.target_label,
            target_mz=target_mz_by_label.get(row.target_label),
            resolver_mode=row.resolver_mode,
            candidate_id=row.candidate_id,
            rt_apex_min=row.rt_apex_min,
            confidence=row.confidence,
            raw_score=row.raw_score,
            reason=row.reason,
        )
        for row in rows
        if row.cwt_only
    )


def _summary(
    rows: tuple[CwtCandidateRow, ...],
    groups: tuple[CwtGroupAuditRow, ...],
    cwt_only_rows: tuple[CwtOnlyAuditRow, ...],
) -> dict[str, int]:
    return {
        "candidate_row_count": len(rows),
        "candidate_group_count": len(groups),
        "cwt_row_count": sum(row.has_cwt for row in rows),
        "cwt_only_row_count": len(cwt_only_rows),
        "selected_cwt_agreed_group_count": _group_class_count(
            groups, "selected_cwt_agreed"
        ),
        "selected_cwt_nearby_group_count": _group_class_count(
            groups, "selected_cwt_nearby"
        ),
        "selected_cwt_far_alternative_group_count": _group_class_count(
            groups, "selected_cwt_far_alternative"
        ),
        "selected_without_cwt_group_count": _group_class_count(
            groups, "selected_without_cwt"
        ),
        "cwt_selected_support_group_count": _conditioned_class_count(
            groups, "cwt_selected_support"
        ),
        "cwt_far_unconfirmed_group_count": _conditioned_class_count(
            groups, "cwt_far_unconfirmed"
        ),
        "cwt_far_chemically_plausible_group_count": _conditioned_class_count(
            groups, "cwt_far_chemically_plausible"
        ),
    }


def _group_class_count(
    groups: tuple[CwtGroupAuditRow, ...],
    cwt_agreement_class: str,
) -> int:
    return sum(row.cwt_agreement_class == cwt_agreement_class for row in groups)


def _conditioned_class_count(
    groups: tuple[CwtGroupAuditRow, ...],
    cwt_conditioned_class: str,
) -> int:
    return sum(row.cwt_conditioned_class == cwt_conditioned_class for row in groups)


def _write_outputs(
    output_dir: Path,
    payload: dict[str, object],
    groups: tuple[CwtGroupAuditRow, ...],
    cwt_only_rows: tuple[CwtOnlyAuditRow, ...],
) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    _write_summary(output_dir / "cwt_peak_candidate_audit_summary.tsv", payload)
    _write_groups(output_dir / "cwt_peak_candidate_groups.tsv", groups)
    _write_groups(
        output_dir / "cwt_peak_candidate_far_alternatives.tsv",
        tuple(
            row
            for row in groups
            if row.cwt_agreement_class == "selected_cwt_far_alternative"
        ),
    )
    _write_cwt_only(output_dir / "cwt_peak_candidate_cwt_only.tsv", cwt_only_rows)
    (output_dir / "cwt_peak_candidate_audit.json").write_text(
        json.dumps(payload, indent=2, sort_keys=True),
        encoding="utf-8",
    )
    (output_dir / "cwt_peak_candidate_audit.md").write_text(
        _markdown(payload),
        encoding="utf-8",
    )


def _write_summary(path: Path, payload: dict[str, object]) -> None:
    summary = payload["summary"]
    if not isinstance(summary, dict):
        raise TypeError("summary payload must be a dictionary")
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=_SUMMARY_COLUMNS, delimiter="\t")
        writer.writeheader()
        writer.writerow(summary)


def _write_groups(path: Path, rows: tuple[CwtGroupAuditRow, ...]) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=_GROUP_COLUMNS, delimiter="\t")
        writer.writeheader()
        writer.writerows(_format_group_row(row) for row in rows)


def _write_cwt_only(path: Path, rows: tuple[CwtOnlyAuditRow, ...]) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=_CWT_ONLY_COLUMNS, delimiter="\t")
        writer.writeheader()
        writer.writerows(_format_cwt_only_row(row) for row in rows)


def _format_group_row(row: CwtGroupAuditRow) -> dict[str, str]:
    return {
        "group_id": row.group_id,
        "sample_name": row.sample_name,
        "target_label": row.target_label,
        "target_mz": _format_optional_float(row.target_mz),
        "resolver_mode": row.resolver_mode,
        "cwt_agreement_class": row.cwt_agreement_class,
        "cwt_conditioned_class": row.cwt_conditioned_class,
        "candidate_count": str(row.candidate_count),
        "cwt_row_count": str(row.cwt_row_count),
        "cwt_only_row_count": str(row.cwt_only_row_count),
        "selected_candidate_id": row.selected_candidate_id,
        "selected_rt_apex_min": _format_optional_float(row.selected_rt_apex_min),
        "selected_proposal_sources": row.selected_proposal_sources,
        "selected_ms2_present": row.selected_ms2_present,
        "selected_nl_match": row.selected_nl_match,
        "selected_ms2_trace_strength": row.selected_ms2_trace_strength,
        "nearest_cwt_candidate_id": row.nearest_cwt_candidate_id,
        "nearest_cwt_rt_apex_min": _format_optional_float(row.nearest_cwt_rt_apex_min),
        "nearest_cwt_delta_min": _format_optional_float(row.nearest_cwt_delta_min),
        "nearest_cwt_ms2_present": row.nearest_cwt_ms2_present,
        "nearest_cwt_nl_match": row.nearest_cwt_nl_match,
        "nearest_cwt_ms2_trace_strength": row.nearest_cwt_ms2_trace_strength,
        "selected_confidence": row.selected_confidence,
        "selected_raw_score": row.selected_raw_score,
        "selected_reason": row.selected_reason,
    }


def _format_cwt_only_row(row: CwtOnlyAuditRow) -> dict[str, str]:
    return {
        "group_id": row.group_id,
        "sample_name": row.sample_name,
        "target_label": row.target_label,
        "target_mz": _format_optional_float(row.target_mz),
        "resolver_mode": row.resolver_mode,
        "candidate_id": row.candidate_id,
        "rt_apex_min": _format_optional_float(row.rt_apex_min),
        "confidence": row.confidence,
        "raw_score": row.raw_score,
        "reason": row.reason,
    }


def _markdown(payload: dict[str, object]) -> str:
    summary = payload["summary"]
    if not isinstance(summary, dict):
        raise TypeError("summary payload must be a dictionary")
    return "\n".join(
        [
            "# CWT Peak Candidate Audit",
            "",
            f"- candidate_row_count: {summary['candidate_row_count']}",
            f"- candidate_group_count: {summary['candidate_group_count']}",
            f"- cwt_row_count: {summary['cwt_row_count']}",
            f"- cwt_only_row_count: {summary['cwt_only_row_count']}",
            "- selected_cwt_agreed_group_count: "
            f"{summary['selected_cwt_agreed_group_count']}",
            "- selected_cwt_nearby_group_count: "
            f"{summary['selected_cwt_nearby_group_count']}",
            "- selected_cwt_far_alternative_group_count: "
            f"{summary['selected_cwt_far_alternative_group_count']}",
            "- selected_without_cwt_group_count: "
            f"{summary['selected_without_cwt_group_count']}",
            "- cwt_selected_support_group_count: "
            f"{summary['cwt_selected_support_group_count']}",
            "- cwt_far_unconfirmed_group_count: "
            f"{summary['cwt_far_unconfirmed_group_count']}",
            "- cwt_far_chemically_plausible_group_count: "
            f"{summary['cwt_far_chemically_plausible_group_count']}",
            "",
        ]
    )


def _float_value(path: Path, row_number: int, column: str, value: str) -> float:
    try:
        return float(value)
    except ValueError as exc:
        message = f"{path}: row {row_number} invalid {column}: {value!r}"
        raise ValueError(message) from exc


def _text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _format_optional_float(value: float | None) -> str:
    if value is None:
        return ""
    return f"{value:.5f}"


if __name__ == "__main__":
    raise SystemExit(main())
