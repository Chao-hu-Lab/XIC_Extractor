"""Root-cause audit for targeted NL dropout review-positive rows."""

from __future__ import annotations

import argparse
import csv
import json
import sys
from collections import Counter
from collections.abc import Mapping, Sequence
from dataclasses import asdict, dataclass
from pathlib import Path

from openpyxl import load_workbook

_RELIABILITY_COLUMNS = (
    "sample_name",
    "target_label",
    "role",
    "reliability_state",
    "risk_reasons",
)

_CANDIDATE_COLUMNS = (
    "sample_name",
    "target_label",
    "resolver_mode",
    "candidate_id",
    "proposal_sources",
    "rt_apex_min",
    "selected",
    "confidence",
    "raw_score",
    "support_labels",
    "concern_labels",
    "quality_flags",
    "ms2_present",
    "nl_match",
    "nl_status",
    "best_loss_ppm",
    "best_ms2_scan_rt_min",
    "apex_ms2_delta_min",
    "best_product_base_ratio",
    "trigger_scan_count",
    "strict_nl_scan_count",
    "ms2_alignment_source",
)

_OPTIONAL_CANDIDATE_COLUMNS = (
    "diagnostic_product_absence_reason",
    "nearest_product_loss_ppm",
    "nearest_product_base_ratio",
    "nearest_product_mz",
)

_SUMMARY_COLUMNS = (
    "rows_checked",
    "review_positive_count",
    "included_count",
    "missing_candidate_count",
    "bucket_counts",
    "target_counts",
    "product_absence_reason_counts",
)

_ROW_COLUMNS = (
    "sample_name",
    "target_label",
    "target_mz",
    "role",
    "reliability_state",
    "targeted_risk_reasons",
    "resolver_mode",
    "selected_candidate_id",
    "selected_rt_apex_min",
    "selected_raw_score",
    "selected_confidence",
    "proposal_sources",
    "support_labels",
    "concern_labels",
    "quality_flags",
    "ms2_present",
    "nl_match",
    "nl_status",
    "best_loss_ppm",
    "best_ms2_scan_rt_min",
    "apex_ms2_delta_min",
    "best_product_base_ratio",
    "trigger_scan_count",
    "strict_nl_scan_count",
    "ms2_alignment_source",
    "diagnostic_product_absence_reason",
    "nearest_product_loss_ppm",
    "nearest_product_base_ratio",
    "nearest_product_mz",
    "root_cause_bucket",
    "root_cause_reason",
)

_HARD_CONFLICT_LABELS = frozenset(
    {
        "hard_quality_flag",
        "hard_nl_conflict",
        "low_scan_support",
        "low_trace_continuity",
        "poor_edge_recovery",
        "rt_centrality_poor",
        "shape_poor",
        "edge_clipped",
    }
)


@dataclass(frozen=True)
class TargetedNLDropoutRootCauseOutputs:
    summary_tsv: Path
    rows_tsv: Path
    json_path: Path
    markdown_path: Path


@dataclass(frozen=True)
class ReliabilityRow:
    sample_name: str
    target_label: str
    role: str
    reliability_state: str
    risk_reasons: tuple[str, ...]


@dataclass(frozen=True)
class CandidateRow:
    sample_name: str
    target_label: str
    resolver_mode: str
    candidate_id: str
    proposal_sources: tuple[str, ...]
    rt_apex_min: float | None
    selected: bool
    confidence: str
    raw_score: float | None
    support_labels: tuple[str, ...]
    concern_labels: tuple[str, ...]
    quality_flags: tuple[str, ...]
    ms2_present: bool | None
    nl_match: bool | None
    nl_status: str
    best_loss_ppm: float | None
    best_ms2_scan_rt_min: float | None
    apex_ms2_delta_min: float | None
    best_product_base_ratio: float | None
    trigger_scan_count: int | None
    strict_nl_scan_count: int | None
    ms2_alignment_source: str
    diagnostic_product_absence_reason: str
    nearest_product_loss_ppm: float | None
    nearest_product_base_ratio: float | None
    nearest_product_mz: float | None


@dataclass(frozen=True)
class RootCauseRow:
    sample_name: str
    target_label: str
    target_mz: float | None
    role: str
    reliability_state: str
    targeted_risk_reasons: str
    resolver_mode: str
    selected_candidate_id: str
    selected_rt_apex_min: float | None
    selected_raw_score: float | None
    selected_confidence: str
    proposal_sources: str
    support_labels: str
    concern_labels: str
    quality_flags: str
    ms2_present: bool | None
    nl_match: bool | None
    nl_status: str
    best_loss_ppm: float | None
    best_ms2_scan_rt_min: float | None
    apex_ms2_delta_min: float | None
    best_product_base_ratio: float | None
    trigger_scan_count: int | None
    strict_nl_scan_count: int | None
    ms2_alignment_source: str
    diagnostic_product_absence_reason: str
    nearest_product_loss_ppm: float | None
    nearest_product_base_ratio: float | None
    nearest_product_mz: float | None
    root_cause_bucket: str
    root_cause_reason: str


@dataclass(frozen=True)
class RootCauseSummary:
    rows_checked: int
    review_positive_count: int
    included_count: int
    missing_candidate_count: int
    bucket_counts: str
    target_counts: str
    product_absence_reason_counts: str


@dataclass(frozen=True)
class TargetedNLDropoutRootCauseResult:
    summary: RootCauseSummary
    rows: tuple[RootCauseRow, ...]


def run_targeted_nl_dropout_root_cause_audit(
    *,
    targeted_reliability_rows_tsv: Path,
    peak_candidates_tsv: Path,
    output_dir: Path,
    targeted_workbook: Path | None = None,
    nl_ppm_max: float = 10.0,
    apex_ms2_delta_max_min: float = 0.08,
    nl_min_intensity_ratio: float = 0.01,
) -> tuple[TargetedNLDropoutRootCauseOutputs, TargetedNLDropoutRootCauseResult]:
    reliability_rows = _read_reliability_rows(targeted_reliability_rows_tsv)
    candidate_rows = _read_candidate_rows(peak_candidates_tsv)
    selected_by_key = _selected_candidates_by_key(candidate_rows)
    target_mz = _read_target_mz(targeted_workbook) if targeted_workbook else {}
    rows = _root_cause_rows(
        reliability_rows,
        selected_by_key,
        target_mz=target_mz,
        nl_ppm_max=nl_ppm_max,
        apex_ms2_delta_max_min=apex_ms2_delta_max_min,
        nl_min_intensity_ratio=nl_min_intensity_ratio,
    )
    result = TargetedNLDropoutRootCauseResult(
        summary=_summary(reliability_rows, rows),
        rows=rows,
    )
    output_dir.mkdir(parents=True, exist_ok=True)
    outputs = TargetedNLDropoutRootCauseOutputs(
        summary_tsv=output_dir / "targeted_nl_dropout_root_cause_summary.tsv",
        rows_tsv=output_dir / "targeted_nl_dropout_root_cause_rows.tsv",
        json_path=output_dir / "targeted_nl_dropout_root_cause.json",
        markdown_path=output_dir / "targeted_nl_dropout_root_cause.md",
    )
    _write_outputs(outputs, result)
    return outputs, result


def main(argv: Sequence[str] | None = None) -> int:
    args = _parse_args(argv)
    try:
        outputs, result = run_targeted_nl_dropout_root_cause_audit(
            targeted_reliability_rows_tsv=args.targeted_reliability_rows_tsv,
            peak_candidates_tsv=args.peak_candidates_tsv,
            output_dir=args.output_dir,
            targeted_workbook=args.targeted_workbook,
            nl_ppm_max=args.nl_ppm_max,
            apex_ms2_delta_max_min=args.apex_ms2_delta_max_min,
            nl_min_intensity_ratio=args.nl_min_intensity_ratio,
        )
    except (FileNotFoundError, KeyError, ValueError) as exc:
        print(str(exc), file=sys.stderr)
        return 2
    print(f"Summary TSV: {outputs.summary_tsv}")
    print(f"Rows TSV: {outputs.rows_tsv}")
    print(f"Root-cause JSON: {outputs.json_path}")
    print(f"Root-cause report: {outputs.markdown_path}")
    print(f"Review-positive rows: {result.summary.review_positive_count}")
    return 0


def _parse_args(argv: Sequence[str] | None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Classify targeted review-positive NL dropout root causes.",
    )
    parser.add_argument("--targeted-reliability-rows-tsv", type=Path, required=True)
    parser.add_argument("--peak-candidates-tsv", type=Path, required=True)
    parser.add_argument("--output-dir", type=Path, required=True)
    parser.add_argument(
        "--targeted-workbook",
        type=Path,
        help="Optional targeted workbook used to include target m/z in rows.",
    )
    parser.add_argument("--nl-ppm-max", type=float, default=10.0)
    parser.add_argument("--apex-ms2-delta-max-min", type=float, default=0.08)
    parser.add_argument("--nl-min-intensity-ratio", type=float, default=0.01)
    return parser.parse_args(argv)


def _read_reliability_rows(path: Path) -> tuple[ReliabilityRow, ...]:
    rows = _read_required_tsv(path, _RELIABILITY_COLUMNS)
    return tuple(
        ReliabilityRow(
            sample_name=row["sample_name"],
            target_label=row["target_label"],
            role=row["role"],
            reliability_state=row["reliability_state"],
            risk_reasons=tuple(_split_labels(row["risk_reasons"])),
        )
        for row in rows
    )


def _read_candidate_rows(path: Path) -> tuple[CandidateRow, ...]:
    rows = _read_required_tsv(path, _CANDIDATE_COLUMNS)
    return tuple(
        CandidateRow(
            sample_name=row["sample_name"],
            target_label=row["target_label"],
            resolver_mode=row["resolver_mode"],
            candidate_id=row["candidate_id"],
            proposal_sources=tuple(_split_labels(row["proposal_sources"])),
            rt_apex_min=_optional_float(row["rt_apex_min"]),
            selected=_bool_value(row["selected"]) is True,
            confidence=row["confidence"],
            raw_score=_optional_float(row["raw_score"]),
            support_labels=tuple(_split_labels(row["support_labels"])),
            concern_labels=tuple(_split_labels(row["concern_labels"])),
            quality_flags=tuple(_split_labels(row["quality_flags"])),
            ms2_present=_bool_value(row["ms2_present"]),
            nl_match=_bool_value(row["nl_match"]),
            nl_status=row["nl_status"],
            best_loss_ppm=_optional_float(row["best_loss_ppm"]),
            best_ms2_scan_rt_min=_optional_float(row["best_ms2_scan_rt_min"]),
            apex_ms2_delta_min=_optional_float(row["apex_ms2_delta_min"]),
            best_product_base_ratio=_optional_float(row["best_product_base_ratio"]),
            trigger_scan_count=_optional_int(row["trigger_scan_count"]),
            strict_nl_scan_count=_optional_int(row["strict_nl_scan_count"]),
            ms2_alignment_source=row["ms2_alignment_source"],
            diagnostic_product_absence_reason=row.get(
                "diagnostic_product_absence_reason",
                "",
            ),
            nearest_product_loss_ppm=_optional_float(
                row.get("nearest_product_loss_ppm", "")
            ),
            nearest_product_base_ratio=_optional_float(
                row.get("nearest_product_base_ratio", "")
            ),
            nearest_product_mz=_optional_float(row.get("nearest_product_mz", "")),
        )
        for row in rows
    )


def _read_required_tsv(
    path: Path,
    required: Sequence[str],
) -> tuple[dict[str, str], ...]:
    if not path.exists():
        raise FileNotFoundError(str(path))
    with path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle, delimiter="\t")
        fieldnames = tuple(reader.fieldnames or ())
        missing = [column for column in required if column not in fieldnames]
        if missing:
            raise ValueError(
                f"{path}: missing required columns: {', '.join(missing)}"
            )
        return tuple(dict(row) for row in reader)


def _read_target_mz(path: Path) -> dict[str, float]:
    if not path.exists():
        raise FileNotFoundError(str(path))
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook["Targets"]
        rows = sheet.iter_rows(values_only=True)
        header = next(rows)
        indexes = {
            str(value).strip(): index for index, value in enumerate(header) if value
        }
        for required in ("Label", "m/z"):
            if required not in indexes:
                raise ValueError(f"Targets is missing required column: {required}")
        values: dict[str, float] = {}
        for row in rows:
            label = _text(row[indexes["Label"]])
            mz = _optional_float(_text(row[indexes["m/z"]]))
            if label and mz is not None:
                values[label] = mz
        return values
    finally:
        workbook.close()


def _selected_candidates_by_key(
    candidate_rows: Sequence[CandidateRow],
) -> dict[tuple[str, str], tuple[CandidateRow, ...]]:
    selected_by_key: dict[tuple[str, str], list[CandidateRow]] = {}
    for row in candidate_rows:
        if not row.selected:
            continue
        selected_by_key.setdefault((row.sample_name, row.target_label), []).append(row)
    return {key: tuple(values) for key, values in selected_by_key.items()}


def _root_cause_rows(
    reliability_rows: Sequence[ReliabilityRow],
    selected_by_key: Mapping[tuple[str, str], tuple[CandidateRow, ...]],
    *,
    target_mz: Mapping[str, float],
    nl_ppm_max: float,
    apex_ms2_delta_max_min: float,
    nl_min_intensity_ratio: float,
) -> tuple[RootCauseRow, ...]:
    rows: list[RootCauseRow] = []
    for reliability in reliability_rows:
        if reliability.reliability_state != "targeted_review_positive":
            continue
        selected_rows = selected_by_key.get(
            (reliability.sample_name, reliability.target_label),
            (),
        )
        selected = selected_rows[0] if len(selected_rows) == 1 else None
        bucket, reason = _classify_root_cause(
            selected_rows,
            nl_ppm_max=nl_ppm_max,
            apex_ms2_delta_max_min=apex_ms2_delta_max_min,
            nl_min_intensity_ratio=nl_min_intensity_ratio,
        )
        rows.append(
            RootCauseRow(
                sample_name=reliability.sample_name,
                target_label=reliability.target_label,
                target_mz=target_mz.get(reliability.target_label),
                role=reliability.role,
                reliability_state=reliability.reliability_state,
                targeted_risk_reasons=";".join(reliability.risk_reasons),
                resolver_mode=selected.resolver_mode if selected else "",
                selected_candidate_id=selected.candidate_id if selected else "",
                selected_rt_apex_min=selected.rt_apex_min if selected else None,
                selected_raw_score=selected.raw_score if selected else None,
                selected_confidence=selected.confidence if selected else "",
                proposal_sources=(
                    ";".join(selected.proposal_sources) if selected else ""
                ),
                support_labels=";".join(selected.support_labels) if selected else "",
                concern_labels=";".join(selected.concern_labels) if selected else "",
                quality_flags=";".join(selected.quality_flags) if selected else "",
                ms2_present=selected.ms2_present if selected else None,
                nl_match=selected.nl_match if selected else None,
                nl_status=selected.nl_status if selected else "",
                best_loss_ppm=selected.best_loss_ppm if selected else None,
                best_ms2_scan_rt_min=(
                    selected.best_ms2_scan_rt_min if selected else None
                ),
                apex_ms2_delta_min=(
                    selected.apex_ms2_delta_min if selected else None
                ),
                best_product_base_ratio=(
                    selected.best_product_base_ratio if selected else None
                ),
                trigger_scan_count=selected.trigger_scan_count if selected else None,
                strict_nl_scan_count=(
                    selected.strict_nl_scan_count if selected else None
                ),
                ms2_alignment_source=(
                    selected.ms2_alignment_source if selected else ""
                ),
                diagnostic_product_absence_reason=(
                    selected.diagnostic_product_absence_reason if selected else ""
                ),
                nearest_product_loss_ppm=(
                    selected.nearest_product_loss_ppm if selected else None
                ),
                nearest_product_base_ratio=(
                    selected.nearest_product_base_ratio if selected else None
                ),
                nearest_product_mz=selected.nearest_product_mz if selected else None,
                root_cause_bucket=bucket,
                root_cause_reason=reason,
            )
        )
    return tuple(rows)


def _classify_root_cause(
    selected_rows: Sequence[CandidateRow],
    *,
    nl_ppm_max: float,
    apex_ms2_delta_max_min: float,
    nl_min_intensity_ratio: float,
) -> tuple[str, str]:
    if not selected_rows:
        return (
            "no_selected_candidate",
            "No selected peak candidate was found for this review-positive row.",
        )
    if len(selected_rows) > 1:
        return (
            "hard_candidate_conflict",
            "More than one selected candidate exists for this sample/target.",
        )
    selected = selected_rows[0]
    hard_labels = set(selected.concern_labels) & _HARD_CONFLICT_LABELS
    if hard_labels or selected.quality_flags:
        reason = (
            f"Hard conflict labels: {','.join(sorted(hard_labels))}"
            if hard_labels
            else "Candidate has quality flags."
        )
        return ("hard_candidate_conflict", reason)
    if selected.ms2_present is not True or selected.trigger_scan_count == 0:
        return (
            "no_ms2_trigger",
            "No usable MS2 trigger was recorded for the selected candidate.",
        )
    if selected.best_loss_ppm is None:
        detail = ""
        if selected.diagnostic_product_absence_reason:
            detail = (
                f" Subcause: {selected.diagnostic_product_absence_reason}."
            )
        nearest = ""
        if selected.nearest_product_loss_ppm is not None:
            nearest = (
                f" Nearest product loss ppm: "
                f"{selected.nearest_product_loss_ppm:.6g}."
            )
        ratio = ""
        if selected.nearest_product_base_ratio is not None:
            ratio = (
                f" Nearest product/base ratio: "
                f"{selected.nearest_product_base_ratio:.6g}."
            )
        return (
            "no_diagnostic_product",
            (
                "No diagnostic product/loss ppm was available from MS2 evidence."
                f"{detail}{nearest}{ratio}"
            ),
        )
    if (
        selected.apex_ms2_delta_min is not None
        and selected.apex_ms2_delta_min > apex_ms2_delta_max_min
    ):
        return (
            "off_apex_ms2",
            (
                f"MS2 scan is {selected.apex_ms2_delta_min:.6g} min from apex, "
                f"above {apex_ms2_delta_max_min:.6g} min."
            ),
        )
    if selected.best_loss_ppm > nl_ppm_max:
        return (
            "ppm_gate_fail",
            (
                f"Best loss ppm {selected.best_loss_ppm:.6g} exceeds "
                f"nl_ppm_max {nl_ppm_max:.6g}."
            ),
        )
    weak_product_threshold = 2 * nl_min_intensity_ratio
    if (
        selected.best_product_base_ratio is not None
        and selected.best_product_base_ratio < weak_product_threshold
    ):
        return (
            "weak_product_ratio",
            (
                f"Product/base ratio {selected.best_product_base_ratio:.6g} is "
                f"below 2 * nl_min_intensity_ratio ({weak_product_threshold:.6g})."
            ),
        )
    return (
        "coherent_ms1_nl_dropout",
        "Selected candidate has coherent MS1 evidence and near-threshold NL dropout.",
    )


def _summary(
    reliability_rows: Sequence[ReliabilityRow],
    rows: Sequence[RootCauseRow],
) -> RootCauseSummary:
    bucket_counts = Counter(row.root_cause_bucket for row in rows)
    target_counts = Counter(row.target_label for row in rows)
    product_absence_counts = Counter(
        row.diagnostic_product_absence_reason
        for row in rows
        if row.root_cause_bucket == "no_diagnostic_product"
        and row.diagnostic_product_absence_reason
    )
    return RootCauseSummary(
        rows_checked=len(reliability_rows),
        review_positive_count=sum(
            row.reliability_state == "targeted_review_positive"
            for row in reliability_rows
        ),
        included_count=len(rows),
        missing_candidate_count=bucket_counts["no_selected_candidate"],
        bucket_counts=_format_counter(bucket_counts),
        target_counts=_format_counter(target_counts),
        product_absence_reason_counts=_format_counter(product_absence_counts),
    )


def _write_outputs(
    outputs: TargetedNLDropoutRootCauseOutputs,
    result: TargetedNLDropoutRootCauseResult,
) -> None:
    _write_tsv(outputs.summary_tsv, _SUMMARY_COLUMNS, [asdict(result.summary)])
    _write_tsv(outputs.rows_tsv, _ROW_COLUMNS, _row_dicts(result.rows))
    outputs.json_path.write_text(
        json.dumps(
            {
                "summary": asdict(result.summary),
                "rows": _row_dicts(result.rows),
            },
            indent=2,
            sort_keys=True,
        ),
        encoding="utf-8",
    )
    outputs.markdown_path.write_text(_markdown(result), encoding="utf-8")


def _row_dicts(rows: Sequence[RootCauseRow]) -> list[dict[str, object]]:
    return [asdict(row) for row in rows]


def _write_tsv(
    path: Path,
    fieldnames: Sequence[str],
    rows: Sequence[Mapping[str, object]],
) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, delimiter="\t")
        writer.writeheader()
        for row in rows:
            writer.writerow(
                {key: _format_value(row.get(key, "")) for key in fieldnames}
            )


def _markdown(result: TargetedNLDropoutRootCauseResult) -> str:
    lines = [
        "# Targeted NL Dropout Root-cause Audit",
        "",
        "This diagnostic classifies targeted_review_positive rows only. It does "
        "not rescan RAW files, change selected peaks, or alter XIC Results.",
        "",
        "## Summary",
        "",
    ]
    summary = asdict(result.summary)
    lines.extend(f"- {key}: {summary[key]}" for key in _SUMMARY_COLUMNS)
    lines.extend(["", "## Rows", ""])
    for row in result.rows:
        mz = "" if row.target_mz is None else f", m/z {row.target_mz:.6g}"
        lines.append(
            f"- {row.root_cause_bucket}: {row.sample_name} / "
            f"{row.target_label}{mz} - {row.root_cause_reason}"
        )
    lines.append("")
    return "\n".join(lines)


def _split_labels(value: str) -> list[str]:
    return [part.strip() for part in value.split(";") if part.strip()]


def _optional_float(value: str) -> float | None:
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _optional_int(value: str) -> int | None:
    numeric = _optional_float(value)
    if numeric is None:
        return None
    return int(numeric)


def _bool_value(value: str) -> bool | None:
    normalized = value.strip().upper()
    if normalized in {"TRUE", "T", "YES", "Y", "1"}:
        return True
    if normalized in {"FALSE", "F", "NO", "N", "0"}:
        return False
    return None


def _format_counter(counter: Counter[str]) -> str:
    return ";".join(
        f"{key}:{counter[key]}"
        for key in sorted(counter)
    )


def _format_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, float):
        return f"{value:.6g}"
    return str(value)


def _text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


if __name__ == "__main__":
    raise SystemExit(main())
