"""Input loaders for targeted peak reliability audit."""

from __future__ import annotations

from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

from tools.diagnostics.diagnostic_io import (
    bool_value as _bool_value,
    optional_float as _float_value,
    optional_int as _int_value,
    read_tsv_required as _read_required_tsv,
    required_indexes as _required_indexes,
    split_semicolon_labels as _split_semicolon_labels,
    text_value as _text,
)
from tools.diagnostics.targeted_peak_reliability_models import (
    _PEAK_CANDIDATE_COLUMNS,
    _CandidateEvidence,
    _ScoreBreakdown,
    _TargetedInputRow,
    _WorksheetLike,
)


def _load_targeted_workbook(
    path: Path,
) -> tuple[tuple[_TargetedInputRow, ...], dict[tuple[str, str], _ScoreBreakdown]]:
    if not path.exists():
        raise FileNotFoundError(str(path))
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        rows = _read_xic_results(workbook["XIC Results"])
        score_by_key = (
            _read_score_breakdown(workbook["Score Breakdown"])
            if "Score Breakdown" in workbook.sheetnames
            else {}
        )
        return rows, score_by_key
    finally:
        workbook.close()


def _read_xic_results(sheet: _WorksheetLike) -> tuple[_TargetedInputRow, ...]:
    iterator = sheet.iter_rows(values_only=True)
    header = next(iterator)
    cols = _required_indexes(
        header,
        (
            "SampleName",
            "Target",
            "Role",
            "RT",
            "Area",
            "NL",
            "Confidence",
        ),
        "XIC Results",
    )
    header_indexes = {
        str(value).strip(): index for index, value in enumerate(header) if value
    }
    reason_index = header_indexes.get("Reason")
    current_sample = ""
    rows: list[_TargetedInputRow] = []
    for row in iterator:
        raw_sample = row[cols["SampleName"]]
        if raw_sample not in (None, ""):
            current_sample = _text(raw_sample)
        if not current_sample:
            continue
        target_label = _text(row[cols["Target"]])
        if not target_label:
            continue
        rows.append(
            _TargetedInputRow(
                sample_name=current_sample,
                target_label=target_label,
                role=_text(row[cols["Role"]]),
                rt=_float_value(row[cols["RT"]]),
                area=_float_value(row[cols["Area"]]),
                confidence=_text(row[cols["Confidence"]]).upper(),
                nl=_text(row[cols["NL"]]).upper(),
                reason=_text(row[reason_index]) if reason_index is not None else "",
            )
        )
    return tuple(rows)


def _read_score_breakdown(
    sheet: _WorksheetLike,
) -> dict[tuple[str, str], _ScoreBreakdown]:
    iterator = sheet.iter_rows(values_only=True)
    header = next(iterator)
    cols = _required_indexes(
        header,
        (
            "SampleName",
            "Target",
            "Prior RT",
            "Prior Source",
            "Total Severity",
            "Quality Flags",
        ),
        "Score Breakdown",
    )
    scores: dict[tuple[str, str], _ScoreBreakdown] = {}
    for row in iterator:
        sample = _text(row[cols["SampleName"]])
        target = _text(row[cols["Target"]])
        if not sample or not target:
            continue
        scores[(sample, target)] = _ScoreBreakdown(
            prior_rt=_float_value(row[cols["Prior RT"]]),
            prior_source=_text(row[cols["Prior Source"]]),
            total_severity=_int_value(row[cols["Total Severity"]]),
            quality_flags=_text(row[cols["Quality Flags"]]),
        )
    return scores


def _load_selected_candidate_evidence(
    path: Path | None,
) -> dict[tuple[str, str], _CandidateEvidence]:
    if path is None:
        return {}
    rows = _read_required_tsv(path, _PEAK_CANDIDATE_COLUMNS)
    selected_by_key: dict[tuple[str, str], list[_CandidateEvidence]] = defaultdict(list)
    for row in rows:
        if _bool_value(row["selected"]) is not True:
            continue
        key = (row["sample_name"], row["target_label"])
        selected_by_key[key].append(
            _CandidateEvidence(
                support_labels=tuple(_split_semicolon_labels(row["support_labels"])),
                concern_labels=tuple(_split_semicolon_labels(row["concern_labels"])),
                proposal_sources=tuple(
                    _split_semicolon_labels(row["proposal_sources"])
                ),
                quality_flags=tuple(_split_semicolon_labels(row["quality_flags"])),
                ms2_present=_bool_value(row["ms2_present"]),
                nl_match=_bool_value(row["nl_match"]),
                raw_score=_float_value(row["raw_score"]),
                diagnostic_product_absence_reason=(
                    row.get(
                        "diagnostic_product_absence_reason",
                        "",
                    )
                    or ""
                ).strip(),
            )
        )
    return {
        key: values[0] for key, values in selected_by_key.items() if len(values) == 1
    }
