"""Strict targeted ISTD benchmark gate for untargeted alignment outputs."""

from __future__ import annotations

import argparse
import csv
import json
import math
import re
import sys
from collections.abc import Mapping, Sequence
from dataclasses import asdict, dataclass
from pathlib import Path
from statistics import median

from openpyxl import load_workbook

ACTIVE_NEUTRAL_LOSS_DA = 116.0474
ACTIVE_NEUTRAL_LOSS_TOLERANCE_DA = 0.01
ISOTOPE_SHIFT_DA = 1.003355
TARGET_MATCH_RT_SEC = 60.0
MEAN_RT_DELTA_MAX_MIN = 0.15
SAMPLE_RT_MEDIAN_ABS_DELTA_MAX_MIN = 0.15
SAMPLE_RT_P95_ABS_DELTA_MAX_MIN = 0.30
LOG_AREA_SPEARMAN_MIN = 0.90
LOG_AREA_PEARSON_MIN = 0.80

SUMMARY_COLUMNS = (
    "target_label",
    "role",
    "active_tag",
    "neutral_loss_da",
    "target_mz",
    "target_rt_min",
    "target_rt_max",
    "targeted_positive_count",
    "targeted_total_count",
    "targeted_mean_rt",
    "candidate_match_count",
    "primary_match_count",
    "primary_feature_ids",
    "selected_feature_id",
    "untargeted_positive_count",
    "coverage_minimum",
    "paired_area_n",
    "log_area_pearson",
    "log_area_spearman",
    "family_mean_rt_delta_min",
    "sample_rt_pair_n",
    "sample_rt_median_abs_delta_min",
    "sample_rt_p95_abs_delta_min",
    "status",
    "failure_modes",
    "note",
)

MATCH_COLUMNS = (
    "target_label",
    "feature_family_id",
    "include_in_primary_matrix",
    "family_center_mz",
    "family_center_rt",
    "family_product_mz",
    "family_observed_neutral_loss_da",
    "mz_delta_ppm",
    "rt_delta_sec",
    "product_delta_ppm",
    "loss_delta_da",
    "mass_shift_da",
    "match_type",
    "distance_score",
)


@dataclass(frozen=True)
class BenchmarkThresholds:
    active_neutral_loss_da: float = ACTIVE_NEUTRAL_LOSS_DA
    additional_active_neutral_loss_das: tuple[float, ...] = ()
    active_neutral_loss_tolerance_da: float = ACTIVE_NEUTRAL_LOSS_TOLERANCE_DA
    default_match_ppm: float = 20.0
    match_rt_sec: float = TARGET_MATCH_RT_SEC
    mean_rt_delta_max_min: float = MEAN_RT_DELTA_MAX_MIN
    sample_rt_median_abs_delta_max_min: float = (
        SAMPLE_RT_MEDIAN_ABS_DELTA_MAX_MIN
    )
    sample_rt_p95_abs_delta_max_min: float = SAMPLE_RT_P95_ABS_DELTA_MAX_MIN
    log_area_spearman_min: float = LOG_AREA_SPEARMAN_MIN
    log_area_pearson_min: float = LOG_AREA_PEARSON_MIN


@dataclass(frozen=True)
class TargetDefinition:
    label: str
    role: str
    mz: float
    rt_min: float
    rt_max: float
    ppm_tol: float
    neutral_loss_da: float
    product_mz: float


@dataclass(frozen=True)
class TargetedPoint:
    sample_stem: str
    target_label: str
    role: str
    rt: float | None
    area: float | None
    nl: str
    confidence: str
    reason: str

    @property
    def positive(self) -> bool:
        return (
            self.area is not None
            and self.area > 0
            and self.rt is not None
        )


@dataclass(frozen=True)
class AlignmentFeature:
    feature_family_id: str
    neutral_loss_tag: str
    family_center_mz: float
    family_center_rt: float
    family_product_mz: float
    family_observed_neutral_loss_da: float
    include_in_primary_matrix: bool


@dataclass(frozen=True)
class AlignmentCell:
    feature_family_id: str
    sample_stem: str
    status: str
    area: float | None
    apex_rt: float | None


@dataclass(frozen=True)
class AlignmentMatrixData:
    areas_by_family: dict[str, dict[str, float]]
    sample_stems: frozenset[str]


@dataclass(frozen=True)
class CandidateMatch:
    target_label: str
    feature_family_id: str
    include_in_primary_matrix: bool
    family_center_mz: float
    family_center_rt: float
    family_product_mz: float
    family_observed_neutral_loss_da: float
    mz_delta_ppm: float
    rt_delta_sec: float
    product_delta_ppm: float
    loss_delta_da: float
    mass_shift_da: float
    match_type: str
    distance_score: float


@dataclass(frozen=True)
class BenchmarkSummary:
    target_label: str
    role: str
    active_tag: bool
    neutral_loss_da: float
    target_mz: float
    target_rt_min: float
    target_rt_max: float
    targeted_positive_count: int
    targeted_total_count: int
    targeted_mean_rt: float | None
    candidate_match_count: int
    primary_match_count: int
    primary_feature_ids: tuple[str, ...]
    selected_feature_id: str
    untargeted_positive_count: int
    coverage_minimum: int
    paired_area_n: int
    log_area_pearson: float | None
    log_area_spearman: float | None
    family_mean_rt_delta_min: float | None
    sample_rt_pair_n: int
    sample_rt_median_abs_delta_min: float | None
    sample_rt_p95_abs_delta_min: float | None
    status: str
    failure_modes: tuple[str, ...]
    note: str


@dataclass(frozen=True)
class BenchmarkOutputs:
    summary_tsv: Path
    matches_tsv: Path
    json_path: Path
    markdown_path: Path


def run_targeted_istd_benchmark(
    *,
    targeted_workbook: Path,
    alignment_dir: Path,
    output_dir: Path,
    thresholds: BenchmarkThresholds = BenchmarkThresholds(),
) -> tuple[BenchmarkOutputs, tuple[BenchmarkSummary, ...]]:
    targets = _read_target_definitions(targeted_workbook)
    targeted_points = _read_targeted_points(targeted_workbook)
    review_rows = _read_alignment_review(alignment_dir / "alignment_review.tsv")
    matrix = _read_alignment_matrix(alignment_dir / "alignment_matrix.tsv")
    cells = _read_alignment_cells(alignment_dir / "alignment_cells.tsv")

    summaries: list[BenchmarkSummary] = []
    matches: list[CandidateMatch] = []
    for target in targets:
        target_points = tuple(
            point
            for point in targeted_points.get(target.label, ())
            if point.sample_stem in matrix.sample_stems
        )
        target_matches = _match_target_to_alignment(
            target,
            review_rows,
            target_points=target_points,
            thresholds=thresholds,
        )
        matches.extend(target_matches)
        summaries.append(
            _summarize_target(
                target,
                target_points,
                target_matches,
                matrix=matrix.areas_by_family,
                cells=cells,
                thresholds=thresholds,
            ),
        )

    output_dir.mkdir(parents=True, exist_ok=True)
    outputs = BenchmarkOutputs(
        summary_tsv=output_dir / "targeted_istd_benchmark_summary.tsv",
        matches_tsv=output_dir / "targeted_istd_benchmark_matches.tsv",
        json_path=output_dir / "targeted_istd_benchmark.json",
        markdown_path=output_dir / "targeted_istd_benchmark.md",
    )
    _write_tsv(outputs.summary_tsv, SUMMARY_COLUMNS, _summary_rows(summaries))
    _write_tsv(outputs.matches_tsv, MATCH_COLUMNS, _match_rows(matches))
    _write_json(outputs.json_path, _json_payload(summaries, thresholds))
    _write_markdown(outputs.markdown_path, summaries)
    return outputs, tuple(summaries)


def main(argv: Sequence[str] | None = None) -> int:
    args = _parse_args(argv)
    thresholds = BenchmarkThresholds(
        active_neutral_loss_da=args.active_neutral_loss_da,
        additional_active_neutral_loss_das=tuple(
            args.additional_active_neutral_loss_da
        ),
        active_neutral_loss_tolerance_da=args.active_neutral_loss_tolerance_da,
        default_match_ppm=args.default_match_ppm,
        match_rt_sec=args.match_rt_sec,
        mean_rt_delta_max_min=args.mean_rt_delta_max_min,
        sample_rt_median_abs_delta_max_min=(
            args.sample_rt_median_abs_delta_max_min
        ),
        sample_rt_p95_abs_delta_max_min=args.sample_rt_p95_abs_delta_max_min,
        log_area_spearman_min=args.log_area_spearman_min,
        log_area_pearson_min=args.log_area_pearson_min,
    )
    try:
        outputs, summaries = run_targeted_istd_benchmark(
            targeted_workbook=args.targeted_workbook.resolve(),
            alignment_dir=args.alignment_dir.resolve(),
            output_dir=args.output_dir.resolve(),
            thresholds=thresholds,
        )
    except (FileNotFoundError, KeyError, ValueError) as exc:
        print(str(exc), file=sys.stderr)
        return 2
    print(f"Summary TSV: {outputs.summary_tsv}")
    print(f"Matches TSV: {outputs.matches_tsv}")
    print(f"Benchmark JSON: {outputs.json_path}")
    print(f"Benchmark report: {outputs.markdown_path}")
    return 1 if any(row.status == "FAIL" for row in summaries) else 0


def _parse_args(argv: Sequence[str] | None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Run strict targeted ISTD benchmark for untargeted alignment.",
    )
    parser.add_argument("--targeted-workbook", type=Path, required=True)
    parser.add_argument("--alignment-dir", type=Path, required=True)
    parser.add_argument("--output-dir", type=Path, required=True)
    parser.add_argument("--active-neutral-loss-da", type=float, default=116.0474)
    parser.add_argument(
        "--additional-active-neutral-loss-da",
        action="append",
        default=[],
        type=float,
        help=(
            "Additional selected neutral-loss masses treated as active in a "
            "multi-tag benchmark."
        ),
    )
    parser.add_argument(
        "--active-neutral-loss-tolerance-da",
        type=float,
        default=0.01,
    )
    parser.add_argument("--default-match-ppm", type=float, default=20.0)
    parser.add_argument("--match-rt-sec", type=float, default=60.0)
    parser.add_argument("--mean-rt-delta-max-min", type=float, default=0.15)
    parser.add_argument(
        "--sample-rt-median-abs-delta-max-min",
        type=float,
        default=0.15,
    )
    parser.add_argument(
        "--sample-rt-p95-abs-delta-max-min",
        type=float,
        default=0.30,
    )
    parser.add_argument("--log-area-spearman-min", type=float, default=0.90)
    parser.add_argument("--log-area-pearson-min", type=float, default=0.80)
    return parser.parse_args(argv)


def _read_target_definitions(path: Path) -> tuple[TargetDefinition, ...]:
    if not path.exists():
        raise FileNotFoundError(str(path))
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook["Targets"]
        rows = sheet.iter_rows(values_only=True)
        header = next(rows)
        cols = _required_indexes(
            header,
            (
                "Label",
                "Role",
                "m/z",
                "RT min",
                "RT max",
                "ppm tol",
                "NL (Da)",
                "Expected product m/z",
            ),
            "Targets",
        )
        targets: list[TargetDefinition] = []
        for row in rows:
            role = _text(row[cols["Role"]])
            if role != "ISTD":
                continue
            label = _text(row[cols["Label"]])
            if not label:
                continue
            targets.append(
                TargetDefinition(
                    label=label,
                    role=role,
                    mz=_required_float(row[cols["m/z"]], "m/z", label),
                    rt_min=_required_float(row[cols["RT min"]], "RT min", label),
                    rt_max=_required_float(row[cols["RT max"]], "RT max", label),
                    ppm_tol=_required_float(row[cols["ppm tol"]], "ppm tol", label),
                    neutral_loss_da=_required_float(
                        row[cols["NL (Da)"]],
                        "NL (Da)",
                        label,
                    ),
                    product_mz=_required_float(
                        row[cols["Expected product m/z"]],
                        "Expected product m/z",
                        label,
                    ),
                )
            )
        return tuple(targets)
    finally:
        workbook.close()


def _read_targeted_points(path: Path) -> dict[str, tuple[TargetedPoint, ...]]:
    if not path.exists():
        raise FileNotFoundError(str(path))
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook["XIC Results"]
        rows = sheet.iter_rows(values_only=True)
        header = next(rows)
        cols = _required_indexes(
            header,
            (
                "SampleName",
                "Target",
                "Role",
                "RT",
                "Area",
                "NL",
                "Confidence",
                "Reason",
            ),
            "XIC Results",
        )
        current_sample = ""
        grouped: dict[str, list[TargetedPoint]] = {}
        for row in rows:
            raw_sample = row[cols["SampleName"]]
            if raw_sample not in (None, ""):
                current_sample = _normalize_sample_id(_text(raw_sample))
            if not current_sample:
                continue
            label = _text(row[cols["Target"]])
            role = _text(row[cols["Role"]])
            if not label or role != "ISTD":
                continue
            grouped.setdefault(label, []).append(
                TargetedPoint(
                    sample_stem=current_sample,
                    target_label=label,
                    role=role,
                    rt=_float_value(row[cols["RT"]]),
                    area=_float_value(row[cols["Area"]]),
                    nl=_text(row[cols["NL"]]),
                    confidence=_text(row[cols["Confidence"]]),
                    reason=_text(row[cols["Reason"]]),
                )
            )
        return {label: tuple(points) for label, points in grouped.items()}
    finally:
        workbook.close()


def _read_alignment_review(path: Path) -> tuple[AlignmentFeature, ...]:
    rows = _read_required_tsv(path)
    _require_fields(
        rows,
        (
            "feature_family_id",
            "neutral_loss_tag",
            "family_center_mz",
            "family_center_rt",
            "family_product_mz",
            "family_observed_neutral_loss_da",
            "include_in_primary_matrix",
        ),
        path,
    )
    return tuple(
        AlignmentFeature(
            feature_family_id=row["feature_family_id"],
            neutral_loss_tag=row["neutral_loss_tag"],
            family_center_mz=_required_float(
                row.get("family_center_mz"),
                "family_center_mz",
                row["feature_family_id"],
            ),
            family_center_rt=_required_float(
                row.get("family_center_rt"),
                "family_center_rt",
                row["feature_family_id"],
            ),
            family_product_mz=_required_float(
                row.get("family_product_mz"),
                "family_product_mz",
                row["feature_family_id"],
            ),
            family_observed_neutral_loss_da=_required_float(
                row.get("family_observed_neutral_loss_da"),
                "family_observed_neutral_loss_da",
                row["feature_family_id"],
            ),
            include_in_primary_matrix=_is_primary_review_row(row),
        )
        for row in rows
    )


def _read_alignment_matrix(path: Path) -> AlignmentMatrixData:
    rows = _read_required_tsv(path)
    _require_fields(rows, ("feature_family_id",), path)
    metadata_columns = {
        "feature_family_id",
        "neutral_loss_tag",
        "family_center_mz",
        "family_center_rt",
        "family_product_mz",
        "family_observed_neutral_loss_da",
    }
    fieldnames = set(rows[0])
    sample_columns = sorted(fieldnames - metadata_columns)
    matrix: dict[str, dict[str, float]] = {}
    normalized_samples = frozenset(
        _normalize_sample_id(sample) for sample in sample_columns
    )
    for row in rows:
        family_id = row["feature_family_id"]
        values: dict[str, float] = {}
        for sample in sample_columns:
            area = _float_value(row.get(sample))
            if area is not None and area > 0:
                values[_normalize_sample_id(sample)] = area
        matrix[family_id] = values
    return AlignmentMatrixData(
        areas_by_family=matrix,
        sample_stems=normalized_samples,
    )


def _is_primary_review_row(row: Mapping[str, str]) -> bool:
    if not _is_trueish(row.get("include_in_primary_matrix")):
        return False
    identity_decision = (row.get("identity_decision") or "").strip()
    if identity_decision and identity_decision != "production_family":
        return False
    return True


def _read_alignment_cells(path: Path) -> dict[tuple[str, str], AlignmentCell]:
    rows = _read_required_tsv(path)
    _require_fields(
        rows,
        ("feature_family_id", "sample_stem", "status", "area", "apex_rt"),
        path,
    )
    cells: dict[tuple[str, str], AlignmentCell] = {}
    for row in rows:
        sample = _normalize_sample_id(row["sample_stem"])
        cell = AlignmentCell(
            feature_family_id=row["feature_family_id"],
            sample_stem=sample,
            status=row.get("status", ""),
            area=_float_value(row.get("area")),
            apex_rt=_float_value(row.get("apex_rt")),
        )
        cells[(cell.feature_family_id, sample)] = cell
    return cells


def _match_target_to_alignment(
    target: TargetDefinition,
    features: tuple[AlignmentFeature, ...],
    *,
    target_points: tuple[TargetedPoint, ...],
    thresholds: BenchmarkThresholds,
) -> tuple[CandidateMatch, ...]:
    target_mean_rt = _mean(
        point.rt for point in target_points if point.positive
    )
    exact_matches = _match_target_to_alignment_with_shift(
        target,
        features,
        target_mean_rt=target_mean_rt,
        thresholds=thresholds,
        mass_shift_da=0.0,
    )
    if (
        any(match.include_in_primary_matrix for match in exact_matches)
        or not _is_active_tag(target, thresholds)
    ):
        return tuple(sorted(exact_matches, key=_match_sort_key))
    isotope_matches = tuple(
        match
        for mass_shift_da in (ISOTOPE_SHIFT_DA, -ISOTOPE_SHIFT_DA)
        for match in _match_target_to_alignment_with_shift(
            target,
            features,
            target_mean_rt=target_mean_rt,
            thresholds=thresholds,
            mass_shift_da=mass_shift_da,
        )
    )
    isotope_matches = _best_isotope_shift_matches(isotope_matches)
    return tuple(sorted((*exact_matches, *isotope_matches), key=_match_sort_key))


def _match_target_to_alignment_with_shift(
    target: TargetDefinition,
    features: tuple[AlignmentFeature, ...],
    *,
    target_mean_rt: float | None,
    thresholds: BenchmarkThresholds,
    mass_shift_da: float,
) -> tuple[CandidateMatch, ...]:
    shifted_mz = target.mz + mass_shift_da
    shifted_product_mz = target.product_mz + mass_shift_da
    match_type = "exact" if mass_shift_da == 0.0 else "isotope_shift"
    matches: list[CandidateMatch] = []
    for feature in features:
        mz_delta_ppm = _ppm_delta(shifted_mz, feature.family_center_mz)
        if abs(mz_delta_ppm) > target.ppm_tol:
            continue
        product_delta_ppm = _ppm_delta(
            shifted_product_mz,
            feature.family_product_mz,
        )
        if abs(product_delta_ppm) > target.ppm_tol:
            continue
        loss_delta_da = (
            feature.family_observed_neutral_loss_da - target.neutral_loss_da
        )
        if abs(loss_delta_da) > thresholds.active_neutral_loss_tolerance_da:
            continue
        rt_delta_sec = _target_rt_delta_sec(
            feature.family_center_rt,
            target,
            target_mean_rt,
        )
        if abs(rt_delta_sec) > thresholds.match_rt_sec:
            continue
        matches.append(
            CandidateMatch(
                target_label=target.label,
                feature_family_id=feature.feature_family_id,
                include_in_primary_matrix=feature.include_in_primary_matrix,
                family_center_mz=feature.family_center_mz,
                family_center_rt=feature.family_center_rt,
                family_product_mz=feature.family_product_mz,
                family_observed_neutral_loss_da=(
                    feature.family_observed_neutral_loss_da
                ),
                mz_delta_ppm=mz_delta_ppm,
                rt_delta_sec=rt_delta_sec,
                product_delta_ppm=product_delta_ppm,
                loss_delta_da=loss_delta_da,
                mass_shift_da=mass_shift_da,
                match_type=match_type,
                distance_score=max(
                    abs(mz_delta_ppm) / target.ppm_tol,
                    abs(product_delta_ppm) / target.ppm_tol,
                    abs(rt_delta_sec) / thresholds.match_rt_sec,
                ),
            )
        )
    return tuple(sorted(matches, key=_match_sort_key))


def _best_isotope_shift_matches(
    matches: tuple[CandidateMatch, ...],
) -> tuple[CandidateMatch, ...]:
    primary_matches = tuple(
        match for match in matches if match.include_in_primary_matrix
    )
    if not primary_matches:
        return matches
    best_shift = min(
        {match.mass_shift_da for match in primary_matches},
        key=lambda shift: (
            min(
                match.distance_score
                for match in primary_matches
                if match.mass_shift_da == shift
            ),
            abs(shift),
        ),
    )
    return tuple(match for match in matches if match.mass_shift_da == best_shift)


def _summarize_target(
    target: TargetDefinition,
    points: tuple[TargetedPoint, ...],
    matches: tuple[CandidateMatch, ...],
    *,
    matrix: Mapping[str, Mapping[str, float]],
    cells: Mapping[tuple[str, str], AlignmentCell],
    thresholds: BenchmarkThresholds,
) -> BenchmarkSummary:
    positives = tuple(point for point in points if point.positive)
    targeted_mean_rt = _mean(point.rt for point in positives)
    primary_matches = tuple(
        match for match in matches if match.include_in_primary_matrix
    )
    selected = primary_matches[0] if len(primary_matches) == 1 else None
    selected_family = selected.feature_family_id if selected else ""
    selected_matrix = matrix.get(selected_family, {})
    active_tag = _is_active_tag(target, thresholds)
    paired_logs: list[tuple[float, float]] = []
    rt_deltas: list[float] = []
    for point in positives:
        area = selected_matrix.get(point.sample_stem)
        if area is not None and area > 0 and point.area is not None:
            paired_logs.append((math.log10(point.area), math.log10(area)))
            cell = cells.get((selected_family, point.sample_stem))
            if cell is not None and cell.apex_rt is not None and point.rt is not None:
                rt_deltas.append(cell.apex_rt - point.rt)

    coverage_minimum = max(0, len(positives) - max(1, math.ceil(len(positives) * 0.02)))
    family_mean_rt_delta = (
        selected.family_center_rt - targeted_mean_rt
        if selected is not None and targeted_mean_rt is not None
        else None
    )
    pearson = _pearson(paired_logs)
    spearman = _spearman(paired_logs)
    median_rt_delta = _median_abs(rt_deltas)
    p95_rt_delta = _percentile_abs(rt_deltas, 0.95)
    failure_modes = _failure_modes(
        active_tag=active_tag,
        primary_matches=primary_matches,
        untargeted_positive_count=len(selected_matrix),
        coverage_minimum=coverage_minimum,
        family_mean_rt_delta=family_mean_rt_delta,
        sample_rt_median_abs_delta=median_rt_delta,
        sample_rt_p95_abs_delta=p95_rt_delta,
        paired_area_n=len(paired_logs),
        pearson=pearson,
        spearman=spearman,
        thresholds=thresholds,
    )
    return BenchmarkSummary(
        target_label=target.label,
        role=target.role,
        active_tag=active_tag,
        neutral_loss_da=target.neutral_loss_da,
        target_mz=target.mz,
        target_rt_min=target.rt_min,
        target_rt_max=target.rt_max,
        targeted_positive_count=len(positives),
        targeted_total_count=len(points),
        targeted_mean_rt=targeted_mean_rt,
        candidate_match_count=len(matches),
        primary_match_count=len(primary_matches),
        primary_feature_ids=tuple(match.feature_family_id for match in primary_matches),
        selected_feature_id=selected_family,
        untargeted_positive_count=len(selected_matrix),
        coverage_minimum=coverage_minimum,
        paired_area_n=len(paired_logs),
        log_area_pearson=pearson,
        log_area_spearman=spearman,
        family_mean_rt_delta_min=family_mean_rt_delta,
        sample_rt_pair_n=len(rt_deltas),
        sample_rt_median_abs_delta_min=median_rt_delta,
        sample_rt_p95_abs_delta_min=p95_rt_delta,
        status="FAIL" if failure_modes else "PASS",
        failure_modes=failure_modes,
        note=_note(active_tag, failure_modes),
    )


def _failure_modes(
    *,
    active_tag: bool,
    primary_matches: tuple[CandidateMatch, ...],
    untargeted_positive_count: int,
    coverage_minimum: int,
    family_mean_rt_delta: float | None,
    sample_rt_median_abs_delta: float | None,
    sample_rt_p95_abs_delta: float | None,
    paired_area_n: int,
    pearson: float | None,
    spearman: float | None,
    thresholds: BenchmarkThresholds,
) -> tuple[str, ...]:
    if not active_tag:
        return ("FALSE_POSITIVE_TAG",) if primary_matches else ()
    failures: list[str] = []
    if len(primary_matches) == 0:
        failures.append("MISS")
    elif len(primary_matches) > 1:
        failures.append("SPLIT")
    if len(primary_matches) != 1:
        return tuple(failures)
    if untargeted_positive_count < coverage_minimum:
        failures.append("COVERAGE")
    if family_mean_rt_delta is None or abs(family_mean_rt_delta) > (
        thresholds.mean_rt_delta_max_min
    ):
        failures.append("DRIFT")
    if sample_rt_median_abs_delta is None or sample_rt_median_abs_delta > (
        thresholds.sample_rt_median_abs_delta_max_min
    ):
        failures.append("DRIFT")
    if sample_rt_p95_abs_delta is None or sample_rt_p95_abs_delta > (
        thresholds.sample_rt_p95_abs_delta_max_min
    ):
        failures.append("DRIFT")
    if paired_area_n < 3:
        failures.append("AREA_INSUFFICIENT")
    elif (
        pearson is None
        or pearson < thresholds.log_area_pearson_min
        or spearman is None
        or spearman < thresholds.log_area_spearman_min
    ):
        failures.append("AREA_MISMATCH")
    return tuple(dict.fromkeys(failures))


def _target_rt_delta_sec(
    rt: float,
    target: TargetDefinition,
    targeted_mean_rt: float | None,
) -> float:
    if target.rt_min <= rt <= target.rt_max:
        return 0.0
    if targeted_mean_rt is not None:
        return (rt - targeted_mean_rt) * 60.0
    if rt < target.rt_min:
        return (rt - target.rt_min) * 60.0
    return (rt - target.rt_max) * 60.0


def _summary_rows(summaries: Sequence[BenchmarkSummary]) -> list[dict[str, object]]:
    return [
        {
            **asdict(summary),
            "active_tag": _bool_text(summary.active_tag),
            "primary_feature_ids": ";".join(summary.primary_feature_ids),
            "failure_modes": ";".join(summary.failure_modes),
        }
        for summary in summaries
    ]


def _match_rows(matches: Sequence[CandidateMatch]) -> list[dict[str, object]]:
    return [asdict(match) for match in matches]


def _json_payload(
    summaries: Sequence[BenchmarkSummary],
    thresholds: BenchmarkThresholds,
) -> dict[str, object]:
    fail_count = sum(summary.status == "FAIL" for summary in summaries)
    active_fail_count = sum(
        summary.status == "FAIL" and summary.active_tag
        for summary in summaries
    )
    false_positive_tag_count = sum(
        "FALSE_POSITIVE_TAG" in summary.failure_modes
        for summary in summaries
    )
    return {
        "overall_status": "FAIL" if fail_count else "PASS",
        "fail_count": fail_count,
        "active_fail_count": active_fail_count,
        "false_positive_tag_count": false_positive_tag_count,
        "thresholds": asdict(thresholds),
        "summaries": _summary_rows(summaries),
    }


def _write_markdown(path: Path, summaries: Sequence[BenchmarkSummary]) -> None:
    fail_count = sum(summary.status == "FAIL" for summary in summaries)
    lines = [
        "# Targeted ISTD Benchmark",
        "",
        f"Overall status: {'FAIL' if fail_count else 'PASS'}",
        "",
        "| Target | Active | Primary hits | Selected | Status | Failure modes |",
        "|---|---:|---:|---|---|---|",
    ]
    for summary in summaries:
        lines.append(
            "| "
            f"{summary.target_label} | "
            f"{_bool_text(summary.active_tag)} | "
            f"{summary.primary_match_count} | "
            f"{summary.selected_feature_id} | "
            f"{summary.status} | "
            f"{';'.join(summary.failure_modes)} |"
        )
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def _write_json(path: Path, payload: Mapping[str, object]) -> None:
    path.write_text(json.dumps(payload, indent=2, sort_keys=True), encoding="utf-8")


def _write_tsv(
    path: Path,
    fieldnames: Sequence[str],
    rows: Sequence[Mapping[str, object]],
) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, delimiter="\t")
        writer.writeheader()
        for row in rows:
            writer.writerow(
                {key: _format_value(row.get(key, "")) for key in fieldnames}
            )


def _read_required_tsv(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        raise FileNotFoundError(str(path))
    with path.open(newline="", encoding="utf-8-sig") as handle:
        return list(csv.DictReader(handle, delimiter="\t"))


def _require_fields(
    rows: list[dict[str, str]],
    required: Sequence[str],
    path: Path,
) -> None:
    if not rows:
        raise ValueError(f"{path} has no data rows")
    fieldnames = set(rows[0])
    missing = [field for field in required if field not in fieldnames]
    if missing:
        raise ValueError(f"{path} is missing required columns: {missing}")


def _required_indexes(
    header: Sequence[object],
    required: Sequence[str],
    sheet_name: str,
) -> dict[str, int]:
    indexes = {str(value).strip(): index for index, value in enumerate(header) if value}
    missing = [field for field in required if field not in indexes]
    if missing:
        raise ValueError(f"{sheet_name} sheet missing required columns: {missing}")
    return indexes


def _required_float(value: object, field: str, label: str) -> float:
    parsed = _float_value(value)
    if parsed is None:
        raise ValueError(f"{label} has invalid {field}: {value!r}")
    return parsed


def _float_value(value: object) -> float | None:
    if value in (None, "") or isinstance(value, bool):
        return None
    if isinstance(value, int | float):
        if math.isfinite(value):
            return float(value)
        return None
    try:
        parsed = float(str(value).strip())
    except ValueError:
        return None
    return parsed if math.isfinite(parsed) else None


def _text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _format_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return _bool_text(value)
    if isinstance(value, float):
        if not math.isfinite(value):
            return ""
        return f"{value:.6g}"
    if isinstance(value, tuple):
        return ";".join(str(part) for part in value)
    return str(value)


def _normalize_sample_id(sample_id: str) -> str:
    value = sample_id.strip()
    return re.sub(
        r"(^|_)QC_(\d+)$",
        lambda match: f"{match.group(1)}QC{match.group(2)}",
        value,
    )


def _is_trueish(value: str | None) -> bool:
    if value is None:
        return False
    return value.strip().lower() in {"1", "true", "t", "yes", "y"}


def _ppm_delta(reference: float, observed: float) -> float:
    denominator = max(abs(reference), 1e-12)
    return (observed - reference) / denominator * 1_000_000.0


def _match_sort_key(match: CandidateMatch) -> tuple[object, ...]:
    return (
        0 if match.include_in_primary_matrix else 1,
        match.distance_score,
        abs(match.rt_delta_sec),
        abs(match.mz_delta_ppm),
        match.feature_family_id,
    )


def _mean(values: Sequence[float | None] | object) -> float | None:
    finite = [
        float(value)
        for value in values
        if isinstance(value, (int, float))
        and not isinstance(value, bool)
        and math.isfinite(value)
    ]
    if not finite:
        return None
    return sum(finite) / len(finite)


def _median_abs(values: Sequence[float]) -> float | None:
    finite = [abs(value) for value in values if math.isfinite(value)]
    if not finite:
        return None
    return float(median(finite))


def _percentile_abs(values: Sequence[float], quantile: float) -> float | None:
    finite = sorted(abs(value) for value in values if math.isfinite(value))
    if not finite:
        return None
    index = math.ceil(len(finite) * quantile) - 1
    return float(finite[min(max(index, 0), len(finite) - 1)])


def _pearson(pairs: Sequence[tuple[float, float]]) -> float | None:
    if len(pairs) < 3:
        return None
    xs = [pair[0] for pair in pairs]
    ys = [pair[1] for pair in pairs]
    x_mean = sum(xs) / len(xs)
    y_mean = sum(ys) / len(ys)
    numerator = sum((x - x_mean) * (y - y_mean) for x, y in pairs)
    x_denominator = math.sqrt(sum((x - x_mean) ** 2 for x in xs))
    y_denominator = math.sqrt(sum((y - y_mean) ** 2 for y in ys))
    denominator = x_denominator * y_denominator
    if denominator == 0:
        return None
    return numerator / denominator


def _spearman(pairs: Sequence[tuple[float, float]]) -> float | None:
    if len(pairs) < 3:
        return None
    x_ranks = _ranks([pair[0] for pair in pairs])
    y_ranks = _ranks([pair[1] for pair in pairs])
    return _pearson(tuple(zip(x_ranks, y_ranks, strict=True)))


def _ranks(values: Sequence[float]) -> list[float]:
    ranked = sorted(enumerate(values), key=lambda item: item[1])
    ranks = [0.0] * len(values)
    index = 0
    while index < len(ranked):
        end = index + 1
        while end < len(ranked) and ranked[end][1] == ranked[index][1]:
            end += 1
        rank = (index + 1 + end) / 2.0
        for original_index, _value in ranked[index:end]:
            ranks[original_index] = rank
        index = end
    return ranks


def _is_active_tag(
    target: TargetDefinition,
    thresholds: BenchmarkThresholds,
) -> bool:
    active_masses = (
        thresholds.active_neutral_loss_da,
        *thresholds.additional_active_neutral_loss_das,
    )
    return any(
        abs(target.neutral_loss_da - active_mass)
        <= thresholds.active_neutral_loss_tolerance_da
        for active_mass in active_masses
    )


def _note(active_tag: bool, failure_modes: tuple[str, ...]) -> str:
    if not active_tag and not failure_modes:
        return "inactive tag excluded"
    if not failure_modes:
        return "strict gate passed"
    return "strict gate failed"


def _bool_text(value: bool) -> str:
    return "TRUE" if value else "FALSE"


if __name__ == "__main__":
    raise SystemExit(main())
