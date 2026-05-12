"""Audit untargeted alignment against a targeted workbook ground truth."""

from __future__ import annotations

import argparse
import csv
import html
import math
from collections import Counter, defaultdict
from collections.abc import Sequence
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

PRODUCTION_STATUSES = {"detected", "rescued"}
PASS_MODE = "PASS"
SPLIT_MODE = "SPLIT"
DRIFT_MODE = "DRIFT"
DUPLICATE_MODE = "DUPLICATE"
MISS_MODE = "MISS"


@dataclass(frozen=True)
class AuditConfig:
    target_workbook: Path
    alignment_run: Path
    target_label: str
    istd_label: str
    target_mz: float
    ppm: float
    pass_rt_sec: float
    drift_rt_sec: float
    output_dir: Path


def main(argv: Sequence[str] | None = None) -> int:
    config = _parse_args(argv)
    target_rows = _load_target_ground_truth(config)
    review_rows = _load_tsv(config.alignment_run / "alignment_review.tsv")
    cell_rows = _load_tsv(config.alignment_run / "alignment_cells.tsv")

    review_in_mz = _filter_review_by_mz(review_rows, config)
    review_index = {
        row["feature_family_id"]: row for row in review_in_mz
    }
    cells_by_sample = _cells_by_sample_in_review_range(cell_rows, review_index)
    comparison = [
        _classify_sample(target, cells_by_sample.get(target.sample_stem, ()), config)
        for target in target_rows
    ]

    config.output_dir.mkdir(parents=True, exist_ok=True)
    _write_dict_csv(config.output_dir / "gt_target.csv", target_rows)
    _write_dict_csv(config.output_dir / "comparison.csv", comparison)
    _write_report(
        config.output_dir / "failure_mode_report.md",
        comparison,
        review_in_mz,
        config,
    )
    _write_svg(
        config.output_dir / "failure_mode_chart.svg",
        comparison,
        config,
    )
    print(f"Wrote {config.output_dir / 'gt_target.csv'}")
    print(f"Wrote {config.output_dir / 'comparison.csv'}")
    print(f"Wrote {config.output_dir / 'failure_mode_report.md'}")
    print(f"Wrote {config.output_dir / 'failure_mode_chart.svg'}")
    return 0


def _parse_args(argv: Sequence[str] | None) -> AuditConfig:
    parser = argparse.ArgumentParser(
        description="Audit untargeted alignment against targeted workbook GT.",
    )
    parser.add_argument("--target-workbook", type=Path, required=True)
    parser.add_argument("--alignment-run", type=Path, required=True)
    parser.add_argument("--target-label", required=True)
    parser.add_argument("--istd-label", required=True)
    parser.add_argument("--target-mz", type=float, required=True)
    parser.add_argument("--ppm", type=float, required=True)
    parser.add_argument("--pass-rt-sec", type=float, required=True)
    parser.add_argument("--drift-rt-sec", type=float, required=True)
    parser.add_argument("--output-dir", type=Path, required=True)
    args = parser.parse_args(argv)
    return AuditConfig(
        target_workbook=args.target_workbook,
        alignment_run=args.alignment_run,
        target_label=args.target_label,
        istd_label=args.istd_label,
        target_mz=args.target_mz,
        ppm=args.ppm,
        pass_rt_sec=args.pass_rt_sec,
        drift_rt_sec=args.drift_rt_sec,
        output_dir=args.output_dir,
    )


@dataclass(frozen=True)
class TargetGroundTruth:
    sample_stem: str
    group: str
    target_mz: float
    target_rt_min: float | None
    target_peak_start_min: float | None
    target_peak_end_min: float | None
    target_peak_width_min: float | None
    target_area: float | None
    target_confidence: str
    target_nl_ok: str
    target_reason: str
    istd_rt_min: float | None
    istd_rt_delta_sec: float | None


def _load_target_ground_truth(config: AuditConfig) -> list[TargetGroundTruth]:
    rows = _target_workbook_rows(config.target_workbook)
    _propagate_sample_context(rows)
    analyte_rows = _rows_by_target_role(rows, config.target_label, "Analyte")
    istd_rows = _rows_by_target_role(rows, config.istd_label, "ISTD")
    targets: list[TargetGroundTruth] = []
    for sample in sorted(analyte_rows):
        analyte = analyte_rows[sample]
        istd = istd_rows.get(sample)
        target_rt = _to_float(analyte.get("RT"))
        istd_rt = _to_float(istd.get("RT")) if istd else None
        rt_delta = (
            (target_rt - istd_rt) * 60.0
            if target_rt is not None and istd_rt is not None
            else None
        )
        targets.append(
            TargetGroundTruth(
                sample_stem=sample,
                group=str(analyte.get("Group") or ""),
                target_mz=config.target_mz,
                target_rt_min=target_rt,
                target_peak_start_min=_to_float(analyte.get("PeakStart")),
                target_peak_end_min=_to_float(analyte.get("PeakEnd")),
                target_peak_width_min=_to_float(analyte.get("PeakWidth")),
                target_area=_to_float(analyte.get("Area")),
                target_confidence=str(analyte.get("Confidence") or ""),
                target_nl_ok=str(analyte.get("NL") or ""),
                target_reason=str(analyte.get("Reason") or ""),
                istd_rt_min=istd_rt,
                istd_rt_delta_sec=rt_delta,
            ),
        )
    return targets


def _target_workbook_rows(path: Path) -> list[dict[str, object]]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    worksheet = workbook["XIC Results"]
    iterator = worksheet.iter_rows(values_only=True)
    headers = [str(value or "") for value in next(iterator)]
    rows: list[dict[str, object]] = []
    for row in iterator:
        record = {
            header: row[index] if index < len(row) else None
            for index, header in enumerate(headers)
        }
        if record.get("SampleName") is None and record.get("Target") is None:
            continue
        rows.append(record)
    return rows


def _propagate_sample_context(rows: list[dict[str, object]]) -> None:
    last_sample: object = None
    last_group: object = None
    for row in rows:
        if row.get("SampleName"):
            last_sample = row["SampleName"]
            last_group = row.get("Group")
        else:
            row["SampleName"] = last_sample
            row["Group"] = last_group


def _rows_by_target_role(
    rows: list[dict[str, object]],
    target: str,
    role: str,
) -> dict[str, dict[str, object]]:
    selected: dict[str, dict[str, object]] = {}
    for row in rows:
        if row.get("Target") != target or row.get("Role") != role:
            continue
        sample = row.get("SampleName")
        if not isinstance(sample, str) or not sample:
            raise ValueError(f"Missing sample for {target}/{role}")
        selected[sample] = row
    return selected


def _load_tsv(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8", newline="") as handle:
        return [
            {
                key: _unescape_excel_formula(value)
                for key, value in row.items()
            }
            for row in csv.DictReader(handle, delimiter="\t")
        ]


def _filter_review_by_mz(
    rows: list[dict[str, str]],
    config: AuditConfig,
) -> list[dict[str, str]]:
    low = config.target_mz * (1.0 - config.ppm * 1e-6)
    high = config.target_mz * (1.0 + config.ppm * 1e-6)
    return [
        row
        for row in rows
        if (mz := _to_float(row.get("family_center_mz"))) is not None
        and low <= mz <= high
    ]


def _cells_by_sample_in_review_range(
    cell_rows: list[dict[str, str]],
    review_index: dict[str, dict[str, str]],
) -> dict[str, tuple[dict[str, str], ...]]:
    grouped: dict[str, list[dict[str, str]]] = defaultdict(list)
    for cell in cell_rows:
        if cell.get("feature_family_id") not in review_index:
            continue
        grouped[cell["sample_stem"]].append(cell)
    return {sample: tuple(rows) for sample, rows in grouped.items()}


def _classify_sample(
    target: TargetGroundTruth,
    cells: tuple[dict[str, str], ...],
    config: AuditConfig,
) -> dict[str, object]:
    production = [
        cell for cell in cells if _status(cell) in PRODUCTION_STATUSES
    ]
    duplicates = [
        cell for cell in cells if _status(cell) == "duplicate_assigned"
    ]
    production_in_window = _production_cells_in_gt_window(target, production)
    closest = _closest_cell(target, production or duplicates or list(cells))
    closest_delta = _rt_delta_sec(target, closest)
    mode = _failure_mode(
        closest=closest,
        closest_delta_sec=closest_delta,
        production_in_window=production_in_window,
        production=production,
        duplicates=duplicates,
        config=config,
    )
    closest_rt = _cell_rt(closest) if closest is not None else None
    return {
        "sample_stem": target.sample_stem,
        "group": target.group,
        "gt_target_rt_min": _format_float(target.target_rt_min, 4),
        "gt_target_confidence": target.target_confidence,
        "gt_peak_start_min": _format_float(target.target_peak_start_min, 4),
        "gt_peak_end_min": _format_float(target.target_peak_end_min, 4),
        "family_count_total": len({cell["feature_family_id"] for cell in cells}),
        "family_ids_all": _join_ids(cell["feature_family_id"] for cell in cells),
        "production_family_ids": _join_ids(
            cell["feature_family_id"] for cell in production
        ),
        "duplicate_family_ids": _join_ids(
            cell["feature_family_id"] for cell in duplicates
        ),
        "production_family_count_in_gt_window": len(
            {cell["feature_family_id"] for cell in production_in_window}
        ),
        "production_family_ids_in_gt_window": _join_ids(
            cell["feature_family_id"] for cell in production_in_window
        ),
        "closest_family_id": closest.get("feature_family_id", "")
        if closest is not None
        else "",
        "closest_family_mz": closest.get("family_center_mz", "")
        if closest is not None
        else "",
        "closest_status": _status(closest) if closest is not None else "",
        "closest_apex_rt_min": _format_float(closest_rt, 4),
        "closest_rt_delta_sec": _format_float(closest_delta, 2),
        "failure_mode": mode,
    }


def _production_cells_in_gt_window(
    target: TargetGroundTruth,
    cells: list[dict[str, str]],
) -> list[dict[str, str]]:
    start = target.target_peak_start_min
    end = target.target_peak_end_min
    if start is None or end is None:
        return []
    return [
        cell for cell in cells
        if (rt := _cell_rt(cell)) is not None and start <= rt <= end
    ]


def _closest_cell(
    target: TargetGroundTruth,
    cells: list[dict[str, str]],
) -> dict[str, str] | None:
    closest: dict[str, str] | None = None
    closest_delta: float | None = None
    for cell in cells:
        delta = _rt_delta_sec(target, cell)
        if delta is None:
            continue
        if closest_delta is None or abs(delta) < abs(closest_delta):
            closest = cell
            closest_delta = delta
    return closest


def _failure_mode(
    *,
    closest: dict[str, str] | None,
    closest_delta_sec: float | None,
    production_in_window: list[dict[str, str]],
    production: list[dict[str, str]],
    duplicates: list[dict[str, str]],
    config: AuditConfig,
) -> str:
    if closest is None or closest_delta_sec is None:
        return MISS_MODE
    if production:
        if abs(closest_delta_sec) <= config.pass_rt_sec:
            if len({cell["feature_family_id"] for cell in production_in_window}) > 1:
                return SPLIT_MODE
            return PASS_MODE
        if abs(closest_delta_sec) <= config.drift_rt_sec:
            return DRIFT_MODE
        return MISS_MODE
    if duplicates and abs(closest_delta_sec) <= config.drift_rt_sec:
        return DUPLICATE_MODE
    return MISS_MODE


def _write_report(
    path: Path,
    comparison: list[dict[str, object]],
    review_rows: list[dict[str, str]],
    config: AuditConfig,
) -> None:
    counts = Counter(str(row["failure_mode"]) for row in comparison)
    total = len(comparison)
    lines = [
        f"# Targeted GT Alignment Audit: {config.target_label}",
        "",
        f"Target m/z: {config.target_mz:.6g}",
        f"Alignment run: `{config.alignment_run}`",
        "",
        "## Summary",
        "",
        "| Mode | Count | Percent |",
        "|---|---:|---:|",
    ]
    for mode in (PASS_MODE, SPLIT_MODE, DRIFT_MODE, DUPLICATE_MODE, MISS_MODE):
        count = counts.get(mode, 0)
        percent = count / total * 100.0 if total else 0.0
        lines.append(f"| {mode} | {count} | {percent:.1f}% |")
    lines.extend(
        [
            "",
            "## Families In Target m/z Range",
            "",
            "| Family | m/z | RT | Anchor | Warning |",
            "|---|---:|---:|---|---|",
        ],
    )
    for row in review_rows:
        lines.append(
            "| "
            f"{row.get('feature_family_id', '')} | "
            f"{row.get('family_center_mz', '')} | "
            f"{row.get('family_center_rt', '')} | "
            f"{row.get('has_anchor', '')} | "
            f"{row.get('warning', '')} |"
        )
    lines.extend(["", "## Per-Sample Detail", ""])
    for row in comparison:
        lines.extend(
            [
                f"### {row['sample_stem']} - `{row['failure_mode']}`",
                "",
                f"- GT RT: {row['gt_target_rt_min']} min",
                f"- Production families: {row['production_family_ids'] or '(none)'}",
                f"- Duplicate families: {row['duplicate_family_ids'] or '(none)'}",
                (
                    "- Closest: "
                    f"{row['closest_family_id'] or '(none)'} "
                    f"{row['closest_status']} "
                    f"delta={row['closest_rt_delta_sec']} sec"
                ),
                "",
            ],
        )
    path.write_text("\n".join(lines), encoding="utf-8")


def _write_svg(
    path: Path,
    comparison: list[dict[str, object]],
    config: AuditConfig,
) -> None:
    row_height = 34
    width = 900
    height = max(160, 80 + len(comparison) * row_height)
    colors = {
        PASS_MODE: "#2ca02c",
        SPLIT_MODE: "#ff7f0e",
        DRIFT_MODE: "#1f77b4",
        DUPLICATE_MODE: "#9467bd",
        MISS_MODE: "#d62728",
    }
    items = [
        _svg_text(20, 30, f"Targeted GT Audit: {config.target_label}", 18),
        _svg_text(20, 55, f"Alignment run: {config.alignment_run}", 11),
    ]
    for index, row in enumerate(comparison):
        y = 90 + index * row_height
        mode = str(row["failure_mode"])
        color = colors.get(mode, "#777777")
        items.append(
            f'<rect x="20" y="{y - 18}" width="18" height="18" fill="{color}" />',
        )
        label = (
            f"{row['sample_stem']}  {mode}  "
            f"closest={row['closest_family_id'] or 'NA'} "
            f"delta={row['closest_rt_delta_sec'] or 'NA'} sec"
        )
        items.append(_svg_text(50, y - 4, label, 12))
    svg = (
        f'<svg xmlns="http://www.w3.org/2000/svg" width="{width}" '
        f'height="{height}" viewBox="0 0 {width} {height}">\n'
        '<rect width="100%" height="100%" fill="white" />\n'
        + "\n".join(items)
        + "\n</svg>\n"
    )
    path.write_text(svg, encoding="utf-8")


def _write_dict_csv(path: Path, rows: list[object]) -> None:
    serialized = [_as_output_dict(row) for row in rows]
    if not serialized:
        path.write_text("", encoding="utf-8")
        return
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(serialized[0].keys()))
        writer.writeheader()
        for row in serialized:
            writer.writerow(
                {key: _escape_excel_formula(value) for key, value in row.items()},
            )


def _as_output_dict(row: object) -> dict[str, object]:
    if isinstance(row, TargetGroundTruth):
        return {
            "sample_stem": row.sample_stem,
            "group": row.group,
            "target_mz": _format_float(row.target_mz, 6),
            "target_rt_min": _format_float(row.target_rt_min, 4),
            "target_peak_start_min": _format_float(row.target_peak_start_min, 4),
            "target_peak_end_min": _format_float(row.target_peak_end_min, 4),
            "target_peak_width_min": _format_float(row.target_peak_width_min, 4),
            "target_area": _format_float(row.target_area, 2),
            "target_confidence": row.target_confidence,
            "target_nl_ok": row.target_nl_ok,
            "target_reason": row.target_reason,
            "istd_rt_min": _format_float(row.istd_rt_min, 4),
            "istd_rt_delta_sec": _format_float(row.istd_rt_delta_sec, 2),
        }
    return dict(row)  # type: ignore[arg-type]


def _status(cell: dict[str, str] | None) -> str:
    return "" if cell is None else (cell.get("status") or "").strip().lower()


def _cell_rt(cell: dict[str, str] | None) -> float | None:
    if cell is None:
        return None
    return _to_float(cell.get("apex_rt")) or _to_float(cell.get("family_center_rt"))


def _rt_delta_sec(
    target: TargetGroundTruth,
    cell: dict[str, str] | None,
) -> float | None:
    if target.target_rt_min is None:
        return None
    rt = _cell_rt(cell)
    if rt is None:
        return None
    return (rt - target.target_rt_min) * 60.0


def _join_ids(values: object) -> str:
    return ";".join(sorted({str(value) for value in values if value}))


def _to_float(value: object) -> float | None:
    if value is None or value == "":
        return None
    try:
        parsed = float(value)
    except (TypeError, ValueError):
        return None
    return parsed if math.isfinite(parsed) else None


def _format_float(value: float | None, places: int) -> str:
    if value is None:
        return ""
    return f"{value:.{places}f}"


def _escape_excel_formula(value: object) -> object:
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        return f"'{value}"
    return value


def _unescape_excel_formula(value: object) -> str:
    if value is None:
        return ""
    text = str(value)
    if len(text) > 1 and text[0] == "'" and text[1] in "=+-@":
        return text[1:]
    return text


def _svg_text(x: int, y: int, text: str, size: int) -> str:
    return (
        f'<text x="{x}" y="{y}" font-family="Arial, sans-serif" '
        f'font-size="{size}">{html.escape(text)}</text>'
    )


if __name__ == "__main__":
    raise SystemExit(main())
