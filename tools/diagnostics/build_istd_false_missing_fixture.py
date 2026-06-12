from __future__ import annotations

import argparse
import re
import sys
from collections.abc import Sequence
from pathlib import Path

from openpyxl import load_workbook

from tools.diagnostics.diagnostic_io import (
    required_indexes,
    write_delimited_rows,
)

OLD_ROW_TO_TARGET = {
    "245.1332/12.28": "d3-5-medC",
    "261.1283/8.97": "d3-5-hmdC",
}

FIELDNAMES = (
    "old_matrix_path",
    "old_matrix_sheet",
    "targeted_workbook_path",
    "targeted_sheet",
    "old_row_coordinate",
    "targeted_identity",
    "old_sample_id",
    "targeted_sample_id",
    "sample_mapping_rule",
    "targeted_rt",
    "targeted_area",
    "targeted_nl",
    "targeted_confidence",
    "targeted_reason",
)


def build_istd_false_missing_fixture(
    *,
    old_matrix_path: Path,
    targeted_workbook_path: Path,
    output_csv: Path,
    old_matrix_sheet: str = "RawIntensity",
    targeted_sheet: str = "XIC Results",
) -> list[dict[str, str]]:
    old_missing = _old_missing_samples(old_matrix_path, old_matrix_sheet)
    targeted_rows = _targeted_rows(targeted_workbook_path, targeted_sheet)
    rows: list[dict[str, str]] = []
    for old_row_coordinate, target_label in OLD_ROW_TO_TARGET.items():
        for old_sample_id in old_missing[old_row_coordinate]:
            targeted_sample_id = _map_sample_id(old_sample_id)
            targeted = targeted_rows.get((target_label, targeted_sample_id))
            if targeted is None:
                raise ValueError(
                    "Missing targeted evidence for "
                    f"{target_label} / {targeted_sample_id}",
                )
            rows.append(
                {
                    "old_matrix_path": str(old_matrix_path),
                    "old_matrix_sheet": old_matrix_sheet,
                    "targeted_workbook_path": str(targeted_workbook_path),
                    "targeted_sheet": targeted_sheet,
                    "old_row_coordinate": old_row_coordinate,
                    "targeted_identity": target_label,
                    "old_sample_id": old_sample_id,
                    "targeted_sample_id": targeted_sample_id,
                    "sample_mapping_rule": "regex:_QC_(number)->_QC(number)",
                    "targeted_rt": _text(targeted["RT"]),
                    "targeted_area": _text(targeted["Area"]),
                    "targeted_nl": _text(targeted["NL"]),
                    "targeted_confidence": _text(targeted["Confidence"]),
                    "targeted_reason": _text(targeted["Reason"]),
                },
            )
    _write_csv(output_csv, rows)
    return rows


def main(argv: Sequence[str] | None = None) -> int:
    args = _parse_args(argv)
    try:
        build_istd_false_missing_fixture(
            old_matrix_path=args.old_matrix,
            targeted_workbook_path=args.targeted_workbook,
            output_csv=args.output_csv,
        )
    except (FileNotFoundError, KeyError, ValueError) as exc:
        print(str(exc), file=sys.stderr)
        return 2
    return 0


def _parse_args(argv: Sequence[str] | None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build ISTD false-missing validation fixture.",
    )
    parser.add_argument("--old-matrix", type=Path, required=True)
    parser.add_argument("--targeted-workbook", type=Path, required=True)
    parser.add_argument("--output-csv", type=Path, required=True)
    return parser.parse_args(argv)


def _old_missing_samples(path: Path, sheet_name: str) -> dict[str, list[str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name]
        header = list(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)))
        result: dict[str, list[str]] = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            coordinate = _text(row[0])
            if coordinate not in OLD_ROW_TO_TARGET:
                continue
            missing: list[str] = []
            for column_name, value in zip(header[1:], row[1:]):
                if column_name is None or str(column_name).startswith("Imputation"):
                    continue
                if value is None or value == "":
                    missing.append(str(column_name))
            result[coordinate] = missing
        for coordinate in OLD_ROW_TO_TARGET:
            if coordinate not in result:
                raise KeyError(f"Old matrix is missing row: {coordinate}")
        return result
    finally:
        workbook.close()


def _targeted_rows(
    path: Path, sheet_name: str
) -> dict[tuple[str, str], dict[str, object]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name]
        header = list(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)))
        required = ("SampleName", "Target", "RT", "Area", "NL", "Confidence", "Reason")
        try:
            indexes = required_indexes(header, required, "XIC Results")
        except ValueError as exc:
            present = {str(name).strip() for name in header if name}
            missing = sorted(set(required) - present)
            raise KeyError(f"Targeted workbook is missing columns: {missing}") from exc
        rows: dict[tuple[str, str], dict[str, object]] = {}
        current_sample = ""
        for row in sheet.iter_rows(min_row=2, values_only=True):
            sample = row[indexes["SampleName"]]
            if sample is not None and sample != "":
                current_sample = str(sample)
            target = row[indexes["Target"]]
            if target is None or target == "":
                continue
            if current_sample == "":
                continue
            rows[(str(target), current_sample)] = {
                "RT": row[indexes["RT"]],
                "Area": row[indexes["Area"]],
                "NL": row[indexes["NL"]],
                "Confidence": row[indexes["Confidence"]],
                "Reason": row[indexes["Reason"]],
            }
        return rows
    finally:
        workbook.close()


def _map_sample_id(sample_id: str) -> str:
    return re.sub(r"_QC_(\d+)$", r"_QC\1", sample_id)


def _write_csv(path: Path, rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    write_delimited_rows(path, rows, FIELDNAMES)


def _text(value: object) -> str:
    if value is None:
        return ""
    return str(value)


if __name__ == "__main__":
    raise SystemExit(main())
