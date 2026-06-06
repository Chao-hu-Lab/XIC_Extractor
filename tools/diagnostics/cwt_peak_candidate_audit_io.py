from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from tools.diagnostics.cwt_peak_candidate_audit_models import (
    _REQUIRED_COLUMNS,
    CwtCandidateRow,
)
from tools.diagnostics.diagnostic_io import (
    read_tsv_required,
)
from tools.diagnostics.diagnostic_io import (
    required_indexes as _required_indexes,
)
from tools.diagnostics.diagnostic_io import (
    text_value as _text,
)


def _read_peak_candidates(path: Path) -> tuple[CwtCandidateRow, ...]:
    rows = read_tsv_required(path, _REQUIRED_COLUMNS)
    return tuple(_row_from_dict(path, index, row) for index, row in enumerate(rows, 2))


def _read_target_mz(path: Path) -> dict[str, float]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if "Targets" not in workbook.sheetnames:
            raise ValueError(f"{path}: missing required sheet: Targets")
        rows = workbook["Targets"].iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            raise ValueError(f"{path}: Targets sheet is empty")
        indexes = _required_indexes(header, ("Label", "m/z"), "Targets")
        target_mz: dict[str, float] = {}
        for row_number, row in enumerate(rows, 2):
            label = _text(row[indexes["Label"]])
            if not label:
                continue
            target_mz[label] = _float_value(
                path,
                row_number,
                "m/z",
                _text(row[indexes["m/z"]]),
            )
        return target_mz
    finally:
        workbook.close()


def _row_from_dict(
    path: Path,
    row_number: int,
    row: dict[str, str],
) -> CwtCandidateRow:
    return CwtCandidateRow(
        sample_name=row["sample_name"],
        target_label=row["target_label"],
        resolver_mode=row["resolver_mode"],
        candidate_id=row["candidate_id"],
        proposal_sources=row["proposal_sources"],
        rt_apex_min=_float_value(path, row_number, "rt_apex_min", row["rt_apex_min"]),
        selected=row["selected"].strip().upper() == "TRUE",
        confidence=row["confidence"],
        raw_score=row["raw_score"],
        reason=row["reason"],
        ms2_present=row.get("ms2_present", ""),
        nl_match=row.get("nl_match", ""),
        ms2_trace_strength=row.get("ms2_trace_strength", ""),
    )


def _float_value(path: Path, row_number: int, column: str, value: str) -> float:
    try:
        return float(value)
    except ValueError as exc:
        message = f"{path}: row {row_number} invalid {column}: {value!r}"
        raise ValueError(message) from exc
