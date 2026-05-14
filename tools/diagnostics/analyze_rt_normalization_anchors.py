"""Anchor-based RT normalization diagnostic for untargeted alignment outputs."""

from __future__ import annotations

import argparse
import csv
import json
import math
import re
import sys
from collections.abc import Iterable, Mapping, Sequence
from dataclasses import asdict, dataclass
from pathlib import Path
from statistics import median

from openpyxl import load_workbook

ACTIVE_NEUTRAL_LOSS_DA = 116.0474
ACTIVE_NEUTRAL_LOSS_TOLERANCE_DA = 0.01

SUMMARY_COLUMNS = ("metric", "value")
SAMPLE_COLUMNS = (
    "sample_stem",
    "model_type",
    "anchor_count",
    "slope",
    "intercept",
    "anchor_median_abs_residual_min",
    "anchor_max_abs_residual_min",
)
ANCHOR_COLUMNS = (
    "sample_stem",
    "target_label",
    "reference_rt_min",
    "observed_rt_min",
    "normalized_rt_min",
    "normalized_residual_min",
)
FAMILY_COLUMNS = (
    "feature_family_id",
    "include_in_primary_matrix",
    "family_center_mz",
    "family_center_rt",
    "raw_cell_count",
    "modelled_cell_count",
    "unmodelled_cell_count",
    "raw_rt_range_min",
    "normalized_rt_range_min",
    "rt_range_improvement_min",
    "raw_rt_median_min",
    "normalized_rt_median_min",
)


@dataclass(frozen=True)
class RtNormalizationOutputs:
    summary_tsv: Path
    sample_tsv: Path
    anchor_tsv: Path
    family_tsv: Path
    json_path: Path
    markdown_path: Path


@dataclass(frozen=True)
class AnchorDefinition:
    label: str
    role: str
    neutral_loss_da: float
    reference_rt_min: float


@dataclass(frozen=True)
class AnchorPoint:
    sample_stem: str
    target_label: str
    observed_rt_min: float
    reference_rt_min: float


@dataclass(frozen=True)
class SampleRtModel:
    sample_stem: str
    model_type: str
    anchor_count: int
    slope: float
    intercept: float
    anchor_median_abs_residual_min: float | None
    anchor_max_abs_residual_min: float | None

    def normalize_rt(self, raw_rt_min: float) -> float:
        return (raw_rt_min - self.intercept) / self.slope


@dataclass(frozen=True)
class AnchorResidual:
    sample_stem: str
    target_label: str
    reference_rt_min: float
    observed_rt_min: float
    normalized_rt_min: float
    normalized_residual_min: float


@dataclass(frozen=True)
class AlignmentFeature:
    feature_family_id: str
    include_in_primary_matrix: bool
    family_center_mz: float | None
    family_center_rt: float | None


@dataclass(frozen=True)
class AlignmentCell:
    feature_family_id: str
    sample_stem: str
    apex_rt: float


@dataclass(frozen=True)
class FamilyRtSummary:
    feature_family_id: str
    include_in_primary_matrix: bool
    family_center_mz: float | None
    family_center_rt: float | None
    raw_cell_count: int
    modelled_cell_count: int
    unmodelled_cell_count: int
    raw_rt_range_min: float | None
    normalized_rt_range_min: float | None
    rt_range_improvement_min: float | None
    raw_rt_median_min: float | None
    normalized_rt_median_min: float | None


@dataclass(frozen=True)
class RtNormalizationResult:
    overall_status: str
    reference_source: str
    anchor_label_count: int
    sample_count: int
    modelled_sample_count: int
    unmodelled_sample_count: int
    family_count: int
    families_improved_count: int
    families_worsened_count: int
    median_raw_rt_range_min: float | None
    median_normalized_rt_range_min: float | None
    median_rt_range_improvement_min: float | None
    samples: tuple[SampleRtModel, ...]
    anchors: tuple[AnchorResidual, ...]
    families: tuple[FamilyRtSummary, ...]


def run_rt_normalization_anchor_diagnostic(
    *,
    targeted_workbook: Path,
    alignment_dir: Path,
    output_dir: Path,
    active_neutral_loss_da: float = ACTIVE_NEUTRAL_LOSS_DA,
    active_neutral_loss_tolerance_da: float = ACTIVE_NEUTRAL_LOSS_TOLERANCE_DA,
    reference_source: str = "observed-median",
) -> tuple[RtNormalizationOutputs, RtNormalizationResult]:
    anchors = _read_anchor_definitions(
        targeted_workbook,
        active_neutral_loss_da=active_neutral_loss_da,
        active_neutral_loss_tolerance_da=active_neutral_loss_tolerance_da,
    )
    points = _read_anchor_points(targeted_workbook, anchors)
    points = _apply_reference_source(points, reference_source)
    models, residuals, sample_count = _fit_sample_models(points)
    features = _read_alignment_review(alignment_dir / "alignment_review.tsv")
    cells = _read_alignment_cells(alignment_dir / "alignment_cells.tsv")
    families = _summarize_families(features, cells, models)
    result = _build_result(
        reference_source=reference_source,
        anchor_label_count=len(anchors),
        sample_count=sample_count,
        models=models,
        residuals=residuals,
        families=families,
    )

    output_dir.mkdir(parents=True, exist_ok=True)
    outputs = RtNormalizationOutputs(
        summary_tsv=output_dir / "rt_normalization_summary.tsv",
        sample_tsv=output_dir / "rt_normalization_samples.tsv",
        anchor_tsv=output_dir / "rt_normalization_anchors.tsv",
        family_tsv=output_dir / "rt_normalization_families.tsv",
        json_path=output_dir / "rt_normalization.json",
        markdown_path=output_dir / "rt_normalization.md",
    )
    _write_tsv(outputs.summary_tsv, SUMMARY_COLUMNS, _summary_rows(result))
    _write_tsv(outputs.sample_tsv, SAMPLE_COLUMNS, _sample_rows(result.samples))
    _write_tsv(outputs.anchor_tsv, ANCHOR_COLUMNS, _anchor_rows(result.anchors))
    _write_tsv(outputs.family_tsv, FAMILY_COLUMNS, _family_rows(result.families))
    _write_json(outputs.json_path, _json_payload(result))
    _write_markdown(outputs.markdown_path, result)
    return outputs, result


def main(argv: Sequence[str] | None = None) -> int:
    args = _parse_args(argv)
    try:
        outputs, result = run_rt_normalization_anchor_diagnostic(
            targeted_workbook=args.targeted_workbook.resolve(),
            alignment_dir=args.alignment_dir.resolve(),
            output_dir=args.output_dir.resolve(),
            active_neutral_loss_da=args.active_neutral_loss_da,
            active_neutral_loss_tolerance_da=(
                args.active_neutral_loss_tolerance_da
            ),
            reference_source=args.reference_source,
        )
    except (FileNotFoundError, KeyError, ValueError) as exc:
        print(str(exc), file=sys.stderr)
        return 2
    print(f"Summary TSV: {outputs.summary_tsv}")
    print(f"Sample models TSV: {outputs.sample_tsv}")
    print(f"Anchor residuals TSV: {outputs.anchor_tsv}")
    print(f"Family RT TSV: {outputs.family_tsv}")
    print(f"Diagnostic JSON: {outputs.json_path}")
    print(f"Diagnostic report: {outputs.markdown_path}")
    return 0 if result.overall_status == "PASS" else 1


def _parse_args(argv: Sequence[str] | None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Analyze anchor-based RT normalization for alignment outputs.",
    )
    parser.add_argument("--targeted-workbook", type=Path, required=True)
    parser.add_argument("--alignment-dir", type=Path, required=True)
    parser.add_argument("--output-dir", type=Path, required=True)
    parser.add_argument("--active-neutral-loss-da", type=float, default=116.0474)
    parser.add_argument(
        "--active-neutral-loss-tolerance-da",
        type=float,
        default=0.01,
    )
    parser.add_argument(
        "--reference-source",
        choices=("observed-median", "target-window"),
        default="observed-median",
    )
    return parser.parse_args(argv)


def _read_anchor_definitions(
    path: Path,
    *,
    active_neutral_loss_da: float,
    active_neutral_loss_tolerance_da: float,
) -> dict[str, AnchorDefinition]:
    if not path.exists():
        raise FileNotFoundError(str(path))
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook["Targets"]
        rows = sheet.iter_rows(values_only=True)
        header = next(rows)
        cols = _required_indexes(
            header,
            ("Label", "Role", "RT min", "RT max", "NL (Da)"),
            "Targets",
        )
        anchors: dict[str, AnchorDefinition] = {}
        for row in rows:
            label = _text(row[cols["Label"]])
            role = _text(row[cols["Role"]])
            if not label or role != "ISTD":
                continue
            neutral_loss = _required_float(row[cols["NL (Da)"]], "NL (Da)", label)
            if abs(neutral_loss - active_neutral_loss_da) > (
                active_neutral_loss_tolerance_da
            ):
                continue
            rt_min = _required_float(row[cols["RT min"]], "RT min", label)
            rt_max = _required_float(row[cols["RT max"]], "RT max", label)
            anchors[label] = AnchorDefinition(
                label=label,
                role=role,
                neutral_loss_da=neutral_loss,
                reference_rt_min=(rt_min + rt_max) / 2.0,
            )
        return anchors
    finally:
        workbook.close()


def _read_anchor_points(
    path: Path,
    anchors: Mapping[str, AnchorDefinition],
) -> tuple[AnchorPoint, ...]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook["XIC Results"]
        rows = sheet.iter_rows(values_only=True)
        header = next(rows)
        cols = _required_indexes(
            header,
            ("SampleName", "Target", "Role", "RT"),
            "XIC Results",
        )
        current_sample = ""
        points: list[AnchorPoint] = []
        for row in rows:
            raw_sample = row[cols["SampleName"]]
            if raw_sample not in (None, ""):
                current_sample = _normalize_sample_id(_text(raw_sample))
            if not current_sample:
                continue
            label = _text(row[cols["Target"]])
            role = _text(row[cols["Role"]])
            if role != "ISTD" or label not in anchors:
                continue
            rt = _float_value(row[cols["RT"]])
            if rt is None:
                continue
            points.append(
                AnchorPoint(
                    sample_stem=current_sample,
                    target_label=label,
                    observed_rt_min=rt,
                    reference_rt_min=anchors[label].reference_rt_min,
                )
            )
        return tuple(points)
    finally:
        workbook.close()


def _fit_sample_models(
    points: Sequence[AnchorPoint],
) -> tuple[dict[str, SampleRtModel], tuple[AnchorResidual, ...], int]:
    points_by_sample: dict[str, list[AnchorPoint]] = {}
    for point in points:
        points_by_sample.setdefault(point.sample_stem, []).append(point)

    models: dict[str, SampleRtModel] = {}
    residuals: list[AnchorResidual] = []
    for sample, sample_points in sorted(points_by_sample.items()):
        deduped = _dedupe_anchor_points(sample_points)
        model = _fit_sample_model(sample, deduped)
        if model is None:
            continue
        sample_residuals: list[float] = []
        for point in deduped:
            normalized_rt = model.normalize_rt(point.observed_rt_min)
            residual = normalized_rt - point.reference_rt_min
            sample_residuals.append(residual)
            residuals.append(
                AnchorResidual(
                    sample_stem=sample,
                    target_label=point.target_label,
                    reference_rt_min=point.reference_rt_min,
                    observed_rt_min=point.observed_rt_min,
                    normalized_rt_min=normalized_rt,
                    normalized_residual_min=residual,
                )
            )
        models[sample] = SampleRtModel(
            sample_stem=model.sample_stem,
            model_type=model.model_type,
            anchor_count=model.anchor_count,
            slope=model.slope,
            intercept=model.intercept,
            anchor_median_abs_residual_min=_median_abs(sample_residuals),
            anchor_max_abs_residual_min=_max_abs(sample_residuals),
        )
    return models, tuple(residuals), len(points_by_sample)


def _apply_reference_source(
    points: tuple[AnchorPoint, ...],
    reference_source: str,
) -> tuple[AnchorPoint, ...]:
    if reference_source == "target-window":
        return points
    if reference_source != "observed-median":
        raise ValueError(f"unsupported reference source: {reference_source}")
    by_label: dict[str, list[float]] = {}
    for point in points:
        by_label.setdefault(point.target_label, []).append(point.observed_rt_min)
    references = {
        label: float(median(values)) for label, values in by_label.items() if values
    }
    return tuple(
        AnchorPoint(
            sample_stem=point.sample_stem,
            target_label=point.target_label,
            observed_rt_min=point.observed_rt_min,
            reference_rt_min=references[point.target_label],
        )
        for point in points
        if point.target_label in references
    )


def _dedupe_anchor_points(points: Sequence[AnchorPoint]) -> tuple[AnchorPoint, ...]:
    latest: dict[str, AnchorPoint] = {}
    for point in points:
        latest[point.target_label] = point
    return tuple(latest[label] for label in sorted(latest))


def _fit_sample_model(
    sample_stem: str,
    points: Sequence[AnchorPoint],
) -> SampleRtModel | None:
    if not points:
        return None
    if len(points) == 1:
        point = points[0]
        return SampleRtModel(
            sample_stem=sample_stem,
            model_type="shift",
            anchor_count=1,
            slope=1.0,
            intercept=point.observed_rt_min - point.reference_rt_min,
            anchor_median_abs_residual_min=None,
            anchor_max_abs_residual_min=None,
        )
    references = [point.reference_rt_min for point in points]
    observed = [point.observed_rt_min for point in points]
    ref_mean = sum(references) / len(references)
    obs_mean = sum(observed) / len(observed)
    denominator = sum((value - ref_mean) ** 2 for value in references)
    if denominator <= 0:
        return None
    slope = sum(
        (reference - ref_mean) * (rt - obs_mean)
        for reference, rt in zip(references, observed, strict=True)
    ) / denominator
    if not math.isfinite(slope) or abs(slope) < 1e-9:
        return None
    intercept = obs_mean - slope * ref_mean
    return SampleRtModel(
        sample_stem=sample_stem,
        model_type="affine",
        anchor_count=len(points),
        slope=slope,
        intercept=intercept,
        anchor_median_abs_residual_min=None,
        anchor_max_abs_residual_min=None,
    )


def _read_alignment_review(path: Path) -> dict[str, AlignmentFeature]:
    rows = _read_required_tsv(path)
    _require_fields(rows, ("feature_family_id",), path)
    features: dict[str, AlignmentFeature] = {}
    for row in rows:
        family_id = row["feature_family_id"]
        features[family_id] = AlignmentFeature(
            feature_family_id=family_id,
            include_in_primary_matrix=_is_trueish(
                row.get("include_in_primary_matrix"),
            ),
            family_center_mz=_float_value(row.get("family_center_mz")),
            family_center_rt=_float_value(row.get("family_center_rt")),
        )
    return features


def _read_alignment_cells(path: Path) -> tuple[AlignmentCell, ...]:
    rows = _read_required_tsv(path)
    _require_fields(
        rows,
        ("feature_family_id", "sample_stem", "apex_rt"),
        path,
    )
    cells: list[AlignmentCell] = []
    for row in rows:
        apex_rt = _float_value(row.get("apex_rt"))
        if apex_rt is None:
            continue
        cells.append(
            AlignmentCell(
                feature_family_id=row["feature_family_id"],
                sample_stem=_normalize_sample_id(row["sample_stem"]),
                apex_rt=apex_rt,
            )
        )
    return tuple(cells)


def _summarize_families(
    features: Mapping[str, AlignmentFeature],
    cells: Sequence[AlignmentCell],
    models: Mapping[str, SampleRtModel],
) -> tuple[FamilyRtSummary, ...]:
    cells_by_family: dict[str, list[AlignmentCell]] = {
        family_id: [] for family_id in features
    }
    for cell in cells:
        cells_by_family.setdefault(cell.feature_family_id, []).append(cell)

    rows: list[FamilyRtSummary] = []
    for family_id in sorted(cells_by_family):
        family_cells = cells_by_family[family_id]
        raw_rts = [cell.apex_rt for cell in family_cells]
        normalized_rts = [
            models[cell.sample_stem].normalize_rt(cell.apex_rt)
            for cell in family_cells
            if cell.sample_stem in models
        ]
        feature = features.get(
            family_id,
            AlignmentFeature(family_id, False, None, None),
        )
        raw_range = _range(raw_rts)
        normalized_range = _range(normalized_rts)
        improvement = (
            raw_range - normalized_range
            if raw_range is not None and normalized_range is not None
            else None
        )
        rows.append(
            FamilyRtSummary(
                feature_family_id=family_id,
                include_in_primary_matrix=feature.include_in_primary_matrix,
                family_center_mz=feature.family_center_mz,
                family_center_rt=feature.family_center_rt,
                raw_cell_count=len(raw_rts),
                modelled_cell_count=len(normalized_rts),
                unmodelled_cell_count=len(raw_rts) - len(normalized_rts),
                raw_rt_range_min=raw_range,
                normalized_rt_range_min=normalized_range,
                rt_range_improvement_min=improvement,
                raw_rt_median_min=_median(raw_rts),
                normalized_rt_median_min=_median(normalized_rts),
            )
        )
    return tuple(rows)


def _build_result(
    *,
    reference_source: str,
    anchor_label_count: int,
    sample_count: int,
    models: Mapping[str, SampleRtModel],
    residuals: tuple[AnchorResidual, ...],
    families: tuple[FamilyRtSummary, ...],
) -> RtNormalizationResult:
    improvements = [
        family.rt_range_improvement_min
        for family in families
        if family.rt_range_improvement_min is not None
    ]
    raw_ranges = [
        family.raw_rt_range_min
        for family in families
        if family.raw_rt_range_min is not None
    ]
    normalized_ranges = [
        family.normalized_rt_range_min
        for family in families
        if family.normalized_rt_range_min is not None
    ]
    overall_status = "PASS" if anchor_label_count >= 2 and models else "FAIL"
    return RtNormalizationResult(
        overall_status=overall_status,
        reference_source=reference_source,
        anchor_label_count=anchor_label_count,
        sample_count=sample_count,
        modelled_sample_count=len(models),
        unmodelled_sample_count=max(sample_count - len(models), 0),
        family_count=len(families),
        families_improved_count=sum(
            1 for value in improvements if value is not None and value > 0
        ),
        families_worsened_count=sum(
            1 for value in improvements if value is not None and value < 0
        ),
        median_raw_rt_range_min=_median(raw_ranges),
        median_normalized_rt_range_min=_median(normalized_ranges),
        median_rt_range_improvement_min=_median(improvements),
        samples=tuple(models[sample] for sample in sorted(models)),
        anchors=residuals,
        families=families,
    )


def _summary_rows(result: RtNormalizationResult) -> list[dict[str, object]]:
    payload = _json_payload(result)
    return [{"metric": key, "value": value} for key, value in payload.items()]


def _sample_rows(samples: Sequence[SampleRtModel]) -> list[dict[str, object]]:
    return [asdict(sample) for sample in samples]


def _anchor_rows(anchors: Sequence[AnchorResidual]) -> list[dict[str, object]]:
    return [asdict(anchor) for anchor in anchors]


def _family_rows(families: Sequence[FamilyRtSummary]) -> list[dict[str, object]]:
    return [asdict(family) for family in families]


def _json_payload(result: RtNormalizationResult) -> dict[str, object]:
    return {
        "overall_status": result.overall_status,
        "reference_source": result.reference_source,
        "anchor_label_count": result.anchor_label_count,
        "sample_count": result.sample_count,
        "modelled_sample_count": result.modelled_sample_count,
        "unmodelled_sample_count": result.unmodelled_sample_count,
        "family_count": result.family_count,
        "families_improved_count": result.families_improved_count,
        "families_worsened_count": result.families_worsened_count,
        "median_raw_rt_range_min": result.median_raw_rt_range_min,
        "median_normalized_rt_range_min": result.median_normalized_rt_range_min,
        "median_rt_range_improvement_min": result.median_rt_range_improvement_min,
    }


def _write_markdown(path: Path, result: RtNormalizationResult) -> None:
    lines = [
        "# RT Normalization Anchor Diagnostic",
        "",
        f"Overall status: {result.overall_status}",
        "",
        "| Metric | Value |",
        "|---|---:|",
    ]
    for row in _summary_rows(result):
        lines.append(f"| {row['metric']} | {_format_value(row['value'])} |")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def _write_json(path: Path, payload: Mapping[str, object]) -> None:
    path.write_text(json.dumps(payload, indent=2, sort_keys=True), encoding="utf-8")


def _write_tsv(
    path: Path,
    fieldnames: Sequence[str],
    rows: Sequence[Mapping[str, object]],
) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, delimiter="\t")
        writer.writeheader()
        for row in rows:
            writer.writerow(
                {key: _format_value(row.get(key, "")) for key in fieldnames}
            )


def _read_required_tsv(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        raise FileNotFoundError(str(path))
    with path.open(newline="", encoding="utf-8-sig") as handle:
        return list(csv.DictReader(handle, delimiter="\t"))


def _require_fields(
    rows: list[dict[str, str]],
    required: Sequence[str],
    path: Path,
) -> None:
    if not rows:
        raise ValueError(f"{path} has no data rows")
    fieldnames = set(rows[0])
    missing = [field for field in required if field not in fieldnames]
    if missing:
        raise ValueError(f"{path} is missing required columns: {missing}")


def _required_indexes(
    header: Sequence[object],
    required: Sequence[str],
    sheet_name: str,
) -> dict[str, int]:
    indexes = {str(value).strip(): index for index, value in enumerate(header) if value}
    missing = [field for field in required if field not in indexes]
    if missing:
        raise ValueError(f"{sheet_name} sheet missing required columns: {missing}")
    return indexes


def _required_float(value: object, field: str, label: str) -> float:
    parsed = _float_value(value)
    if parsed is None:
        raise ValueError(f"{label} has invalid {field}: {value!r}")
    return parsed


def _float_value(value: object) -> float | None:
    if value in (None, "") or isinstance(value, bool):
        return None
    if isinstance(value, int | float):
        return float(value) if math.isfinite(value) else None
    try:
        parsed = float(str(value).strip())
    except ValueError:
        return None
    return parsed if math.isfinite(parsed) else None


def _text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_sample_id(sample_id: str) -> str:
    value = sample_id.strip()
    return re.sub(
        r"(^|_)QC_(\d+)$",
        lambda match: f"{match.group(1)}QC{match.group(2)}",
        value,
    )


def _is_trueish(value: str | None) -> bool:
    if value is None:
        return False
    return value.strip().lower() in {"1", "true", "t", "yes", "y"}


def _median(values: Iterable[float | None]) -> float | None:
    finite = [
        float(value)
        for value in values
        if isinstance(value, (int, float))
        and not isinstance(value, bool)
        and math.isfinite(value)
    ]
    if not finite:
        return None
    return float(median(finite))


def _median_abs(values: Iterable[float]) -> float | None:
    finite = [abs(value) for value in values if math.isfinite(value)]
    if not finite:
        return None
    return float(median(finite))


def _max_abs(values: Iterable[float]) -> float | None:
    finite = [abs(value) for value in values if math.isfinite(value)]
    if not finite:
        return None
    return max(finite)


def _range(values: Sequence[float]) -> float | None:
    finite = [value for value in values if math.isfinite(value)]
    if len(finite) < 2:
        return None
    return max(finite) - min(finite)


def _format_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, float):
        if not math.isfinite(value):
            return ""
        return f"{value:.6g}"
    return str(value)


if __name__ == "__main__":
    raise SystemExit(main())
