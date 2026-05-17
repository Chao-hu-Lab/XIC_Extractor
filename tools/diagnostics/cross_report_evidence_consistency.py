"""Cross-report consistency diagnostic for targeted evidence surfaces."""

from __future__ import annotations

import argparse
import csv
import json
import sys
from collections import Counter
from collections.abc import Mapping, Sequence
from dataclasses import asdict, dataclass
from pathlib import Path

from openpyxl import load_workbook

from xic_extractor.evidence_semantics import (
    EvidenceSignalSet,
    classify_evidence_consistency,
)

_RELIABILITY_COLUMNS = (
    "sample_name",
    "target_label",
    "reliability_state",
    "risk_reasons",
)

_CANDIDATE_COLUMNS = (
    "sample_name",
    "target_label",
    "resolver_mode",
    "candidate_id",
    "proposal_sources",
    "rt_apex_min",
    "selected",
    "confidence",
    "raw_score",
    "support_labels",
    "concern_labels",
    "quality_flags",
    "ms2_present",
    "nl_match",
)

_SUMMARY_COLUMNS = (
    "rows_checked",
    "consistent_count",
    "mismatch_count",
    "missing_candidate_count",
    "missing_reliability_count",
    "issue_counts",
)

_ROW_COLUMNS = (
    "sample_name",
    "target_label",
    "target_mz",
    "reliability_state",
    "targeted_risk_reasons",
    "resolver_mode",
    "selected_candidate_id",
    "selected_rt_apex_min",
    "selected_raw_score",
    "selected_confidence",
    "targeted_area_to_median_ratio",
    "candidate_support_labels",
    "candidate_concern_labels",
    "candidate_consistency_labels",
    "consistency_status",
    "issue_type",
    "reason",
)


@dataclass(frozen=True)
class CrossReportConsistencyOutputs:
    summary_tsv: Path
    rows_tsv: Path
    json_path: Path
    markdown_path: Path


@dataclass(frozen=True)
class ReliabilityRow:
    sample_name: str
    target_label: str
    reliability_state: str
    risk_reasons: tuple[str, ...]
    area_to_target_median_ratio: float | None = None


@dataclass(frozen=True)
class CandidateRow:
    sample_name: str
    target_label: str
    resolver_mode: str
    candidate_id: str
    proposal_sources: tuple[str, ...]
    rt_apex_min: float | None
    selected: bool
    confidence: str
    raw_score: float | None
    support_labels: tuple[str, ...]
    concern_labels: tuple[str, ...]
    quality_flags: tuple[str, ...]
    ms2_present: bool | None
    nl_match: bool | None


@dataclass(frozen=True)
class ConsistencyRow:
    sample_name: str
    target_label: str
    target_mz: float | None
    reliability_state: str
    targeted_risk_reasons: str
    resolver_mode: str
    selected_candidate_id: str
    selected_rt_apex_min: float | None
    selected_raw_score: float | None
    selected_confidence: str
    targeted_area_to_median_ratio: float | None
    candidate_support_labels: str
    candidate_concern_labels: str
    candidate_consistency_labels: str
    consistency_status: str
    issue_type: str
    reason: str


@dataclass(frozen=True)
class ConsistencySummary:
    rows_checked: int
    consistent_count: int
    mismatch_count: int
    missing_candidate_count: int
    missing_reliability_count: int
    issue_counts: str


@dataclass(frozen=True)
class ConsistencyResult:
    summary: ConsistencySummary
    rows: tuple[ConsistencyRow, ...]


def run_cross_report_evidence_consistency(
    *,
    targeted_reliability_rows_tsv: Path,
    peak_candidates_tsv: Path,
    output_dir: Path,
    targeted_workbook: Path | None = None,
) -> tuple[CrossReportConsistencyOutputs, ConsistencyResult]:
    reliability_rows = _read_reliability_rows(targeted_reliability_rows_tsv)
    candidate_rows = _read_candidate_rows(peak_candidates_tsv)
    target_mz = _read_target_mz(targeted_workbook) if targeted_workbook else {}
    rows = _consistency_rows(reliability_rows, candidate_rows, target_mz=target_mz)
    result = ConsistencyResult(summary=_summary(rows), rows=rows)
    output_dir.mkdir(parents=True, exist_ok=True)
    outputs = CrossReportConsistencyOutputs(
        summary_tsv=output_dir / "cross_report_evidence_consistency_summary.tsv",
        rows_tsv=output_dir / "cross_report_evidence_consistency_rows.tsv",
        json_path=output_dir / "cross_report_evidence_consistency.json",
        markdown_path=output_dir / "cross_report_evidence_consistency.md",
    )
    _write_outputs(outputs, result)
    return outputs, result


def main(argv: Sequence[str] | None = None) -> int:
    args = _parse_args(argv)
    try:
        outputs, result = run_cross_report_evidence_consistency(
            targeted_reliability_rows_tsv=args.targeted_reliability_rows_tsv,
            peak_candidates_tsv=args.peak_candidates_tsv,
            output_dir=args.output_dir,
            targeted_workbook=args.targeted_workbook,
        )
    except (FileNotFoundError, KeyError, ValueError) as exc:
        print(str(exc), file=sys.stderr)
        return 2
    print(f"Summary TSV: {outputs.summary_tsv}")
    print(f"Rows TSV: {outputs.rows_tsv}")
    print(f"Consistency JSON: {outputs.json_path}")
    print(f"Consistency report: {outputs.markdown_path}")
    return 1 if result.summary.mismatch_count else 0


def _parse_args(argv: Sequence[str] | None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Compare targeted reliability and peak candidate evidence.",
    )
    parser.add_argument("--targeted-reliability-rows-tsv", type=Path, required=True)
    parser.add_argument("--peak-candidates-tsv", type=Path, required=True)
    parser.add_argument("--output-dir", type=Path, required=True)
    parser.add_argument(
        "--targeted-workbook",
        type=Path,
        help="Optional targeted workbook used to include target m/z in review rows.",
    )
    return parser.parse_args(argv)


def _read_reliability_rows(path: Path) -> tuple[ReliabilityRow, ...]:
    rows = _read_required_tsv(path, _RELIABILITY_COLUMNS)
    return tuple(
        ReliabilityRow(
            sample_name=row["sample_name"],
            target_label=row["target_label"],
            reliability_state=row["reliability_state"],
            risk_reasons=tuple(_split_labels(row["risk_reasons"])),
            area_to_target_median_ratio=_optional_float(
                row.get("area_to_target_median_ratio", "")
            ),
        )
        for row in rows
    )


def _read_candidate_rows(path: Path) -> tuple[CandidateRow, ...]:
    rows = _read_required_tsv(path, _CANDIDATE_COLUMNS)
    return tuple(
        CandidateRow(
            sample_name=row["sample_name"],
            target_label=row["target_label"],
            resolver_mode=row["resolver_mode"],
            candidate_id=row["candidate_id"],
            proposal_sources=tuple(_split_labels(row["proposal_sources"])),
            rt_apex_min=_optional_float(row["rt_apex_min"]),
            selected=_bool_value(row["selected"]) is True,
            confidence=row["confidence"],
            raw_score=_optional_float(row["raw_score"]),
            support_labels=tuple(_split_labels(row["support_labels"])),
            concern_labels=tuple(_split_labels(row["concern_labels"])),
            quality_flags=tuple(_split_labels(row["quality_flags"])),
            ms2_present=_bool_value(row["ms2_present"]),
            nl_match=_bool_value(row["nl_match"]),
        )
        for row in rows
    )


def _read_required_tsv(
    path: Path,
    required: Sequence[str],
) -> tuple[dict[str, str], ...]:
    if not path.exists():
        raise FileNotFoundError(str(path))
    with path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle, delimiter="\t")
        fieldnames = tuple(reader.fieldnames or ())
        missing = [column for column in required if column not in fieldnames]
        if missing:
            raise ValueError(
                f"{path}: missing required columns: {', '.join(missing)}"
            )
        return tuple(dict(row) for row in reader)


def _read_target_mz(path: Path) -> dict[str, float]:
    if not path.exists():
        raise FileNotFoundError(str(path))
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook["Targets"]
        rows = sheet.iter_rows(values_only=True)
        header = next(rows)
        indexes = {
            str(value).strip(): index for index, value in enumerate(header) if value
        }
        for required in ("Label", "m/z"):
            if required not in indexes:
                raise ValueError(f"Targets is missing required column: {required}")
        values: dict[str, float] = {}
        for row in rows:
            label = _text(row[indexes["Label"]])
            mz = _optional_float(_text(row[indexes["m/z"]]))
            if label and mz is not None:
                values[label] = mz
        return values
    finally:
        workbook.close()


def _consistency_rows(
    reliability_rows: tuple[ReliabilityRow, ...],
    candidate_rows: tuple[CandidateRow, ...],
    *,
    target_mz: Mapping[str, float],
) -> tuple[ConsistencyRow, ...]:
    reliability_by_key = {
        (row.sample_name, row.target_label): row for row in reliability_rows
    }
    selected_by_key: dict[tuple[str, str], list[CandidateRow]] = {}
    for row in candidate_rows:
        if row.selected:
            selected_by_key.setdefault((row.sample_name, row.target_label), []).append(
                row
            )
    keys = sorted(set(reliability_by_key) | set(selected_by_key))
    rows: list[ConsistencyRow] = []
    for key in keys:
        reliability = reliability_by_key.get(key)
        selected_rows = selected_by_key.get(key, [])
        if not selected_rows:
            if reliability is not None:
                rows.append(
                    _consistency_row(
                        reliability,
                        None,
                        target_mz=target_mz.get(reliability.target_label),
                    )
                )
            continue
        if reliability is None:
            for selected in selected_rows:
                rows.append(
                    _consistency_row(
                        None,
                        selected,
                        target_mz=target_mz.get(selected.target_label),
                    )
                )
            continue
        for selected in selected_rows:
            rows.append(
                _consistency_row(
                    reliability,
                    selected,
                    target_mz=target_mz.get(reliability.target_label),
                    multiple_selected=len(selected_rows) > 1,
                )
            )
    return tuple(rows)


def _consistency_row(
    reliability: ReliabilityRow | None,
    selected: CandidateRow | None,
    *,
    target_mz: float | None,
    multiple_selected: bool = False,
) -> ConsistencyRow:
    state = reliability.reliability_state if reliability is not None else ""
    risk_reasons = reliability.risk_reasons if reliability is not None else ()
    consistency = _candidate_consistency(selected)
    status, issue, reason = _classify_consistency(
        reliability,
        selected,
        consistency,
        multiple_selected=multiple_selected,
    )
    sample = (
        reliability.sample_name
        if reliability is not None
        else selected.sample_name
        if selected is not None
        else ""
    )
    target = (
        reliability.target_label
        if reliability is not None
        else selected.target_label
        if selected is not None
        else ""
    )
    return ConsistencyRow(
        sample_name=sample,
        target_label=target,
        target_mz=target_mz,
        reliability_state=state,
        targeted_risk_reasons=";".join(risk_reasons),
        resolver_mode=selected.resolver_mode if selected is not None else "",
        selected_candidate_id=selected.candidate_id if selected is not None else "",
        selected_rt_apex_min=selected.rt_apex_min if selected is not None else None,
        selected_raw_score=selected.raw_score if selected is not None else None,
        selected_confidence=selected.confidence if selected is not None else "",
        targeted_area_to_median_ratio=(
            reliability.area_to_target_median_ratio
            if reliability is not None
            else None
        ),
        candidate_support_labels=(
            ";".join(selected.support_labels) if selected is not None else ""
        ),
        candidate_concern_labels=(
            ";".join(selected.concern_labels) if selected is not None else ""
        ),
        candidate_consistency_labels=";".join(consistency),
        consistency_status=status,
        issue_type=issue,
        reason=reason,
    )


def _candidate_consistency(selected: CandidateRow | None) -> tuple[str, ...]:
    if selected is None:
        return ()
    return classify_evidence_consistency(
        EvidenceSignalSet(
            support_labels=selected.support_labels,
            concern_labels=selected.concern_labels,
            proposal_sources=selected.proposal_sources,
            quality_flags=selected.quality_flags,
            ms2_present=selected.ms2_present,
            nl_match=selected.nl_match,
            raw_score=selected.raw_score,
        )
    )


def _classify_consistency(
    reliability: ReliabilityRow | None,
    selected: CandidateRow | None,
    consistency: tuple[str, ...],
    *,
    multiple_selected: bool,
) -> tuple[str, str, str]:
    labels = set(consistency)
    if reliability is None:
        return (
            "mismatch",
            "missing_targeted_reliability",
            "Selected candidate has no matching targeted reliability row.",
        )
    if selected is None:
        if reliability.reliability_state == "targeted_negative":
            return (
                "consistent",
                "",
                "",
            )
        return (
            "mismatch",
            "missing_selected_candidate",
            "Targeted reliability row has no selected peak candidate row.",
        )
    if multiple_selected:
        return (
            "mismatch",
            "multiple_selected_candidates",
            "More than one selected candidate exists for this sample/target.",
        )
    state = reliability.reliability_state
    if state == "benchmark_eligible" and labels & {
        "hard_nl_conflict",
        "missing_ms2",
    }:
        return (
            "mismatch",
            "targeted_clean_candidate_conflict",
            "Targeted reliability says clean, but selected candidate has "
            "hard conflict labels.",
        )
    if state == "targeted_review_positive" and "plausible_nl_dropout" not in labels:
        return (
            "mismatch",
            "review_positive_not_supported_by_candidate",
            "Targeted review-positive state is not supported by selected "
            "candidate consistency labels.",
        )
    if (
        state == "targeted_review"
        and "plausible_nl_dropout" in labels
        and "hard_nl_conflict" not in labels
    ):
        if _has_review_positive_blocker(reliability.risk_reasons):
            return ("consistent", "", "")
        return (
            "mismatch",
            "targeted_review_candidate_suggests_dropout",
            "Targeted review row may be a stronger review-positive dropout case.",
        )
    if state == "targeted_negative" and "ms1_coherent" in labels:
        return (
            "mismatch",
            "targeted_negative_candidate_has_peak",
            "Targeted negative row has a coherent selected candidate peak.",
        )
    return ("consistent", "", "")


def _has_review_positive_blocker(risk_reasons: Sequence[str]) -> bool:
    blockers = {
        "hard_nl_conflict",
        "no_ms2",
        "quality_flags",
        "weak_area_rank",
    }
    return any(reason in blockers for reason in risk_reasons)


def _summary(rows: tuple[ConsistencyRow, ...]) -> ConsistencySummary:
    issues = Counter(row.issue_type for row in rows if row.issue_type)
    return ConsistencySummary(
        rows_checked=len(rows),
        consistent_count=sum(row.consistency_status == "consistent" for row in rows),
        mismatch_count=sum(row.consistency_status == "mismatch" for row in rows),
        missing_candidate_count=issues["missing_selected_candidate"],
        missing_reliability_count=issues["missing_targeted_reliability"],
        issue_counts=";".join(f"{issue}:{count}" for issue, count in issues.items()),
    )


def _write_outputs(
    outputs: CrossReportConsistencyOutputs,
    result: ConsistencyResult,
) -> None:
    _write_tsv(outputs.summary_tsv, _SUMMARY_COLUMNS, [asdict(result.summary)])
    _write_tsv(outputs.rows_tsv, _ROW_COLUMNS, _row_dicts(result.rows))
    outputs.json_path.write_text(
        json.dumps(
            {
                "summary": asdict(result.summary),
                "rows": _row_dicts(result.rows),
            },
            indent=2,
            sort_keys=True,
        ),
        encoding="utf-8",
    )
    outputs.markdown_path.write_text(_markdown(result), encoding="utf-8")


def _row_dicts(rows: Sequence[ConsistencyRow]) -> list[dict[str, object]]:
    return [asdict(row) for row in rows]


def _write_tsv(
    path: Path,
    fieldnames: Sequence[str],
    rows: Sequence[Mapping[str, object]],
) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, delimiter="\t")
        writer.writeheader()
        for row in rows:
            writer.writerow(
                {key: _format_value(row.get(key, "")) for key in fieldnames}
            )


def _markdown(result: ConsistencyResult) -> str:
    lines = [
        "# Cross-report Evidence Consistency",
        "",
        "This diagnostic compares targeted reliability rows with selected "
        "peak-candidate evidence. It does not change production selection.",
        "",
        "## Summary",
        "",
    ]
    summary = asdict(result.summary)
    lines.extend(f"- {key}: {summary[key]}" for key in _SUMMARY_COLUMNS)
    lines.extend(["", "## Top Issues", ""])
    for row in result.rows:
        if row.consistency_status != "mismatch":
            continue
        mz = "" if row.target_mz is None else f", m/z {row.target_mz:.6g}"
        lines.append(
            f"- {row.issue_type}: {row.sample_name} / {row.target_label}{mz}"
        )
    lines.append("")
    return "\n".join(lines)


def _split_labels(value: str) -> list[str]:
    return [part.strip() for part in value.split(";") if part.strip()]


def _optional_float(value: str) -> float | None:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number


def _bool_value(value: str) -> bool | None:
    normalized = value.strip().upper()
    if normalized in {"TRUE", "T", "YES", "Y", "1"}:
        return True
    if normalized in {"FALSE", "F", "NO", "N", "0"}:
        return False
    return None


def _format_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        return f"{value:.6g}"
    return str(value)


def _text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


if __name__ == "__main__":
    raise SystemExit(main())
