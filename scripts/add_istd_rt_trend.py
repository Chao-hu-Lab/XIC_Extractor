"""
scripts/add_istd_rt_trend.py

為已存在的 XIC 結果 Excel 新增 ISTD_RT_Trend sheet。

Usage:
    uv run python scripts/add_istd_rt_trend.py <xic_output.xlsx> <injection_order.xlsx>

輸出：直接修改 xic_output.xlsx（in-place），加入 ISTD_RT_Trend sheet。

Sheet 格式：
    - 每列一個樣本，依進樣順序排列
    - 欄位：InjectionOrder, SampleType, SampleName, <ISTD1>_RT, <ISTD2>_RT, ...
    - 顏色編碼依偏離中位數的程度：綠 (< 0.2 min)、黃 (0.2–0.5 min)、紅 (> 0.5 min)
    - 無法配對的樣本標灰色
"""

from __future__ import annotations

import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ── RT 偏移色碼閾值 ──────────────────────────────────────
_OK_DELTA_MIN: float = 0.2  # < 0.2 min → 綠
_WARN_DELTA_MIN: float = 0.5  # 0.2–0.5 min → 黃；> 0.5 min → 紅

_FILL_HEADER = PatternFill("solid", fgColor="2F4F7F")
_FILL_GREEN = PatternFill("solid", fgColor="C6EFCE")
_FILL_YELLOW = PatternFill("solid", fgColor="FFEB9C")
_FILL_RED = PatternFill("solid", fgColor="FFC7CE")
_FILL_GRAY = PatternFill("solid", fgColor="D9D9D9")
_FILL_QC = PatternFill("solid", fgColor="EAF2FB")
_FONT_HEADER = Font(bold=True, color="FFFFFF")
_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT = Alignment(horizontal="left", vertical="center")


# ── 名稱配對（結構化 + token fallback）─────────────────

# 樣本類型前綴對應表：injection order 用字 → raw stem 用字（小寫）
_TYPE_ALIASES: dict[str, str] = {
    "tumor": "tumor",
    "normal": "normal",
    "benign": "benign",
    "benignfat": "benign",
    "qc": "qc",
    "breast": "qc",
}


def _extract_bc_id(name: str) -> str:
    """擷取 BC+數字 編號，如 'BC2257'、'BC0979'，不區分大小寫，統一大寫回傳。"""
    m = re.search(r"BC\d+", name, re.IGNORECASE)
    return m.group().upper() if m else ""


def _extract_type(name: str) -> str:
    """從名稱判斷樣本類型（tumor / normal / benign / qc）。"""
    n = name.lower()
    for key, canonical in _TYPE_ALIASES.items():
        if key in n:
            return canonical
    return ""


def _extract_qc_number(name: str) -> str:
    """從 QC 樣本名稱擷取編號，如 'QC_1'、'QC1'、'QC 1' → '1'。"""
    m = re.search(r"QC[_\s\*]*(\d+)", name, re.IGNORECASE)
    return m.group(1) if m else ""


def _match_score(inj_name: str, raw_stem: str) -> float:
    """結構化配對分數（0–1）。

    優先以 BC 編號（主鍵）+ 樣本類型（次鍵）配對；
    QC 樣本以 QC 編號配對；
    其餘 fallback 到 token overlap。
    """
    inj_bc = _extract_bc_id(inj_name)
    raw_bc = _extract_bc_id(raw_stem)

    if inj_bc and raw_bc:
        if inj_bc != raw_bc:
            return 0.0
        inj_type = _extract_type(inj_name)
        raw_type = _extract_type(raw_stem)
        if inj_type and raw_type:
            return 1.0 if inj_type == raw_type else 0.0
        return 0.8  # BC 相符但 type 無法判斷

    inj_type = _extract_type(inj_name)
    raw_type = _extract_type(raw_stem)
    if inj_type == "qc" and raw_type == "qc":
        inj_qc = _extract_qc_number(inj_name)
        raw_qc = _extract_qc_number(raw_stem)
        if inj_qc and raw_qc:
            return 1.0 if inj_qc == raw_qc else 0.0

    # fallback：token overlap
    def _tokens(s: str) -> set[str]:
        s = re.sub(r"[_\-\s]+", " ", s.lower().strip())
        return {t for t in s.split() if len(t) >= 2}

    inj_t = _tokens(inj_name)
    if not inj_t:
        return 0.0
    return len(inj_t & _tokens(raw_stem)) / len(inj_t)


def _build_name_map(
    inj_names: list[str], raw_stems: list[str], threshold: float = 0.75
) -> dict[str, str]:
    """回傳 {inj_name: raw_stem}，最佳配對分數 < threshold 視為無配對。"""
    result: dict[str, str] = {}
    for inj in inj_names:
        scores = {s: _match_score(inj, s) for s in raw_stems}
        best_stem = max(scores, key=scores.__getitem__)
        result[inj] = best_stem if scores[best_stem] >= threshold else ""
    return result


# ── 讀取 injection order Excel ──────────────────────────


def _read_injection_order(path: Path) -> list[dict[str, object]]:
    """回傳 [{"inj_name": str, "sample_type": str, "injection_order": int}, ...]。"""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        raise ValueError(f"{path}: empty file")
    header = [str(c).strip().lower().replace(" ", "_") if c else "" for c in rows[0]]
    try:
        name_col = header.index("sample_name")
        type_col = header.index("sample_type")
        order_col = header.index("injection_order")
    except ValueError as exc:
        raise ValueError(
            f"{path}: expected columns Sample_Name / Sample_Type / "
            f"Injection_Order; got {rows[0]}"
        ) from exc
    result = []
    for row in rows[1:]:
        if not row[name_col]:
            continue
        result.append(
            {
                "inj_name": str(row[name_col]).strip(),
                "sample_type": str(row[type_col]).strip() if row[type_col] else "",
                "injection_order": int(row[order_col])
                if row[order_col] is not None
                else 0,
            }
        )
    return sorted(result, key=lambda r: r["injection_order"])


# ── 從 XIC Results sheet 讀 ISTD RT ─────────────────────


def _read_istd_rts(ws) -> dict[str, dict[str, str]]:
    """回傳 {sample_name: {istd_label: rt_str}}。

    XIC Results sheet 有 merged cells；只有合併範圍第一列有值，
    其餘 MergedCell 需從上一個有效值繼承。
    """
    rows = list(ws.iter_rows(values_only=False))
    if not rows:
        return {}
    header_row = rows[0]
    headers = [c.value for c in header_row]
    try:
        col_sample = headers.index("SampleName")
        col_target = headers.index("Target")
        col_role = headers.index("Role")
        col_rt = headers.index("RT")
    except ValueError as exc:
        raise ValueError(f"XIC Results sheet missing expected column: {exc}") from exc

    result: dict[str, dict[str, str]] = {}
    last_sample = ""
    for row in rows[1:]:
        sample_cell = row[col_sample].value
        if sample_cell:
            last_sample = str(sample_cell).strip()
        sample = last_sample

        role = row[col_role].value
        if str(role).strip() != "ISTD":
            continue
        target = str(row[col_target].value or "").strip()
        rt_val = row[col_rt].value
        rt_str = (
            f"{float(rt_val):.4f}" if rt_val not in (None, "", "ND", "ERROR") else ""
        )
        result.setdefault(sample, {})[target] = rt_str

    return result


# ── 組建 trend 表 ────────────────────────────────────────


def _rt_fill(rt_str: str, median_rt: float | None) -> PatternFill:
    if not rt_str or median_rt is None:
        return _FILL_GRAY
    try:
        delta = abs(float(rt_str) - median_rt)
    except ValueError:
        return _FILL_GRAY
    if delta < _OK_DELTA_MIN:
        return _FILL_GREEN
    if delta < _WARN_DELTA_MIN:
        return _FILL_YELLOW
    return _FILL_RED


def _compute_medians(
    istd_labels: list[str],
    ordered_rows: list[dict],
    inj_to_raw: dict[str, str],
    istd_data: dict[str, dict[str, str]],
) -> dict[str, float | None]:
    """每個 ISTD 計算所有可用 RT 值的中位數（只用 QC 樣本；若 QC 不足則用全部）。"""
    from statistics import median

    medians: dict[str, float | None] = {}
    for label in istd_labels:
        qc_vals: list[float] = []
        all_vals: list[float] = []
        for r in ordered_rows:
            raw = inj_to_raw.get(r["inj_name"], "")
            if not raw:
                continue
            rt_str = istd_data.get(raw, {}).get(label, "")
            if not rt_str:
                continue
            try:
                val = float(rt_str)
            except ValueError:
                continue
            all_vals.append(val)
            if r["sample_type"] == "QC":
                qc_vals.append(val)
        pool = qc_vals if len(qc_vals) >= 3 else all_vals
        medians[label] = median(pool) if pool else None
    return medians


def _build_trend_sheet(
    ws,
    ordered_rows: list[dict],
    inj_to_raw: dict[str, str],
    istd_data: dict[str, dict[str, str]],
    istd_labels: list[str],
) -> None:
    medians = _compute_medians(istd_labels, ordered_rows, inj_to_raw, istd_data)

    # ── 標頭列 ──
    meta_headers = ["InjectionOrder", "SampleType", "SampleName"]
    all_headers = meta_headers + [f"{lbl}\nRT (min)" for lbl in istd_labels]
    for col_idx, hdr in enumerate(all_headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=hdr)
        cell.fill = _FILL_HEADER
        cell.font = _FONT_HEADER
        cell.alignment = _ALIGN_CENTER

    # ── 中位數列 ──
    ws.cell(row=2, column=1, value="Median (RT)").font = Font(bold=True, italic=True)
    ws.cell(row=2, column=2, value="—")
    ws.cell(row=2, column=3, value="—")
    for col_idx, label in enumerate(istd_labels, start=4):
        med = medians.get(label)
        cell = ws.cell(row=2, column=col_idx, value=round(med, 4) if med else "")
        cell.font = Font(bold=True, italic=True)
        cell.alignment = _ALIGN_CENTER

    # ── 資料列 ──
    for row_idx, r in enumerate(ordered_rows, start=3):
        raw_stem = inj_to_raw.get(r["inj_name"], "")
        is_qc = r["sample_type"] == "QC"

        meta_vals = [r["injection_order"], r["sample_type"], r["inj_name"]]
        for col_idx, val in enumerate(meta_vals, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = _ALIGN_LEFT
            if is_qc:
                cell.fill = _FILL_QC

        for col_idx, label in enumerate(istd_labels, start=4):
            rt_str = istd_data.get(raw_stem, {}).get(label, "") if raw_stem else ""
            cell = ws.cell(
                row=row_idx, column=col_idx, value=float(rt_str) if rt_str else ""
            )
            cell.alignment = _ALIGN_CENTER
            cell.fill = (
                _FILL_QC
                if is_qc and not rt_str
                else _rt_fill(rt_str, medians.get(label))
                if rt_str
                else _FILL_GRAY
            )
            if not raw_stem:
                cell.fill = _FILL_GRAY

    # ── 欄寬 & 凍結 ──
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 36
    for col_idx in range(4, 4 + len(istd_labels)):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18
    ws.freeze_panes = "D3"
    ws.row_dimensions[1].height = 36
    ws.auto_filter.ref = (
        f"A1:{get_column_letter(3 + len(istd_labels))}{2 + len(ordered_rows)}"
    )


# ── 主流程 ───────────────────────────────────────────────


def add_istd_rt_trend(excel_path: Path, injection_order_path: Path) -> None:
    print(f"讀取 XIC 結果：{excel_path}")
    wb = openpyxl.load_workbook(excel_path)
    if "XIC Results" not in wb.sheetnames:
        raise ValueError(f"{excel_path}: 找不到 'XIC Results' sheet")
    istd_data = _read_istd_rts(wb["XIC Results"])
    raw_stems = sorted(istd_data.keys())

    print(f"讀取進樣順序：{injection_order_path}")
    ordered_rows = _read_injection_order(injection_order_path)
    inj_names = [r["inj_name"] for r in ordered_rows]

    print(f"Lazy 配對 {len(inj_names)} 個進樣名稱 → {len(raw_stems)} 個樣本…")
    inj_to_raw = _build_name_map(inj_names, raw_stems)
    matched = sum(1 for v in inj_to_raw.values() if v)
    unmatched = [k for k, v in inj_to_raw.items() if not v]
    print(f"  配對成功：{matched}/{len(inj_names)}")
    if unmatched:
        print(f"  ⚠ 無法配對（{len(unmatched)} 個）：")
        for name in unmatched:
            print(f"    · {name}")

    # 蒐集所有 ISTD labels（依第一個出現的樣本排序，保持一致）
    istd_labels: list[str] = []
    seen: set[str] = set()
    for sample_rts in istd_data.values():
        for lbl in sample_rts:
            if lbl not in seen:
                istd_labels.append(lbl)
                seen.add(lbl)

    # 移除舊 sheet（若存在）再重建
    if "ISTD_RT_Trend" in wb.sheetnames:
        del wb["ISTD_RT_Trend"]
    ws_trend = wb.create_sheet("ISTD_RT_Trend")

    _build_trend_sheet(ws_trend, ordered_rows, inj_to_raw, istd_data, istd_labels)

    # 移到第二個位置（Summary 前）
    wb.move_sheet("ISTD_RT_Trend", offset=-(len(wb.sheetnames) - 1))

    wb.save(excel_path)
    print(f"✓ 已寫入 ISTD_RT_Trend sheet → {excel_path}")


def main() -> None:
    if len(sys.argv) != 3:
        print(
            "Usage: uv run python scripts/add_istd_rt_trend.py "
            "<xic_output.xlsx> <injection_order.xlsx>"
        )
        sys.exit(1)
    add_istd_rt_trend(Path(sys.argv[1]), Path(sys.argv[2]))


if __name__ == "__main__":
    main()
