import argparse
import math
import sys
from collections.abc import Sequence
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

DEFAULT_NUMERIC_TOLERANCE = 1e-9
COMPARE_SHEETS = (
    "Overview",
    "XIC Results",
    "Summary",
    "Review Queue",
    "Targets",
    "Diagnostics",
    "Run Metadata",
)
OPTIONAL_COMPARE_SHEETS = ("Score Breakdown",)
IGNORED_METADATA_KEYS = {
    "generated_at",
    "elapsed",
    "elapsed_seconds",
    "runtime",
    "runtime_seconds",
    "output_path",
    "output_workbook",
    "workbook_path",
    "output_dir",
}


@dataclass(frozen=True)
class WorkbookCompareResult:
    matched: bool
    differences: list[str]


def compare_workbooks(
    left_path: Path,
    right_path: Path,
    *,
    numeric_tolerance: float = DEFAULT_NUMERIC_TOLERANCE,
) -> WorkbookCompareResult:
    left = load_workbook(left_path, data_only=True)
    right = load_workbook(right_path, data_only=True)
    differences: list[str] = []

    for sheet_name in _sheets_to_compare(left.sheetnames, right.sheetnames):
        if sheet_name not in left.sheetnames:
            differences.append(f"{left_path}: missing sheet {sheet_name!r}")
            continue
        if sheet_name not in right.sheetnames:
            differences.append(f"{right_path}: missing sheet {sheet_name!r}")
            continue
        differences.extend(
            _compare_sheet(
                sheet_name,
                _sheet_rows(left[sheet_name], sheet_name),
                _sheet_rows(right[sheet_name], sheet_name),
                numeric_tolerance,
            )
        )

    return WorkbookCompareResult(matched=not differences, differences=differences)


def _sheets_to_compare(
    left_sheetnames: Sequence[str], right_sheetnames: Sequence[str]
) -> tuple[str, ...]:
    optional = tuple(
        sheet
        for sheet in OPTIONAL_COMPARE_SHEETS
        if sheet in left_sheetnames or sheet in right_sheetnames
    )
    return (*COMPARE_SHEETS, *optional)


def _sheet_rows(sheet, sheet_name: str) -> list[tuple[Any, ...]]:
    rows: list[tuple[Any, ...]] = []
    for raw_row in sheet.iter_rows(values_only=True):
        row = _trim_trailing_empty(raw_row)
        if sheet_name == "Run Metadata" and _is_ignored_metadata_row(row):
            continue
        rows.append(row)
    while rows and not rows[-1]:
        rows.pop()
    return rows


def _trim_trailing_empty(row: Sequence[Any]) -> tuple[Any, ...]:
    values = list(row)
    while values and values[-1] is None:
        values.pop()
    return tuple(values)


def _is_ignored_metadata_row(row: tuple[Any, ...]) -> bool:
    if not row:
        return False
    key = str(row[0]).strip().lower()
    return key in IGNORED_METADATA_KEYS


def _compare_sheet(
    sheet_name: str,
    left_rows: list[tuple[Any, ...]],
    right_rows: list[tuple[Any, ...]],
    numeric_tolerance: float,
) -> list[str]:
    differences: list[str] = []
    if len(left_rows) != len(right_rows):
        differences.append(
            f"{sheet_name}: row count differs ({len(left_rows)} != {len(right_rows)})"
        )

    for row_number, (left_row, right_row) in enumerate(
        zip(left_rows, right_rows, strict=False), start=1
    ):
        width = max(len(left_row), len(right_row))
        for column_number in range(width):
            left_value = (
                left_row[column_number] if column_number < len(left_row) else None
            )
            right_value = (
                right_row[column_number] if column_number < len(right_row) else None
            )
            if _values_match(left_value, right_value, numeric_tolerance):
                continue
            differences.append(
                f"{sheet_name}!R{row_number}C{column_number + 1}: "
                f"{left_value!r} != {right_value!r}"
            )
    return differences


def _values_match(left: Any, right: Any, numeric_tolerance: float) -> bool:
    if left == right:
        return True
    if isinstance(left, bool) or isinstance(right, bool):
        return False
    if isinstance(left, int | float) and isinstance(right, int | float):
        if not math.isfinite(float(left)) or not math.isfinite(float(right)):
            return False
        return abs(float(left) - float(right)) <= numeric_tolerance
    return False


def main(argv: Sequence[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Compare XIC result workbooks.")
    parser.add_argument("left", type=Path)
    parser.add_argument("right", type=Path)
    args = parser.parse_args(argv)

    result = compare_workbooks(args.left, args.right)
    if result.matched:
        print("Workbook compare passed.")
        return 0
    for difference in result.differences:
        print(difference, file=sys.stderr)
    return 1


if __name__ == "__main__":
    raise SystemExit(main())
