"""
02_csv_to_excel.py
將 xic_results.csv 轉換為格式化 Excel。
欄位結構從 config/targets.csv 自動推導，支援任意數量的特徵。
"""

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parent.parent
CONFIG_DIR = ROOT / "config"
OUTPUT_DIR = ROOT / "output"
CSV_PATH = OUTPUT_DIR / "xic_results.csv"
EXCEL_PATH = OUTPUT_DIR / "xic_results.xlsx"

# ── Colour palette (MS1 target pairs, cycling) ────────────────────────────────
_MS1_PALETTES = [
    {"header": "1B5E20", "light": "C8E6C9", "alt": "A5D6A7"},  # green
    {"header": "0D47A1", "light": "BBDEFB", "alt": "90CAF9"},  # blue
    {"header": "4A148C", "light": "E1BEE7", "alt": "CE93D8"},  # purple
    {"header": "004D40", "light": "B2DFDB", "alt": "80CBC4"},  # teal
    {"header": "E65100", "light": "FFE0B2", "alt": "FFCC80"},  # orange
    {"header": "880E4F", "light": "FCE4EC", "alt": "F48FB1"},  # pink
]
_MS2_HEADER = "37474F"  # blue-grey
_MS2_OK = "C8E6C9"  # green  ✓
_MS2_WARN = "FFF9C4"  # yellow ⚠
_MS2_ND = "FFCDD2"  # red    ✗

_SAMPLE_HEADER = "2E4057"
WHITE = "FFFFFF"
GREY = "F5F5F5"
_THIN = Side(style="thin", color="BDBDBD")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
CENTER = Alignment(horizontal="center", vertical="center")
CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
# PS1 outputs ASCII tokens; map them to display symbols here
_NL_DISPLAY = {"OK": "✓", "ND": "✗"}  # WARN_Xppm → "⚠ Xppm" handled in _nl_display()
ND_ERROR = {"ND", "ERROR"}


def _fill(hex6: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex6)


def _apply(cell, **kw) -> None:
    for k, v in kw.items():
        setattr(cell, k, v)


def _header_style(hex6: str) -> dict:
    return dict(
        font=Font(bold=True, color="FFFFFF", size=11),
        fill=_fill(hex6),
        alignment=CENTER_WRAP,
        border=BORDER,
    )


def _safe_float(s: str) -> float | None:
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


# ── Parse targets.csv to understand column types ───────────────────────────────
def _load_column_meta() -> dict:
    """
    Returns dict: csv_col_name → {type, palette, label, nl_col}
    type = 'sample' | 'ms1_rt' | 'ms1_int' | 'ms2_nl'
    nl_col: for MS1 columns, the CSV key of the confirming NL column (or None).
            Linkage by prefix convention: "258.1085_NL116" confirms "258.1085".
    """
    meta: dict[str, dict] = {"SampleName": {"type": "sample"}}
    palette_idx = 0
    ms1_labels: list[str] = []

    targets_path = CONFIG_DIR / "targets.csv"
    if not targets_path.exists():
        return meta

    with open(targets_path, newline="", encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            label = row["label"].strip()
            ms_level = row.get("ms_level", "1").strip()

            if ms_level == "1":
                pal = _MS1_PALETTES[palette_idx % len(_MS1_PALETTES)]
                palette_idx += 1
                ms1_labels.append(label)
                meta[f"{label}_RT"] = {
                    "type": "ms1_rt",
                    "palette": pal,
                    "label": label,
                    "nl_col": None,
                }
                meta[f"{label}_Int"] = {
                    "type": "ms1_int",
                    "palette": pal,
                    "label": label,
                    "nl_col": None,
                }
            else:
                meta[label] = {"type": "ms2_nl", "label": label}
                # Back-link to the MS1 RT/Int entries whose label is a prefix of this NL label
                linked = next(
                    (l for l in ms1_labels if label.startswith(l + "_")), None
                )
                if linked:
                    for suffix in ("_RT", "_Int"):
                        if (k := linked + suffix) in meta:
                            meta[k]["nl_col"] = label

    return meta


def _nl_to_display(value: str) -> str:
    """Convert PS1 ASCII token to Unicode display string."""
    if value == "OK":
        return "✓"
    if value == "ND":
        return "✗"
    if value.startswith("WARN_"):
        return "⚠ " + value[5:]  # WARN_12.3ppm → ⚠ 12.3ppm
    return value


def _nl_cell_fill(value: str) -> PatternFill:
    """Return fill colour based on PS1 ASCII token."""
    if value == "OK":
        return _fill(_MS2_OK)
    if value.startswith("WARN_"):
        return _fill(_MS2_WARN)
    return _fill(_MS2_ND)  # ND, ERROR, anything else


# ── Build main data sheet ──────────────────────────────────────────────────────
def _build_data_sheet(
    ws, rows: list[dict], col_meta: dict, csv_keys: list[str]
) -> None:
    # Header row
    for col_idx, key in enumerate(csv_keys, start=1):
        meta = col_meta.get(key, {"type": "unknown"})
        t = meta.get("type", "")

        if t == "sample":
            label, hex6 = "Sample Name", _SAMPLE_HEADER
        elif t == "ms1_rt":
            label = f"{meta['label']}\nRT (min)"
            hex6 = meta["palette"]["header"]
        elif t == "ms1_int":
            label = f"{meta['label']}\nIntensity"
            hex6 = meta["palette"]["header"]
        elif t == "ms2_nl":
            label = f"{meta['label']}\nNL check"
            hex6 = _MS2_HEADER
        else:
            label, hex6 = key, _SAMPLE_HEADER

        _apply(ws.cell(row=1, column=col_idx, value=label), **_header_style(hex6))

    ws.row_dimensions[1].height = 40

    # Data rows
    for row_idx, row in enumerate(rows, start=2):
        alt = row_idx % 2 == 0
        for col_idx, key in enumerate(csv_keys, start=1):
            raw_val = row.get(key, "")
            meta = col_meta.get(key, {"type": "unknown"})
            t = meta.get("type", "")

            cell = ws.cell(row=row_idx, column=col_idx)

            # MS2 NL columns: convert ASCII token → Unicode + colour
            if t == "ms2_nl":
                cell.value = _nl_to_display(raw_val)
                _apply(
                    cell, fill=_nl_cell_fill(raw_val), alignment=CENTER, border=BORDER
                )
                continue

            # MS1 / sample: ND or ERROR → orange warning cell
            if raw_val in ND_ERROR:
                cell.value = raw_val
                _apply(cell, fill=_fill("FFE0B2"), alignment=CENTER, border=BORDER)
                continue

            # Numeric conversion for MS1 columns
            value = _safe_float(raw_val) if t in ("ms1_rt", "ms1_int") else raw_val
            cell.value = value

            if t == "sample":
                fill = _fill(GREY) if alt else _fill(WHITE)
            elif t == "ms1_rt":
                fill = _fill(
                    meta["palette"]["alt"] if alt else meta["palette"]["light"]
                )
            elif t == "ms1_int":
                fill = _fill(
                    meta["palette"]["alt"] if alt else meta["palette"]["light"]
                )
            else:
                fill = _fill(GREY) if alt else _fill(WHITE)

            _apply(cell, fill=fill, alignment=CENTER, border=BORDER)

            if isinstance(value, float):
                cell.number_format = "0.0000" if t == "ms1_rt" else "#,##0"

    # Column widths
    for col_idx, key in enumerate(csv_keys, start=1):
        t = col_meta.get(key, {}).get("type", "")
        ws.column_dimensions[get_column_letter(col_idx)].width = (
            30 if t == "sample" else 16 if t in ("ms1_rt", "ms1_int") else 18
        )

    ws.freeze_panes = "B2"


# ── Build summary sheet ────────────────────────────────────────────────────────
def _build_summary_sheet(
    ws, rows: list[dict], col_meta: dict, csv_keys: list[str]
) -> None:
    n = len(rows)
    # Collect MS1 data groups and MS2 NL groups
    ms1_keys = [
        k for k in csv_keys if col_meta.get(k, {}).get("type") in ("ms1_rt", "ms1_int")
    ]
    ms2_keys = [k for k in csv_keys if col_meta.get(k, {}).get("type") == "ms2_nl"]

    # Header
    sum_headers = ["Metric"] + ms1_keys + ms2_keys
    for col_idx, h in enumerate(sum_headers, start=1):
        meta = col_meta.get(h, {})
        t = meta.get("type", "")
        hex6 = (
            (meta.get("palette", {}).get("header") or _MS2_HEADER)
            if t != "sample"
            else _SAMPLE_HEADER
        )
        _apply(
            ws.cell(row=1, column=col_idx, value=h if t else h),
            **_header_style(_SAMPLE_HEADER if col_idx == 1 else hex6),
        )
    ws.row_dimensions[1].height = 30

    metrics = ["N (detected)", "N (ND/ERROR)", "Mean", "Min", "Max"]
    for row_idx, metric in enumerate(metrics, start=2):
        ws.cell(row=row_idx, column=1, value=metric)
        _apply(
            ws.cell(row=row_idx, column=1),
            fill=_fill(GREY if row_idx % 2 == 0 else WHITE),
            alignment=CENTER,
            border=BORDER,
        )

        col_offset = 2
        for key in ms1_keys:
            nl_col = col_meta.get(key, {}).get("nl_col")
            if nl_col:
                # Only include NL-confirmed rows (OK or WARN); ND = false positive, excluded
                _src = [
                    r
                    for r in rows
                    if r.get(nl_col, "") == "OK"
                    or r.get(nl_col, "").startswith("WARN_")
                ]
            else:
                _src = rows
            vals = [v for r in _src if (v := _safe_float(r.get(key, ""))) is not None]
            cell = ws.cell(row=row_idx, column=col_offset)
            t = col_meta.get(key, {}).get("type", "")
            if metric == "N (detected)":
                val = len(vals)
            elif metric == "N (ND/ERROR)":
                val = n - len(vals)
            elif metric == "Mean":
                val = (
                    f"{sum(vals) / len(vals):.4f}"
                    if vals and t == "ms1_rt"
                    else (f"{sum(vals) / len(vals):,.0f}" if vals else "—")
                )
            elif metric == "Min":
                val = (
                    f"{min(vals):.4f}"
                    if vals and t == "ms1_rt"
                    else (f"{min(vals):,.0f}" if vals else "—")
                )
            elif metric == "Max":
                val = (
                    f"{max(vals):.4f}"
                    if vals and t == "ms1_rt"
                    else (f"{max(vals):,.0f}" if vals else "—")
                )
            else:
                val = "—"
            _apply(
                cell,
                value=val,
                fill=_fill(GREY if row_idx % 2 == 0 else WHITE),
                alignment=CENTER,
                border=BORDER,
            )
            col_offset += 1

        for key in ms2_keys:
            vals = [r.get(key, "") for r in rows]  # ASCII tokens: OK / WARN_Xppm / ND
            cell = ws.cell(row=row_idx, column=col_offset)
            if metric == "N (detected)":
                val = sum(1 for v in vals if v == "OK")
            elif metric == "N (ND/ERROR)":
                val = sum(1 for v in vals if v == "ND")
            elif metric == "Mean":
                val = f"⚠: {sum(1 for v in vals if v.startswith('WARN_'))}"
            else:
                val = "—"
            _apply(
                cell,
                value=val,
                fill=_fill(GREY if row_idx % 2 == 0 else WHITE),
                alignment=CENTER,
                border=BORDER,
            )
            col_offset += 1

    for col_idx in range(1, len(sum_headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20


# ── Entry point ────────────────────────────────────────────────────────────────
def main() -> None:
    with open(CSV_PATH, newline="", encoding="utf-8-sig") as f:
        rows = list(csv.DictReader(f))

    if not rows:
        print("CSV is empty.")
        return

    csv_keys = list(rows[0].keys())
    col_meta = _load_column_meta()

    wb = Workbook()
    ws_data = wb.active
    ws_data.title = "XIC Results"
    _build_data_sheet(ws_data, rows, col_meta, csv_keys)

    ws_sum = wb.create_sheet("Summary")
    _build_summary_sheet(ws_sum, rows, col_meta, csv_keys)

    wb.save(EXCEL_PATH)

    ms1_cols = [k for k in csv_keys if col_meta.get(k, {}).get("type") == "ms1_rt"]
    ms2_cols = [k for k in csv_keys if col_meta.get(k, {}).get("type") == "ms2_nl"]
    print(f"Saved : {EXCEL_PATH}")
    print(f"Rows  : {len(rows)}")
    for k in ms1_cols:
        nl_col = col_meta.get(k, {}).get("nl_col")
        if nl_col:
            n_det = sum(
                1
                for r in rows
                if _safe_float(r.get(k, "")) is not None
                and (r.get(nl_col, "") == "OK" or r.get(nl_col, "").startswith("WARN_"))
            )
            print(
                f"  {k.replace('_RT', '')} detected (NL confirmed): {n_det}/{len(rows)}"
            )
        else:
            n_det = sum(1 for r in rows if _safe_float(r.get(k, "")) is not None)
            print(f"  {k.replace('_RT', '')} detected: {n_det}/{len(rows)}")
    for k in ms2_cols:
        n_ok = sum(1 for r in rows if r.get(k, "") == "OK")
        n_warn = sum(1 for r in rows if r.get(k, "").startswith("WARN_"))
        n_nd = len(rows) - n_ok - n_warn
        print(f"  {k}  OK:{n_ok}  WARN:{n_warn}  ND:{n_nd}")


if __name__ == "__main__":
    main()
