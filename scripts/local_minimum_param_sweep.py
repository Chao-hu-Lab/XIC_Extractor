from __future__ import annotations

import argparse
import csv
import multiprocessing
import os
import shutil
import sys
from collections.abc import Sequence
from dataclasses import dataclass
from itertools import product
from pathlib import Path
from statistics import median
from typing import Callable, Iterable

if __package__ in {None, ""}:
    sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from xic_extractor import extractor
from xic_extractor.config import ConfigError, Target, load_config
from xic_extractor.raw_reader import RawReaderError
from xic_extractor.settings_schema import (
    CANONICAL_SETTINGS_DEFAULTS,
    CANONICAL_SETTINGS_DESCRIPTIONS,
)


@dataclass(frozen=True)
class ManualTruthRow:
    sheet: str
    sample_name: str
    target: str
    manual_rt: float | None
    manual_height: float | None
    manual_area: float | None
    manual_width: float | None
    manual_shape: str


@dataclass(frozen=True)
class ProgramPeakRow:
    sample_name: str
    target: str
    is_istd: bool
    rt: float | None
    height: float | None
    area: float | None
    detected: bool


@dataclass(frozen=True)
class ParameterSet:
    name: str
    settings_overrides: dict[str, str]


@dataclass(frozen=True)
class SweepCase:
    name: str
    sample_name: str
    raw_path: Path
    targets_path: Path


@dataclass(frozen=True)
class PerTargetScoreRow:
    name: str
    sample_name: str
    target: str
    is_istd: bool
    manual_rt: float | None
    program_rt: float | None
    rt_abs_delta_min: float | None
    manual_height: float | None
    program_height: float | None
    height_abs_pct_error: float | None
    manual_area: float | None
    program_area: float | None
    area_abs_pct_error: float | None
    missing_peak: bool
    manual_shape: str


@dataclass(frozen=True)
class ParameterSetScore:
    name: str
    settings_overrides: dict[str, str]
    detected_rows: int
    scored_rows: int
    missing_manual_peaks: int
    istd_misses: int
    area_median_abs_pct_error: float | None
    height_median_abs_pct_error: float | None
    rt_median_abs_delta_min: float | None
    rt_max_abs_delta_min: float | None
    area_within_10pct: int
    area_within_20pct: int
    large_area_misses: int
    per_target_rows: list[PerTargetScoreRow]


@dataclass(frozen=True)
class SweepResult:
    scores: list[ParameterSetScore]


SweepRunner = Callable[[ParameterSet], list[ProgramPeakRow]]


_STATIC_LOCAL_MINIMUM_PARAMS = {
    "resolver_min_relative_height": "0.0",
    "resolver_min_absolute_height": "25.0",
    "resolver_peak_duration_min": "0.0",
    "resolver_peak_duration_max": "2.0",
}

_CALIBRATION_V1_OVERRIDES = (
    ("local_minimum_duration_1p5", {"resolver_peak_duration_max": "1.5"}),
    ("local_minimum_duration_2p0", {"resolver_peak_duration_max": "2.0"}),
    ("local_minimum_duration_3p0", {"resolver_peak_duration_max": "3.0"}),
    (
        "local_minimum_search_0p05",
        {"resolver_min_search_range_min": "0.05"},
    ),
    (
        "local_minimum_search_0p04",
        {"resolver_min_search_range_min": "0.04"},
    ),
    (
        "local_minimum_search_0p05_duration_2p0",
        {
            "resolver_min_search_range_min": "0.05",
            "resolver_peak_duration_max": "2.0",
        },
    ),
    (
        "local_minimum_search_0p04_duration_2p0",
        {
            "resolver_min_search_range_min": "0.04",
            "resolver_peak_duration_max": "2.0",
        },
    ),
    (
        "local_minimum_search_0p05_duration_1p5",
        {
            "resolver_min_search_range_min": "0.05",
            "resolver_peak_duration_max": "1.5",
        },
    ),
    (
        "local_minimum_search_0p05_duration_2p0_edge_1p5",
        {
            "resolver_min_search_range_min": "0.05",
            "resolver_peak_duration_max": "2.0",
            "resolver_min_ratio_top_edge": "1.5",
        },
    ),
)

_CALIBRATION_V2_OVERRIDES = (
    (
        "local_minimum_min_duration_0p02",
        {"resolver_peak_duration_min": "0.02"},
    ),
    (
        "local_minimum_min_duration_0p03",
        {"resolver_peak_duration_min": "0.03"},
    ),
    (
        "local_minimum_rel_height_0p01",
        {"resolver_min_relative_height": "0.01"},
    ),
    (
        "local_minimum_rel_height_0p02",
        {"resolver_min_relative_height": "0.02"},
    ),
    (
        "local_minimum_rel_height_0p03",
        {"resolver_min_relative_height": "0.03"},
    ),
    (
        "local_minimum_min_duration_0p02_rel_height_0p01",
        {
            "resolver_peak_duration_min": "0.02",
            "resolver_min_relative_height": "0.01",
        },
    ),
    (
        "local_minimum_min_duration_0p02_rel_height_0p02",
        {
            "resolver_peak_duration_min": "0.02",
            "resolver_min_relative_height": "0.02",
        },
    ),
    (
        "local_minimum_min_duration_0p03_rel_height_0p01",
        {
            "resolver_peak_duration_min": "0.03",
            "resolver_min_relative_height": "0.01",
        },
    ),
)

_GRID_VALUES = {
    "quick": {
        "resolver_chrom_threshold": ("0.03", "0.05"),
        "resolver_min_search_range_min": ("0.05", "0.08"),
        "resolver_min_ratio_top_edge": ("1.5", "1.7"),
        "resolver_min_scans": ("3", "5"),
    },
    "standard": {
        "resolver_chrom_threshold": ("0.03", "0.05", "0.08"),
        "resolver_min_search_range_min": ("0.05", "0.08", "0.12"),
        "resolver_min_ratio_top_edge": ("1.3", "1.5", "1.7", "2.0"),
        "resolver_min_scans": ("3", "5"),
    },
}

GRID_CHOICES = (*_GRID_VALUES.keys(), "calibration-v1", "calibration-v2")

_THIN = Side(style="thin", color="BDBDBD")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center")
_HEADER_FILL = PatternFill("solid", fgColor="37474F")
_FAIL_FILL = PatternFill("solid", fgColor="FFCDD2")


def build_parameter_sets(*, grid: str) -> list[ParameterSet]:
    if grid not in GRID_CHOICES:
        raise ValueError(f"Unknown grid: {grid}")

    parameter_sets = [
        ParameterSet("legacy_savgol", {"resolver_mode": "legacy_savgol"}),
        ParameterSet("local_minimum_current", {"resolver_mode": "local_minimum"}),
    ]
    if grid == "calibration-v1":
        parameter_sets.extend(
            ParameterSet(
                name,
                {
                    "resolver_mode": "local_minimum",
                    **_STATIC_LOCAL_MINIMUM_PARAMS,
                    **overrides,
                },
            )
            for name, overrides in _CALIBRATION_V1_OVERRIDES
        )
        return parameter_sets

    if grid == "calibration-v2":
        parameter_sets.extend(
            ParameterSet(
                name,
                {
                    "resolver_mode": "local_minimum",
                    **_STATIC_LOCAL_MINIMUM_PARAMS,
                    **overrides,
                },
            )
            for name, overrides in _CALIBRATION_V2_OVERRIDES
        )
        return parameter_sets

    grid_values = _GRID_VALUES[grid]
    keys = list(grid_values.keys())
    for idx, values in enumerate(product(*(grid_values[key] for key in keys)), start=1):
        settings = {
            "resolver_mode": "local_minimum",
            **_STATIC_LOCAL_MINIMUM_PARAMS,
            **dict(zip(keys, values)),
        }
        parameter_sets.append(
            ParameterSet(f"local_minimum_grid_{idx:03d}", settings)
        )
    return parameter_sets


def main(argv: Sequence[str] | None = None) -> int:
    multiprocessing.freeze_support()
    args = _parse_args(argv)
    output_dir = args.output_dir.resolve()
    try:
        truth_rows = read_manual_truth(args.manual_workbook.resolve())
        parameter_sets = _apply_execution_overrides(
            build_parameter_sets(grid=args.grid),
            parallel_mode=args.parallel_mode,
            parallel_workers=args.parallel_workers,
        )
        cases = _build_cases(args)
        result = run_sweep(
            truth_rows,
            parameter_sets,
            lambda parameter_set: _run_parameter_set(
                parameter_set, cases, output_dir
            ),
        )
        output_path = write_sweep_workbook(
            output_dir / "local_minimum_param_sweep_summary.xlsx",
            result.scores,
        )
    except (ConfigError, RawReaderError, OSError, ValueError) as exc:
        print(str(exc), file=sys.stderr)
        return 2

    _print_summary(output_path, result)
    return 0


def read_manual_truth(path: Path) -> list[ManualTruthRow]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    rows: list[ManualTruthRow] = []
    for sheet_name in ("DNA", "RNA"):
        if sheet_name not in workbook.sheetnames:
            continue
        worksheet = workbook[sheet_name]
        rows.extend(_read_sheet_truth(worksheet))
    workbook.close()
    return rows


def _parse_args(argv: Sequence[str] | None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Sweep local_minimum resolver parameters against manual truth."
    )
    parser.add_argument("--manual-workbook", type=Path, required=True)
    parser.add_argument("--nosplit-raw", type=Path, required=True)
    parser.add_argument("--split-raw", type=Path, required=True)
    parser.add_argument("--nosplit-targets", type=Path, required=True)
    parser.add_argument("--split-targets", type=Path, required=True)
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path("output/local_minimum_param_sweep_manual"),
    )
    parser.add_argument(
        "--grid",
        choices=GRID_CHOICES,
        default="quick",
    )
    parser.add_argument(
        "--parallel-mode",
        choices=("serial", "process"),
        default=None,
        help="Optional execution backend override for every sweep run.",
    )
    parser.add_argument(
        "--parallel-workers",
        type=_positive_int,
        default=None,
        help="Optional process worker override for every sweep run.",
    )
    return parser.parse_args(argv)


def _apply_execution_overrides(
    parameter_sets: list[ParameterSet],
    *,
    parallel_mode: str | None,
    parallel_workers: int | None,
) -> list[ParameterSet]:
    overrides: dict[str, str] = {}
    if parallel_mode is not None:
        overrides["parallel_mode"] = parallel_mode
    if parallel_workers is not None:
        overrides["parallel_workers"] = str(parallel_workers)
    if not overrides:
        return parameter_sets
    return [
        ParameterSet(
            parameter_set.name,
            {**parameter_set.settings_overrides, **overrides},
        )
        for parameter_set in parameter_sets
    ]


def _positive_int(value: str) -> int:
    parsed = int(value)
    if parsed < 1:
        raise argparse.ArgumentTypeError("parallel-workers must be >= 1")
    return parsed


def _build_cases(args: argparse.Namespace) -> list[SweepCase]:
    return [
        SweepCase(
            name="nosplit",
            sample_name=args.nosplit_raw.resolve().stem,
            raw_path=args.nosplit_raw.resolve(),
            targets_path=args.nosplit_targets.resolve(),
        ),
        SweepCase(
            name="split",
            sample_name=args.split_raw.resolve().stem,
            raw_path=args.split_raw.resolve(),
            targets_path=args.split_targets.resolve(),
        ),
    ]


def _run_parameter_set(
    parameter_set: ParameterSet,
    cases: list[SweepCase],
    output_dir: Path,
) -> list[ProgramPeakRow]:
    rows: list[ProgramPeakRow] = []
    for case in cases:
        rows.extend(_run_case(parameter_set, case, output_dir))
    return rows


def _run_case(
    parameter_set: ParameterSet,
    case: SweepCase,
    output_dir: Path,
) -> list[ProgramPeakRow]:
    data_dir = _stage_raw_case(case, output_dir / "staged_raw" / case.name)
    config_dir = output_dir / "runs" / parameter_set.name / case.name / "config"
    _write_case_config(
        config_dir,
        data_dir=data_dir,
        targets_path=case.targets_path,
        settings_overrides=parameter_set.settings_overrides,
    )
    config, targets = load_config(config_dir)
    run_output = extractor.run(config, targets)
    return _collect_program_rows(run_output, targets)


def _stage_raw_case(case: SweepCase, data_dir: Path) -> Path:
    if not case.raw_path.exists():
        raise FileNotFoundError(f"{case.raw_path}: raw file is missing")
    data_dir.mkdir(parents=True, exist_ok=True)
    staged_raw = data_dir / case.raw_path.name
    if staged_raw.exists():
        return data_dir
    try:
        os.link(case.raw_path, staged_raw)
    except OSError:
        shutil.copy2(case.raw_path, staged_raw)
    return data_dir


def _write_case_config(
    config_dir: Path,
    *,
    data_dir: Path,
    targets_path: Path,
    settings_overrides: dict[str, str],
) -> None:
    if not targets_path.exists():
        raise FileNotFoundError(f"{targets_path}: targets file is missing")
    config_dir.mkdir(parents=True, exist_ok=True)
    shutil.copy2(targets_path, config_dir / "targets.csv")
    settings = {
        **CANONICAL_SETTINGS_DEFAULTS,
        "data_dir": str(data_dir),
        **settings_overrides,
    }
    _write_settings_csv(config_dir / "settings.csv", settings)


def _write_settings_csv(path: Path, settings: dict[str, str]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=["key", "value", "description"])
        writer.writeheader()
        for key, value in settings.items():
            writer.writerow(
                {
                    "key": key,
                    "value": value,
                    "description": CANONICAL_SETTINGS_DESCRIPTIONS.get(key, ""),
                }
            )


def _collect_program_rows(
    run_output: extractor.RunOutput,
    targets: list[Target],
) -> list[ProgramPeakRow]:
    rows: list[ProgramPeakRow] = []
    for file_result in run_output.file_results:
        for target in targets:
            result = file_result.results.get(target.label)
            peak = result.peak if result is not None else None
            rows.append(
                ProgramPeakRow(
                    sample_name=file_result.sample_name,
                    target=target.label,
                    is_istd=target.is_istd,
                    rt=result.reported_rt if result is not None else None,
                    height=(
                        peak.intensity_smoothed
                        if peak is not None and peak.intensity_smoothed is not None
                        else peak.intensity if peak is not None else None
                    ),
                    area=peak.area if peak is not None else None,
                    detected=peak is not None,
                )
            )
    return rows


def _print_summary(output_path: Path, result: SweepResult) -> None:
    ranked = _rank_scores(result.scores)
    best = ranked[0] if ranked else None
    print(f"Summary workbook: {output_path}")
    if best is not None:
        print(
            "Best: "
            f"{best.name}, area_median_abs_pct_error="
            f"{best.area_median_abs_pct_error}, "
            f"missing_manual_peaks={best.missing_manual_peaks}"
        )


def run_sweep(
    truth_rows: list[ManualTruthRow],
    parameter_sets: list[ParameterSet],
    runner: SweepRunner,
    *,
    istd_targets: set[str] | None = None,
) -> SweepResult:
    scores: list[ParameterSetScore] = []
    for parameter_set in parameter_sets:
        program_rows = runner(parameter_set)
        scores.append(
            score_parameter_set(
                parameter_set.name,
                parameter_set.settings_overrides,
                truth_rows,
                program_rows,
                istd_targets=istd_targets,
            )
        )
    return SweepResult(scores=scores)


def write_sweep_workbook(output_path: Path, scores: list[ParameterSetScore]) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()

    summary = workbook.active
    summary.title = "Summary"
    _write_summary_sheet(summary, scores)

    per_target = workbook.create_sheet("PerTarget")
    _write_per_target_sheet(per_target, scores)

    failures = workbook.create_sheet("Failures")
    _write_failures_sheet(failures, scores)

    run_config = workbook.create_sheet("RunConfig")
    _write_run_config_sheet(run_config, scores)

    workbook.save(output_path)
    return output_path


def score_parameter_set(
    name: str,
    settings_overrides: dict[str, str],
    truth_rows: list[ManualTruthRow],
    program_rows: list[ProgramPeakRow],
    *,
    istd_targets: set[str] | None = None,
) -> ParameterSetScore:
    program_by_key = {
        (row.sample_name, row.target): row
        for row in program_rows
    }
    istd_target_names = set(istd_targets or ())
    per_target_rows: list[PerTargetScoreRow] = []
    for truth in truth_rows:
        program = program_by_key.get((truth.sample_name, truth.target))
        detected = _program_peak_detected(program)
        is_istd = (
            truth.target in istd_target_names
            or (program.is_istd if program is not None else False)
        )
        rt_abs_delta = _abs_delta(program.rt, truth.manual_rt) if detected else None
        height_abs_pct_error = (
            _abs_pct_error(program.height, truth.manual_height) if detected else None
        )
        area_abs_pct_error = (
            _abs_pct_error(program.area, truth.manual_area) if detected else None
        )
        per_target_rows.append(
            PerTargetScoreRow(
                name=name,
                sample_name=truth.sample_name,
                target=truth.target,
                is_istd=is_istd,
                manual_rt=truth.manual_rt,
                program_rt=program.rt if detected else None,
                rt_abs_delta_min=_round_optional(rt_abs_delta),
                manual_height=truth.manual_height,
                program_height=program.height if detected else None,
                height_abs_pct_error=_round_optional(height_abs_pct_error),
                manual_area=truth.manual_area,
                program_area=program.area if detected else None,
                area_abs_pct_error=_round_optional(area_abs_pct_error),
                missing_peak=not detected,
                manual_shape=truth.manual_shape,
            )
        )

    area_errors = [
        row.area_abs_pct_error
        for row in per_target_rows
        if row.area_abs_pct_error is not None
    ]
    height_errors = [
        row.height_abs_pct_error
        for row in per_target_rows
        if row.height_abs_pct_error is not None
    ]
    rt_deltas = [
        row.rt_abs_delta_min
        for row in per_target_rows
        if row.rt_abs_delta_min is not None
    ]
    missing_rows = [row for row in per_target_rows if row.missing_peak]
    return ParameterSetScore(
        name=name,
        settings_overrides=settings_overrides,
        detected_rows=sum(not row.missing_peak for row in per_target_rows),
        scored_rows=len(area_errors),
        missing_manual_peaks=len(missing_rows),
        istd_misses=sum(row.is_istd for row in missing_rows),
        area_median_abs_pct_error=_rounded_median(area_errors),
        height_median_abs_pct_error=_rounded_median(height_errors),
        rt_median_abs_delta_min=_rounded_median(rt_deltas),
        rt_max_abs_delta_min=round(max(rt_deltas), 6) if rt_deltas else None,
        area_within_10pct=sum(error <= 0.10 + 1e-12 for error in area_errors),
        area_within_20pct=sum(error <= 0.20 + 1e-12 for error in area_errors),
        large_area_misses=sum(error > 0.20 + 1e-12 for error in area_errors),
        per_target_rows=per_target_rows,
    )


def _read_sheet_truth(worksheet: Worksheet) -> list[ManualTruthRow]:
    values = list(worksheet.iter_rows(values_only=True))
    if len(values) < 3:
        return []

    raw_blocks = _iter_raw_blocks(values[0])
    rows: list[ManualTruthRow] = []
    for row in values[2:]:
        target = _cell_text(_row_value(row, 1))
        if not target:
            continue
        for sample_name, start_idx in raw_blocks:
            manual_rt = _safe_float(_row_value(row, start_idx))
            manual_height = _safe_float(_row_value(row, start_idx + 1))
            manual_area = _safe_float(_row_value(row, start_idx + 2))
            manual_width = _safe_float(_row_value(row, start_idx + 3))
            manual_shape = _cell_text(_row_value(row, start_idx + 4))
            if manual_rt is None and manual_area is None:
                continue
            rows.append(
                ManualTruthRow(
                    sheet=worksheet.title,
                    sample_name=sample_name,
                    target=target,
                    manual_rt=manual_rt,
                    manual_height=manual_height,
                    manual_area=manual_area,
                    manual_width=manual_width,
                    manual_shape=manual_shape,
                )
            )
    return rows


def _iter_raw_blocks(header_row: Iterable[object]) -> list[tuple[str, int]]:
    blocks: list[tuple[str, int]] = []
    for idx, value in enumerate(header_row):
        if idx < 3:
            continue
        sample_name = _cell_text(value)
        if sample_name:
            blocks.append((_sample_stem(sample_name), idx))
    return blocks


def _sample_stem(value: str) -> str:
    return Path(value).stem


def _row_value(row: tuple[object, ...], idx: int) -> object | None:
    return row[idx] if idx < len(row) else None


def _safe_float(value: object | None) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _cell_text(value: object | None) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _write_summary_sheet(worksheet: Worksheet, scores: list[ParameterSetScore]) -> None:
    headers = [
        "Rank",
        "Name",
        "GuardrailStatus",
        "DetectedRows",
        "ScoredRows",
        "MissingManualPeaks",
        "ISTDMisses",
        "AreaMedianAbsPctError",
        "AreaWithin10Pct",
        "AreaWithin20Pct",
        "LargeAreaMisses",
        "HeightMedianAbsPctError",
        "RTMedianAbsDeltaMin",
        "RTMaxAbsDeltaMin",
    ]
    _write_header(worksheet, headers)
    for rank, score in enumerate(_rank_scores(scores), start=1):
        _write_row(
            worksheet,
            rank + 1,
            [
                rank,
                score.name,
                "PASS" if _guardrail_passes(score) else "FAIL",
                score.detected_rows,
                score.scored_rows,
                score.missing_manual_peaks,
                score.istd_misses,
                score.area_median_abs_pct_error,
                score.area_within_10pct,
                score.area_within_20pct,
                score.large_area_misses,
                score.height_median_abs_pct_error,
                score.rt_median_abs_delta_min,
                score.rt_max_abs_delta_min,
            ],
            failed=not _guardrail_passes(score),
        )
    _autosize(worksheet, len(headers))


def _write_per_target_sheet(
    worksheet: Worksheet, scores: list[ParameterSetScore]
) -> None:
    headers = [
        "Name",
        "SampleName",
        "Target",
        "ISTD",
        "ManualRT",
        "ProgramRT",
        "RTAbsDeltaMin",
        "ManualHeight",
        "ProgramHeight",
        "HeightAbsPctError",
        "ManualArea",
        "ProgramArea",
        "AreaAbsPctError",
        "MissingPeak",
        "ManualShape",
    ]
    _write_header(worksheet, headers)
    row_idx = 2
    for score in scores:
        for row in score.per_target_rows:
            _write_row(
                worksheet,
                row_idx,
                [
                    row.name,
                    row.sample_name,
                    row.target,
                    row.is_istd,
                    row.manual_rt,
                    row.program_rt,
                    row.rt_abs_delta_min,
                    row.manual_height,
                    row.program_height,
                    row.height_abs_pct_error,
                    row.manual_area,
                    row.program_area,
                    row.area_abs_pct_error,
                    row.missing_peak,
                    row.manual_shape,
                ],
                failed=_target_row_failed(row),
            )
            row_idx += 1
    _autosize(worksheet, len(headers))


def _write_failures_sheet(
    worksheet: Worksheet, scores: list[ParameterSetScore]
) -> None:
    headers = ["Name", "SampleName", "Target", "Issue", "MetricValue", "Limit"]
    _write_header(worksheet, headers)
    row_idx = 2
    for score in scores:
        for row in score.per_target_rows:
            for issue, metric_value, limit in _failure_records(row):
                _write_row(
                    worksheet,
                    row_idx,
                    [
                        score.name,
                        row.sample_name,
                        row.target,
                        issue,
                        metric_value,
                        limit,
                    ],
                    failed=True,
                )
                row_idx += 1
    _autosize(worksheet, len(headers))


def _write_run_config_sheet(
    worksheet: Worksheet, scores: list[ParameterSetScore]
) -> None:
    headers = ["Name", "Key", "Value"]
    _write_header(worksheet, headers)
    row_idx = 2
    for score in scores:
        for key, value in sorted(score.settings_overrides.items()):
            _write_row(worksheet, row_idx, [score.name, key, value], failed=False)
            row_idx += 1
    _autosize(worksheet, len(headers))


def _rank_scores(scores: list[ParameterSetScore]) -> list[ParameterSetScore]:
    return sorted(
        scores,
        key=lambda score: (
            not _guardrail_passes(score),
            score.missing_manual_peaks,
            _none_last(score.area_median_abs_pct_error),
            _none_last(score.rt_median_abs_delta_min),
            score.name,
        ),
    )


def _guardrail_passes(score: ParameterSetScore) -> bool:
    return (
        score.missing_manual_peaks == 0
        and score.istd_misses == 0
        and score.large_area_misses == 0
        and (
            score.rt_median_abs_delta_min is None
            or score.rt_median_abs_delta_min <= 0.05
        )
        and (score.rt_max_abs_delta_min is None or score.rt_max_abs_delta_min <= 0.20)
    )


def _target_row_failed(row: PerTargetScoreRow) -> bool:
    return any(_failure_records(row))


def _failure_records(row: PerTargetScoreRow) -> list[tuple[str, object, str]]:
    failures: list[tuple[str, object, str]] = []
    if row.missing_peak:
        failures.append(("MISSING_PEAK", "", "manual peak must be detected"))
    if row.rt_abs_delta_min is not None and row.rt_abs_delta_min > 0.20:
        failures.append(("RT_MAX_DELTA", row.rt_abs_delta_min, "<=0.20"))
    if row.area_abs_pct_error is not None and row.area_abs_pct_error > 0.20:
        failures.append(("AREA_ERROR", row.area_abs_pct_error, "<=0.20"))
    return failures


def _none_last(value: float | None) -> float:
    return float("inf") if value is None else value


def _write_header(worksheet: Worksheet, headers: list[str]) -> None:
    for col_idx, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=col_idx, value=header)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border = _BORDER
    worksheet.freeze_panes = "A2"


def _write_row(
    worksheet: Worksheet, row_idx: int, values: list[object], *, failed: bool
) -> None:
    for col_idx, value in enumerate(values, start=1):
        cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
        cell.font = Font(name="Arial")
        cell.alignment = _CENTER
        cell.border = _BORDER
        if failed:
            cell.fill = _FAIL_FILL


def _autosize(worksheet: Worksheet, n_cols: int) -> None:
    for col_idx in range(1, n_cols + 1):
        letter = get_column_letter(col_idx)
        width = max(
            12,
            min(
                60,
                max(
                    len(str(worksheet.cell(row=row_idx, column=col_idx).value or ""))
                    for row_idx in range(1, worksheet.max_row + 1)
                )
                + 2,
            ),
        )
        worksheet.column_dimensions[letter].width = width


def _program_peak_detected(row: ProgramPeakRow | None) -> bool:
    return bool(
        row is not None
        and row.detected
        and row.rt is not None
        and row.area is not None
    )


def _abs_delta(left: float | None, right: float | None) -> float | None:
    if left is None or right is None:
        return None
    return abs(left - right)


def _abs_pct_error(value: float | None, truth: float | None) -> float | None:
    if value is None or truth is None or truth <= 0:
        return None
    return abs(value - truth) / truth


def _rounded_median(values: list[float]) -> float | None:
    if not values:
        return None
    return round(median(values), 6)


def _round_optional(value: float | None) -> float | None:
    if value is None:
        return None
    return round(value, 6)


if __name__ == "__main__":
    raise SystemExit(main())
