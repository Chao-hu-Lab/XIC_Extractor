"""
將 xic_results.csv 轉換為格式化 Excel。

欄位結構從 Target 設定推導，支援 apex RT、raw intensity、integrated
area、peak boundary、四態 neutral-loss 結果與 diagnostics sheet。
"""

from __future__ import annotations

import csv
import statistics
from collections.abc import Iterable
from datetime import datetime
from pathlib import Path
from typing import overload

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from xic_extractor.config import ExtractionConfig, Target, load_config
from xic_extractor.output import metadata
from xic_extractor.output.review_metrics import ReviewMetrics, build_review_metrics
from xic_extractor.output.schema import (
    DIAGNOSTIC_HEADERS as _DIAGNOSTIC_HEADERS,
)
from xic_extractor.output.schema import (
    LONG_ADVANCED_HEADERS as _ADVANCED_HEADERS,
)
from xic_extractor.output.schema import (
    LONG_HEADERS as _LONG_HEADERS,
)
from xic_extractor.sample_groups import classify_sample_group

_MS2_HEADER = "37474F"
_MS2_OK = "C8E6C9"
_MS2_WARN = "FFF9C4"
_MS2_FAIL = "FFCDD2"
_MS2_NO_MS2 = "E0E0E0"
_SAMPLE_HEADER = "2E4057"
_OVERVIEW_HEADER_FILL = "1F4E5F"
_DAILY_REVIEW_TAB = "1F4E5F"
_RESULT_TAB = "5B7C99"
_TECHNICAL_TAB = "B0BEC5"
WHITE = "FFFFFF"
GREY = "F5F5F5"
_THIN = Side(style="thin", color="BDBDBD")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
CENTER = Alignment(horizontal="center", vertical="center")
CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
ND_ERROR = {"ND", "ERROR"}
_FORMULA_PREFIXES = ("=", "+", "-", "@")
_TARGETS_HEADERS = [
    "Label",
    "Role",
    "ISTD Pair",
    "m/z",
    "RT min",
    "RT max",
    "ppm tol",
    "NL (Da)",
    "Expected product m/z",
    "NL ppm warn",
    "NL ppm max",
]
_SUMMARY_HEADERS = [
    "Target",
    "Role",
    "ISTD Pair",
    "Flagged Rows",
    "Flagged %",
    "MS2/NL Flags",
    "Low Confidence Rows",
    "Detected",
    "Total",
    "Detection %",
    "Mean RT",
    "Median Area (detected)",
    "Area / ISTD ratio (paired detected)",
    "NL OK",
    "NL WARN",
    "NL FAIL",
    "NO MS2",
    "RT Delta vs ISTD",
    "Confidence HIGH",
    "Confidence MEDIUM",
    "Confidence LOW",
    "Confidence VERY_LOW",
]
_REVIEW_HEADERS = [
    "Priority",
    "Sample",
    "Target",
    "Role",
    "Status",
    "Why",
    "RT",
    "Area",
    "Action",
    "Issue Count",
    "Evidence",
]


def _fill(hex6: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex6)


def _apply(cell, **kw: object) -> None:
    for key, value in kw.items():
        setattr(cell, key, value)


def _header_style(hex6: str) -> dict[str, object]:
    return {
        "font": Font(bold=True, color="FFFFFF", size=11),
        "fill": _fill(hex6),
        "alignment": CENTER_WRAP,
        "border": BORDER,
    }


def _safe_float(value: object) -> float | None:
    if not isinstance(value, str | int | float):
        return None
    try:
        return float(value)
    except ValueError:
        return None


def _excel_text(value: object) -> object:
    if not isinstance(value, str):
        return value
    if value.startswith(_FORMULA_PREFIXES):
        return "'" + value
    return value


def _nl_to_display(value: str) -> str:
    if value == "OK":
        return "✓"
    if value.startswith("WARN_"):
        return "⚠ " + value[5:]
    if value == "NL_FAIL":
        return "✗ NL"
    if value == "NO_MS2":
        return "— MS2"
    if value == "ND":
        return "✗"
    return value


def _nl_cell_fill(value: str) -> PatternFill:
    if value == "OK":
        return _fill(_MS2_OK)
    if value.startswith("WARN_"):
        return _fill(_MS2_WARN)
    if value == "NO_MS2":
        return _fill(_MS2_NO_MS2)
    return _fill(_MS2_FAIL)


def _long_cell_value(header: str, raw_val: str) -> object:
    if header in {"SampleName", "Target", "ISTD Pair", "Confidence", "Reason"}:
        return _excel_text(raw_val)
    if header == "NL":
        return _nl_to_display(raw_val) if raw_val else ""
    if header in {"RT", "Area", "Int", "PeakStart", "PeakEnd", "PeakWidth"}:
        if raw_val in ND_ERROR:
            return raw_val
        parsed = _safe_float(raw_val)
        return parsed if parsed is not None else raw_val
    return raw_val


def _long_cell_fill(header: str, raw_val: str, alt: bool) -> PatternFill:
    if raw_val in ND_ERROR:
        return _fill("FFE0B2" if raw_val == "ND" else "FF7043")
    if header == "NL" and raw_val:
        return _nl_cell_fill(raw_val)
    if header in _ADVANCED_HEADERS:
        return _fill("ECEFF1")
    return _fill(GREY if alt else WHITE)


def _long_number_format(header: str) -> str:
    if header in {"RT", "PeakStart", "PeakEnd"}:
        return "0.0000"
    if header in {"Area", "Int"}:
        return "0.00E+00"
    if header == "PeakWidth":
        return "0.0000"
    return "#,##0"


def _long_column_width(header: str) -> int:
    return {
        "SampleName": 28,
        "Group": 14,
        "Target": 24,
        "Role": 12,
        "ISTD Pair": 24,
        "RT": 12,
        "Area": 16,
        "NL": 14,
        "Int": 14,
        "PeakStart": 14,
        "PeakEnd": 14,
        "PeakWidth": 14,
        "Confidence": 14,
        "Reason": 42,
    }[header]


def _build_data_sheet(ws, rows: list[dict[str, str]]) -> None:
    for col_idx, header in enumerate(_LONG_HEADERS, start=1):
        color = _MS2_HEADER if header in _ADVANCED_HEADERS else _SAMPLE_HEADER
        _apply(ws.cell(row=1, column=col_idx, value=header), **_header_style(color))
    ws.row_dimensions[1].height = 30

    for row_idx, row in enumerate(rows, start=2):
        alt = row_idx % 2 == 0
        for col_idx, header in enumerate(_LONG_HEADERS, start=1):
            raw_val = row.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx)
            value = _long_cell_value(header, raw_val)
            cell.value = value
            _apply(
                cell,
                fill=_long_cell_fill(header, raw_val, alt),
                alignment=CENTER,
                border=BORDER,
            )
            if isinstance(value, float):
                cell.number_format = _long_number_format(header)

    ws.auto_filter.ref = (
        f"A1:{get_column_letter(len(_LONG_HEADERS))}{max(1, len(rows) + 1)}"
    )
    ws.freeze_panes = "A2"
    for col_idx, header in enumerate(_LONG_HEADERS, start=1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = _long_column_width(header)
        if header in _ADVANCED_HEADERS:
            ws.column_dimensions[letter].hidden = True
            ws.column_dimensions[letter].outlineLevel = 1
    _merge_repeated_identity_cells(ws, rows)


def _merge_repeated_identity_cells(ws, rows: list[dict[str, str]]) -> None:
    if not rows:
        return
    block_start = 2
    previous = _identity_key(rows[0])
    for offset, row in enumerate(rows[1:], start=3):
        current = _identity_key(row)
        if current == previous:
            continue
        _merge_identity_block(ws, block_start, offset - 1)
        block_start = offset
        previous = current
    _merge_identity_block(ws, block_start, len(rows) + 1)


def _identity_key(row: dict[str, str]) -> tuple[str, str]:
    return row.get("SampleName", ""), row.get("Group", "")


def _merge_identity_block(ws, start_row: int, end_row: int) -> None:
    if end_row <= start_row:
        return
    for column in (1, 2):
        ws.merge_cells(
            start_row=start_row,
            start_column=column,
            end_row=end_row,
            end_column=column,
        )
        ws.cell(row=start_row, column=column).alignment = CENTER_WRAP


def _sample_group(name: str) -> str:
    return classify_sample_group(name)


def _build_overview_sheet(
    ws,
    rows: list[dict[str, str]],
    diagnostics: list[dict[str, str]],
    review_rows: list[dict[str, str]],
) -> None:
    ws.title = "Overview"
    ws.merge_cells("A1:D1")
    _apply(
        ws["A1"],
        value="XIC Review Overview",
        font=Font(bold=True, color="FFFFFF", size=14),
        fill=_fill(_OVERVIEW_HEADER_FILL),
        alignment=Alignment(horizontal="left", vertical="center"),
        border=BORDER,
    )
    ws.row_dimensions[1].height = 28

    metrics = [
        ("Samples", len(_distinct_values(rows, "SampleName"))),
        ("Targets", len(_distinct_values(rows, "Target"))),
        ("Review Items", len(review_rows)),
        ("Diagnostics", len(diagnostics)),
    ]
    for row_idx, (label, value) in enumerate(metrics, start=3):
        _apply(
            ws.cell(row=row_idx, column=1, value=label),
            font=Font(bold=True),
            fill=_fill(GREY),
            alignment=CENTER,
            border=BORDER,
        )
        _apply(
            ws.cell(row=row_idx, column=2, value=value),
            fill=_fill(WHITE),
            alignment=CENTER,
            border=BORDER,
        )

    next_row = _write_overview_count_section(
        ws,
        8,
        "Top Targets",
        "Target",
        _top_review_counts(review_rows, "Target"),
    )
    _write_overview_count_section(
        ws,
        next_row + 1,
        "Top Samples",
        "Sample",
        _top_review_counts(review_rows, "Sample"),
    )

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18


def _apply_sheet_role_styles(wb: Workbook) -> None:
    role_colors = {
        "Overview": _DAILY_REVIEW_TAB,
        "Review Queue": _DAILY_REVIEW_TAB,
        "XIC Results": _RESULT_TAB,
        "Summary": _RESULT_TAB,
        "Targets": _TECHNICAL_TAB,
        "Diagnostics": _TECHNICAL_TAB,
        "Run Metadata": _TECHNICAL_TAB,
        "Score Breakdown": _TECHNICAL_TAB,
    }
    for name, color in role_colors.items():
        if name in wb.sheetnames:
            wb[name].sheet_properties.tabColor = color


def _distinct_values(rows: list[dict[str, str]], key: str) -> set[str]:
    return {value for row in rows if (value := row.get(key, ""))}


def _top_review_counts(
    review_rows: list[dict[str, str]], key: str
) -> list[tuple[str, int]]:
    counts: dict[str, int] = {}
    for row in review_rows:
        value = row.get(key, "")
        if not value:
            continue
        counts[value] = counts.get(value, 0) + 1
    return sorted(counts.items(), key=lambda item: (-item[1], item[0]))[:5]


def _write_overview_count_section(
    ws,
    start_row: int,
    title: str,
    label_header: str,
    counts: list[tuple[str, int]],
) -> int:
    _apply(
        ws.cell(row=start_row, column=1, value=title),
        font=Font(bold=True, color="FFFFFF"),
        fill=_fill(_OVERVIEW_HEADER_FILL),
        alignment=CENTER,
        border=BORDER,
    )
    _apply(
        ws.cell(row=start_row + 1, column=1, value=label_header),
        font=Font(bold=True),
        fill=_fill(GREY),
        alignment=CENTER,
        border=BORDER,
    )
    _apply(
        ws.cell(row=start_row + 1, column=2, value="Review Items"),
        font=Font(bold=True),
        fill=_fill(GREY),
        alignment=CENTER,
        border=BORDER,
    )

    if not counts:
        _apply(
            ws.cell(row=start_row + 2, column=1, value="None"),
            fill=_fill(WHITE),
            alignment=CENTER,
            border=BORDER,
        )
        _apply(
            ws.cell(row=start_row + 2, column=2, value=0),
            fill=_fill(WHITE),
            alignment=CENTER,
            border=BORDER,
        )
        return start_row + 3

    for offset, (label, count) in enumerate(counts, start=2):
        row_idx = start_row + offset
        fill_hex = GREY if row_idx % 2 == 0 else WHITE
        _apply(
            ws.cell(row=row_idx, column=1, value=_excel_text(label)),
            fill=_fill(fill_hex),
            alignment=CENTER,
            border=BORDER,
        )
        _apply(
            ws.cell(row=row_idx, column=2, value=count),
            fill=_fill(fill_hex),
            alignment=CENTER,
            border=BORDER,
        )
    return start_row + len(counts) + 2


def _build_summary_sheet(
    ws,
    rows: list[dict[str, str]],
    count_no_ms2_as_detected: bool = False,
    review_rows: list[dict[str, str]] | None = None,
) -> None:
    for col_idx, header in enumerate(_SUMMARY_HEADERS, start=1):
        _apply(
            ws.cell(row=1, column=col_idx, value=header),
            **_header_style(_MS2_HEADER),
        )
    ws.row_dimensions[1].height = 30

    metrics = build_review_metrics(
        rows,
        diagnostics=[],
        review_rows=review_rows or [],
        count_no_ms2_as_detected=count_no_ms2_as_detected,
    )
    for row_idx, target in enumerate(_target_summaries(rows), start=2):
        fill_hex = GREY if row_idx % 2 == 0 else WHITE
        values = _summary_row_values(
            target,
            rows,
            count_no_ms2_as_detected,
            metrics,
        )
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            _apply(
                cell,
                fill=_fill(fill_hex),
                alignment=CENTER,
                border=BORDER,
            )
            if _SUMMARY_HEADERS[col_idx - 1] == "Median Area (detected)":
                cell.number_format = "0.00E+00"

    widths = [
        24,
        12,
        24,
        14,
        14,
        12,
        14,
        12,
        10,
        12,
        12,
        16,
        24,
        10,
        10,
        10,
        10,
        22,
        14,
        16,
        12,
        14,
    ]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.auto_filter.ref = (
        f"A1:{get_column_letter(len(_SUMMARY_HEADERS))}"
        f"{max(1, len(_target_summaries(rows)) + 1)}"
    )
    ws.freeze_panes = "A2"


def _build_review_queue_sheet(
    ws,
    queue_rows: list[dict[str, str]],
) -> None:
    for col_idx, header in enumerate(_REVIEW_HEADERS, start=1):
        _apply(
            ws.cell(row=1, column=col_idx, value=header),
            **_header_style(_MS2_HEADER),
        )
    ws.row_dimensions[1].height = 30

    for row_idx, row in enumerate(queue_rows, start=2):
        fill_hex = _review_priority_fill(row["Priority"])
        for col_idx, header in enumerate(_REVIEW_HEADERS, start=1):
            raw_value = row.get(header, "")
            value = _review_cell_value(header, raw_value)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            _apply(
                cell,
                fill=_review_cell_fill(header, raw_value, fill_hex),
                alignment=(
                    CENTER_WRAP
                    if header in {"Action", "Evidence"}
                    else CENTER
                ),
                border=BORDER,
            )
            if isinstance(value, float):
                cell.number_format = _review_number_format(header)

    widths = [10, 28, 24, 12, 12, 28, 12, 16, 38, 12, 72]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.auto_filter.ref = (
        f"A1:{get_column_letter(len(_REVIEW_HEADERS))}"
        f"{max(1, len(queue_rows) + 1)}"
    )
    ws.freeze_panes = "A2"


def _review_queue_rows(
    rows: list[dict[str, str]],
    diagnostics: list[dict[str, str]],
) -> list[dict[str, str]]:
    diagnostics_by_key = _diagnostics_grouped_by_key(diagnostics)
    rows_by_key = _results_grouped_by_key(rows)
    queue: list[dict[str, str]] = []
    for key, key_rows in rows_by_key.items():
        row_diagnostics = diagnostics_by_key.get(key, [])
        row, issue = _primary_review_row(key_rows, row_diagnostics)
        if not issue:
            continue
        evidence = _review_evidence(row, row_diagnostics)
        queue.append(
            {
                "Priority": str(_review_priority(row, issue)),
                "Sample": row.get("SampleName", ""),
                "Target": row.get("Target", ""),
                "Role": row.get("Role", ""),
                "Status": _review_status(issue, row.get("Confidence", "")),
                "Why": _review_why(issue, row.get("Reason", ""), evidence),
                "RT": row.get("RT", ""),
                "Area": row.get("Area", ""),
                "Action": _suggested_action(issue, row),
                "Issue Count": str(len(row_diagnostics) if row_diagnostics else 1),
                "Evidence": evidence,
            }
        )
    queue.sort(
        key=lambda item: (
            int(item["Priority"]),
            item["Sample"],
            item["Target"],
            item["Why"],
        )
    )
    return queue


def _results_grouped_by_key(
    rows: list[dict[str, str]],
) -> dict[tuple[str, str], list[dict[str, str]]]:
    by_key: dict[tuple[str, str], list[dict[str, str]]] = {}
    for row in rows:
        key = (row.get("SampleName", ""), row.get("Target", ""))
        by_key.setdefault(key, []).append(row)
    return by_key


def _diagnostics_grouped_by_key(
    diagnostics: list[dict[str, str]],
) -> dict[tuple[str, str], list[dict[str, str]]]:
    by_key: dict[tuple[str, str], list[dict[str, str]]] = {}
    for row in diagnostics:
        key = (row.get("SampleName", ""), row.get("Target", ""))
        by_key.setdefault(key, []).append(row)
    return by_key


def _primary_review_row(
    rows: list[dict[str, str]],
    diagnostics: list[dict[str, str]],
) -> tuple[dict[str, str], str]:
    ranked: list[tuple[int, int, dict[str, str], str]] = []
    for index, row in enumerate(rows):
        issue = _primary_review_issue(row, diagnostics)
        if issue:
            ranked.append((_review_priority(row, issue), index, row, issue))
    if not ranked:
        return rows[0], ""
    _, _, row, issue = min(ranked, key=lambda item: (item[0], item[1]))
    return row, issue


def _primary_review_issue(
    row: dict[str, str], diagnostics: list[dict[str, str]]
) -> str:
    candidates = [diagnostic.get("Issue", "") for diagnostic in diagnostics]
    row_issue = _row_review_issue(row)
    if row_issue:
        candidates.append(row_issue)
    candidates = [issue for issue in candidates if issue]
    if not candidates:
        return ""
    return min(
        enumerate(candidates),
        key=lambda item: (_review_priority(row, item[1]), item[0]),
    )[1]


def _review_evidence(
    row: dict[str, str], diagnostics: list[dict[str, str]]
) -> str:
    if diagnostics:
        parts = [
            diagnostic.get("Reason", "") or diagnostic.get("Issue", "")
            for diagnostic in diagnostics
        ]
    else:
        parts = [row.get("Reason", "") or _row_review_issue(row)]
    return "; ".join(_dedupe_text(part for part in parts if part))


def _dedupe_text(values: Iterable[str]) -> list[str]:
    seen: set[str] = set()
    deduped: list[str] = []
    for value in values:
        if value in seen:
            continue
        seen.add(value)
        deduped.append(value)
    return deduped


def _row_review_issue(row: dict[str, str]) -> str:
    nl = row.get("NL", "")
    if nl == "NL_FAIL":
        return "NL_FAIL"
    if nl == "NO_MS2":
        return "NO_MS2"
    if nl.startswith("WARN_"):
        return "NL_WARN"
    confidence = row.get("Confidence", "")
    if confidence in {"LOW", "VERY_LOW"}:
        return f"CONFIDENCE_{confidence}"
    return ""


def _review_priority(row: dict[str, str], issue: str) -> int:
    confidence = row.get("Confidence", "")
    if issue in {"PEAK_NOT_FOUND", "NO_SIGNAL", "FILE_ERROR", "NL_FAIL"}:
        return 1
    if confidence == "VERY_LOW":
        return 1
    if issue in {"NO_MS2", "NL_WARN"} or confidence == "LOW":
        return 2
    return 3


def _review_status(issue: str, confidence: str) -> str:
    if issue in {"NL_FAIL", "PEAK_NOT_FOUND", "NO_SIGNAL", "FILE_ERROR"}:
        return "Review"
    if issue in {"NO_MS2", "NL_WARN"} or confidence in {"LOW", "VERY_LOW"}:
        return "Check"
    return "Info"


def _review_why(issue: str, reason: str, evidence: str) -> str:
    if issue == "NL_FAIL":
        return "NL support failed"
    if issue == "NO_MS2":
        return "MS2 trigger missing"
    if issue == "NL_WARN":
        return "NL support is borderline"
    if issue in {"PEAK_NOT_FOUND", "NO_SIGNAL"}:
        return "Peak not found"
    if issue.startswith("CONFIDENCE_"):
        parsed = _first_concern(reason)
        if parsed and parsed != "all checks passed":
            return parsed
        return issue.removeprefix("CONFIDENCE_") + " confidence"
    parsed = _first_concern(reason) or _first_concern(evidence)
    return parsed if parsed and parsed != "all checks passed" else issue


def _first_concern(text: str) -> str:
    if not text:
        return ""
    if text.startswith(_FORMULA_PREFIXES):
        return ""
    normalized = text.removeprefix("concerns:").strip()
    first = normalized.split(";", 1)[0].strip()
    if first.startswith("weak candidate:"):
        first = first.removeprefix("weak candidate:").strip()
    return first


def _suggested_action(issue: str, row: dict[str, str]) -> str:
    if issue in {"NL_FAIL", "NL_WARN"}:
        return "Check MS2 / NL evidence near selected RT"
    if issue == "NO_MS2":
        return "Check whether missing DDA trigger is acceptable"
    if issue in {"PEAK_NOT_FOUND", "NO_SIGNAL"}:
        return "Inspect XIC trace and target RT window"
    if issue == "FILE_ERROR":
        return "Check RAW file readability"
    if row.get("Confidence") in {"LOW", "VERY_LOW"}:
        return "Review peak shape, RT prior, and ISTD pairing"
    return "Review diagnostic detail"


def _review_cell_value(header: str, raw_value: str) -> object:
    if header in {"Priority", "Issue Count"}:
        parsed = _safe_float(raw_value)
        return int(parsed) if parsed is not None else raw_value
    if header in {"RT", "Area"}:
        if raw_value in ND_ERROR:
            return raw_value
        parsed = _safe_float(raw_value)
        return parsed if parsed is not None else raw_value
    if header in {
        "Sample",
        "Target",
        "Why",
        "Action",
        "Evidence",
    }:
        return _excel_text(raw_value)
    return raw_value


def _review_cell_fill(header: str, raw_value: str, row_fill_hex: str) -> PatternFill:
    return _fill(row_fill_hex)


def _review_priority_fill(priority: str) -> str:
    return {"1": _MS2_FAIL, "2": _MS2_WARN}.get(priority, GREY)


def _review_number_format(header: str) -> str:
    if header == "RT":
        return "0.0000"
    if header == "Area":
        return "0.00E+00"
    return "#,##0"


def _target_summaries(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    seen: set[str] = set()
    summaries: list[dict[str, str]] = []
    for row in rows:
        target = row.get("Target", "")
        if not target or target in seen:
            continue
        seen.add(target)
        summaries.append(row)
    return summaries


def _summary_row_values(
    target_row: dict[str, str],
    rows: list[dict[str, str]],
    count_no_ms2_as_detected: bool,
    metrics: ReviewMetrics,
) -> list[object]:
    target = target_row["Target"]
    target_rows = [row for row in rows if row.get("Target") == target]
    detected_rows = [
        row for row in target_rows if _is_long_detected(row, count_no_ms2_as_detected)
    ]
    total = len(target_rows)
    detected = len(detected_rows)
    nl_counts = _long_nl_counts(target_rows)
    confidence_counts = _long_confidence_counts(target_rows)
    target_metrics = metrics.targets[target]
    return [
        _excel_text(target),
        target_row.get("Role", ""),
        _excel_text(target_row.get("ISTD Pair", "")),
        target_metrics.flagged_rows,
        target_metrics.flagged_percent,
        target_metrics.ms2_nl_flags,
        target_metrics.low_confidence_rows,
        detected,
        total,
        f"{detected / total * 100:.0f}%" if total else "0%",
        _long_mean_rt(detected_rows),
        _long_median_area(detected_rows),
        _long_area_ratio(target_row, rows, count_no_ms2_as_detected),
        nl_counts["OK"],
        nl_counts["WARN"],
        nl_counts["NL_FAIL"],
        nl_counts["NO_MS2"],
        _long_rt_delta(target_row, rows, count_no_ms2_as_detected),
        confidence_counts["HIGH"],
        confidence_counts["MEDIUM"],
        confidence_counts["LOW"],
        confidence_counts["VERY_LOW"],
    ]


def _long_confidence_counts(target_rows: list[dict[str, str]]) -> dict[str, int]:
    counts = {"HIGH": 0, "MEDIUM": 0, "LOW": 0, "VERY_LOW": 0}
    for row in target_rows:
        confidence = row.get("Confidence", "HIGH")
        if confidence in counts:
            counts[confidence] += 1
    return counts


def _is_long_detected(
    row: dict[str, str], count_no_ms2_as_detected: bool = False
) -> bool:
    if _safe_float(row.get("RT", "")) is None:
        return False
    if _safe_float(row.get("Area", "")) is None:
        return False
    nl = row.get("NL", "")
    if nl == "NO_MS2":
        return count_no_ms2_as_detected
    return True


def _long_mean_rt(rows: list[dict[str, str]]) -> str:
    values = [
        value for row in rows if (value := _safe_float(row.get("RT", ""))) is not None
    ]
    return f"{sum(values) / len(values):.4f}" if values else "—"


def _long_median_area(rows: list[dict[str, str]]) -> float | str:
    values = [
        value for row in rows if (value := _safe_float(row.get("Area", ""))) is not None
    ]
    return statistics.median(values) if values else "—"


def _long_nl_counts(rows: list[dict[str, str]]) -> dict[str, int]:
    values = [row.get("NL", "") for row in rows]
    ok = sum(1 for value in values if value == "OK")
    warn = sum(1 for value in values if value.startswith("WARN_"))
    no_ms2 = sum(1 for value in values if value == "NO_MS2")
    fail = sum(1 for value in values if value == "NL_FAIL")
    return {"OK": ok, "WARN": warn, "NL_FAIL": fail, "NO_MS2": no_ms2}


def _long_area_ratio(
    target_row: dict[str, str],
    rows: list[dict[str, str]],
    count_no_ms2_as_detected: bool,
) -> str:
    istd_pair = target_row.get("ISTD Pair", "")
    if not istd_pair:
        return "—"
    ratios: list[float] = []
    rows_by_sample = _rows_by_sample_and_target(rows)
    for row in rows:
        if row.get("Target") != target_row["Target"]:
            continue
        if classify_sample_group(row.get("SampleName", "")) != "QC":
            continue
        if not _is_long_detected(row, count_no_ms2_as_detected):
            continue
        istd = rows_by_sample.get((row.get("SampleName", ""), istd_pair))
        if istd is None or not _is_long_detected(istd, count_no_ms2_as_detected):
            continue
        analyte_area = _safe_float(row.get("Area", ""))
        istd_area = _safe_float(istd.get("Area", ""))
        if analyte_area is None or istd_area is None or istd_area == 0:
            continue
        ratios.append(analyte_area / istd_area)
    if not ratios:
        return "—"
    mean_ratio = sum(ratios) / len(ratios)
    sd_ratio = statistics.stdev(ratios) if len(ratios) > 1 else 0.0
    cv_text = (
        f"{sd_ratio / mean_ratio * 100:.1f}%"
        if len(ratios) > 1 and mean_ratio != 0
        else "—"
    )
    return f"{mean_ratio:.4f}±{sd_ratio:.4f}, CV={cv_text} (n={len(ratios)})"


def _long_rt_delta(
    target_row: dict[str, str],
    rows: list[dict[str, str]],
    count_no_ms2_as_detected: bool,
) -> str:
    istd_pair = target_row.get("ISTD Pair", "")
    if not istd_pair:
        return "—"
    deltas: list[float] = []
    rows_by_sample = _rows_by_sample_and_target(rows)
    for row in rows:
        if row.get("Target") != target_row["Target"]:
            continue
        if not _is_long_detected(row, count_no_ms2_as_detected):
            continue
        istd = rows_by_sample.get((row.get("SampleName", ""), istd_pair))
        if istd is None or not _is_long_detected(istd, count_no_ms2_as_detected):
            continue
        rt_analyte = _safe_float(row.get("RT", ""))
        rt_istd = _safe_float(istd.get("RT", ""))
        if rt_analyte is None or rt_istd is None:
            continue
        deltas.append(abs(rt_analyte - rt_istd))
    if not deltas:
        return "—"
    mean_delta = sum(deltas) / len(deltas)
    sd_delta = statistics.stdev(deltas) if len(deltas) > 1 else 0.0
    return f"{mean_delta:.4f}±{sd_delta:.4f} min (n={len(deltas)})"


def _rows_by_sample_and_target(
    rows: list[dict[str, str]],
) -> dict[tuple[str, str], dict[str, str]]:
    return {(row.get("SampleName", ""), row.get("Target", "")): row for row in rows}


def _build_targets_sheet(ws, targets: list[Target]) -> None:
    for col_idx, header in enumerate(_TARGETS_HEADERS, start=1):
        _apply(
            ws.cell(row=1, column=col_idx, value=header),
            **_header_style(_SAMPLE_HEADER),
        )
    ws["I1"].comment = Comment(
        "Nominal target product m/z = target m/z - neutral_loss_da. "
        "Strict observed-loss NL evidence is evaluated from each MS2 scan precursor "
        "and selected candidate alignment.",
        "XIC Extractor",
    )
    ws.row_dimensions[1].height = 30

    for row_idx, target in enumerate(targets, start=2):
        fill_hex = GREY if row_idx % 2 == 0 else WHITE
        expected_product = (
            target.mz - target.neutral_loss_da
            if target.neutral_loss_da is not None
            else None
        )
        values: list[object] = [
            _excel_text(target.label),
            "ISTD" if target.is_istd else "Analyte",
            _excel_text(target.istd_pair),
            target.mz,
            target.rt_min,
            target.rt_max,
            target.ppm_tol,
            target.neutral_loss_da if target.neutral_loss_da is not None else "—",
            expected_product if expected_product is not None else "—",
            target.nl_ppm_warn if target.nl_ppm_warn is not None else "—",
            target.nl_ppm_max if target.nl_ppm_max is not None else "—",
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            _apply(cell, fill=_fill(fill_hex), alignment=CENTER, border=BORDER)
            header = _TARGETS_HEADERS[col_idx - 1]
            if isinstance(value, float):
                if header in {"m/z", "Expected product m/z", "NL (Da)"}:
                    cell.number_format = "0.0000"
                elif header in {"RT min", "RT max"}:
                    cell.number_format = "0.00"

    widths = [24, 12, 24, 14, 10, 10, 10, 12, 22, 14, 14]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.auto_filter.ref = f"A1:K{max(1, len(targets) + 1)}"
    ws.freeze_panes = "A2"


def _build_metadata_sheet(ws, config: ExtractionConfig) -> None:
    for col_idx, header in enumerate(("Key", "Value"), start=1):
        _apply(
            ws.cell(row=1, column=col_idx, value=header),
            **_header_style(_SAMPLE_HEADER),
        )

    for row_idx, (key, value) in enumerate(
        metadata.build_metadata_rows(config), start=2
    ):
        fill_hex = GREY if row_idx % 2 == 0 else WHITE
        _apply(
            ws.cell(row=row_idx, column=1, value=key),
            fill=_fill(fill_hex),
            alignment=CENTER,
            border=BORDER,
        )
        _apply(
            ws.cell(row=row_idx, column=2, value=value),
            fill=_fill(fill_hex),
            alignment=CENTER,
            border=BORDER,
        )

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 48
    ws.freeze_panes = "A2"


def _read_diagnostics(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open(newline="", encoding="utf-8-sig") as handle:
        return list(csv.DictReader(handle))


def _build_diagnostics_sheet(ws, rows: list[dict[str, str]]) -> None:
    for col_idx, header in enumerate(_DIAGNOSTIC_HEADERS, start=1):
        _apply(
            ws.cell(row=1, column=col_idx, value=header),
            **_header_style(_MS2_HEADER),
        )
    for row_idx, row in enumerate(rows, start=2):
        issue = row.get("Issue", "")
        for col_idx, header in enumerate(_DIAGNOSTIC_HEADERS, start=1):
            raw = row.get(header, "")
            value = (
                _excel_text(raw)
                if header in {"SampleName", "Target", "Reason"}
                else raw
            )
            _apply(
                ws.cell(row=row_idx, column=col_idx, value=value),
                fill=_fill(_diagnostic_fill(issue)),
                alignment=CENTER,
                border=BORDER,
            )
    ws.auto_filter.ref = f"A1:D{max(1, len(rows) + 1)}"
    widths = [24, 20, 18, 60]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"


def _diagnostic_fill(issue: str) -> str:
    if issue in {"NO_MS2", "WINDOW_TOO_SHORT"}:
        return _MS2_NO_MS2
    if issue in {"NL_FAIL", "PEAK_NOT_FOUND", "NO_SIGNAL", "FILE_ERROR"}:
        return _MS2_FAIL
    return _MS2_WARN


def _build_score_breakdown_sheet(ws, rows: list[dict[str, str]]) -> None:
    if not rows:
        return
    headers = list(rows[0].keys())
    for col_idx, header in enumerate(headers, start=1):
        _apply(
            ws.cell(row=1, column=col_idx, value=header),
            **_header_style(_MS2_HEADER),
        )
    for row_idx, row in enumerate(rows, start=2):
        fill_hex = GREY if row_idx % 2 == 0 else WHITE
        for col_idx, header in enumerate(headers, start=1):
            raw_value = row.get(header, "")
            value = (
                _safe_float(raw_value)
                if header
                in {
                    "symmetry",
                    "local_sn",
                    "nl_support",
                    "rt_prior",
                    "rt_centrality",
                    "noise_shape",
                    "peak_width",
                    "Quality Penalty",
                    "Total Severity",
                    "Prior RT",
                }
                else _excel_text(raw_value)
            )
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            _apply(cell, fill=_fill(fill_hex), alignment=CENTER, border=BORDER)
            if isinstance(value, float) and header == "Prior RT":
                cell.number_format = "0.0000"
    ws.auto_filter.ref = (
        f"A1:{get_column_letter(len(headers))}{max(1, len(rows) + 1)}"
    )
    ws.freeze_panes = "A2"


@overload
def run(base_or_config: Path) -> Path: ...


@overload
def run(base_or_config: ExtractionConfig, targets: list[Target]) -> Path: ...


def run(
    base_or_config: Path | ExtractionConfig,
    targets: list[Target] | None = None,
) -> Path:
    if isinstance(base_or_config, Path):
        config, loaded_targets = load_config(base_or_config / "config")
        return run(config, loaded_targets)
    if targets is None:
        raise TypeError("targets are required when run() receives ExtractionConfig")
    return _run_with_config(base_or_config, targets)


def _run_with_config(config: ExtractionConfig, targets: list[Target]) -> Path:
    output_dir = config.output_csv.parent
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    excel_path = output_dir / f"xic_results_{timestamp}.xlsx"

    rows = _read_long_results(config, targets)
    if not rows:
        print("CSV is empty.")
        return excel_path

    diagnostics = _read_diagnostics(config.diagnostics_csv)
    score_breakdown = _read_score_breakdown(config)
    review_rows = _review_queue_rows(rows, diagnostics)

    wb = Workbook()
    ws_overview = wb.active
    _build_overview_sheet(ws_overview, rows, diagnostics, review_rows)

    ws_review = wb.create_sheet("Review Queue")
    _build_review_queue_sheet(ws_review, review_rows)

    ws_data = wb.create_sheet("XIC Results")
    _build_data_sheet(ws_data, rows)

    ws_summary = wb.create_sheet("Summary")
    _build_summary_sheet(
        ws_summary,
        rows,
        count_no_ms2_as_detected=config.count_no_ms2_as_detected,
        review_rows=review_rows,
    )

    ws_targets = wb.create_sheet("Targets")
    _build_targets_sheet(ws_targets, targets)

    ws_diagnostics = wb.create_sheet("Diagnostics")
    _build_diagnostics_sheet(ws_diagnostics, diagnostics)
    ws_metadata = wb.create_sheet("Run Metadata")
    _build_metadata_sheet(ws_metadata, config)
    if config.emit_score_breakdown and score_breakdown:
        ws_breakdown = wb.create_sheet("Score Breakdown")
        _build_score_breakdown_sheet(ws_breakdown, score_breakdown)
    wb.active = wb.index(ws_overview)
    _apply_sheet_role_styles(wb)

    wb.save(excel_path)
    _print_summary(excel_path, rows, config.count_no_ms2_as_detected)
    return excel_path


def _read_long_results(
    config: ExtractionConfig, targets: list[Target]
) -> list[dict[str, str]]:
    long_path = config.output_csv.with_name("xic_results_long.csv")
    if long_path.exists():
        return _read_results(long_path)
    return _wide_to_long_rows(_read_results(config.output_csv), targets)


def _wide_to_long_rows(
    rows: list[dict[str, str]], targets: list[Target]
) -> list[dict[str, str]]:
    long_rows: list[dict[str, str]] = []
    for row in rows:
        sample_name = row.get("SampleName", "")
        for target in targets:
            long_rows.append(
                {
                    "SampleName": sample_name,
                    "Group": _sample_group(sample_name),
                    "Target": target.label,
                    "Role": "ISTD" if target.is_istd else "Analyte",
                    "ISTD Pair": target.istd_pair,
                    "RT": row.get(f"{target.label}_RT", ""),
                    "Area": row.get(f"{target.label}_Area", ""),
                    "NL": row.get(f"{target.label}_NL", "")
                    if target.neutral_loss_da is not None
                    else "",
                    "Int": row.get(f"{target.label}_Int", ""),
                    "PeakStart": row.get(f"{target.label}_PeakStart", ""),
                    "PeakEnd": row.get(f"{target.label}_PeakEnd", ""),
                    "PeakWidth": _legacy_peak_width(row, target.label),
                    "Confidence": "HIGH",
                    "Reason": "",
                }
            )
    return long_rows


def _read_score_breakdown(config: ExtractionConfig) -> list[dict[str, str]]:
    path = config.output_csv.with_name("xic_score_breakdown.csv")
    if not path.exists():
        return []
    return _read_results(path)


def _legacy_peak_width(row: dict[str, str], label: str) -> str:
    existing = row.get(f"{label}_PeakWidth", "")
    if existing:
        return existing
    start = row.get(f"{label}_PeakStart", "")
    end = row.get(f"{label}_PeakEnd", "")
    if start in ND_ERROR or end in ND_ERROR:
        return start if start == end else ""
    start_value = _safe_float(start)
    end_value = _safe_float(end)
    if start_value is None or end_value is None:
        return ""
    return f"{abs(end_value - start_value):.4f}"


def _read_results(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        return list(csv.DictReader(handle))


def _print_summary(
    excel_path: Path,
    rows: list[dict[str, str]],
    count_no_ms2_as_detected: bool,
) -> None:
    print(f"Saved : {excel_path}")
    sample_count = len({row.get("SampleName", "") for row in rows})
    print(f"Rows  : {sample_count}")
    for target in _target_summaries(rows):
        label = target["Target"]
        target_rows = [row for row in rows if row.get("Target") == label]
        detected = sum(
            1 for row in target_rows if _is_long_detected(row, count_no_ms2_as_detected)
        )
        note = (
            " (NL confirmed)" if any(row.get("NL", "") for row in target_rows) else ""
        )
        print(f"  {label} detected{note}: {detected}/{len(target_rows)}")
    for target in _target_summaries(rows):
        label = target["Target"]
        target_rows = [row for row in rows if row.get("Target") == label]
        if not any(row.get("NL", "") for row in target_rows):
            continue
        counts = _long_nl_counts(target_rows)
        print(
            f"  {label}_NL  OK:{counts['OK']}  WARN:{counts['WARN']}  "
            f"FAIL:{counts['NL_FAIL']}  NO_MS2:{counts['NO_MS2']}"
        )
    for target in _target_summaries(rows):
        if target.get("Role") != "ISTD":
            continue
        label = target["Target"]
        target_rows = [row for row in rows if row.get("Target") == label]
        detected = sum(
            1 for row in target_rows if _is_long_detected(row, count_no_ms2_as_detected)
        )
        if detected < len(target_rows):
            print(f"ISTD_ND: {label} {detected}/{len(target_rows)}")


def main() -> None:
    base_dir = Path(__file__).parent.parent
    run(base_dir)


if __name__ == "__main__":
    main()
