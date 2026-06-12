from __future__ import annotations

import argparse
import math
import sys
from collections.abc import Sequence
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from xic_extractor.tabular_io import write_tsv

DEFAULT_NUMERIC_TOLERANCE = 1e-9
COMPARE_SHEET = "Matrix"


@dataclass(frozen=True)
class AlignmentWorkbookCompareResult:
    matched: bool
    differences: tuple[str, ...]


def compare_alignment_workbooks(
    left_path: Path,
    right_path: Path,
    *,
    numeric_tolerance: float = DEFAULT_NUMERIC_TOLERANCE,
) -> AlignmentWorkbookCompareResult:
    left = load_workbook(left_path, data_only=True)
    right = load_workbook(right_path, data_only=True)
    differences: list[str] = []
    if COMPARE_SHEET not in left.sheetnames:
        differences.append(f"{left_path}: missing sheet {COMPARE_SHEET!r}")
    if COMPARE_SHEET not in right.sheetnames:
        differences.append(f"{right_path}: missing sheet {COMPARE_SHEET!r}")
    if differences:
        return AlignmentWorkbookCompareResult(False, tuple(differences))

    differences.extend(
        _compare_rows(
            _sheet_rows(left[COMPARE_SHEET]),
            _sheet_rows(right[COMPARE_SHEET]),
            numeric_tolerance,
        )
    )
    return AlignmentWorkbookCompareResult(not differences, tuple(differences))


def _sheet_rows(sheet: Any) -> list[tuple[Any, ...]]:
    rows: list[tuple[Any, ...]] = []
    for raw_row in sheet.iter_rows(values_only=True):
        row = _trim_trailing_empty(raw_row)
        rows.append(row)
    while rows and not rows[-1]:
        rows.pop()
    return rows


def _trim_trailing_empty(row: Sequence[Any]) -> tuple[Any, ...]:
    values = list(row)
    while values and values[-1] is None:
        values.pop()
    return tuple(values)


def _compare_rows(
    left_rows: list[tuple[Any, ...]],
    right_rows: list[tuple[Any, ...]],
    numeric_tolerance: float,
) -> list[str]:
    differences: list[str] = []
    if len(left_rows) != len(right_rows):
        differences.append(
            f"{COMPARE_SHEET}: row count differs "
            f"({len(left_rows)} != {len(right_rows)})"
        )
    for row_number, (left_row, right_row) in enumerate(
        zip(left_rows, right_rows, strict=False),
        start=1,
    ):
        width = max(len(left_row), len(right_row))
        for column_number in range(width):
            left_value = (
                left_row[column_number] if column_number < len(left_row) else None
            )
            right_value = (
                right_row[column_number] if column_number < len(right_row) else None
            )
            if _values_match(left_value, right_value, numeric_tolerance):
                continue
            differences.append(
                f"{COMPARE_SHEET}!R{row_number}C{column_number + 1}: "
                f"{left_value!r} != {right_value!r}"
            )
    return differences


def _values_match(left: Any, right: Any, numeric_tolerance: float) -> bool:
    if left == right:
        return True
    if isinstance(left, bool) or isinstance(right, bool):
        return False
    if isinstance(left, int | float) and isinstance(right, int | float):
        if not math.isfinite(float(left)) or not math.isfinite(float(right)):
            return False
        return abs(float(left) - float(right)) <= numeric_tolerance
    return False


def main(argv: Sequence[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Compare alignment workbook Matrix sheets.",
    )
    parser.add_argument("left", type=Path)
    parser.add_argument("right", type=Path)
    parser.add_argument(
        "--numeric-tolerance",
        type=float,
        default=DEFAULT_NUMERIC_TOLERANCE,
    )
    parser.add_argument("--output-tsv", type=Path)
    parser.add_argument("--output-report", type=Path)
    args = parser.parse_args(argv)

    result = compare_alignment_workbooks(
        args.left,
        args.right,
        numeric_tolerance=args.numeric_tolerance,
    )
    if args.output_tsv is not None:
        write_compare_tsv(args.output_tsv, result)
    if args.output_report is not None:
        write_compare_report(args.output_report, result)
    if result.matched:
        print("Alignment workbook Matrix compare passed.")
        return 0
    for difference in result.differences:
        print(difference, file=sys.stderr)
    return 1


def write_compare_tsv(
    path: Path,
    result: AlignmentWorkbookCompareResult,
) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    rows = (
        [{"status": "PASS", "difference": ""}]
        if result.matched
        else [
            {"status": "FAIL", "difference": difference}
            for difference in result.differences
        ]
    )
    write_tsv(path, rows, ("status", "difference"), lineterminator="\n")
    return path


def write_compare_report(
    path: Path,
    result: AlignmentWorkbookCompareResult,
) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    lines = [
        "Alignment workbook Matrix compare",
        f"Status: {'PASS' if result.matched else 'FAIL'}",
    ]
    if result.differences:
        lines.append("")
        lines.extend(result.differences)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return path


if __name__ == "__main__":
    raise SystemExit(main())
