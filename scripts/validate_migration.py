from __future__ import annotations

import argparse
import csv
import shutil
import subprocess
from dataclasses import dataclass
from pathlib import Path
from statistics import median
from typing import Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from xic_extractor.config import Target, load_config
from xic_extractor.extractor import RunOutput
from xic_extractor.extractor import run as run_extractor

RT_MEDIAN_LIMIT = 0.003
RT_MAX_LIMIT = 0.010
SMOOTHED_RATIO_MIN = 0.95
SMOOTHED_RATIO_MAX = 1.05
SMOOTHED_MAX_DEVIATION_LIMIT = 0.20
NL_AGREEMENT_MIN_PCT = 95.0

_VALIDATION_HEADERS = [
    "SampleName",
    "Target",
    "RT_Old",
    "Int_Old",
    "NL_Old",
    "RT_New",
    "Int_New_Raw",
    "Int_New_Smoothed",
    "Area_New",
    "PeakStart_New",
    "PeakEnd_New",
    "NL_New",
]
_FAIL_HEADERS = [
    "Target",
    "SampleName",
    "Issue",
    "Reason",
    "MetricValue",
    "Limit",
    "OverrideDecision",
    "OverrideReason",
    "Reviewer",
    "ScreenshotPath",
]
_THIN = Side(style="thin", color="BDBDBD")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center")
_HEADER_FILL = PatternFill("solid", fgColor="37474F")
_FAIL_FILL = PatternFill("solid", fgColor="FFCDD2")
_PASS_FILL = PatternFill("solid", fgColor="C8E6C9")


@dataclass(frozen=True)
class ValidationCase:
    name: str
    path: Path


@dataclass(frozen=True)
class ValidationRow:
    sample_name: str
    target: str
    rt_old: float | None
    int_old: float | None
    nl_old: str
    rt_new: float | None
    int_new_raw: float | None
    int_new_smoothed: float | None
    area_new: float | None
    peak_start_new: float | None
    peak_end_new: float | None
    nl_new: str
    new_row_present: bool = True


@dataclass(frozen=True)
class TargetReport:
    target: str
    status: str
    n_rows: int
    median_rt_delta: float | None
    max_rt_delta: float | None
    smoothed_median_ratio: float | None
    smoothed_max_deviation: float | None
    nl_agreement_pct: float | None


@dataclass(frozen=True)
class FailureRecord:
    target: str
    sample_name: str
    issue: str
    reason: str
    metric_value: str
    limit: str


@dataclass(frozen=True)
class ValidationReport:
    targets: list[TargetReport]
    failures: list[FailureRecord]

    @property
    def failed(self) -> bool:
        return bool(self.failures)


def parse_case_arg(value: str) -> ValidationCase:
    name, sep, path_text = value.partition("=")
    if not sep or not name.strip() or not path_text.strip():
        raise argparse.ArgumentTypeError("case must use NAME=PATH")
    return ValidationCase(name=name.strip(), path=Path(path_text.strip()))


def compare_validation_rows(rows: list[ValidationRow]) -> ValidationReport:
    reports: list[TargetReport] = []
    failures: list[FailureRecord] = []
    for target in sorted({row.target for row in rows}):
        target_rows = [row for row in rows if row.target == target]
        target_failures = _target_failures(target, target_rows)
        failures.extend(target_failures)
        reports.append(_target_report(target, target_rows, target_failures))
    return ValidationReport(targets=reports, failures=failures)


def strict_exit_code(
    report: ValidationReport, *, strict: bool, allow_overrides: bool
) -> int:
    if not strict:
        return 0
    if report.failed and not allow_overrides:
        return 1
    return 0


def write_validation_workbook(
    output_path: Path,
    rows: list[ValidationRow],
    report: ValidationReport,
) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    _write_summary_sheet(ws_summary, report)

    ws_rows = wb.create_sheet("PerTarget")
    _write_rows_sheet(ws_rows, rows)

    ws_fail = wb.create_sheet("FAIL")
    _write_fail_sheet(ws_fail, report.failures)

    wb.save(output_path)
    return output_path


def collect_new_validation_rows(
    output: RunOutput,
    targets: list[Target],
) -> list[ValidationRow]:
    rows: list[ValidationRow] = []
    for file_result in output.file_results:
        for target in targets:
            result = file_result.results.get(target.label)
            if result is None:
                rows.append(
                    ValidationRow(
                        sample_name=file_result.sample_name,
                        target=target.label,
                        rt_old=None,
                        int_old=None,
                        nl_old="",
                        rt_new=None,
                        int_new_raw=None,
                        int_new_smoothed=None,
                        area_new=None,
                        peak_start_new=None,
                        peak_end_new=None,
                        nl_new="ERROR" if file_result.error else "ND",
                    )
                )
                continue

            peak = result.peak_result.peak
            rows.append(
                ValidationRow(
                    sample_name=file_result.sample_name,
                    target=target.label,
                    rt_old=None,
                    int_old=None,
                    nl_old="",
                    rt_new=peak.rt if peak else None,
                    int_new_raw=peak.intensity if peak else None,
                    int_new_smoothed=peak.intensity_smoothed if peak else None,
                    area_new=peak.area if peak else None,
                    peak_start_new=peak.peak_start if peak else None,
                    peak_end_new=peak.peak_end if peak else None,
                    nl_new=result.nl.to_token() if result.nl else "",
                )
            )
    return rows


def merge_old_new_rows(
    old_rows: list[ValidationRow],
    new_rows: list[ValidationRow],
) -> list[ValidationRow]:
    old_by_key = {(row.sample_name, row.target): row for row in old_rows}
    new_by_key = {(row.sample_name, row.target): row for row in new_rows}
    merged: list[ValidationRow] = []
    for key in sorted(old_by_key.keys() | new_by_key.keys()):
        old = old_by_key.get(key)
        new = new_by_key.get(key)
        merged.append(
            ValidationRow(
                sample_name=new.sample_name if new else old.sample_name,
                target=new.target if new else old.target,
                rt_old=old.rt_old if old else None,
                int_old=old.int_old if old else None,
                nl_old=old.nl_old if old else "",
                rt_new=new.rt_new if new else None,
                int_new_raw=new.int_new_raw if new else None,
                int_new_smoothed=new.int_new_smoothed if new else None,
                area_new=new.area_new if new else None,
                peak_start_new=new.peak_start_new if new else None,
                peak_end_new=new.peak_end_new if new else None,
                nl_new=new.nl_new if new else "",
                new_row_present=new is not None,
            )
        )
    return merged


def read_old_results(path: Path) -> list[ValidationRow]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        csv_rows = list(csv.DictReader(handle))
    rows: list[ValidationRow] = []
    for csv_row in csv_rows:
        sample_name = csv_row.get("SampleName", "")
        for key in csv_row:
            if not key.endswith("_RT"):
                continue
            target = key.removesuffix("_RT")
            rows.append(
                ValidationRow(
                    sample_name=sample_name,
                    target=target,
                    rt_old=_safe_float(csv_row.get(f"{target}_RT", "")),
                    int_old=_safe_float(csv_row.get(f"{target}_Int", "")),
                    nl_old=csv_row.get(f"{target}_NL", ""),
                    rt_new=None,
                    int_new_raw=None,
                    int_new_smoothed=None,
                    area_new=None,
                    peak_start_new=None,
                    peak_end_new=None,
                    nl_new="",
                )
            )
    return rows


def main(argv: Sequence[str] | None = None) -> int:
    args = _parse_args(argv)
    old_worktree = args.old_worktree.resolve()
    new_worktree = args.new_worktree.resolve()
    output_path = args.output.resolve()

    cases = list(args.case) + [
        ValidationCase(name=raw_file.stem, path=raw_file) for raw_file in args.raw_file
    ]
    if cases:
        _stage_cases(old_worktree, cases)
        _stage_cases(new_worktree, cases)

    old_csv = _run_old_pipeline(old_worktree)
    old_rows = read_old_results(old_csv)
    config, targets = load_config(new_worktree / "config")
    new_output = run_extractor(config, targets)
    new_rows = collect_new_validation_rows(new_output, targets)
    rows = merge_old_new_rows(old_rows, new_rows)
    report = compare_validation_rows(rows)
    write_validation_workbook(output_path, rows, report)
    _print_report(output_path, report)
    return strict_exit_code(
        report,
        strict=args.strict,
        allow_overrides=args.allow_overrides,
    )


def _parse_args(argv: Sequence[str] | None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Validate Python area extraction against the legacy PS1 pipeline."
    )
    parser.add_argument("--old-worktree", type=Path, required=True)
    parser.add_argument("--new-worktree", type=Path, default=Path.cwd())
    parser.add_argument("--raw-file", type=Path, action="append", default=[])
    parser.add_argument("--case", type=parse_case_arg, action="append", default=[])
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("docs/superpowers/specs/2026-04-13-migration-validation.xlsx"),
    )
    parser.add_argument("--strict", action="store_true")
    parser.add_argument("--allow-overrides", action="store_true")
    return parser.parse_args(argv)


def _target_failures(
    target: str,
    rows: list[ValidationRow],
) -> list[FailureRecord]:
    failures: list[FailureRecord] = []
    missing_new_rows = [
        row for row in rows if _old_peak_ok(row) and not row.new_row_present
    ]
    for row in missing_new_rows:
        failures.append(
            FailureRecord(
                target=target,
                sample_name=row.sample_name,
                issue="NEW_ROW_MISSING",
                reason="Old pipeline produced this sample-target row, but new did not.",
                metric_value="",
                limit="required",
            )
        )

    missing_new_peaks = [
        row
        for row in rows
        if _old_peak_ok(row) and row.new_row_present and not _new_peak_ok(row)
    ]
    for row in missing_new_peaks:
        failures.append(
            FailureRecord(
                target=target,
                sample_name=row.sample_name,
                issue="NEW_PEAK_MISSING",
                reason="Old pipeline detected a peak, but new pipeline did not.",
                metric_value=(
                    f"old_rt={row.rt_old:.4f}; old_int={row.int_old:.0f}"
                    if row.rt_old is not None and row.int_old is not None
                    else ""
                ),
                limit="new peak required when old peak exists",
            )
        )

    missing_smoothed = [
        row for row in rows if _new_peak_ok(row) and row.int_new_smoothed is None
    ]
    for row in missing_smoothed:
        failures.append(
            FailureRecord(
                target=target,
                sample_name=row.sample_name,
                issue="MISSING_SMOOTHED_INTENSITY",
                reason="Int_New_Smoothed is required for every OK new peak.",
                metric_value="",
                limit="required",
            )
        )
    if missing_smoothed:
        return failures

    rt_deltas = _rt_deltas(rows)
    if rt_deltas:
        median_delta = median(value for _row, value in rt_deltas)
        max_delta = max(value for _row, value in rt_deltas)
        if median_delta > RT_MEDIAN_LIMIT or max_delta > RT_MAX_LIMIT:
            failures.append(
                FailureRecord(
                    target=target,
                    sample_name="",
                    issue="RT_DRIFT",
                    reason="RT drift exceeded migration threshold.",
                    metric_value=(f"median={median_delta:.4f}; max={max_delta:.4f}"),
                    limit=(f"median<={RT_MEDIAN_LIMIT:.3f}; max<={RT_MAX_LIMIT:.3f}"),
                )
            )

    smoothed_ratios = _smoothed_ratios(rows)
    if smoothed_ratios:
        ratios = [value for _row, value in smoothed_ratios]
        median_ratio = median(ratios)
        max_deviation = max(abs(value - 1.0) for value in ratios)
        if (
            not SMOOTHED_RATIO_MIN <= median_ratio <= SMOOTHED_RATIO_MAX
            or max_deviation >= SMOOTHED_MAX_DEVIATION_LIMIT
        ):
            failures.append(
                FailureRecord(
                    target=target,
                    sample_name="",
                    issue="SMOOTHED_INTENSITY_DRIFT",
                    reason=(
                        "Like-for-like smoothed apex intensity drift exceeded "
                        "threshold."
                    ),
                    metric_value=(
                        f"median_ratio={median_ratio:.4f}; "
                        f"max_deviation={max_deviation:.4f}"
                    ),
                    limit=(
                        f"median in [{SMOOTHED_RATIO_MIN:.2f}, "
                        f"{SMOOTHED_RATIO_MAX:.2f}]; "
                        f"max_deviation<{SMOOTHED_MAX_DEVIATION_LIMIT:.2f}"
                    ),
                )
            )

    nl_pct = _nl_agreement_pct(rows)
    if nl_pct is not None and nl_pct < NL_AGREEMENT_MIN_PCT:
        failures.append(
            FailureRecord(
                target=target,
                sample_name="",
                issue="NL_STATUS_MISMATCH",
                reason="NL status agreement below threshold after legacy mapping.",
                metric_value=f"{nl_pct:.1f}%",
                limit=f">={NL_AGREEMENT_MIN_PCT:.1f}%",
            )
        )
    return failures


def _target_report(
    target: str,
    rows: list[ValidationRow],
    failures: list[FailureRecord],
) -> TargetReport:
    rt_values = [value for _row, value in _rt_deltas(rows)]
    smoothed_values = [value for _row, value in _smoothed_ratios(rows)]
    return TargetReport(
        target=target,
        status="FAIL" if failures else "PASS",
        n_rows=len(rows),
        median_rt_delta=_rounded_median(rt_values),
        max_rt_delta=round(max(rt_values), 4) if rt_values else None,
        smoothed_median_ratio=_rounded_median(smoothed_values),
        smoothed_max_deviation=(
            round(max(abs(value - 1.0) for value in smoothed_values), 4)
            if smoothed_values
            else None
        ),
        nl_agreement_pct=_nl_agreement_pct(rows),
    )


def _rt_deltas(rows: list[ValidationRow]) -> list[tuple[ValidationRow, float]]:
    return [
        (row, abs(row.rt_new - row.rt_old))
        for row in rows
        if row.rt_new is not None and row.rt_old is not None
    ]


def _smoothed_ratios(rows: list[ValidationRow]) -> list[tuple[ValidationRow, float]]:
    return [
        (row, row.int_new_smoothed / row.int_old)
        for row in rows
        if row.int_new_smoothed is not None
        and row.int_old is not None
        and row.int_old > 0
    ]


def _nl_agreement_pct(rows: list[ValidationRow]) -> float | None:
    comparable = [
        row for row in rows if row.nl_old != "" and row.nl_new not in {"", "ERROR"}
    ]
    if not comparable:
        return None
    agree = sum(
        1 for row in comparable if _legacy_nl(row.nl_new) == _legacy_nl(row.nl_old)
    )
    return round(agree / len(comparable) * 100.0, 1)


def _new_peak_ok(row: ValidationRow) -> bool:
    return row.rt_new is not None and row.int_new_raw is not None


def _old_peak_ok(row: ValidationRow) -> bool:
    return row.rt_old is not None and row.int_old is not None


def _legacy_nl(token: str) -> str:
    if token in {"NL_FAIL", "NO_MS2"}:
        return "ND"
    if token.startswith("WARN_"):
        return "WARN"
    return token


def _rounded_median(values: list[float]) -> float | None:
    return round(median(values), 4) if values else None


def _safe_float(value: str | None) -> float | None:
    if value is None or value in {"", "ND", "ERROR"}:
        return None
    try:
        return float(value)
    except ValueError:
        return None


def _write_summary_sheet(ws, report: ValidationReport) -> None:
    headers = [
        "Target",
        "Status",
        "Rows",
        "Median RT Δ",
        "Max RT Δ",
        "Smoothed Median Ratio",
        "Smoothed Max Deviation",
        "NL Agreement %",
    ]
    _write_header(ws, headers)
    for row_idx, target in enumerate(report.targets, start=2):
        values = [
            target.target,
            target.status,
            target.n_rows,
            target.median_rt_delta,
            target.max_rt_delta,
            target.smoothed_median_ratio,
            target.smoothed_max_deviation,
            target.nl_agreement_pct,
        ]
        _write_row(ws, row_idx, values, target.status == "FAIL")
    _autosize(ws, len(headers))


def _write_rows_sheet(ws, rows: list[ValidationRow]) -> None:
    _write_header(ws, _VALIDATION_HEADERS)
    for row_idx, row in enumerate(rows, start=2):
        _write_row(
            ws,
            row_idx,
            [
                row.sample_name,
                row.target,
                row.rt_old,
                row.int_old,
                row.nl_old,
                row.rt_new,
                row.int_new_raw,
                row.int_new_smoothed,
                row.area_new,
                row.peak_start_new,
                row.peak_end_new,
                row.nl_new,
            ],
            False,
        )
    ws.auto_filter.ref = f"A1:L{max(1, len(rows) + 1)}"
    _autosize(ws, len(_VALIDATION_HEADERS))


def _write_fail_sheet(ws, failures: list[FailureRecord]) -> None:
    _write_header(ws, _FAIL_HEADERS)
    for row_idx, failure in enumerate(failures, start=2):
        _write_row(
            ws,
            row_idx,
            [
                failure.target,
                failure.sample_name,
                failure.issue,
                failure.reason,
                failure.metric_value,
                failure.limit,
                "",
                "",
                "",
                "",
            ],
            True,
        )
    ws.auto_filter.ref = f"A1:J{max(1, len(failures) + 1)}"
    _autosize(ws, len(_FAIL_HEADERS))


def _write_header(ws, headers: list[str]) -> None:
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border = _BORDER
    ws.freeze_panes = "A2"


def _write_row(ws, row_idx: int, values: list[object], failed: bool) -> None:
    fill = _FAIL_FILL if failed else _PASS_FILL
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = Font(name="Arial")
        cell.alignment = _CENTER
        cell.border = _BORDER
        if failed:
            cell.fill = fill


def _autosize(ws, n_cols: int) -> None:
    for col_idx in range(1, n_cols + 1):
        letter = get_column_letter(col_idx)
        width = max(
            12,
            min(
                60,
                max(
                    len(str(ws.cell(row=row_idx, column=col_idx).value or ""))
                    for row_idx in range(1, ws.max_row + 1)
                )
                + 2,
            ),
        )
        ws.column_dimensions[letter].width = width


def _stage_cases(worktree: Path, cases: list[ValidationCase]) -> Path:
    data_dir = worktree / "local_validation_raw"
    data_dir.mkdir(parents=True, exist_ok=True)
    for stale_raw in data_dir.glob("*.raw"):
        stale_raw.unlink()
    for case in cases:
        dst = data_dir / f"{case.name}{case.path.suffix}"
        shutil.copy2(case.path, dst)
    _update_settings_value(worktree / "config" / "settings.csv", "data_dir", data_dir)
    return data_dir


def _update_settings_value(path: Path, key: str, value: Path) -> None:
    if not path.exists():
        raise FileNotFoundError(f"{path}: settings.csv is required for validation")
    with path.open(newline="", encoding="utf-8-sig") as handle:
        rows = list(csv.DictReader(handle))
        fieldnames = list(rows[0].keys()) if rows else ["key", "value", "description"]
    found = False
    for row in rows:
        if row.get("key") == key:
            row["value"] = str(value)
            found = True
            break
    if not found:
        rows.append({"key": key, "value": str(value), "description": ""})
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def _run_old_pipeline(old_worktree: Path) -> Path:
    script = old_worktree / "scripts" / "01_extract_xic.ps1"
    completed = subprocess.run(
        [
            "powershell",
            "-ExecutionPolicy",
            "Bypass",
            "-File",
            str(script),
            "-RootDir",
            str(old_worktree),
        ],
        cwd=old_worktree,
        text=True,
        capture_output=True,
        check=False,
    )
    if completed.returncode != 0:
        raise RuntimeError(
            "old PS1 pipeline failed:\n"
            f"STDOUT:\n{completed.stdout}\nSTDERR:\n{completed.stderr}"
        )
    return old_worktree / "output" / "xic_results.csv"


def _print_report(output_path: Path, report: ValidationReport) -> None:
    print(f"Validation workbook: {output_path}")
    for target in report.targets:
        print(f"{target.status}: {target.target}")
    if report.failed:
        print("Failures:")
        for failure in report.failures:
            print(f"  {failure.target} {failure.issue}: {failure.metric_value}")


if __name__ == "__main__":
    raise SystemExit(main())
