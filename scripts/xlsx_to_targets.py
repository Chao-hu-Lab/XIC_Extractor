"""Convert RT-check Excel tables to XIC Extractor targets.csv format.

Usage examples
--------------
# DNA sheet only (default)
uv run python scripts/xlsx_to_targets.py path/to/RT_check.xlsx

# RNA sheet
uv run python scripts/xlsx_to_targets.py path/to/RT_check.xlsx --sheet RNA

# Both sheets together
uv run python scripts/xlsx_to_targets.py path/to/RT_check.xlsx --sheet both

# Custom output path and RT window
uv run python scripts/xlsx_to_targets.py path/to/RT_check.xlsx --output config/my_targets.csv --rt-window 1.0
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
from difflib import SequenceMatcher
from pathlib import Path
from typing import Optional

import openpyxl

# Neutral-loss masses
NL_DR: float = 116.0474  # deoxyribose  – DNA nucleosides
NL_R: float = 132.0423  # ribose        – RNA nucleosides
NL_MER: float = 146.0579  # 2′-O-methylribose – 2′-OMe RNA nucleosides

# Default target.csv field values
DEFAULT_PPM_TOL: int = 20
DEFAULT_NL_PPM_WARN: int = 20
DEFAULT_NL_PPM_MAX: int = 50
DEFAULT_RT_WINDOW: float = 1.5  # minutes

# Common ISTD label prefixes (deuterium, 15N, 13C labels)
_ISTD_PREFIXES: tuple[str, ...] = (
    "d2-",
    "d3-",
    "d4-",
    "d5-",
    "d6-",
    "d7-",
    "d8-",
    "d9-",
    "15N",
    "13C",
    "[13C",
    "[15N",
)

TargetRow = dict[str, str | float | int | bool]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _is_istd(label: str) -> bool:
    """Return True if the label looks like an internal standard."""
    return any(label.startswith(prefix) for prefix in _ISTD_PREFIXES)


def _is_number(val: object) -> bool:
    """Return True if *val* is a usable numeric value (not None or '-')."""
    return isinstance(val, (int, float))


def _find_header_row(
    rows: list[tuple],
    required_cols: list[str],
) -> Optional[int]:
    """Return 0-based row index of the first row containing all *required_cols*."""
    for i, row in enumerate(rows):
        row_strs = {str(v) for v in row if v is not None}
        if all(col in row_strs for col in required_cols):
            return i
    return None


def _col_index(header: tuple, name: str) -> Optional[int]:
    """Return 0-based index of the first occurrence of *name* in *header*."""
    for i, val in enumerate(header):
        if val == name:
            return i
    return None


def _all_col_indices(header: tuple, name: str) -> list[int]:
    """Return all 0-based indices of *name* in *header*."""
    return [i for i, v in enumerate(header) if v == name]


# ---------------------------------------------------------------------------
# ISTD pair matching
# ---------------------------------------------------------------------------


def _strip_istd_prefix(label: str) -> str:
    """Remove isotope/deuterium prefix to expose the core compound name.

    Handles formats like:
    - ``d3-5-hmdC``          → ``5-hmdC``
    - ``15N5-8-oxodG``       → ``8-oxodG``
    - ``[13C,15N2]-8-oxo-Guo`` → ``8-oxo-Guo``
    """
    # Bracket-style: [13C,15N2]-
    if label.startswith("["):
        idx = label.find("]-")
        if idx >= 0:
            return label[idx + 2 :]
    # Prefix-style: d3-, d4-, 15N5-, 13C2-, etc.
    m = re.match(r"^(?:d\d+|(?:\d+[CN]\d*-))+", label)
    if m:
        return label[m.end() :]
    return label


def _normalize_name(name: str) -> str:
    """Strip all non-letter characters and lowercase for fuzzy matching."""
    return re.sub(r"[^a-zA-Z]", "", name).lower()


_MAX_RT_DIFF_MIN: float = (
    3.0  # hard RT cutoff – reject any pair further apart than this
)


def _assign_istd_pairs(targets: list[TargetRow]) -> None:
    """Fill ``istd_pair`` on standard rows by matching against ISTD names.

    Algorithm:
    1. For each ISTD, strip isotope prefix and normalize the name.
    2. Reject candidates whose RT differs by more than ``_MAX_RT_DIFF_MIN``.
    3. Score remaining standards by name similarity (SequenceMatcher).
    4. Assign if the best score meets the 0.70 threshold.
    5. Process ISTDs in descending name-score order so clear matches claim
       their standard before ambiguous ones compete.

    Modifications are made in-place; ISTD rows keep ``istd_pair`` empty.
    Unmatched ISTDs are printed as warnings so the user can pair them manually.
    """
    standards = [t for t in targets if t["is_istd"] == "false"]
    istds = [t for t in targets if t["is_istd"] == "true"]

    std_norms = [_normalize_name(str(s["label"])) for s in standards]

    def _best_name_score(istd: TargetRow) -> float:
        core = _normalize_name(_strip_istd_prefix(str(istd["label"])))
        return max(
            (SequenceMatcher(None, core, n).ratio() for n in std_norms), default=0.0
        )

    for istd in sorted(istds, key=_best_name_score, reverse=True):
        core = _strip_istd_prefix(str(istd["label"]))
        norm_core = _normalize_name(core)
        istd_rt = (float(istd["rt_min"]) + float(istd["rt_max"])) / 2

        best_std: Optional[TargetRow] = None
        best_score = 0.0

        for std, norm_std in zip(standards, std_norms):
            if std["istd_pair"]:  # already claimed by another ISTD
                continue
            std_rt = (float(std["rt_min"]) + float(std["rt_max"])) / 2
            if abs(istd_rt - std_rt) > _MAX_RT_DIFF_MIN:  # hard RT gate
                continue
            name_sim = SequenceMatcher(None, norm_core, norm_std).ratio()
            if name_sim < 0.70:
                continue
            if name_sim > best_score:
                best_score = name_sim
                best_std = std

        if best_std is not None:
            best_std["istd_pair"] = str(istd["label"])
        else:
            print(
                f"  Warning: no standard matched for ISTD '{istd['label']}' "
                f"(RT {istd_rt:.2f} min) – fill istd_pair manually",
                file=sys.stderr,
            )


# ---------------------------------------------------------------------------
# DNA parser
# ---------------------------------------------------------------------------


def parse_dna_sheet(
    ws: openpyxl.worksheet.worksheet.Worksheet, rt_window: float
) -> list[TargetRow]:
    """Parse the DNA sheet and return target rows.

    Expected header columns (in any block):
    - ``m/z``               → precursor m/z
    - ``Anticipated adducts`` → compound label
    - ``RT``                → retention time (min)

    Neutral loss is always dR (116.0474 Da) for DNA nucleosides.
    """
    rows = list(ws.iter_rows(values_only=True))
    header_idx = _find_header_row(rows, ["m/z", "Anticipated adducts"])
    if header_idx is None:
        print(
            "Warning: DNA sheet – could not find header row with 'm/z' and 'Anticipated adducts'",
            file=sys.stderr,
        )
        return []

    header = rows[header_idx]
    adducts_cols = _all_col_indices(header, "Anticipated adducts")
    mz_all = _all_col_indices(header, "m/z")
    rt_all = _all_col_indices(header, "RT")

    targets: list[TargetRow] = []

    for adducts_col in adducts_cols:
        # Nearest m/z column *before* this block
        mz_col = max((c for c in mz_all if c < adducts_col), default=None)
        # First RT column *after* this block
        rt_col = min((c for c in rt_all if c > adducts_col), default=None)

        if mz_col is None or rt_col is None:
            continue

        max_col = max(adducts_col, mz_col, rt_col)
        for row in rows[header_idx + 1 :]:
            if len(row) <= max_col:
                continue
            label_raw = row[adducts_col]
            mz_raw = row[mz_col]
            rt_raw = row[rt_col]

            if not label_raw or not _is_number(mz_raw) or not _is_number(rt_raw):
                continue

            label = str(label_raw).strip()
            mz = float(mz_raw)
            rt = float(rt_raw)

            targets.append(
                {
                    "label": label,
                    "mz": round(mz, 6),
                    "rt_min": round(rt - rt_window, 2),
                    "rt_max": round(rt + rt_window, 2),
                    "ppm_tol": DEFAULT_PPM_TOL,
                    "neutral_loss_da": NL_DR,
                    "nl_ppm_warn": DEFAULT_NL_PPM_WARN,
                    "nl_ppm_max": DEFAULT_NL_PPM_MAX,
                    "is_istd": "true" if _is_istd(label) else "false",
                    "istd_pair": "",
                }
            )

    return targets


# ---------------------------------------------------------------------------
# RNA parser
# ---------------------------------------------------------------------------


def parse_rna_sheet(
    ws: openpyxl.worksheet.worksheet.Worksheet, rt_window: float
) -> list[TargetRow]:
    """Parse the RNA sheet and return target rows.

    Expected header columns:
    - ``Abbre.``         → compound abbreviation (label)
    - ``[M+H]+``         → precursor m/z
    - ``[M+H-R]+``       → ribose NL product (132.0423 Da) – numeric if applicable
    - ``[M+H-R-CH2]+``   → 2′-OMe-ribose NL product (146.0579 Da) – numeric if applicable
    - ``RT``             → first RT column

    Compounds with neither NL column populated are skipped (cannot be
    validated by MS2 diagnostic).
    """
    rows = list(ws.iter_rows(values_only=True))
    header_idx = _find_header_row(rows, ["Abbre.", "[M+H]+"])
    if header_idx is None:
        print(
            "Warning: RNA sheet – could not find header row with 'Abbre.' and '[M+H]+'",
            file=sys.stderr,
        )
        return []

    header = rows[header_idx]
    abbre_col = _col_index(header, "Abbre.")
    mh_col = _col_index(header, "[M+H]+")
    mhr_col = _col_index(header, "[M+H-R]+")
    mhrcm_col = _col_index(header, "[M+H-R-CH2]+")
    rt_cols = _all_col_indices(header, "RT")
    rt_col = min(rt_cols) if rt_cols else None

    if None in (abbre_col, mh_col, rt_col):
        print("Warning: RNA sheet – missing required columns", file=sys.stderr)
        return []

    max_col = max(
        c for c in [abbre_col, mh_col, mhr_col, mhrcm_col, rt_col] if c is not None
    )
    targets: list[TargetRow] = []

    for row in rows[header_idx + 1 :]:
        if len(row) <= max_col:
            continue
        label_raw = row[abbre_col]
        mz_raw = row[mh_col]
        rt_raw = row[rt_col]

        if not label_raw or not _is_number(mz_raw) or not _is_number(rt_raw):
            continue

        label = str(label_raw).strip()
        mz = float(mz_raw)
        rt = float(rt_raw)

        # Determine neutral loss: R (132.0423) or MeR (146.0579)
        nl: Optional[float] = None
        if mhr_col is not None and _is_number(row[mhr_col]):
            nl = NL_R
        elif mhrcm_col is not None and _is_number(row[mhrcm_col]):
            nl = NL_MER

        if nl is None:
            # No MS2 diagnostic NL available – skip
            continue

        targets.append(
            {
                "label": label,
                "mz": round(mz, 6),
                "rt_min": round(rt - rt_window, 2),
                "rt_max": round(rt + rt_window, 2),
                "ppm_tol": DEFAULT_PPM_TOL,
                "neutral_loss_da": nl,
                "nl_ppm_warn": DEFAULT_NL_PPM_WARN,
                "nl_ppm_max": DEFAULT_NL_PPM_MAX,
                "is_istd": "true" if _is_istd(label) else "false",
                "istd_pair": "",
            }
        )

    return targets


# ---------------------------------------------------------------------------
# Writer
# ---------------------------------------------------------------------------


_FIELDNAMES = [
    "label",
    "mz",
    "rt_min",
    "rt_max",
    "ppm_tol",
    "neutral_loss_da",
    "nl_ppm_warn",
    "nl_ppm_max",
    "is_istd",
    "istd_pair",
]


def write_targets_csv(targets: list[TargetRow], output_path: Path) -> None:
    """Write targets list to a CSV file compatible with XIC Extractor."""
    with output_path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=_FIELDNAMES)
        writer.writeheader()
        writer.writerows(targets)
    print(f"Written {len(targets)} targets to {output_path}")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def _pick_xlsx_via_dialog() -> Path:
    """Open a file-chooser dialog and return the selected path.

    Falls back to a terminal prompt when tkinter is unavailable.
    """
    try:
        import tkinter as tk
        from tkinter import filedialog

        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        path = filedialog.askopenfilename(
            title="選擇 RT 標準品 Excel 檔案",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")],
        )
        root.destroy()
        if not path:
            print("未選擇檔案，結束。", file=sys.stderr)
            sys.exit(0)
        return Path(path)
    except Exception:
        path_str = input("請輸入 Excel 檔案路徑：").strip().strip('"')
        return Path(path_str)


def _build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="Convert RT-check Excel table to XIC Extractor targets.csv",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    p.add_argument(
        "xlsx",
        nargs="?",
        type=Path,
        help="Input Excel file (.xlsx)；省略時跳出選檔視窗",
    )
    p.add_argument(
        "--sheet",
        choices=["DNA", "RNA", "both"],
        default="DNA",
        help="Which sheet to parse",
    )
    p.add_argument(
        "--output",
        type=Path,
        default=None,
        help=(
            "Output CSV path.  Default: same folder as the input xlsx, "
            "named <stem>_targets_<SHEET>.csv"
        ),
    )
    p.add_argument(
        "--rt-window",
        type=float,
        default=DEFAULT_RT_WINDOW,
        metavar="MIN",
        help="RT window (± minutes) around observed RT",
    )
    p.add_argument(
        "--ppm-tol",
        type=int,
        default=DEFAULT_PPM_TOL,
        help="XIC precursor m/z tolerance (ppm)",
    )
    return p


def main() -> None:
    parser = _build_parser()
    args = parser.parse_args()

    # If no xlsx supplied, open a file-chooser dialog
    xlsx_path: Path = args.xlsx if args.xlsx is not None else _pick_xlsx_via_dialog()
    if not xlsx_path.exists():
        print(f"Error: file not found: {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    print(f"Reading {xlsx_path} …")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=False)

    targets: list[TargetRow] = []

    if args.sheet in ("DNA", "both"):
        if "DNA" not in wb.sheetnames:
            print("Warning: 'DNA' sheet not found in workbook", file=sys.stderr)
        else:
            dna_targets = parse_dna_sheet(wb["DNA"], args.rt_window)
            print(f"  DNA: {len(dna_targets)} compounds found")
            targets.extend(dna_targets)

    if args.sheet in ("RNA", "both"):
        if "RNA" not in wb.sheetnames:
            print("Warning: 'RNA' sheet not found in workbook", file=sys.stderr)
        else:
            rna_targets = parse_rna_sheet(wb["RNA"], args.rt_window)
            print(f"  RNA: {len(rna_targets)} compounds found")
            targets.extend(rna_targets)

    if not targets:
        print("No targets extracted. Check warnings above.", file=sys.stderr)
        sys.exit(1)

    _assign_istd_pairs(targets)
    paired = sum(1 for t in targets if t["istd_pair"])
    print(f"  ISTD pairs matched: {paired}")

    # Resolve output path:
    #   explicit --output  →  use as-is
    #   default            →  project config/ dir, named after the xlsx stem + sheet
    #                         (always lands next to settings.csv, never overwrites targets.csv)
    if args.output is not None:
        output_path = args.output
    else:
        config_dir = Path(__file__).resolve().parent.parent / "config"
        output_path = config_dir / f"{xlsx_path.stem}_targets_{args.sheet}.csv"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    write_targets_csv(targets, output_path)


if __name__ == "__main__":
    main()
