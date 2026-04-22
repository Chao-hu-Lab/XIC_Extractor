import csv
import gc
from collections.abc import Callable
from dataclasses import dataclass, replace
from pathlib import Path
from statistics import median
from typing import Any, Literal, cast

import numpy as np
from openpyxl import Workbook
from scipy.signal import peak_widths

from xic_extractor.config import ExtractionConfig, Target
from xic_extractor.injection_rolling import read_injection_order, rolling_median_rt
from xic_extractor.neutral_loss import NLResult, check_nl, find_nl_anchor_rt
from xic_extractor.peak_scoring import (
    ScoringContext,
    candidate_quality_penalty,
    compute_local_sn_cache,
)
from xic_extractor.raw_reader import RawReaderError, open_raw, preflight_raw_reader
from xic_extractor.rt_prior_library import LibraryEntry, load_library
from xic_extractor.sample_groups import classify_sample_group
from xic_extractor.signal_processing import (
    PeakCandidate,
    PeakDetectionResult,
    PeakResult,
    find_peak_and_area,
)

DiagnosticIssue = Literal[
    "PEAK_NOT_FOUND",
    "NO_SIGNAL",
    "WINDOW_TOO_SHORT",
    "NL_FAIL",
    "NO_MS2",
    "FILE_ERROR",
    "MULTI_PEAK",
    "TAILING",
    "NL_ANCHOR_FALLBACK",
    "ISTD_ANCHOR_MISSING",
    "ANCHOR_RT_MISMATCH",
    "ISTD_CONFIDENCE",
]

# Asymmetry ratio (right / left half-width at 5 % peak height) above this
# threshold triggers a TAILING warning.  USP guideline is 2.0 for quantification.
_TAILING_THRESHOLD: float = 2.0
# paired analyte 有自己的 NL anchor 時，以 target anchor 作為最直接的證據。
_PAIRED_TARGET_ANCHOR_PEAK_DELTA_MAX_MIN: float = 0.25
# target anchor 缺失時才退回 ISTD anchor，門檻較寬但仍擋掉明顯旁峰。
_PAIRED_FALLBACK_ISTD_PEAK_DELTA_MAX_MIN: float = 0.5
# 非 paired/非拒絕情境下，選出的峰 RT 距 NL anchor 超過此距離時發出警告。
_ANCHOR_PEAK_DELTA_WARN_MIN: float = 0.5

_MS1_SUFFIXES = ("RT", "Int", "Area", "PeakStart", "PeakEnd", "PeakWidth")
_DIAGNOSTIC_FIELDS = ("SampleName", "Target", "Issue", "Reason")
_LONG_OUTPUT_FIELDS = (
    "SampleName",
    "Group",
    "Target",
    "Role",
    "ISTD Pair",
    "RT",
    "Area",
    "NL",
    "Int",
    "PeakStart",
    "PeakEnd",
    "PeakWidth",
    "Confidence",
    "Reason",
)
_SCORE_BREAKDOWN_FIELDS = (
    "SampleName",
    "Target",
    "symmetry",
    "local_sn",
    "nl_support",
    "rt_prior",
    "rt_centrality",
    "noise_shape",
    "peak_width",
    "Quality Penalty",
    "Quality Flags",
    "Total Severity",
    "Confidence",
    "Prior RT",
    "Prior Source",
)


@dataclass(frozen=True)
class DiagnosticRecord:
    sample_name: str
    target_label: str
    issue: DiagnosticIssue
    reason: str


@dataclass(frozen=True)
class ExtractionResult:
    peak_result: PeakDetectionResult
    nl: NLResult | None
    target_label: str = ""
    role: str = ""
    istd_pair: str = ""
    confidence: str = ""
    reason: str = ""
    severities: tuple[tuple[int, str], ...] = ()
    prior_rt: float | None = None
    prior_source: str = ""
    quality_penalty: int = 0
    quality_flags: tuple[str, ...] = ()

    @property
    def peak(self) -> PeakResult | None:
        return self.peak_result.peak

    @property
    def nl_result(self) -> NLResult | None:
        return self.nl

    @property
    def total_severity(self) -> int:
        return sum(severity for severity, _ in self.severities) + self.quality_penalty


@dataclass
class FileResult:
    sample_name: str
    results: dict[str, ExtractionResult]
    group: str | None = None
    error: str | None = None

    @property
    def extraction_results(self) -> list[ExtractionResult]:
        return list(self.results.values())


@dataclass
class RunOutput:
    file_results: list[FileResult]
    diagnostics: list[DiagnosticRecord]


def run(
    config: ExtractionConfig,
    targets: list[Target],
    progress_callback: Callable[[int, int, str], None] | None = None,
    should_stop: Callable[[], bool] | None = None,
    injection_order: dict[str, int] | None = None,
    rt_prior_library: dict[tuple[str, str], LibraryEntry] | None = None,
) -> RunOutput:
    reader_errors = preflight_raw_reader(config.dll_dir)
    if reader_errors:
        raise RawReaderError(" ".join(reader_errors))

    raw_paths = sorted(config.data_dir.glob("*.raw"))
    resolved_injection_order = _resolve_injection_order(
        config, raw_paths, injection_order
    )
    resolved_rt_prior_library = _resolve_rt_prior_library(config, rt_prior_library)
    istd_targets = [target for target in targets if target.is_istd]
    istd_rts_by_sample: dict[str, dict[str, float]] = {}
    for raw_path in raw_paths:
        if should_stop is not None and should_stop():
            break
        prepass = _extract_istd_anchors_only(config, istd_targets, raw_path)
        if prepass is None:
            continue
        anchors, _, _, _ = prepass
        for istd_label, anchor_rt in anchors.items():
            istd_rts_by_sample.setdefault(istd_label, {})[raw_path.stem] = anchor_rt

    scoring_context_factory = _build_scoring_context_factory(
        config=config,
        injection_order=(
            resolved_injection_order
            if resolved_injection_order is not None
            else _fallback_injection_order_from_mtime(raw_paths)
        ),
        istd_rts_by_sample=istd_rts_by_sample,
        rt_prior_library=resolved_rt_prior_library or {},
    )

    file_results: list[FileResult] = []
    diagnostics: list[DiagnosticRecord] = []
    total = len(raw_paths)

    for index, raw_path in enumerate(raw_paths, start=1):
        if should_stop is not None and should_stop():
            break

        file_result, file_diagnostics = _process_file(
            config,
            targets,
            raw_path,
            scoring_context_factory=scoring_context_factory,
        )
        file_results.append(file_result)
        diagnostics.extend(file_diagnostics)

        if progress_callback is not None:
            progress_callback(index, total, raw_path.name)
        if index % 50 == 0:
            gc.collect()

    output = RunOutput(file_results=file_results, diagnostics=diagnostics)
    _write_output_csv(config, targets, file_results)
    _write_long_output_csv(config, targets, file_results)
    _write_diagnostics_csv(config, diagnostics)
    if config.emit_score_breakdown:
        _write_score_breakdown_csv(config, file_results)
    return output


def _process_file(
    config: ExtractionConfig,
    targets: list[Target],
    raw_path: Path,
    *,
    scoring_context_factory: Callable[..., Any] | None = None,
    precomputed_istd_results: dict[str, ExtractionResult] | None = None,
    precomputed_istd_diagnostics: list[DiagnosticRecord] | None = None,
    precomputed_istd_anchor_rts: dict[str, float] | None = None,
    precomputed_istd_shape_metrics: dict[str, tuple[float, float | None]] | None = None,
) -> tuple[FileResult, list[DiagnosticRecord]]:
    sample_name = raw_path.stem
    try:
        with open_raw(raw_path, config.dll_dir) as raw:
            results: dict[str, ExtractionResult] = dict(precomputed_istd_results or {})
            diagnostics: list[DiagnosticRecord] = list(
                precomputed_istd_diagnostics or []
            )
            istd_shape_metrics_by_label: dict[str, tuple[float, float | None]] = dict(
                precomputed_istd_shape_metrics or {}
            )
            istd_confidence_by_label = {
                label: result.peak_result.confidence
                for label, result in results.items()
                if result.peak_result.confidence is not None
            }

            if precomputed_istd_anchor_rts is None:
                istd_anchor_rts: dict[str, float] = {}
                for target in targets:
                    if not target.is_istd:
                        continue
                    anchor_rt = _extract_one_target(
                        raw,
                        config,
                        sample_name,
                        target,
                        reference_rt=None,
                        strict_preferred_rt=False,
                        results=results,
                        diagnostics=diagnostics,
                        scoring_context_factory=scoring_context_factory,
                        shape_metrics_by_label=istd_shape_metrics_by_label,
                    )
                    confidence = results[target.label].peak_result.confidence
                    if confidence is not None:
                        istd_confidence_by_label[target.label] = confidence
                    if anchor_rt is not None:
                        istd_anchor_rts[target.label] = anchor_rt
            else:
                istd_anchor_rts = dict(precomputed_istd_anchor_rts)

            sample_drift = _estimate_sample_drift(targets, istd_anchor_rts)

            # Pass 2: analytes — reference_rt 依 ISTD pair 決定
            # 有 ISTD pair 且 anchor 成功 → 選最靠近 ISTD anchor_rt 的 scan
            # 有 ISTD pair 但 anchor 失敗 → 發 ISTD_ANCHOR_MISSING
            # 無 ISTD pair → reference_rt=None，選 base_peak 最高的 scan
            for target in targets:
                if target.is_istd:
                    continue
                if target.istd_pair and target.istd_pair in istd_anchor_rts:
                    reference_rt: float | None = istd_anchor_rts[target.istd_pair]
                else:
                    if target.istd_pair:
                        diagnostics.append(
                            DiagnosticRecord(
                                sample_name=sample_name,
                                target_label=target.label,
                                issue="ISTD_ANCHOR_MISSING",
                                reason=(
                                    f"ISTD '{target.istd_pair}' yielded no "
                                    f"NL-confirmed anchor in this sample; "
                                    f"falling back to highest base_peak — isobar "
                                    f"discrimination disabled"
                                ),
                            )
                        )
                    reference_rt = None
                _extract_one_target(
                    raw,
                    config,
                    sample_name,
                    target,
                    reference_rt=reference_rt,
                    sample_drift=sample_drift,
                    strict_preferred_rt=reference_rt is not None,
                    results=results,
                    diagnostics=diagnostics,
                    scoring_context_factory=scoring_context_factory,
                    istd_confidence_note=_istd_confidence_note(
                        istd_confidence_by_label.get(target.istd_pair)
                    ),
                    istd_rt_in_this_sample=istd_anchor_rts.get(target.istd_pair),
                    paired_istd_fwhm=_paired_istd_fwhm(
                        target,
                        istd_shape_metrics_by_label,
                    ),
                )

            return FileResult(sample_name=sample_name, results=results), diagnostics
    except Exception as exc:
        reason = f"Failed to open .raw: {type(exc).__name__}: {exc}"
        return (
            FileResult(sample_name=sample_name, results={}, error=reason),
            [
                DiagnosticRecord(
                    sample_name=sample_name,
                    target_label="",
                    issue="FILE_ERROR",
                    reason=reason,
                )
            ],
        )


def _extract_one_target(
    raw: Any,
    config: ExtractionConfig,
    sample_name: str,
    target: Target,
    *,
    reference_rt: float | None,
    sample_drift: float = 0.0,
    strict_preferred_rt: bool = False,
    results: dict[str, ExtractionResult],
    diagnostics: list[DiagnosticRecord],
    scoring_context_factory: Callable[..., Any] | None = None,
    istd_confidence_note: str | None = None,
    istd_rt_in_this_sample: float | None = None,
    paired_istd_fwhm: float | None = None,
    shape_metrics_by_label: dict[str, tuple[float, float | None]] | None = None,
) -> float | None:
    """處理單一 target 並將結果寫入 results/diagnostics。

    回傳 anchor_rt（若無則 None）。
    """
    rt_min, rt_max, anchor_used, anchor_rt = _get_rt_window(
        raw, target, config, reference_rt=reference_rt, sample_drift=sample_drift
    )
    rt, intensity = raw.extract_xic(target.mz, rt_min, rt_max, target.ppm_tol)
    nl_result = _check_target_nl(raw, target, config)
    scoring_context_builder = None
    if scoring_context_factory is not None:
        scoring_context_builder = scoring_context_factory(
            target=target,
            sample_name=sample_name,
            rt=rt,
            intensity=intensity,
            istd_rt_in_this_sample=istd_rt_in_this_sample,
            paired_istd_fwhm=paired_istd_fwhm,
            nl_result=nl_result,
        )
    if scoring_context_builder is not None:
        try:
            peak_result = find_peak_and_area(
                rt,
                intensity,
                config,
                preferred_rt=anchor_rt,
                strict_preferred_rt=strict_preferred_rt,
                scoring_context_builder=scoring_context_builder,
                istd_confidence_note=istd_confidence_note,
            )
        except TypeError as exc:
            if (
                "scoring_context_builder" not in str(exc)
                and "istd_confidence_note" not in str(exc)
            ):
                raise
            peak_result = find_peak_and_area(
                rt,
                intensity,
                config,
                preferred_rt=anchor_rt,
                strict_preferred_rt=strict_preferred_rt,
            )
    else:
        peak_result = find_peak_and_area(
            rt,
            intensity,
            config,
            preferred_rt=anchor_rt,
            strict_preferred_rt=strict_preferred_rt,
        )
    if (
        peak_result.status == "PEAK_NOT_FOUND"
        and peak_result.peak is None
        and target.is_istd
        and anchor_used
        and anchor_rt is not None
    ):
        recovered_peak_result = _recover_istd_peak_with_wider_anchor_window(
            raw,
            config,
            target,
            anchor_rt=anchor_rt,
            scoring_context_factory=scoring_context_factory,
            sample_name=sample_name,
            nl_result=nl_result,
            istd_confidence_note=istd_confidence_note,
            istd_rt_in_this_sample=istd_rt_in_this_sample,
            paired_istd_fwhm=paired_istd_fwhm,
        )
        if recovered_peak_result is not None:
            peak_result = recovered_peak_result
    paired_rejection = _paired_anchor_mismatch_diagnostic(
        sample_name,
        target,
        peak_result,
        reference_rt=reference_rt,
        anchor_rt=anchor_rt,
        strict_preferred_rt=strict_preferred_rt,
    )
    if paired_rejection is not None:
        peak_result = _apply_anchor_mismatch_penalty(
            peak_result,
            paired_rejection.reason,
        )
    shape_metrics = _selected_shape_metrics(intensity, peak_result)
    selected_candidate = _selected_candidate(peak_result)
    quality_penalty = 0
    quality_flags: tuple[str, ...] = ()
    if selected_candidate is not None:
        quality_penalty, _ = candidate_quality_penalty(selected_candidate)
        quality_flags = tuple(
            str(flag) for flag in getattr(selected_candidate, "quality_flags", ())
        )
    if shape_metrics_by_label is not None and shape_metrics is not None:
        shape_metrics_by_label[target.label] = shape_metrics

    result = ExtractionResult(
        peak_result=peak_result,
        nl=nl_result,
        target_label=target.label,
        role="ISTD" if target.is_istd else "Analyte",
        istd_pair=target.istd_pair,
        confidence=(
            peak_result.confidence or "HIGH"
            if peak_result.peak is not None
            else ""
        ),
        reason=peak_result.reason or "",
        severities=peak_result.severities,
        prior_rt=getattr(scoring_context_builder, "rt_prior", None),
        prior_source=getattr(scoring_context_builder, "prior_source", ""),
        quality_penalty=quality_penalty,
        quality_flags=quality_flags,
    )
    results[target.label] = result
    diagnostics.extend(_build_diagnostics(sample_name, target, result, config))
    if paired_rejection is not None:
        diagnostics.append(paired_rejection)

    # 問題 7：anchor 找到，但選出的峰 RT 與 anchor 相差過大（可能 anchor 是雜訊）
    if (
        paired_rejection is None
        and anchor_rt is not None
        and peak_result.peak is not None
    ):
        delta = abs(peak_result.peak.rt - anchor_rt)
        if delta > _ANCHOR_PEAK_DELTA_WARN_MIN:
            diagnostics.append(
                DiagnosticRecord(
                    sample_name=sample_name,
                    target_label=target.label,
                    issue="ANCHOR_RT_MISMATCH",
                    reason=(
                        f"Peak RT {peak_result.peak.rt:.3f} min deviates "
                        f"{delta:.2f} min "
                        f"from NL anchor at {anchor_rt:.3f} min "
                        f"(threshold {_ANCHOR_PEAK_DELTA_WARN_MIN} min); "
                        f"anchor scan may be noise — verify manually"
                    ),
                )
            )

    # 問題 8：anchor 找不到時的 fallback；若 NL check 也失敗，一併提示
    if not anchor_used and target.neutral_loss_da is not None:
        rt_center = (target.rt_min + target.rt_max) / 2.0
        nl_note = ""
        if result.nl is not None and result.nl.status in {"NL_FAIL", "NO_MS2"}:
            nl_note = f"; NL check also {result.nl.status} within fallback window"
        diagnostics.append(
            DiagnosticRecord(
                sample_name=sample_name,
                target_label=target.label,
                issue="NL_ANCHOR_FALLBACK",
                reason=(
                    f"No NL-confirmed MS2 within RT center "
                    f"{rt_center:.2f} ± {config.nl_rt_anchor_search_margin_min} min; "
                    f"fallback window [{rt_min:.2f}, {rt_max:.2f}]{nl_note}"
                ),
            )
        )
    return anchor_rt


def _get_rt_window(
    raw: Any,
    target: Target,
    config: ExtractionConfig,
    *,
    reference_rt: float | None,
    sample_drift: float = 0.0,
) -> tuple[float, float, bool, float | None]:
    """回傳 (rt_min, rt_max, anchor_used, anchor_rt)。

    reference_rt 控制 anchor 選擇邏輯（見 find_nl_anchor_rt）：
    - None → 選最高 base_peak（ISTD 及無 ISTD pair 的 analyte）
    - float → 選最靠近 reference_rt 的 scan
      （有 ISTD pair 的 analyte，傳 ISTD anchor_rt）

    sample_drift 是從本樣本所有成功 ISTD 估計的整體 RT 偏移量，用於校正 anchor
    搜尋中心與 fallback 窗口中心，讓兩者跟著樣本實際 RT 移動。
    """
    if target.neutral_loss_da is None or target.nl_ppm_max is None:
        return target.rt_min, target.rt_max, False, None

    rt_center = (target.rt_min + target.rt_max) / 2.0 + sample_drift
    anchor_rt = find_nl_anchor_rt(
        raw,
        precursor_mz=target.mz,
        rt_center=rt_center,
        search_margin_min=config.nl_rt_anchor_search_margin_min,
        neutral_loss_da=target.neutral_loss_da,
        nl_ppm_max=target.nl_ppm_max,
        ms2_precursor_tol_da=config.ms2_precursor_tol_da,
        nl_min_intensity_ratio=config.nl_min_intensity_ratio,
        reference_rt=reference_rt,
    )
    if (
        target.is_istd
        and reference_rt is None
        and anchor_rt is not None
        and abs(anchor_rt - rt_center) > config.nl_rt_anchor_half_window_min
    ):
        centered_anchor_rt = find_nl_anchor_rt(
            raw,
            precursor_mz=target.mz,
            rt_center=rt_center,
            search_margin_min=config.nl_rt_anchor_search_margin_min,
            neutral_loss_da=target.neutral_loss_da,
            nl_ppm_max=target.nl_ppm_max,
            ms2_precursor_tol_da=config.ms2_precursor_tol_da,
            nl_min_intensity_ratio=config.nl_min_intensity_ratio,
            reference_rt=rt_center,
        )
        if (
            centered_anchor_rt is not None
            and abs(centered_anchor_rt - rt_center) < abs(anchor_rt - rt_center)
        ):
            anchor_rt = centered_anchor_rt

    if anchor_rt is not None:
        half = config.nl_rt_anchor_half_window_min
        return max(0.0, anchor_rt - half), anchor_rt + half, True, anchor_rt

    half = config.nl_fallback_half_window_min
    return max(0.0, rt_center - half), rt_center + half, False, None


def _recover_istd_peak_with_wider_anchor_window(
    raw: Any,
    config: ExtractionConfig,
    target: Target,
    *,
    anchor_rt: float,
    scoring_context_factory: Callable[..., Any] | None,
    sample_name: str,
    nl_result: NLResult | None,
    istd_confidence_note: str | None,
    istd_rt_in_this_sample: float | None,
    paired_istd_fwhm: float | None,
) -> PeakDetectionResult | None:
    wider_half_window = max(
        config.nl_fallback_half_window_min,
        config.nl_rt_anchor_half_window_min,
    )
    if wider_half_window <= config.nl_rt_anchor_half_window_min:
        return None

    rt_min = max(0.0, anchor_rt - wider_half_window)
    rt_max = anchor_rt + wider_half_window
    rt, intensity = raw.extract_xic(target.mz, rt_min, rt_max, target.ppm_tol)
    scoring_context_builder = None
    if scoring_context_factory is not None:
        scoring_context_builder = scoring_context_factory(
            target=target,
            sample_name=sample_name,
            rt=rt,
            intensity=intensity,
            istd_rt_in_this_sample=istd_rt_in_this_sample,
            paired_istd_fwhm=paired_istd_fwhm,
            nl_result=nl_result,
        )
    if scoring_context_builder is not None:
        try:
            peak_result = find_peak_and_area(
                rt,
                intensity,
                config,
                preferred_rt=anchor_rt,
                strict_preferred_rt=False,
                scoring_context_builder=scoring_context_builder,
                istd_confidence_note=istd_confidence_note,
            )
        except TypeError as exc:
            if (
                "scoring_context_builder" not in str(exc)
                and "istd_confidence_note" not in str(exc)
            ):
                raise
            peak_result = find_peak_and_area(
                rt,
                intensity,
                config,
                preferred_rt=anchor_rt,
                strict_preferred_rt=False,
            )
    else:
        peak_result = find_peak_and_area(
            rt,
            intensity,
            config,
            preferred_rt=anchor_rt,
            strict_preferred_rt=False,
        )
    if peak_result.peak is None:
        return None
    return peak_result


def _estimate_sample_drift(
    targets: list[Target], istd_anchor_rts: dict[str, float]
) -> float:
    """從本樣本成功定位的 ISTD anchor RT 估計整體 RT 偏移量。

    每個 ISTD 的偏移 = anchor_rt − configured_rt_center。
    取中位數作為 sample-level 估計，用於校正無 ISTD pair analyte 的搜尋中心。
    若無任何 ISTD anchor 可用，回傳 0.0（不校正）。
    """
    deltas: list[float] = []
    for target in targets:
        if not target.is_istd:
            continue
        anchor_rt = istd_anchor_rts.get(target.label)
        if anchor_rt is None:
            continue
        rt_center = (target.rt_min + target.rt_max) / 2.0
        deltas.append(anchor_rt - rt_center)
    return median(deltas) if deltas else 0.0


def _check_target_nl(
    raw: Any,
    target: Target,
    config: ExtractionConfig,
) -> NLResult | None:
    """NL 品質評估用 target 原始窗口，確保 matched_scan_count 完整。"""
    if target.neutral_loss_da is None:
        return None
    if target.nl_ppm_warn is None or target.nl_ppm_max is None:
        return None
    return check_nl(
        raw,
        precursor_mz=target.mz,
        rt_min=target.rt_min,
        rt_max=target.rt_max,
        neutral_loss_da=target.neutral_loss_da,
        nl_ppm_warn=target.nl_ppm_warn,
        nl_ppm_max=target.nl_ppm_max,
        ms2_precursor_tol_da=config.ms2_precursor_tol_da,
        nl_min_intensity_ratio=config.nl_min_intensity_ratio,
    )


def _paired_anchor_mismatch_diagnostic(
    sample_name: str,
    target: Target,
    peak_result: PeakDetectionResult,
    *,
    reference_rt: float | None,
    anchor_rt: float | None,
    strict_preferred_rt: bool,
) -> DiagnosticRecord | None:
    if (
        not strict_preferred_rt
        or reference_rt is None
        or peak_result.peak is None
    ):
        return None

    peak = peak_result.peak
    if anchor_rt is not None:
        expected_rt = anchor_rt
        anchor_label = "target NL anchor"
        allowed_delta = _PAIRED_TARGET_ANCHOR_PEAK_DELTA_MAX_MIN
        secondary_note = f"; ISTD anchor at {reference_rt:.3f} min"
    else:
        expected_rt = reference_rt
        anchor_label = "ISTD anchor"
        allowed_delta = _PAIRED_FALLBACK_ISTD_PEAK_DELTA_MAX_MIN
        secondary_note = "; no target NL anchor"

    delta = abs(peak.rt - expected_rt)
    if delta <= allowed_delta:
        return None

    return DiagnosticRecord(
        sample_name=sample_name,
        target_label=target.label,
        issue="ANCHOR_RT_MISMATCH",
        reason=(
            f"Paired analyte peak RT {peak.rt:.3f} min deviates "
            f"{delta:.2f} min from {anchor_label} at "
            f"{expected_rt:.3f} min (allowed ±{allowed_delta:.2f} min)"
            f"{secondary_note}; retained with downgraded confidence for manual review"
        ),
    )


def _apply_anchor_mismatch_penalty(
    peak_result: PeakDetectionResult,
    mismatch_reason: str,
) -> PeakDetectionResult:
    reason = f"anchor mismatch; {mismatch_reason}"
    if peak_result.reason:
        reason = f"{peak_result.reason}; {reason}"
    confidence = _anchor_mismatch_confidence(peak_result.confidence)
    return replace(peak_result, confidence=confidence, reason=reason)


def _anchor_mismatch_confidence(confidence: str | None) -> str:
    if confidence == "VERY_LOW":
        return "VERY_LOW"
    return "LOW"


def _build_diagnostics(
    sample_name: str,
    target: Target,
    result: ExtractionResult,
    config: ExtractionConfig,
) -> list[DiagnosticRecord]:
    records: list[DiagnosticRecord] = []
    if result.peak_result.status != "OK":
        records.append(
            DiagnosticRecord(
                sample_name=sample_name,
                target_label=target.label,
                issue=result.peak_result.status,
                reason=_peak_reason(target, result.peak_result, config),
            )
        )

    if result.nl is not None and result.nl.status in {"NL_FAIL", "NO_MS2"}:
        records.append(
            DiagnosticRecord(
                sample_name=sample_name,
                target_label=target.label,
                issue=cast(DiagnosticIssue, result.nl.status),
                reason=_nl_reason(target, result.nl, config),
            )
        )

    if result.peak_result.status == "OK" and result.peak_result.peak is not None:
        peak = result.peak_result.peak
        if result.peak_result.n_prominent_peaks > 1:
            records.append(
                DiagnosticRecord(
                    sample_name=sample_name,
                    target_label=target.label,
                    issue="MULTI_PEAK",
                    reason=_multi_peak_reason(target, result.peak_result, result.nl),
                )
            )
        left_half = peak.rt - peak.peak_start
        right_half = peak.peak_end - peak.rt
        if left_half > 0 and (right_half / left_half) > _TAILING_THRESHOLD:
            records.append(
                DiagnosticRecord(
                    sample_name=sample_name,
                    target_label=target.label,
                    issue="TAILING",
                    reason=_tailing_reason(peak),
                )
            )

        istd_confidence = _istd_confidence_diagnostic(sample_name, target, result)
        if istd_confidence is not None:
            records.append(istd_confidence)

    return records


def _peak_reason(
    target: Target, peak_result: PeakDetectionResult, config: ExtractionConfig
) -> str:
    if peak_result.status == "NO_SIGNAL":
        return (
            f"XIC empty in window [{target.rt_min}, {target.rt_max}] for "
            f"m/z {target.mz} +/- {target.ppm_tol:g} ppm"
        )
    if peak_result.status == "WINDOW_TOO_SHORT":
        return (
            f"Only {peak_result.n_points} scans in window; "
            f"savgol requires >= {config.smooth_window}"
        )
    max_value = _format_optional_number(peak_result.max_smoothed)
    prominence_pct = config.peak_min_prominence_ratio * 100
    return (
        "No peak met prominence >= "
        f"{prominence_pct:g}% of max smoothed (max={max_value})"
    )


def _nl_reason(target: Target, nl: NLResult, config: ExtractionConfig) -> str:
    if nl.status == "NO_MS2":
        return (
            f"No MS2 scan targeting precursor {target.mz} +/- "
            f"{config.ms2_precursor_tol_da:g} Da within RT "
            f"[{target.rt_min}, {target.rt_max}]; "
            f"{nl.valid_ms2_scan_count} valid MS2 scans in window "
            f"({nl.parse_error_count} parse errors)"
        )

    limit = target.nl_ppm_max if target.nl_ppm_max is not None else 0.0
    assert (
        target.neutral_loss_da is not None
    )  # NL is only checked when neutral_loss_da is set
    nl_da = target.neutral_loss_da
    expected_product = target.mz - nl_da

    if nl.best_ppm is not None:
        rt_info = (
            f" at scan RT {nl.best_scan_rt:.3f} min"
            if nl.best_scan_rt is not None
            else ""
        )
        return (
            f"Precursor {target.mz} triggered {nl.matched_scan_count} MS2 scans; "
            f"NL {nl_da:g} Da → expected product m/z {expected_product:.4f}; "
            f"best match {nl.best_ppm:.1f} ppm (limit {limit:g} ppm){rt_info}"
        )

    return (
        f"Precursor {target.mz} triggered {nl.matched_scan_count} MS2 scans; "
        f"NL {nl_da:g} Da → expected product m/z {expected_product:.4f}; "
        f"not detected in any matched scan"
    )


def _multi_peak_reason(
    target: Target, peak_result: PeakDetectionResult, nl: NLResult | None
) -> str:
    base = (
        f"{peak_result.n_prominent_peaks} prominent peaks detected in window "
        f"[{target.rt_min}, {target.rt_max}]; tallest peak selected — "
        f"verify integration window or split into separate targets"
    )
    if nl is not None and nl.best_scan_rt is not None:
        base += f"; NL-confirmed MS2 at RT {nl.best_scan_rt:.3f} min"
    return base


def _tailing_reason(peak: PeakResult) -> str:
    left_half = peak.rt - peak.peak_start
    right_half = peak.peak_end - peak.rt
    ratio = right_half / left_half if left_half > 0 else float("inf")
    return (
        f"Asymmetry ratio {ratio:.2f} (right/left half-width at 5% peak height, "
        f"USP limit 2.0); apex RT {peak.rt:.4f} min, "
        f"peak [{peak.peak_start:.4f}, {peak.peak_end:.4f}]"
    )


def _istd_confidence_diagnostic(
    sample_name: str, target: Target, result: ExtractionResult
) -> DiagnosticRecord | None:
    if not target.is_istd:
        return None

    flags: list[str] = []
    confidence = "HIGH"
    if result.nl is not None:
        if result.nl.status == "NO_MS2":
            flags.append("NO_MS2")
            confidence = "MEDIUM"
        elif result.nl.status == "NL_FAIL":
            flags.append("NL_FAIL")
            confidence = "LOW"
        elif result.nl.status == "WARN":
            flags.append("NL_WARN")
            confidence = "MEDIUM"

    if not flags:
        return None

    return DiagnosticRecord(
        sample_name=sample_name,
        target_label=target.label,
        issue="ISTD_CONFIDENCE",
        reason=(
            f"ISTD confidence={confidence}; flags={','.join(flags)}; "
            "MS1 peak retained because ISTD NL evidence is diagnostic support, "
            "not a hard detection gate"
        ),
    )


def _write_output_csv(
    config: ExtractionConfig, targets: list[Target], file_results: list[FileResult]
) -> None:
    config.output_csv.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = _output_fieldnames(targets)
    with config.output_csv.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for file_result in file_results:
            writer.writerow(_output_row(file_result, targets))


def _write_diagnostics_csv(
    config: ExtractionConfig, diagnostics: list[DiagnosticRecord]
) -> None:
    config.diagnostics_csv.parent.mkdir(parents=True, exist_ok=True)
    with config.diagnostics_csv.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=_DIAGNOSTIC_FIELDS)
        writer.writeheader()
        for record in diagnostics:
            writer.writerow(
                {
                    "SampleName": record.sample_name,
                    "Target": record.target_label,
                    "Issue": record.issue,
                    "Reason": record.reason,
                }
            )


def _write_long_output_csv(
    config: ExtractionConfig, targets: list[Target], file_results: list[FileResult]
) -> None:
    path = config.output_csv.with_name("xic_results_long.csv")
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=_LONG_OUTPUT_FIELDS)
        writer.writeheader()
        for file_result in file_results:
            writer.writerows(_long_output_rows(file_result, targets))


def _long_output_rows(
    file_result: FileResult, targets: list[Target]
) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for target in targets:
        row = {
            "SampleName": file_result.sample_name,
            "Group": _sample_group(file_result.sample_name),
            "Target": target.label,
            "Role": "ISTD" if target.is_istd else "Analyte",
            "ISTD Pair": target.istd_pair,
            "RT": "",
            "Area": "",
            "NL": "",
            "Int": "",
            "PeakStart": "",
            "PeakEnd": "",
            "PeakWidth": "",
            "Confidence": "",
            "Reason": "",
        }
        if file_result.error is not None:
            _set_long_ms1_values(row, "ERROR")
            row["NL"] = "ERROR" if target.neutral_loss_da is not None else ""
        else:
            result = file_result.results[target.label]
            _set_long_peak_values(row, result.peak_result.peak)
            row["NL"] = (
                result.nl.to_token()
                if target.neutral_loss_da is not None and result.nl is not None
                else ""
            )
            row["Confidence"] = result.confidence
            row["Reason"] = result.reason
        rows.append(row)
    return rows


def _write_score_breakdown_csv(
    config: ExtractionConfig, file_results: list[FileResult]
) -> None:
    path = config.output_csv.with_name("xic_score_breakdown.csv")
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=_SCORE_BREAKDOWN_FIELDS)
        writer.writeheader()
        for file_result in file_results:
            for result in file_result.extraction_results:
                severities = {label: severity for severity, label in result.severities}
                writer.writerow(
                    {
                        "SampleName": file_result.sample_name,
                        "Target": result.target_label,
                        "symmetry": _format_optional_severity(
                            severities.get("symmetry")
                        ),
                        "local_sn": _format_optional_severity(
                            severities.get("local_sn")
                        ),
                        "nl_support": _format_optional_severity(
                            severities.get("nl_support")
                        ),
                        "rt_prior": _format_optional_severity(
                            severities.get("rt_prior")
                        ),
                        "rt_centrality": _format_optional_severity(
                            severities.get("rt_centrality")
                        ),
                        "noise_shape": _format_optional_severity(
                            severities.get("noise_shape")
                        ),
                        "peak_width": _format_optional_severity(
                            severities.get("peak_width")
                        ),
                        "Quality Penalty": str(result.quality_penalty),
                        "Quality Flags": ",".join(result.quality_flags),
                        "Total Severity": str(result.total_severity),
                        "Confidence": result.confidence,
                        "Prior RT": _format_optional_number(result.prior_rt),
                        "Prior Source": result.prior_source,
                    }
                )


def _format_optional_severity(value: int | None) -> str:
    if value is None:
        return ""
    return str(value)


def _set_long_ms1_values(row: dict[str, str], value: str) -> None:
    row["RT"] = value
    row["Area"] = value
    row["Int"] = value
    row["PeakStart"] = value
    row["PeakEnd"] = value
    row["PeakWidth"] = value


def _set_long_peak_values(row: dict[str, str], peak: PeakResult | None) -> None:
    if peak is None:
        _set_long_ms1_values(row, "ND")
        return
    row["RT"] = f"{peak.rt:.4f}"
    row["Area"] = f"{peak.area:.2f}"
    row["Int"] = f"{peak.intensity:.0f}"
    row["PeakStart"] = f"{peak.peak_start:.4f}"
    row["PeakEnd"] = f"{peak.peak_end:.4f}"
    row["PeakWidth"] = _format_peak_width(peak)


def _output_fieldnames(targets: list[Target]) -> list[str]:
    fieldnames = ["SampleName"]
    for target in targets:
        fieldnames.extend(_target_fieldnames(target))
    return fieldnames


def _target_fieldnames(target: Target) -> list[str]:
    fieldnames = [f"{target.label}_{suffix}" for suffix in _MS1_SUFFIXES]
    if target.neutral_loss_da is not None:
        fieldnames.append(f"{target.label}_NL")
    return fieldnames


def _output_row(file_result: FileResult, targets: list[Target]) -> dict[str, str]:
    row = {"SampleName": file_result.sample_name}
    for target in targets:
        if file_result.error is not None:
            _set_target_values(row, target, "ERROR")
            continue

        result = file_result.results[target.label]
        _set_peak_values(row, target, result.peak_result.peak)
        if target.neutral_loss_da is not None:
            row[f"{target.label}_NL"] = result.nl.to_token() if result.nl else "ND"
    return row


def _set_target_values(row: dict[str, str], target: Target, value: str) -> None:
    for suffix in _MS1_SUFFIXES:
        row[f"{target.label}_{suffix}"] = value
    if target.neutral_loss_da is not None:
        row[f"{target.label}_NL"] = value


def _set_peak_values(
    row: dict[str, str], target: Target, peak: PeakResult | None
) -> None:
    if peak is None:
        for suffix in _MS1_SUFFIXES:
            row[f"{target.label}_{suffix}"] = "ND"
        return

    row[f"{target.label}_RT"] = f"{peak.rt:.4f}"
    row[f"{target.label}_Int"] = f"{peak.intensity:.0f}"
    row[f"{target.label}_Area"] = f"{peak.area:.2f}"
    row[f"{target.label}_PeakStart"] = f"{peak.peak_start:.4f}"
    row[f"{target.label}_PeakEnd"] = f"{peak.peak_end:.4f}"
    row[f"{target.label}_PeakWidth"] = _format_peak_width(peak)


def _format_peak_width(peak: PeakResult) -> str:
    return f"{abs(peak.peak_end - peak.peak_start):.4f}"


def _format_optional_number(value: float | None) -> str:
    if value is None:
        return "NA"
    return f"{value:g}"


def _sample_group(name: str) -> str:
    return classify_sample_group(name)


def _write_xlsx(
    output_path: Path,
    run_output: RunOutput,
    targets: list[Target],
    emit_score_breakdown: bool,
) -> None:
    workbook = Workbook()
    results_sheet = workbook.active
    results_sheet.title = "XIC Results"
    _write_xic_results_sheet(results_sheet, run_output, targets)

    summary_sheet = workbook.create_sheet("Summary")
    _write_summary_sheet(summary_sheet, run_output, targets)

    if emit_score_breakdown:
        breakdown_sheet = workbook.create_sheet("Score Breakdown")
        _write_score_breakdown_sheet(breakdown_sheet, run_output)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    workbook.close()


def _write_xic_results_sheet(
    sheet: Any, run_output: RunOutput, targets: list[Target]
) -> None:
    headers = [
        "SampleName",
        "Group",
        "Target",
        "Role",
        "ISTD Pair",
        "RT",
        "Area",
        "NL",
        "Int",
        "PeakStart",
        "PeakEnd",
        "PeakWidth",
        "Confidence",
        "Reason",
    ]
    sheet.append(headers)
    for file_result, extraction_result in _iter_output_rows(run_output, targets):
        peak = extraction_result.peak
        sheet.append(
            [
                file_result.sample_name,
                file_result.group,
                extraction_result.target_label,
                extraction_result.role,
                extraction_result.istd_pair,
                peak.rt if peak is not None else None,
                peak.area if peak is not None else None,
                (
                    extraction_result.nl_result.to_token()
                    if extraction_result.nl_result is not None
                    else None
                ),
                peak.intensity if peak is not None else None,
                peak.peak_start if peak is not None else None,
                peak.peak_end if peak is not None else None,
                abs(peak.peak_end - peak.peak_start) if peak is not None else None,
                extraction_result.confidence,
                extraction_result.reason,
            ]
        )


def _write_summary_sheet(
    sheet: Any, run_output: RunOutput, targets: list[Target]
) -> None:
    headers = [
        "Target",
        "Role",
        "Confidence HIGH",
        "Confidence MEDIUM",
        "Confidence LOW",
        "Confidence VERY_LOW",
    ]
    sheet.append(headers)
    counts: dict[tuple[str, str], dict[str, int]] = {}
    for _, extraction_result in _iter_output_rows(run_output, targets):
        key = (extraction_result.target_label, extraction_result.role)
        target_counts = counts.setdefault(
            key,
            {"HIGH": 0, "MEDIUM": 0, "LOW": 0, "VERY_LOW": 0},
        )
        if extraction_result.confidence in target_counts:
            target_counts[extraction_result.confidence] += 1

    for (target_label, role), target_counts in counts.items():
        sheet.append(
            [
                target_label,
                role,
                target_counts["HIGH"],
                target_counts["MEDIUM"],
                target_counts["LOW"],
                target_counts["VERY_LOW"],
            ]
        )


def _write_score_breakdown_sheet(sheet: Any, run_output: RunOutput) -> None:
    headers = [
        "SampleName",
        "Target",
        "symmetry",
        "local_sn",
        "nl_support",
        "rt_prior",
        "rt_centrality",
        "noise_shape",
        "peak_width",
        "Quality Penalty",
        "Quality Flags",
        "Total Severity",
        "Confidence",
        "Prior RT",
        "Prior Source",
    ]
    sheet.append(headers)
    for file_result in run_output.file_results:
        for extraction_result in file_result.extraction_results:
            severities = {
                label: severity
                for severity, label in extraction_result.severities
            }
            sheet.append(
                [
                    file_result.sample_name,
                    extraction_result.target_label,
                    severities.get("symmetry"),
                    severities.get("local_sn"),
                    severities.get("nl_support"),
                    severities.get("rt_prior"),
                    severities.get("rt_centrality"),
                    severities.get("noise_shape"),
                    severities.get("peak_width"),
                    extraction_result.quality_penalty,
                    ",".join(extraction_result.quality_flags),
                    extraction_result.total_severity,
                    extraction_result.confidence,
                    extraction_result.prior_rt,
                    extraction_result.prior_source,
                ]
            )


def _iter_output_rows(
    run_output: RunOutput, targets: list[Target]
) -> list[tuple[FileResult, ExtractionResult]]:
    rows: list[tuple[FileResult, ExtractionResult]] = []
    for file_result in run_output.file_results:
        if targets:
            for target in targets:
                extraction_result = file_result.results.get(target.label)
                if extraction_result is not None:
                    rows.append((file_result, extraction_result))
            continue
        for extraction_result in file_result.extraction_results:
            rows.append((file_result, extraction_result))
    return rows


def _istd_confidence_note(istd_confidence: str | None) -> str | None:
    if istd_confidence in (None, "HIGH", "MEDIUM"):
        return None
    return f"ISTD anchor was {istd_confidence}"


def _resolve_injection_order(
    config: ExtractionConfig,
    raw_paths: list[Path],
    injection_order: dict[str, int] | None,
) -> dict[str, int] | None:
    if injection_order is not None:
        return injection_order
    if config.injection_order_source is not None:
        return read_injection_order(config.injection_order_source)
    if not raw_paths:
        return None
    return _fallback_injection_order_from_mtime(raw_paths)


def _resolve_rt_prior_library(
    config: ExtractionConfig,
    rt_prior_library: dict[tuple[str, str], LibraryEntry] | None,
) -> dict[tuple[str, str], LibraryEntry]:
    if rt_prior_library is not None:
        return rt_prior_library
    if config.rt_prior_library_path is None:
        return {}
    return load_library(config.rt_prior_library_path, config.config_hash)


def _extract_istd_anchors_only(
    config: ExtractionConfig, istd_targets: list[Target], raw_path: Path
) -> tuple[
    dict[str, float],
    dict[str, ExtractionResult],
    list[DiagnosticRecord],
    dict[str, tuple[float, float | None]],
] | None:
    if not istd_targets:
        return {}, {}, [], {}
    try:
        with open_raw(raw_path, config.dll_dir) as raw:
            results: dict[str, ExtractionResult] = {}
            diagnostics: list[DiagnosticRecord] = []
            anchors: dict[str, float] = {}
            shape_metrics_by_label: dict[str, tuple[float, float | None]] = {}
            for target in istd_targets:
                anchor_rt = _extract_one_target(
                    raw,
                    config,
                    raw_path.stem,
                    target,
                    reference_rt=None,
                    strict_preferred_rt=False,
                    results=results,
                    diagnostics=diagnostics,
                    shape_metrics_by_label=shape_metrics_by_label,
                )
                result = results.get(target.label)
                if (
                    anchor_rt is not None
                    and result is not None
                    and _allow_prepass_anchor(result)
                ):
                    anchors[target.label] = anchor_rt
            return anchors, results, diagnostics, shape_metrics_by_label
    except Exception:
        return None


def _build_scoring_context_factory(
    *,
    config: ExtractionConfig,
    injection_order: dict[str, int],
    istd_rts_by_sample: dict[str, dict[str, float]],
    rt_prior_library: dict[tuple[str, str], LibraryEntry],
) -> Callable[..., Callable[[PeakCandidate], ScoringContext]]:
    def _factory(
        *,
        target: Target,
        sample_name: str,
        rt: np.ndarray,
        intensity: np.ndarray,
        istd_rt_in_this_sample: float | None,
        paired_istd_fwhm: float | None,
        nl_result: NLResult | None,
    ) -> Callable[[PeakCandidate], ScoringContext]:
        rt_prior: float | None = None
        rt_prior_sigma: float | None = None
        prior_source = ""

        if target.is_istd:
            rt_map = istd_rts_by_sample.get(target.label, {})
            rt_prior = rolling_median_rt(
                target.label,
                sample_name,
                rt_map,
                injection_order,
                window=config.rolling_window_size,
            )
            if rt_prior is not None:
                prior_source = "rolling_median"
            else:
                library_entry = rt_prior_library.get((target.label, "ISTD"))
                if (
                    library_entry is not None
                    and library_entry.median_abs_rt is not None
                ):
                    rt_prior = library_entry.median_abs_rt
                    rt_prior_sigma = library_entry.sigma_abs_rt
                    prior_source = "library_abs"
        else:
            library_entry = rt_prior_library.get((target.label, "analyte"))
            if (
                library_entry is not None
                and istd_rt_in_this_sample is not None
                and library_entry.median_delta_rt is not None
            ):
                rt_prior = istd_rt_in_this_sample + library_entry.median_delta_rt
                rt_prior_sigma = library_entry.sigma_delta_rt
                prior_source = "delta_rt_library"

        rt_values = np.asarray(rt, dtype=float)
        intensity_values = np.asarray(intensity, dtype=float)
        baseline_array, residual_mad = compute_local_sn_cache(intensity_values)
        ms2_present = nl_result is not None and nl_result.matched_scan_count > 0
        nl_match = nl_result is not None and nl_result.status in {"OK", "WARN"}
        prefer_rt_prior_tiebreak = (
            not target.is_istd
            and bool(target.istd_pair)
            and istd_rt_in_this_sample is not None
            and rt_prior is not None
            and prior_source == "delta_rt_library"
        )

        def builder(candidate: PeakCandidate) -> ScoringContext:
            half_width_ratio, fwhm = _compute_shape_metrics(
                intensity_values,
                candidate.smoothed_apex_index,
            )
            fwhm_ratio: float | None = None
            if (
                not target.is_istd
                and fwhm is not None
                and paired_istd_fwhm is not None
                and paired_istd_fwhm > 0
            ):
                fwhm_ratio = fwhm / paired_istd_fwhm
            return ScoringContext(
                rt_array=rt_values,
                intensity_array=intensity_values,
                apex_index=candidate.smoothed_apex_index,
                half_width_ratio=half_width_ratio,
                fwhm_ratio=fwhm_ratio,
                ms2_present=ms2_present,
                nl_match=nl_match,
                rt_prior=rt_prior,
                rt_prior_sigma=rt_prior_sigma,
                rt_min=target.rt_min,
                rt_max=target.rt_max,
                dirty_matrix=config.dirty_matrix_mode,
                baseline_array=baseline_array,
                residual_mad=residual_mad,
                prefer_rt_prior_tiebreak=prefer_rt_prior_tiebreak,
            )

        setattr(builder, "rt_prior", rt_prior)
        setattr(builder, "prior_source", prior_source)
        return builder

    return _factory


def _make_scoring_context_factory(
    *,
    config: ExtractionConfig,
    injection_order: dict[str, int],
    istd_rts_by_sample: dict[str, dict[str, float]],
    rt_prior_library: dict[tuple[str, str], Any],
) -> Callable[..., Any]:
    return _build_scoring_context_factory(
        config=config,
        injection_order=injection_order,
        istd_rts_by_sample=istd_rts_by_sample,
        rt_prior_library=rt_prior_library,
    )


def _compute_shape_metrics(
    intensity: np.ndarray,
    apex_index: int,
) -> tuple[float, float | None]:
    values = np.asarray(intensity, dtype=float)
    if len(values) == 0 or apex_index < 0 or apex_index >= len(values):
        return 1.0, None
    widths, _, left_ips, right_ips = peak_widths(values, [apex_index], rel_height=0.5)
    if len(widths) == 0:
        return 1.0, None
    fwhm = float(widths[0])
    left = apex_index - float(left_ips[0])
    right = float(right_ips[0]) - apex_index
    if left <= 0 or right <= 0:
        return 1.0, fwhm
    return float(left / right), fwhm


def _selected_shape_metrics(
    intensity: np.ndarray,
    peak_result: PeakDetectionResult,
) -> tuple[float, float | None] | None:
    candidate = _selected_candidate(peak_result)
    if candidate is None:
        return None
    return _compute_shape_metrics(intensity, candidate.smoothed_apex_index)


def _selected_candidate(peak_result: PeakDetectionResult) -> PeakCandidate | None:
    if peak_result.peak is None:
        return None
    for candidate in peak_result.candidates:
        if candidate.peak == peak_result.peak:
            return candidate
    return None


def _allow_prepass_anchor(result: ExtractionResult) -> bool:
    candidate = _selected_candidate(result.peak_result)
    if candidate is None:
        return False
    return not bool(getattr(candidate, "quality_flags", ()))


def _paired_istd_fwhm(
    target: Target,
    istd_shape_metrics_by_label: dict[str, tuple[float, float | None]],
) -> float | None:
    if not target.istd_pair:
        return None
    metrics = istd_shape_metrics_by_label.get(target.istd_pair)
    if metrics is None:
        return None
    return metrics[1]


def _fallback_injection_order_from_mtime(raw_paths: list[Path]) -> dict[str, int]:
    ordered_paths = sorted(
        raw_paths,
        key=lambda path: (path.stat().st_mtime, path.name),
    )
    return {path.stem: index for index, path in enumerate(ordered_paths, start=1)}
