from __future__ import annotations

import csv
import re
import statistics
from pathlib import Path

from openpyxl import load_workbook

_MIN_WINDOW_SAMPLES = 3


def read_injection_order(path: Path) -> dict[str, int]:
    """Read CSV/XLSX with Sample_Name and Injection_Order columns."""
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return _read_xlsx(path)
    if suffix == ".csv":
        return _read_csv(path)
    raise ValueError(f"Unsupported injection-order file type: {suffix}")


def _read_csv(path: Path) -> dict[str, int]:
    out: dict[str, int] = {}
    with path.open(encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            name = (row.get("Sample_Name") or "").strip()
            order = row.get("Injection_Order")
            if not name or order in (None, ""):
                continue
            _add_sample_order(out, name, int(str(order).strip()))
    return out


def _read_xlsx(path: Path) -> dict[str, int]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        header = next(rows)
        cols = {str(value): i for i, value in enumerate(header) if value is not None}
        name_i = cols["Sample_Name"]
        order_i = cols["Injection_Order"]
        out: dict[str, int] = {}
        for row in rows:
            name = row[name_i]
            order = row[order_i]
            if name is None or order is None:
                continue
            _add_sample_order(out, str(name), int(str(order).strip()))
        return out
    finally:
        wb.close()


def _add_sample_order(out: dict[str, int], name: str, order: int) -> None:
    trimmed = name.strip()
    _set_sample_order(out, trimmed, order)
    for alias in _canonical_sample_aliases(trimmed):
        _set_sample_order(out, alias, order)


def _canonical_sample_aliases(name: str) -> tuple[str, ...]:
    aliases: list[str] = []
    alias = _canonical_sample_name(name)
    if alias != name:
        aliases.append(alias)
    qc4_alias = alias.replace("_QC_4", "_QC4")
    if qc4_alias != alias:
        aliases.append(qc4_alias)
    return tuple(dict.fromkeys(aliases))


def _set_sample_order(out: dict[str, int], name: str, order: int) -> None:
    existing = out.get(name)
    if existing is not None and existing != order:
        raise ValueError(
            f"Conflicting injection order for sample name {name!r}: "
            f"{existing} != {order}"
        )
    out[name] = order


def _canonical_sample_name(name: str) -> str:
    match = re.fullmatch(r"(Tumor|Normal) tissue (BC\d+)_DNA", name)
    if match:
        tissue, case_id = match.groups()
        return f"{tissue}{case_id}_DNA"

    match = re.fullmatch(r"Benign fat (BC\d+)_DNA", name)
    if match:
        return f"Benignfat{match.group(1)}_DNA"

    match = re.fullmatch(
        r"(Tumor|Normal) tissue (BC\d+)(?:\*|_)?\s*DNA\s*\+RNA",
        name,
    )
    if match:
        tissue, case_id = match.groups()
        return f"{tissue}{case_id}_DNAandRNA"

    if name.startswith("Breast Cancer Tissue"):
        normalized = name.replace("*", "")
        normalized = re.sub(r"\s+", "_", normalized)
        normalized = re.sub(r"_+", "_", normalized)
        match = re.search(r"_QC_(\d+)$", normalized)
        if match and match.group(1) != "4":
            normalized = normalized[: match.start()] + f"_QC{match.group(1)}"
        return normalized

    return name


def rolling_median_rt(
    istd_label: str,
    target_sample: str,
    rt_by_sample: dict[str, float],
    injection_order: dict[str, int],
    window: int,
) -> float | None:
    """Return the median RT for samples within the target injection window."""
    _ = istd_label
    target_order = injection_order.get(target_sample)
    if target_order is None:
        return None
    lo = target_order - window
    hi = target_order + window
    values: list[float] = []
    for sample, rt in rt_by_sample.items():
        order = injection_order.get(sample)
        if order is None:
            continue
        if lo <= order <= hi:
            values.append(rt)
    if len(values) < _MIN_WINDOW_SAMPLES:
        return None
    return float(statistics.median(values))
