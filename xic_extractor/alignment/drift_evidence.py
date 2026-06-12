from __future__ import annotations

import statistics
from dataclasses import dataclass
from functools import cached_property
from pathlib import Path
from typing import Literal

from openpyxl import load_workbook

from xic_extractor.injection_rolling import read_injection_order, rolling_median_rt

DriftPriorSource = Literal["targeted_istd_trend", "batch_istd_trend", "none"]


@dataclass(frozen=True)
class SampleDriftEvidence:
    sample_stem: str
    injection_order: int
    trend_id: str
    istd_rt_min: float
    local_trend_rt_min: float
    rt_drift_delta_min: float
    source: Literal["targeted_istd_trend", "batch_istd_trend"]


@dataclass(frozen=True)
class DriftEvidenceLookup:
    points: tuple[SampleDriftEvidence, ...]

    @property
    def source(self) -> DriftPriorSource:
        if not self.points:
            return "none"
        if any(point.source == "targeted_istd_trend" for point in self.points):
            return "targeted_istd_trend"
        return "batch_istd_trend"

    def sample_delta_min(self, sample_stem: str) -> float | None:
        return self._sample_delta_medians.get(sample_stem)

    def injection_order(self, sample_stem: str) -> int | None:
        orders = self._injection_orders_by_sample.get(sample_stem)
        if not orders:
            return None
        if len(orders) == 1:
            return orders[0]
        raise ValueError(
            "conflicting injection order for "
            f"sample {sample_stem!r}: {list(orders)}"
        )

    @cached_property
    def _sample_delta_medians(self) -> dict[str, float]:
        by_sample: dict[str, list[float]] = {}
        for point in self.points:
            by_sample.setdefault(point.sample_stem, []).append(point.rt_drift_delta_min)
        return {
            sample_stem: float(statistics.median(deltas))
            for sample_stem, deltas in by_sample.items()
        }

    @cached_property
    def _injection_orders_by_sample(self) -> dict[str, tuple[int, ...]]:
        by_sample: dict[str, set[int]] = {}
        for point in self.points:
            by_sample.setdefault(point.sample_stem, set()).add(point.injection_order)
        return {
            sample_stem: tuple(sorted(orders))
            for sample_stem, orders in by_sample.items()
        }


def read_targeted_istd_drift_evidence(
    targeted_workbook: Path,
    sample_info: Path,
    local_window: int = 4,
) -> DriftEvidenceLookup:
    injection_order = read_injection_order(sample_info)
    rt_by_label = _read_istd_rt_by_label(targeted_workbook)
    ordered_labels = tuple(sorted(rt_by_label))
    trend_ids = {
        label: f"trend-{index:03d}"
        for index, label in enumerate(ordered_labels, start=1)
    }
    points: list[SampleDriftEvidence] = []
    for label in ordered_labels:
        rt_by_sample = rt_by_label[label]
        for sample_stem in sorted(
            rt_by_sample,
            key=lambda sample: (injection_order.get(sample, 10**12), sample),
        ):
            local_median = rolling_median_rt(
                label,
                sample_stem,
                rt_by_sample,
                injection_order,
                window=local_window,
            )
            sample_order = injection_order.get(sample_stem)
            if local_median is None or sample_order is None:
                continue
            istd_rt = rt_by_sample[sample_stem]
            points.append(
                SampleDriftEvidence(
                    sample_stem=sample_stem,
                    injection_order=sample_order,
                    trend_id=trend_ids[label],
                    istd_rt_min=istd_rt,
                    local_trend_rt_min=local_median,
                    rt_drift_delta_min=istd_rt - local_median,
                    source="targeted_istd_trend",
                )
            )
    return DriftEvidenceLookup(points=tuple(points))


def _read_istd_rt_by_label(targeted_workbook: Path) -> dict[str, dict[str, float]]:
    wb = load_workbook(targeted_workbook, read_only=True, data_only=True)
    try:
        ws = wb["XIC Results"]
        rows = ws.iter_rows(values_only=True)
        header = next(rows)
        cols = _header_indexes(header)
        sample_name = ""
        out: dict[str, dict[str, float]] = {}
        for row in rows:
            raw_sample = row[cols["SampleName"]]
            if raw_sample is not None:
                candidate_sample = str(raw_sample).strip()
                if candidate_sample:
                    sample_name = candidate_sample
            if not sample_name:
                continue
            role = row[cols["Role"]]
            if str(role).strip() != "ISTD":
                continue
            rt = _to_float(row[cols["RT"]])
            if rt is None:
                continue
            label = str(row[cols["Target"]]).strip()
            if not label:
                continue
            out.setdefault(label, {})[sample_name] = rt
        return out
    finally:
        wb.close()


def _header_indexes(header: tuple[object, ...]) -> dict[str, int]:
    cols = {
        str(value).strip(): i
        for i, value in enumerate(header)
        if value is not None
    }
    required = ("SampleName", "Target", "Role", "RT")
    missing = [name for name in required if name not in cols]
    if missing:
        raise ValueError(f"XIC Results sheet missing required columns: {missing}")
    return cols


def _to_float(value: object) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, int | float):
        return float(value)
    try:
        return float(str(value).strip())
    except ValueError:
        return None
