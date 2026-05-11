from __future__ import annotations

import csv
import math
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


@dataclass(frozen=True)
class LoadedFeature:
    feature_id: str
    mz: float
    rt_min: float
    sample_areas: dict[str, float | None]
    metadata: dict[str, str]


@dataclass(frozen=True)
class LoadedMatrix:
    source: str
    features: tuple[LoadedFeature, ...]
    sample_order: tuple[str, ...]


def normalize_sample_name(value: str) -> str:
    name = Path(str(value)).name
    peak_area_suffix = ".mzML Peak area"
    if name.endswith(peak_area_suffix):
        name = name[: -len(peak_area_suffix)]
    for suffix in (".raw", ".mzML", ".RAW", ".MZML"):
        if name.endswith(suffix):
            name = name[: -len(suffix)]
    for prefix in ("program2_DNA_program1_", "program2_RNA_program1_"):
        if name.startswith(prefix):
            name = name[len(prefix) :]
    name = name.replace(
        "Breast_Cancer_Tissue_pooled_QC_",
        "Breast_Cancer_Tissue_pooled_QC",
    )
    return name.strip()


def load_xic_alignment(review_tsv: Path, matrix_tsv: Path) -> LoadedMatrix:
    review_rows, review_columns = _read_delimited(review_tsv, delimiter="\t")
    matrix_rows, matrix_columns = _read_delimited(matrix_tsv, delimiter="\t")
    _require_columns(
        review_tsv,
        review_columns,
        ("cluster_id", "cluster_center_mz", "cluster_center_rt"),
    )
    _require_columns(
        matrix_tsv,
        matrix_columns,
        ("cluster_id", "cluster_center_mz", "cluster_center_rt"),
    )

    review_by_id: dict[str, dict[str, str]] = {}
    for row_number, row in review_rows:
        cluster_id = row.get("cluster_id", "")
        if cluster_id in review_by_id:
            raise ValueError(
                f"{review_tsv}: row {row_number}: duplicate cluster_id {cluster_id!r}"
            )
        review_by_id[cluster_id] = row

    sample_columns = tuple(
        normalize_sample_name(column) for column in matrix_columns[4:]
    )
    features: list[LoadedFeature] = []
    seen_matrix_ids: set[str] = set()
    for row_number, row in matrix_rows:
        cluster_id = row.get("cluster_id", "")
        if cluster_id in seen_matrix_ids:
            raise ValueError(
                f"{matrix_tsv}: row {row_number}: duplicate cluster_id {cluster_id!r}"
            )
        seen_matrix_ids.add(cluster_id)
        review_row = review_by_id.get(cluster_id)
        if review_row is None:
            raise ValueError(
                f"{matrix_tsv}: row {row_number}: "
                f"review row missing for cluster_id {cluster_id!r}"
            )
        matrix_mz = _parse_float(
            matrix_tsv,
            row_number,
            row.get("cluster_center_mz", ""),
            "cluster_center_mz",
        )
        matrix_rt = _parse_float(
            matrix_tsv,
            row_number,
            row.get("cluster_center_rt", ""),
            "cluster_center_rt",
        )
        review_mz = _parse_float(
            review_tsv,
            row_number,
            review_row.get("cluster_center_mz", ""),
            "cluster_center_mz",
        )
        review_rt = _parse_float(
            review_tsv,
            row_number,
            review_row.get("cluster_center_rt", ""),
            "cluster_center_rt",
        )
        if not math.isclose(matrix_mz, review_mz, rel_tol=0.0, abs_tol=1e-9):
            raise ValueError(
                f"{matrix_tsv}: row {row_number}: "
                "cluster_center_mz disagrees with review"
            )
        if not math.isclose(matrix_rt, review_rt, rel_tol=0.0, abs_tol=1e-9):
            raise ValueError(
                f"{matrix_tsv}: row {row_number}: "
                "cluster_center_rt disagrees with review"
            )
        sample_areas = {
            sample: _parse_area(row.get(column, ""))
            for column, sample in zip(matrix_columns[4:], sample_columns, strict=True)
        }
        features.append(
            LoadedFeature(
                feature_id=cluster_id,
                mz=matrix_mz,
                rt_min=matrix_rt,
                sample_areas=sample_areas,
                metadata={key: value for key, value in review_row.items() if key},
            )
        )
    if not features:
        raise ValueError(f"{matrix_tsv}: XIC alignment has zero clusters")
    return LoadedMatrix(
        source="xic_alignment",
        features=tuple(features),
        sample_order=sample_columns,
    )


def load_fh_alignment_tsv(path: Path) -> LoadedMatrix:
    rows, columns = _read_delimited(path, delimiter="\t")
    _require_columns(path, columns, ("alignment_id", "Mz", "RT"))
    sample_columns = tuple(columns[3:])
    return _load_tabular_matrix(
        path,
        rows,
        feature_id_column="alignment_id",
        mz_column="Mz",
        rt_column="RT",
        sample_columns=sample_columns,
        source="fh_alignment",
    )


def load_metabcombiner_tsv(path: Path) -> tuple[LoadedMatrix, ...]:
    rows, columns = _read_delimited(path, delimiter="\t")
    _require_columns(path, columns, ("Mz", "RT"))
    mzmine_marker_index = (
        columns.index("MZmine ID") if "MZmine ID" in columns else len(columns)
    )
    fh_sample_columns = tuple(columns[2:mzmine_marker_index])
    matrices = [
        _load_tabular_matrix(
            path,
            rows,
            feature_id_column=None,
            mz_column="Mz",
            rt_column="RT",
            sample_columns=fh_sample_columns,
            source="metabcombiner_fh_block",
            feature_id_prefix="metabcombiner_fh",
        )
    ]
    if {"MZmine ID", "MZmine m/z", "MZmine RT (min)"}.issubset(columns):
        mzmine_sample_columns = tuple(
            column for column in columns if column.endswith(".mzML Peak area")
        )
        matrices.append(
            _load_metabcombiner_mzmine_matrix(
                path,
                rows,
                sample_columns=mzmine_sample_columns,
            )
        )
    return tuple(matrices)


def _load_metabcombiner_mzmine_matrix(
    path: Path,
    rows: list[tuple[int, dict[str, str]]],
    *,
    sample_columns: tuple[str, ...],
) -> LoadedMatrix:
    sample_order = tuple(normalize_sample_name(column) for column in sample_columns)
    grouped_rows: dict[str, list[tuple[int, dict[str, str]]]] = {}
    for row_number, row in rows:
        if row.get("MZmine m/z", "") == "" or row.get("MZmine RT (min)", "") == "":
            continue
        mzmine_id = row.get("MZmine ID", "")
        feature_id = (
            f"metabcombiner_mzmine:{mzmine_id}"
            if mzmine_id
            else f"metabcombiner_mzmine:{row_number:06d}"
        )
        grouped_rows.setdefault(feature_id, []).append((row_number, row))

    features: list[LoadedFeature] = []
    for feature_id, row_group in grouped_rows.items():
        first_row_number, first_row = row_group[0]
        sample_areas = {
            sample: _max_positive_area(
                row.get(column, "") for _row_number, row in row_group
            )
            for column, sample in zip(sample_columns, sample_order, strict=True)
        }
        features.append(
            LoadedFeature(
                feature_id=feature_id,
                mz=_parse_float(
                    path,
                    first_row_number,
                    first_row.get("MZmine m/z", ""),
                    "MZmine m/z",
                ),
                rt_min=_parse_float(
                    path,
                    first_row_number,
                    first_row.get("MZmine RT (min)", ""),
                    "MZmine RT (min)",
                ),
                sample_areas=sample_areas,
                metadata={
                    "duplicate_row_count": str(len(row_group)),
                    "source_row_numbers": ";".join(
                        str(row_number) for row_number, _row in row_group
                    ),
                },
            )
        )
    return LoadedMatrix(
        source="metabcombiner_mzmine_block",
        features=tuple(features),
        sample_order=sample_order,
    )


def load_combine_fix_xlsx(path: Path, sheet_name: str | None = None) -> LoadedMatrix:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name] if sheet_name else workbook.worksheets[0]
        rows_iter = sheet.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if header is None:
            raise ValueError(f"{path}: workbook sheet is empty")
        columns = tuple("" if value is None else str(value) for value in header)
        _require_columns(path, columns, ("Mz", "RT"))
        rt_index = columns.index("RT")
        sample_columns = tuple(
            column
            for column in columns[rt_index + 1 :]
            if not _is_combine_fix_metadata_column(column)
        )
        rows = [
            (
                row_number,
                {
                    column: "" if value is None else str(value)
                    for column, value in zip(columns, values, strict=False)
                },
            )
            for row_number, values in enumerate(rows_iter, start=2)
        ]
        return _load_tabular_matrix(
            path,
            rows,
            feature_id_column=None,
            mz_column="Mz",
            rt_column="RT",
            sample_columns=sample_columns,
            source="combine_fix",
            feature_id_prefix="combine_fix",
        )
    finally:
        workbook.close()


def _load_tabular_matrix(
    path: Path,
    rows: list[tuple[int, dict[str, str]]],
    *,
    feature_id_column: str | None,
    mz_column: str,
    rt_column: str,
    sample_columns: tuple[str, ...],
    source: str,
    feature_id_prefix: str | None = None,
) -> LoadedMatrix:
    sample_order = tuple(normalize_sample_name(column) for column in sample_columns)
    features: list[LoadedFeature] = []
    seen_feature_ids: set[str] = set()
    for row_number, row in rows:
        feature_id = _feature_id(row, row_number, feature_id_column, feature_id_prefix)
        if feature_id in seen_feature_ids:
            raise ValueError(
                f"{path}: row {row_number}: duplicate feature_id {feature_id!r}"
            )
        seen_feature_ids.add(feature_id)
        sample_areas = {
            sample: _parse_area(row.get(column, ""))
            for column, sample in zip(sample_columns, sample_order, strict=True)
        }
        features.append(
            LoadedFeature(
                feature_id=feature_id,
                mz=_parse_float(path, row_number, row.get(mz_column, ""), mz_column),
                rt_min=_parse_float(
                    path, row_number, row.get(rt_column, ""), rt_column
                ),
                sample_areas=sample_areas,
                metadata={},
            )
        )
    return LoadedMatrix(
        source=source, features=tuple(features), sample_order=sample_order
    )


def _feature_id(
    row: dict[str, str],
    row_number: int,
    feature_id_column: str | None,
    feature_id_prefix: str | None,
) -> str:
    if feature_id_column is not None:
        value = row.get(feature_id_column, "")
        if value:
            if feature_id_prefix is None:
                return value
            return f"{feature_id_prefix}:{value}"
    if feature_id_prefix is None:
        raise ValueError(
            "feature_id_column is required when feature_id_prefix is absent"
        )
    return f"{feature_id_prefix}:{row_number:06d}"


def _read_delimited(
    path: Path,
    *,
    delimiter: str,
) -> tuple[list[tuple[int, dict[str, str]]], tuple[str, ...]]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle, delimiter=delimiter)
        columns = tuple(reader.fieldnames or ())
        rows = [(index, row) for index, row in enumerate(reader, start=2)]
    return rows, columns


def _require_columns(
    path: Path, columns: tuple[str, ...], required: tuple[str, ...]
) -> None:
    missing = [column for column in required if column not in columns]
    if missing:
        raise ValueError(f"{path}: missing required columns: {', '.join(missing)}")


def _parse_float(path: Path, row_number: int, value: str, column: str) -> float:
    try:
        parsed = float(value)
    except ValueError as exc:
        raise ValueError(
            f"{path}: row {row_number}: {column} must be a finite number: {value!r}"
        ) from exc
    if not math.isfinite(parsed):
        raise ValueError(
            f"{path}: row {row_number}: {column} must be a finite number: {value!r}"
        )
    return parsed


def _parse_area(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        parsed = float(value)
    except (TypeError, ValueError):
        return None
    if not math.isfinite(parsed) or parsed <= 0:
        return None
    return parsed


def _max_positive_area(values) -> float | None:
    positive_values = [
        area for area in (_parse_area(value) for value in values) if area is not None
    ]
    return max(positive_values) if positive_values else None


def _is_combine_fix_metadata_column(column: str) -> bool:
    return column.startswith("MZmine") or column.endswith(".mzML Peak area")
