from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook

from xic_extractor.alignment.config import AlignmentConfig
from xic_extractor.alignment.matrix import AlignedCell, AlignmentMatrix
from xic_extractor.alignment.output_rows import (
    cells_by_cluster,
    count_status,
    escape_excel_formula,
    production_matrix_area,
    row_id,
)
from xic_extractor.alignment.production_decisions import (
    ProductionCellDecision,
    ProductionDecisionSet,
    build_production_decisions,
)


def write_alignment_results_xlsx(
    path: Path,
    matrix: AlignmentMatrix,
    *,
    metadata: dict[str, str],
    alignment_config: AlignmentConfig | None = None,
) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    decisions = build_production_decisions(
        matrix,
        alignment_config or AlignmentConfig(),
    )
    workbook = Workbook()
    matrix_sheet = workbook.active
    matrix_sheet.title = "Matrix"
    _write_matrix_sheet(matrix_sheet, matrix, decisions)
    _write_review_sheet(workbook.create_sheet("Review"), matrix, decisions)
    _write_audit_sheet(workbook.create_sheet("Audit"), matrix, decisions)
    _write_metadata_sheet(workbook.create_sheet("Metadata"), metadata)
    workbook.save(path)
    return path


def _write_matrix_sheet(
    sheet: Any,
    matrix: AlignmentMatrix,
    decisions: ProductionDecisionSet,
) -> None:
    headers = [
        "feature_family_id",
        "neutral_loss_tag",
        "family_center_mz",
        "family_center_rt",
        *matrix.sample_order,
    ]
    _append_xlsx_row(sheet, headers)
    grouped_cells = cells_by_cluster(matrix)
    for cluster in matrix.clusters:
        cluster_id = row_id(cluster)
        if not decisions.row(cluster_id).include_in_primary_matrix:
            continue
        cells = {
            cell.sample_stem: cell for cell in grouped_cells.get(cluster_id, ())
        }
        _append_xlsx_row(
            sheet,
            [
                cluster_id,
                cluster.neutral_loss_tag,
                _family_center_mz(cluster),
                _family_center_rt(cluster),
                *[
                    _xlsx_area(
                        decisions.cell(cluster_id, sample)
                        if cells.get(sample) is not None
                        else None
                    )
                    for sample in matrix.sample_order
                ],
            ],
        )


def _write_review_sheet(
    sheet: Any,
    matrix: AlignmentMatrix,
    decisions: ProductionDecisionSet,
) -> None:
    _append_xlsx_row(
        sheet,
        [
            "feature_family_id",
            "neutral_loss_tag",
            "detected_count",
            "rescued_count",
            "accepted_cell_count",
            "accepted_rescue_count",
            "review_rescue_count",
            "absent_count",
            "unchecked_count",
            "duplicate_assigned_count",
            "ambiguous_ms1_owner_count",
            "include_in_primary_matrix",
            "row_flags",
        ],
    )
    grouped_cells = cells_by_cluster(matrix)
    for cluster in matrix.clusters:
        cluster_id = row_id(cluster)
        cells = grouped_cells.get(cluster_id, ())
        row_decision = decisions.row(cluster_id)
        _append_xlsx_row(
            sheet,
            [
                cluster_id,
                cluster.neutral_loss_tag,
                count_status(cells, "detected"),
                count_status(cells, "rescued"),
                row_decision.accepted_cell_count,
                row_decision.accepted_rescue_count,
                row_decision.review_rescue_count,
                count_status(cells, "absent"),
                count_status(cells, "unchecked"),
                count_status(cells, "duplicate_assigned"),
                count_status(cells, "ambiguous_ms1_owner"),
                row_decision.include_in_primary_matrix,
                ";".join(row_decision.row_flags),
            ],
        )


def _write_audit_sheet(
    sheet: Any,
    matrix: AlignmentMatrix,
    decisions: ProductionDecisionSet,
) -> None:
    _append_xlsx_row(
        sheet,
        [
            "feature_family_id",
            "sample_stem",
            "neutral_loss_tag",
            "family_center_mz",
            "family_center_rt",
            "raw_status",
            "production_status",
            "rescue_tier",
            "write_matrix_value",
            "blank_reason",
            "area",
            "apex_rt",
            "rt_delta_sec",
            "claim_state",
            "row_flags",
            "reason",
        ],
    )
    clusters = {row_id(cluster): cluster for cluster in matrix.clusters}
    for cell in matrix.cells:
        cluster = clusters.get(cell.cluster_id)
        if cluster is None:
            raise ValueError(
                f"Alignment cell references unknown cluster: {cell.cluster_id}"
            )
        decision = decisions.cell(cell.cluster_id, cell.sample_stem)
        row_decision = decisions.row(cell.cluster_id)
        _append_xlsx_row(
            sheet,
            [
                cell.cluster_id,
                cell.sample_stem,
                cluster.neutral_loss_tag,
                _family_center_mz(cluster),
                _family_center_rt(cluster),
                decision.raw_status,
                decision.production_status,
                decision.rescue_tier,
                decision.write_matrix_value,
                decision.blank_reason,
                cell.area,
                cell.apex_rt,
                cell.rt_delta_sec,
                _claim_state(cell),
                ";".join(row_decision.row_flags),
                cell.reason,
            ],
        )


def _write_metadata_sheet(sheet: Any, metadata: dict[str, str]) -> None:
    _append_xlsx_row(sheet, ["key", "value"])
    for key in sorted(metadata):
        _append_xlsx_row(sheet, [key, metadata[key]])


def _append_xlsx_row(sheet: Any, values: list[object]) -> None:
    sheet.append([_xlsx_value(value) for value in values])


def _xlsx_value(value: object) -> object:
    if isinstance(value, str):
        return escape_excel_formula(value)
    return value


def _xlsx_area(decision: ProductionCellDecision | None) -> float | None:
    text = production_matrix_area(decision)
    return float(text) if text else None


def _claim_state(cell: AlignedCell) -> str:
    if cell.status == "duplicate_assigned":
        return "loser"
    if cell.status in {"detected", "rescued"}:
        return "winner_or_unclaimed"
    return ""


def _family_center_mz(row: Any) -> float:
    if hasattr(row, "family_center_mz"):
        return row.family_center_mz
    return row.cluster_center_mz


def _family_center_rt(row: Any) -> float:
    if hasattr(row, "family_center_rt"):
        return row.family_center_rt
    return row.cluster_center_rt
