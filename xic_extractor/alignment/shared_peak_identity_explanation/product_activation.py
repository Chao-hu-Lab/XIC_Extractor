from __future__ import annotations

import csv
import math
from collections import Counter
from collections.abc import Mapping, Sequence
from dataclasses import dataclass
from pathlib import Path

from xic_extractor.alignment.primary_matrix_area import (
    ASLS_PRIMARY_MATRIX_AREA_SOURCE,
    MISSING_ASLS_PRIMARY_AREA,
)

from .schema import (
    ACTIVATION_APPLICATION_SCHEMA_VERSION,
    ACTIVATION_APPLICATION_SUMMARY_COLUMNS,
    ACTIVATION_CELL_AUDIT_COLUMNS,
    ACTIVATION_REVIEW_AUDIT_COLUMNS,
    ACTIVATION_VALUE_DELTA_COLUMNS,
    ACTIVATION_VALUE_DELTA_SCHEMA_VERSION,
)


@dataclass(frozen=True)
class ActivationApplicationOutputs:
    matrix_tsv: Path
    review_tsv: Path
    cells_tsv: Path
    summary_tsv: Path
    value_delta_tsv: Path


@dataclass(frozen=True)
class LegacyRtRowReference:
    row_number: int
    mz: float
    rt: float

    @property
    def peak_hypothesis_id(self) -> str:
        return (
            f"mzmine_rtrow_{self.row_number}_mz{self.mz:.4f}_rt{self.rt:.2f}min"
        )


@dataclass(frozen=True)
class FormalMatrixStats:
    family_projection_rows: int = 0
    legacy_rt_row_context_rows: int = 0
    matrix_value_conflict_cells: int = 0
    family_projection_rows_excluded: int = 0
    family_projection_cells_excluded: int = 0


def apply_activation_to_alignment_outputs(
    *,
    activation_decisions_tsv: Path,
    activation_acceptance_tsv: Path,
    alignment_matrix_tsv: Path,
    alignment_review_tsv: Path,
    alignment_cells_tsv: Path,
    output_dir: Path,
    require_acceptance_pass: bool = True,
    output_mode: str = "activated-copy",
    allow_overwrite_source: bool = False,
    legacy_rt_row_oracle_xlsx: Path | None = None,
    legacy_rt_row_oracle_mz_ppm: float = 20.0,
    legacy_rt_row_oracle_rt_tolerance_min: float = 1.0,
    require_complete_peak_hypothesis_identity: bool = False,
    exclude_family_projections: bool = False,
) -> ActivationApplicationOutputs:
    _validate_output_mode(output_mode)
    if exclude_family_projections and output_mode != "formal":
        raise ValueError("--exclude-family-projections requires --output-mode formal")
    decisions = _read_tsv(activation_decisions_tsv)
    _validate_decisions_for_product_application(decisions)
    acceptance_rows = _read_tsv(activation_acceptance_tsv)
    if len(acceptance_rows) != 1:
        raise ValueError("activation acceptance TSV must contain exactly one row")
    acceptance = acceptance_rows[0]
    if require_acceptance_pass and acceptance.get("acceptance_status") != "pass":
        raise ValueError("activation acceptance must pass before product application")

    matrix_header, matrix_rows = _read_tsv_with_header(alignment_matrix_tsv)
    review_header, review_rows = _read_tsv_with_header(alignment_review_tsv)
    cells_header, cell_rows = _read_tsv_with_header(alignment_cells_tsv)
    _require_columns(matrix_header, ("feature_family_id",))
    _require_columns(
        review_header,
        (
            "feature_family_id",
            "neutral_loss_tag",
            "family_center_mz",
            "family_center_rt",
            "include_in_primary_matrix",
        ),
    )
    _require_columns(cells_header, ("feature_family_id", "sample_stem"))

    sample_columns = tuple(column for column in matrix_header if column not in _META)
    review_by_family = {row["feature_family_id"]: row for row in review_rows}
    cells_by_key = {
        (row["feature_family_id"], row["sample_stem"]): row for row in cell_rows
    }
    cell_status_by_key = {
        key: row.get("status", "") for key, row in cells_by_key.items()
    }
    matrix_by_family = {
        row["feature_family_id"]: dict(row) for row in matrix_rows
    }
    original_matrix_by_family = {
        family_id: dict(row) for family_id, row in matrix_by_family.items()
    }
    original_matrix_families = set(matrix_by_family)

    decision_by_key = {
        (row.get("feature_family_id", ""), row.get("sample_id", "")): row
        for row in decisions
        if row.get("sample_id") and row.get("sample_id") != "__family_context__"
    }
    matrix_effects: dict[tuple[str, str], str] = {}
    family_notes: dict[str, list[str]] = {}
    families_added: set[str] = set()
    family_blocked: set[str] = set()
    legacy_rt_row_oracle = _load_legacy_rt_row_oracle(legacy_rt_row_oracle_xlsx)

    for decision in decisions:
        family_id = decision.get("feature_family_id", "")
        sample_id = decision.get("sample_id", "")
        status = decision.get("activation_status", "")
        effect = decision.get("product_effect", "")
        if status == "auto_block" and effect == "block_family_promotion":
            family_blocked.add(family_id)
            matrix_by_family.pop(family_id, None)
            family_notes.setdefault(family_id, []).append("family_promotion_blocked")
            continue
        if sample_id == "__family_context__" or not sample_id:
            continue
        key = (family_id, sample_id)
        if status == "auto_block" and effect == "block_rescue_cell":
            row = matrix_by_family.get(family_id)
            if row is not None and sample_id in row and row.get(sample_id):
                row[sample_id] = ""
                matrix_effects[key] = "blanked"
            else:
                matrix_effects[key] = "block_no_existing_matrix_value"
            family_notes.setdefault(family_id, []).append("rescue_cell_blocked")
        elif status == "auto_activate":
            row = matrix_by_family.get(family_id)
            if row is None:
                row = _new_matrix_row(
                    family_id,
                    sample_columns=sample_columns,
                    review_row=review_by_family.get(family_id, {}),
                )
                matrix_by_family[family_id] = row
                families_added.add(family_id)
            source_cell = cells_by_key.get(key)
            value = _matrix_value_for_activation(source_cell)
            previous = row.get(sample_id, "")
            if previous:
                matrix_effects[key] = "unchanged"
                family_notes.setdefault(family_id, []).append("cell_auto_activated")
            elif value:
                row[sample_id] = value
                matrix_effects[key] = "written"
                family_notes.setdefault(family_id, []).append("cell_auto_activated")
            elif _has_untrusted_primary_or_raw_area(source_cell):
                matrix_effects[key] = MISSING_ASLS_PRIMARY_AREA
            else:
                matrix_effects[key] = "no_cell_area_available"

    matrix_by_family = {
        family_id: row
        for family_id, row in matrix_by_family.items()
        if _has_any_sample_value(row, sample_columns)
    }
    output_matrix_header: Sequence[str]
    formal_matrix_stats = FormalMatrixStats()
    if output_mode == "formal":
        output_matrix_header = _formal_matrix_header(sample_columns)
        output_matrix_rows, formal_matrix_stats = (
            _peak_hypothesis_matrix_rows(
                matrix_by_family=matrix_by_family,
                review_by_family=review_by_family,
                decisions_by_key=decision_by_key,
                sample_columns=sample_columns,
                legacy_rt_row_oracle=legacy_rt_row_oracle,
                legacy_rt_row_oracle_mz_ppm=legacy_rt_row_oracle_mz_ppm,
                legacy_rt_row_oracle_rt_tolerance_min=(
                    legacy_rt_row_oracle_rt_tolerance_min
                ),
                include_family_projections=not exclude_family_projections,
            )
        )
    else:
        output_matrix_header = matrix_header
        output_matrix_rows = [
            matrix_by_family[family_id] for family_id in sorted(matrix_by_family)
        ]
    output_review_rows = _activated_review_rows(
        review_rows=review_rows,
        review_header=review_header,
        decisions=decisions,
        matrix_by_family=matrix_by_family,
        sample_columns=sample_columns,
        cell_status_by_key=cell_status_by_key,
        matrix_effects=matrix_effects,
        family_notes=family_notes,
        family_blocked=family_blocked,
    )
    output_cell_rows = _activated_cell_rows(
        cell_rows=cell_rows,
        decisions_by_key=decision_by_key,
        matrix_effects=matrix_effects,
    )
    value_delta_rows = _value_delta_rows(
        decisions=decisions,
        original_matrix_by_family=original_matrix_by_family,
        matrix_by_family=matrix_by_family,
        cells_by_key=cells_by_key,
        sample_columns=sample_columns,
        matrix_effects=matrix_effects,
    )
    summary_row = _summary_row(
        acceptance=acceptance,
        decisions=decisions,
        output_mode=output_mode,
        input_matrix_rows=len(matrix_rows),
        output_matrix_rows=len(output_matrix_rows),
        formal_matrix_stats=formal_matrix_stats,
        matrix_effects=matrix_effects,
        families_added=families_added,
        families_removed=original_matrix_families - set(matrix_by_family),
    )
    if (
        require_complete_peak_hypothesis_identity
        and output_mode != "formal"
    ):
        raise ValueError(
            "--require-complete-peak-hypothesis-identity requires "
            "--output-mode formal"
        )
    if (
        require_complete_peak_hypothesis_identity
        and summary_row["canonical_row_identity_ready"] != "TRUE"
    ):
        raise ValueError(
            "complete PeakHypothesis identity requires canonical row identity "
            f"readiness; blockers={summary_row['canonical_row_identity_blockers']}"
        )

    output_dir.mkdir(parents=True, exist_ok=True)
    matrix_path, review_path, cells_path = _activated_output_paths(
        output_dir,
        output_mode=output_mode,
    )
    summary_path = output_dir / "activation_application_summary.tsv"
    value_delta_path = output_dir / "activation_value_delta.tsv"
    if not allow_overwrite_source:
        _reject_source_overwrite(
            output_paths=(matrix_path, review_path, cells_path),
            source_paths=(
                alignment_matrix_tsv,
                alignment_review_tsv,
                alignment_cells_tsv,
            ),
        )
    _write_tsv(matrix_path, output_matrix_header, output_matrix_rows)
    review_output_header = (
        tuple(review_header)
        if output_mode == "formal"
        else (*review_header, *ACTIVATION_REVIEW_AUDIT_COLUMNS)
    )
    cells_output_header = (
        tuple(cells_header)
        if output_mode == "formal"
        else (*cells_header, *ACTIVATION_CELL_AUDIT_COLUMNS)
    )
    _write_tsv(review_path, review_output_header, output_review_rows)
    _write_tsv(cells_path, cells_output_header, output_cell_rows)
    _write_tsv(summary_path, ACTIVATION_APPLICATION_SUMMARY_COLUMNS, [summary_row])
    _write_tsv(value_delta_path, ACTIVATION_VALUE_DELTA_COLUMNS, value_delta_rows)
    return ActivationApplicationOutputs(
        matrix_tsv=matrix_path,
        review_tsv=review_path,
        cells_tsv=cells_path,
        summary_tsv=summary_path,
        value_delta_tsv=value_delta_path,
    )


_META = frozenset(
    {"feature_family_id", "neutral_loss_tag", "family_center_mz", "family_center_rt"}
)

_FORMAL_MATRIX_META = (
    "peak_hypothesis_id",
    "feature_family_id",
    "candidate_container_id",
    "row_identity_basis",
    "legacy_rt_row_context_id",
    "neutral_loss_tag",
    "family_center_mz",
    "family_center_rt",
)

_OUTPUT_MODES = frozenset({"activated-copy", "formal"})


def _validate_output_mode(output_mode: str) -> None:
    if output_mode not in _OUTPUT_MODES:
        choices = ", ".join(sorted(_OUTPUT_MODES))
        raise ValueError(f"activation output mode must be one of: {choices}")


def _activated_output_paths(
    output_dir: Path,
    *,
    output_mode: str,
) -> tuple[Path, Path, Path]:
    if output_mode == "formal":
        return (
            output_dir / "alignment_matrix.tsv",
            output_dir / "alignment_review.tsv",
            output_dir / "alignment_cells.tsv",
        )
    return (
        output_dir / "alignment_matrix_activated.tsv",
        output_dir / "alignment_review_activated.tsv",
        output_dir / "alignment_cells_activated.tsv",
    )


def _validate_decisions_for_product_application(
    decisions: Sequence[Mapping[str, str]],
) -> None:
    for decision in decisions:
        if decision.get("activation_status") != "auto_activate":
            continue
        if (
            not decision.get("peak_hypothesis_id")
            or decision.get("activation_unit_scope") != "peak_hypothesis"
        ):
            family_id = decision.get("feature_family_id", "")
            sample_id = decision.get("sample_id", "")
            raise ValueError(
                "auto_activate decisions require peak_hypothesis_id and "
                "activation_unit_scope=peak_hypothesis before product "
                f"application: {family_id}/{sample_id}"
            )


def _formal_matrix_header(sample_columns: Sequence[str]) -> tuple[str, ...]:
    return (*_FORMAL_MATRIX_META, *sample_columns)


def _peak_hypothesis_matrix_rows(
    *,
    matrix_by_family: Mapping[str, Mapping[str, str]],
    review_by_family: Mapping[str, Mapping[str, str]],
    decisions_by_key: Mapping[tuple[str, str], Mapping[str, str]],
    sample_columns: Sequence[str],
    legacy_rt_row_oracle: Sequence[LegacyRtRowReference],
    legacy_rt_row_oracle_mz_ppm: float,
    legacy_rt_row_oracle_rt_tolerance_min: float,
    include_family_projections: bool = True,
) -> tuple[list[dict[str, str]], FormalMatrixStats]:
    rows_by_hypothesis: dict[str, dict[str, str]] = {}
    family_projection_ids: set[str] = set()
    excluded_family_projection_ids: set[str] = set()
    excluded_family_projection_cells = 0
    legacy_rt_row_context_row_ids: set[str] = set()
    matrix_value_conflict_cells = 0
    for family_id in sorted(matrix_by_family):
        family_row = matrix_by_family[family_id]
        review_row = review_by_family.get(family_id, {})
        for sample_id in sample_columns:
            value = family_row.get(sample_id, "")
            if not value:
                continue
            decision = decisions_by_key.get((family_id, sample_id), {})
            peak_hypothesis_id = _decision_peak_hypothesis_id(decision)
            row_identity_basis = "activation_peak_hypothesis"
            candidate_container_id = decision.get("candidate_container_id", family_id)
            if not peak_hypothesis_id:
                projection_id = _legacy_projection_hypothesis_id(family_id)
                if not include_family_projections:
                    excluded_family_projection_ids.add(projection_id)
                    excluded_family_projection_cells += 1
                    continue
                legacy_reference = _match_legacy_rt_row_reference(
                    review_row,
                    legacy_rt_row_oracle,
                    mz_ppm=legacy_rt_row_oracle_mz_ppm,
                    rt_tolerance_min=legacy_rt_row_oracle_rt_tolerance_min,
                )
                peak_hypothesis_id = projection_id
                row_identity_basis = "family_projection_no_split_evidence"
                family_projection_ids.add(peak_hypothesis_id)
                legacy_rt_row_context_id = ""
                if legacy_reference is not None:
                    legacy_rt_row_context_id = legacy_reference.peak_hypothesis_id
                    legacy_rt_row_context_row_ids.add(peak_hypothesis_id)
                candidate_container_id = family_id
            else:
                legacy_rt_row_context_id = ""
            row = rows_by_hypothesis.setdefault(
                peak_hypothesis_id,
                _new_peak_hypothesis_matrix_row(
                    peak_hypothesis_id,
                    family_id=family_id,
                    candidate_container_id=candidate_container_id,
                    row_identity_basis=row_identity_basis,
                    legacy_rt_row_context_id=legacy_rt_row_context_id,
                    review_row=review_row,
                    sample_columns=sample_columns,
                ),
            )
            _append_unique_cell(row, "feature_family_id", family_id)
            _append_unique_cell(row, "candidate_container_id", candidate_container_id)
            _append_unique_cell(
                row,
                "legacy_rt_row_context_id",
                legacy_rt_row_context_id,
            )
            previous = row.get(sample_id, "")
            if previous and previous != value:
                value = _select_conflicting_matrix_value(previous, value)
                matrix_value_conflict_cells += 1
            row[sample_id] = value
    return (
        [rows_by_hypothesis[key] for key in sorted(rows_by_hypothesis)],
        FormalMatrixStats(
            family_projection_rows=len(family_projection_ids),
            legacy_rt_row_context_rows=len(legacy_rt_row_context_row_ids),
            matrix_value_conflict_cells=matrix_value_conflict_cells,
            family_projection_rows_excluded=len(excluded_family_projection_ids),
            family_projection_cells_excluded=excluded_family_projection_cells,
        ),
    )


def _decision_peak_hypothesis_id(decision: Mapping[str, str]) -> str:
    if decision.get("activation_unit_scope") != "peak_hypothesis":
        return ""
    if decision.get("activation_status") != "auto_activate":
        return ""
    return decision.get("peak_hypothesis_id", "")


def _legacy_projection_hypothesis_id(family_id: str) -> str:
    return f"{family_id}::family_projection"


def _load_legacy_rt_row_oracle(
    path: Path | None,
) -> tuple[LegacyRtRowReference, ...]:
    if path is None:
        return ()
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency is declared.
        raise ValueError(
            "legacy RT-row oracle requires openpyxl; install project dependencies"
        ) from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    header = tuple(cell.value for cell in next(sheet.iter_rows(max_row=1)))
    try:
        mz_index = header.index("Mz")
        rt_index = header.index("RT")
    except ValueError as exc:
        raise ValueError(
            "legacy RT-row oracle workbook must contain Mz and RT"
        ) from exc
    rows: list[LegacyRtRowReference] = []
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
        mz = _to_float(row[mz_index])
        rt = _to_float(row[rt_index])
        if mz is None or rt is None:
            continue
        rows.append(LegacyRtRowReference(row_number=row_number, mz=mz, rt=rt))
    return tuple(rows)


def _match_legacy_rt_row_reference(
    review_row: Mapping[str, str],
    oracle_rows: Sequence[LegacyRtRowReference],
    *,
    mz_ppm: float,
    rt_tolerance_min: float,
) -> LegacyRtRowReference | None:
    if not oracle_rows:
        return None
    mz = _to_float(review_row.get("family_center_mz"))
    rt = _to_float(review_row.get("family_center_rt"))
    if mz is None or rt is None:
        return None
    matches: list[tuple[float, float, LegacyRtRowReference]] = []
    for oracle_row in oracle_rows:
        ppm = abs(oracle_row.mz - mz) / mz * 1_000_000
        rt_delta = abs(oracle_row.rt - rt)
        if ppm <= mz_ppm and rt_delta <= rt_tolerance_min:
            matches.append((rt_delta, ppm, oracle_row))
    if not matches:
        return None
    matches.sort(key=lambda item: (item[0], item[1], item[2].row_number))
    if len(matches) > 1 and _legacy_oracle_match_is_ambiguous(matches):
        return None
    return matches[0][2]


def _legacy_oracle_match_is_ambiguous(
    matches: Sequence[tuple[float, float, LegacyRtRowReference]],
) -> bool:
    best = matches[0]
    second = matches[1]
    return abs(best[0] - second[0]) <= 0.05 and abs(best[1] - second[1]) <= 5.0


def _to_float(value: object) -> float | None:
    try:
        number = float(str(value).strip())
    except (TypeError, ValueError):
        return None
    if not math.isfinite(number):
        return None
    return number


def _append_unique_cell(row: dict[str, str], column: str, value: str) -> None:
    if not value:
        return
    existing = tuple(part for part in row.get(column, "").split(";") if part)
    if value in existing:
        return
    row[column] = ";".join((*existing, value))


def _select_conflicting_matrix_value(existing: str, incoming: str) -> str:
    existing_value = _to_float(existing)
    incoming_value = _to_float(incoming)
    if existing_value is None or incoming_value is None:
        return existing
    return incoming if incoming_value > existing_value else existing


def _new_peak_hypothesis_matrix_row(
    peak_hypothesis_id: str,
    *,
    family_id: str,
    candidate_container_id: str,
    row_identity_basis: str,
    legacy_rt_row_context_id: str,
    review_row: Mapping[str, str],
    sample_columns: Sequence[str],
) -> dict[str, str]:
    row = {
        "peak_hypothesis_id": peak_hypothesis_id,
        "feature_family_id": family_id,
        "candidate_container_id": candidate_container_id or family_id,
        "row_identity_basis": row_identity_basis,
        "legacy_rt_row_context_id": legacy_rt_row_context_id,
        "neutral_loss_tag": review_row.get("neutral_loss_tag", ""),
        "family_center_mz": review_row.get("family_center_mz", ""),
        "family_center_rt": review_row.get("family_center_rt", ""),
    }
    row.update({sample: "" for sample in sample_columns})
    return row


def _reject_source_overwrite(
    *,
    output_paths: Sequence[Path],
    source_paths: Sequence[Path],
) -> None:
    output_resolved = {path.resolve() for path in output_paths}
    source_resolved = {path.resolve() for path in source_paths}
    overlap = sorted(str(path) for path in output_resolved & source_resolved)
    if overlap:
        raise ValueError(
            "activation product output would overwrite source alignment artifacts; "
            "write formal outputs to a separate directory or pass "
            "--allow-overwrite-source. Overlap: "
            + ", ".join(overlap)
        )


def _activated_review_rows(
    *,
    review_rows: Sequence[Mapping[str, str]],
    review_header: Sequence[str],
    decisions: Sequence[Mapping[str, str]],
    matrix_by_family: Mapping[str, Mapping[str, str]],
    sample_columns: Sequence[str],
    cell_status_by_key: Mapping[tuple[str, str], str],
    matrix_effects: Mapping[tuple[str, str], str],
    family_notes: Mapping[str, Sequence[str]],
    family_blocked: set[str],
) -> list[dict[str, str]]:
    counts_by_family: dict[str, Counter[str]] = {}
    rules_by_family: dict[str, list[str]] = {}
    for decision in decisions:
        family_id = decision.get("feature_family_id", "")
        counts_by_family.setdefault(family_id, Counter())[
            decision.get("activation_status", "")
        ] += 1
        rule = decision.get("contract_rule_id", "")
        if rule:
            rules_by_family.setdefault(family_id, []).append(rule)

    rows: list[dict[str, str]] = []
    for review_row in review_rows:
        family_id = review_row["feature_family_id"]
        row = {column: review_row.get(column, "") for column in review_header}
        accepted_samples = _accepted_samples(
            matrix_by_family.get(family_id),
            sample_columns,
        )
        accepted_rescue_count = sum(
            1
            for sample in accepted_samples
            if cell_status_by_key.get((family_id, sample)) == "rescued"
        )
        family_counts = counts_by_family.get(family_id, Counter())
        if family_id in family_blocked:
            row["identity_decision"] = "audit_family"
            row["identity_confidence"] = "review"
            row["identity_reason"] = "activation_family_required_tag_gate"
        elif family_id in matrix_by_family and family_counts["auto_activate"]:
            if row.get("include_in_primary_matrix") != "TRUE":
                row["identity_decision"] = (
                    row.get("identity_decision") or "provisional_discovery"
                )
                row["identity_confidence"] = "medium"
                row["identity_reason"] = "activation_peak_hypothesis_candidate"
        row["include_in_primary_matrix"] = (
            "TRUE" if family_id in matrix_by_family else "FALSE"
        )
        row["accepted_cell_count"] = str(len(accepted_samples))
        row["accepted_rescue_count"] = str(accepted_rescue_count)
        row["activation_auto_activate_count"] = str(
            family_counts["auto_activate"]
        )
        row["activation_auto_block_count"] = str(
            family_counts["auto_block"]
        )
        row["activation_review_required_count"] = str(
            family_counts["review_required"]
        )
        row["activation_blocked_cell_count"] = str(
            sum(
                1
                for (effect_family, _sample), effect in matrix_effects.items()
                if effect_family == family_id and effect == "blanked"
            )
        )
        row["activation_written_cell_count"] = str(
            sum(
                1
                for (effect_family, _sample), effect in matrix_effects.items()
                if effect_family == family_id and effect == "written"
            )
        )
        row["activation_rules"] = ";".join(
            dict.fromkeys(rules_by_family.get(family_id, ()))
        )
        row["activation_application_note"] = ";".join(
            dict.fromkeys(family_notes.get(family_id, ()))
        )
        rows.append(row)
    return rows


def _activated_cell_rows(
    *,
    cell_rows: Sequence[Mapping[str, str]],
    decisions_by_key: Mapping[tuple[str, str], Mapping[str, str]],
    matrix_effects: Mapping[tuple[str, str], str],
) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for cell in cell_rows:
        key = (cell["feature_family_id"], cell["sample_stem"])
        decision = decisions_by_key.get(key, {})
        row = dict(cell)
        row.update(
            {
                "activation_status": decision.get("activation_status", ""),
                "activation_action": decision.get("activation_action", ""),
                "activation_product_effect": decision.get("product_effect", ""),
                "activation_contract_rule_id": decision.get("contract_rule_id", ""),
                "activation_peak_hypothesis_id": decision.get(
                    "peak_hypothesis_id",
                    "",
                ),
                "activation_unit_scope": decision.get("activation_unit_scope", ""),
                "activation_matrix_value_effect": matrix_effects.get(key, "unchanged"),
                "activation_reason": decision.get("activation_reason", ""),
            }
        )
        rows.append(row)
    return rows


def _value_delta_rows(
    *,
    decisions: Sequence[Mapping[str, str]],
    original_matrix_by_family: Mapping[str, Mapping[str, str]],
    matrix_by_family: Mapping[str, Mapping[str, str]],
    cells_by_key: Mapping[tuple[str, str], Mapping[str, str]],
    sample_columns: Sequence[str],
    matrix_effects: Mapping[tuple[str, str], str],
) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    sample_column_set = set(sample_columns)
    for decision in decisions:
        family_id = decision.get("feature_family_id", "")
        sample_id = decision.get("sample_id", "")
        if not sample_id or sample_id == "__family_context__":
            continue
        key = (family_id, sample_id)
        original_row = original_matrix_by_family.get(family_id)
        activated_row = matrix_by_family.get(family_id)
        original_value = (
            original_row.get(sample_id, "")
            if original_row is not None and sample_id in sample_column_set
            else ""
        )
        activated_value = (
            activated_row.get(sample_id, "")
            if activated_row is not None and sample_id in sample_column_set
            else ""
        )
        cell = cells_by_key.get(key, {})
        rows.append(
            {
                "activation_value_delta_schema_version": (
                    ACTIVATION_VALUE_DELTA_SCHEMA_VERSION
                ),
                "feature_family_id": family_id,
                "candidate_container_id": decision.get(
                    "candidate_container_id",
                    family_id,
                ),
                "sample_id": sample_id,
                "peak_hypothesis_id": decision.get("peak_hypothesis_id", ""),
                "activation_unit_scope": decision.get("activation_unit_scope", ""),
                "activation_status": decision.get("activation_status", ""),
                "product_effect": decision.get("product_effect", ""),
                "contract_rule_id": decision.get("contract_rule_id", ""),
                "original_matrix_value": original_value,
                "activated_matrix_value": activated_value,
                "source_cell_status": cell.get("status", ""),
                "source_cell_area": cell.get("area", ""),
                "matrix_value_effect": matrix_effects.get(key, "unchanged"),
                "value_changed": (
                    "TRUE" if original_value != activated_value else "FALSE"
                ),
                "activation_reason": decision.get("activation_reason", ""),
            }
        )
    return rows


def _summary_row(
    *,
    acceptance: Mapping[str, str],
    decisions: Sequence[Mapping[str, str]],
    output_mode: str,
    input_matrix_rows: int,
    output_matrix_rows: int,
    formal_matrix_stats: FormalMatrixStats,
    matrix_effects: Mapping[tuple[str, str], str],
    families_added: set[str],
    families_removed: set[str],
) -> dict[str, str]:
    decision_counts = Counter(row.get("activation_status", "") for row in decisions)
    blanked = sum(1 for effect in matrix_effects.values() if effect == "blanked")
    written = sum(1 for effect in matrix_effects.values() if effect == "written")
    return {
        "activation_application_schema_version": (
            ACTIVATION_APPLICATION_SCHEMA_VERSION
        ),
        "application_status": "applied",
        "activation_output_mode": output_mode,
        "acceptance_status": acceptance.get("acceptance_status", ""),
        "blast_radius_current": acceptance.get("blast_radius_current", ""),
        "decision_rows_total": str(len(decisions)),
        "input_matrix_rows": str(input_matrix_rows),
        "output_matrix_rows": str(output_matrix_rows),
        "matrix_row_identity": (
            "peak_hypothesis_id" if output_mode == "formal" else "feature_family_id"
        ),
        "canonical_row_identity_ready": _canonical_row_identity_ready(
            output_mode,
            formal_matrix_stats,
        ),
        "canonical_row_identity_blockers": _canonical_row_identity_blocker(
            output_mode,
            formal_matrix_stats,
        ),
        "canonical_row_identity_scope": _canonical_row_identity_scope(
            output_mode,
            formal_matrix_stats,
        ),
        "family_projection_semantics": _family_projection_semantics(
            output_mode,
            formal_matrix_stats,
        ),
        "legacy_rt_row_context_authority": (
            "context_only_not_identity_authority"
            if formal_matrix_stats.legacy_rt_row_context_rows
            else "not_applicable"
        ),
        "all_family_split_science_ready": "FALSE",
        "legacy_rt_row_context_rows": str(
            formal_matrix_stats.legacy_rt_row_context_rows
        ),
        "family_projection_rows": str(formal_matrix_stats.family_projection_rows),
        "family_projection_rows_excluded": str(
            formal_matrix_stats.family_projection_rows_excluded
        ),
        "family_projection_cells_excluded": str(
            formal_matrix_stats.family_projection_cells_excluded
        ),
        "matrix_value_conflict_cells": str(
            formal_matrix_stats.matrix_value_conflict_cells
        ),
        "matrix_value_conflict_policy": (
            "max_area_pending_baseline"
            if formal_matrix_stats.matrix_value_conflict_cells
            else "not_applicable"
        ),
        "auto_activate_count": str(decision_counts["auto_activate"]),
        "auto_block_count": str(decision_counts["auto_block"]),
        "matrix_cells_written": str(written),
        "matrix_cells_blanked": str(blanked),
        "families_added_to_matrix": str(len(families_added)),
        "families_removed_from_matrix": str(len(families_removed)),
        "summary_reason": "explicit_activation_sidecar_applied",
    }


def _family_projection_semantics(
    output_mode: str,
    formal_matrix_stats: FormalMatrixStats,
) -> str:
    if output_mode != "formal":
        return "not_applicable"
    if formal_matrix_stats.family_projection_rows:
        return "projection_not_split_proof"
    if formal_matrix_stats.family_projection_rows_excluded:
        return "excluded_from_canonical_output"
    return "explicit_hypothesis_only"


def _canonical_row_identity_ready(
    output_mode: str,
    formal_matrix_stats: FormalMatrixStats,
) -> str:
    if output_mode != "formal":
        return "FALSE"
    if (
        formal_matrix_stats.family_projection_rows
        or formal_matrix_stats.family_projection_rows_excluded
    ):
        return "FALSE"
    return "TRUE"


def _canonical_row_identity_blocker(
    output_mode: str,
    formal_matrix_stats: FormalMatrixStats,
) -> str:
    if output_mode != "formal":
        return "formal_output_not_requested"
    if formal_matrix_stats.family_projection_rows:
        return "family_projection_present"
    if formal_matrix_stats.family_projection_rows_excluded:
        return "family_projection_excluded_incomplete_scope"
    return "none"


def _canonical_row_identity_scope(
    output_mode: str,
    formal_matrix_stats: FormalMatrixStats,
) -> str:
    if output_mode != "formal":
        return "legacy_feature_family_row"
    if formal_matrix_stats.family_projection_rows:
        return "partial_peak_hypothesis_with_family_projections"
    if formal_matrix_stats.family_projection_rows_excluded:
        return "partial_canonical_peak_hypothesis_rows_only"
    return "formal_peak_hypothesis_identity"


def _matrix_value_for_activation(cell: Mapping[str, str] | None) -> str:
    if cell is None:
        return ""
    if cell.get("primary_matrix_area_source", "") != ASLS_PRIMARY_MATRIX_AREA_SOURCE:
        return ""
    return cell.get("primary_matrix_area", "")


def _has_untrusted_primary_or_raw_area(cell: Mapping[str, str] | None) -> bool:
    return cell is not None and bool(
        cell.get("primary_matrix_area", "") or cell.get("area", "")
    )


def _new_matrix_row(
    family_id: str,
    *,
    sample_columns: Sequence[str],
    review_row: Mapping[str, str],
) -> dict[str, str]:
    row = {
        "feature_family_id": family_id,
        "neutral_loss_tag": review_row.get("neutral_loss_tag", ""),
        "family_center_mz": review_row.get("family_center_mz", ""),
        "family_center_rt": review_row.get("family_center_rt", ""),
    }
    row.update({sample: "" for sample in sample_columns})
    return row


def _accepted_samples(
    matrix_row: Mapping[str, str] | None,
    sample_columns: Sequence[str],
) -> tuple[str, ...]:
    if matrix_row is None:
        return ()
    return tuple(sample for sample in sample_columns if matrix_row.get(sample))


def _has_any_sample_value(
    row: Mapping[str, str],
    sample_columns: Sequence[str],
) -> bool:
    return any(row.get(sample) for sample in sample_columns)


def _require_columns(header: Sequence[str], columns: Sequence[str]) -> None:
    missing = [column for column in columns if column not in header]
    if missing:
        raise ValueError(f"missing required columns: {','.join(missing)}")


def _read_tsv(path: Path) -> tuple[dict[str, str], ...]:
    _header, rows = _read_tsv_with_header(path)
    return rows


def _read_tsv_with_header(
    path: Path,
) -> tuple[tuple[str, ...], tuple[dict[str, str], ...]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle, delimiter="\t")
        header = tuple(reader.fieldnames or ())
        return header, tuple(dict(row) for row in reader)


def _write_tsv(
    path: Path,
    fieldnames: Sequence[str],
    rows: Sequence[Mapping[str, str]],
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=fieldnames,
            delimiter="\t",
            lineterminator="\n",
        )
        writer.writeheader()
        for row in rows:
            writer.writerow({column: row.get(column, "") for column in fieldnames})
