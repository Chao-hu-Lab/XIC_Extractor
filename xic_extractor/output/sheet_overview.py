from __future__ import annotations

from openpyxl.styles import Alignment, Font

from xic_extractor.output.workbook_styles import (
    _OVERVIEW_HEADER_FILL,
    BORDER,
    CENTER,
    GREY,
    WHITE,
    _apply,
    _excel_text,
    _fill,
)


def _build_overview_sheet(
    ws,
    rows: list[dict[str, str]],
    diagnostics: list[dict[str, str]],
    review_rows: list[dict[str, str]],
) -> None:
    ws.title = "Overview"
    ws.merge_cells("A1:D1")
    _apply(
        ws["A1"],
        value="XIC Review Overview",
        font=Font(bold=True, color="FFFFFF", size=14),
        fill=_fill(_OVERVIEW_HEADER_FILL),
        alignment=Alignment(horizontal="left", vertical="center"),
        border=BORDER,
    )
    ws.row_dimensions[1].height = 28

    metrics = [
        ("Samples", len(_distinct_values(rows, "SampleName"))),
        ("Targets", len(_distinct_values(rows, "Target"))),
        ("Review Items", len(review_rows)),
        ("Diagnostics", len(diagnostics)),
    ]
    for row_idx, (label, value) in enumerate(metrics, start=3):
        _apply(
            ws.cell(row=row_idx, column=1, value=label),
            font=Font(bold=True),
            fill=_fill(GREY),
            alignment=CENTER,
            border=BORDER,
        )
        _apply(
            ws.cell(row=row_idx, column=2, value=value),
            fill=_fill(WHITE),
            alignment=CENTER,
            border=BORDER,
        )

    next_row = _write_overview_count_section(
        ws,
        8,
        "Top Targets",
        "Target",
        _top_review_counts(review_rows, "Target"),
    )
    next_row = _write_overview_count_section(
        ws,
        next_row + 1,
        "Top Samples",
        "Sample",
        _top_review_counts(review_rows, "Sample"),
    )
    _write_overview_how_to_read(ws, next_row + 1)

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18

def _distinct_values(rows: list[dict[str, str]], key: str) -> set[str]:
    return {value for row in rows if (value := row.get(key, ""))}


def _top_review_counts(
    review_rows: list[dict[str, str]], key: str
) -> list[tuple[str, int]]:
    counts: dict[str, int] = {}
    for row in review_rows:
        value = row.get(key, "")
        if not value:
            continue
        counts[value] = counts.get(value, 0) + 1
    return sorted(counts.items(), key=lambda item: (-item[1], item[0]))[:5]


def _write_overview_count_section(
    ws,
    start_row: int,
    title: str,
    label_header: str,
    counts: list[tuple[str, int]],
) -> int:
    _apply(
        ws.cell(row=start_row, column=1, value=title),
        font=Font(bold=True, color="FFFFFF"),
        fill=_fill(_OVERVIEW_HEADER_FILL),
        alignment=CENTER,
        border=BORDER,
    )
    _apply(
        ws.cell(row=start_row + 1, column=1, value=label_header),
        font=Font(bold=True),
        fill=_fill(GREY),
        alignment=CENTER,
        border=BORDER,
    )
    _apply(
        ws.cell(row=start_row + 1, column=2, value="Review Items"),
        font=Font(bold=True),
        fill=_fill(GREY),
        alignment=CENTER,
        border=BORDER,
    )

    if not counts:
        _apply(
            ws.cell(row=start_row + 2, column=1, value="None"),
            fill=_fill(WHITE),
            alignment=CENTER,
            border=BORDER,
        )
        _apply(
            ws.cell(row=start_row + 2, column=2, value=0),
            fill=_fill(WHITE),
            alignment=CENTER,
            border=BORDER,
        )
        return start_row + 3

    for offset, (label, count) in enumerate(counts, start=2):
        row_idx = start_row + offset
        fill_hex = GREY if row_idx % 2 == 0 else WHITE
        _apply(
            ws.cell(row=row_idx, column=1, value=_excel_text(label)),
            fill=_fill(fill_hex),
            alignment=CENTER,
            border=BORDER,
        )
        _apply(
            ws.cell(row=row_idx, column=2, value=count),
            fill=_fill(fill_hex),
            alignment=CENTER,
            border=BORDER,
        )
    return start_row + len(counts) + 2


def _write_overview_how_to_read(ws, start_row: int) -> int:
    _apply(
        ws.cell(row=start_row, column=1, value="How to read"),
        font=Font(bold=True, color="FFFFFF"),
        fill=_fill(_OVERVIEW_HEADER_FILL),
        alignment=CENTER,
        border=BORDER,
    )
    notes = [
        (
            "Start with Summary Detection % to find targets with systematic "
            "detection loss."
        ),
        "Review Queue has one row per sample-target needing attention.",
        "Flagged % is review workload, not detection failure.",
        (
            "NL_FAIL rows are review evidence, not counted detections or "
            "Summary analytical aggregates."
        ),
        "Diagnostics is a hidden technical log for debugging verbose evidence.",
        "Score Breakdown is a technical audit sheet when enabled.",
        "HTML Review Report is for visual batch QA when enabled.",
    ]
    for offset, note in enumerate(notes, start=1):
        row_idx = start_row + offset
        ws.merge_cells(
            start_row=row_idx,
            start_column=1,
            end_row=row_idx,
            end_column=4,
        )
        _apply(
            ws.cell(row=row_idx, column=1, value=note),
            fill=_fill(WHITE),
            alignment=Alignment(horizontal="left", vertical="center", wrap_text=True),
            border=BORDER,
        )
    return start_row + len(notes) + 1
