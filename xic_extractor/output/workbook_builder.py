from __future__ import annotations

from collections.abc import Callable
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook

from xic_extractor.config import ExtractionConfig, Target
from xic_extractor.injection_rolling import read_injection_order
from xic_extractor.output.review_queue_model import _review_queue_rows
from xic_extractor.output.review_report import (
    review_report_path_for_excel,
    write_review_report,
)
from xic_extractor.output.sheet_diagnostics import _build_diagnostics_sheet
from xic_extractor.output.sheet_metadata import _build_metadata_sheet
from xic_extractor.output.sheet_overview import _build_overview_sheet
from xic_extractor.output.sheet_results import _build_data_sheet
from xic_extractor.output.sheet_review_queue import _build_review_queue_sheet
from xic_extractor.output.sheet_score_breakdown import _build_score_breakdown_sheet
from xic_extractor.output.sheet_summary import (
    _build_summary_sheet,
    _is_long_detected,
    _long_nl_counts,
    _target_summaries,
)
from xic_extractor.output.sheet_targets import _build_targets_sheet
from xic_extractor.output.workbook_inputs import (
    _read_diagnostics,
    _read_long_results,
    _read_score_breakdown,
)
from xic_extractor.output.workbook_styles import _apply_sheet_role_styles

ReviewReportWriter = Callable[..., Path]


def write_workbook_from_rows(
    config: ExtractionConfig,
    targets: list[Target],
    rows: list[dict[str, str]],
    *,
    diagnostics: list[dict[str, str]],
    score_breakdown: list[dict[str, str]] | None = None,
    output_path: Path,
    report_writer: ReviewReportWriter = write_review_report,
) -> Path:
    review_rows = _review_queue_rows(rows, diagnostics)
    wb = Workbook()
    ws_overview = wb.active
    _build_overview_sheet(ws_overview, rows, diagnostics, review_rows)

    ws_review = wb.create_sheet("Review Queue")
    _build_review_queue_sheet(ws_review, review_rows)

    ws_data = wb.create_sheet("XIC Results")
    _build_data_sheet(ws_data, rows)

    ws_summary = wb.create_sheet("Summary")
    _build_summary_sheet(
        ws_summary,
        rows,
        count_no_ms2_as_detected=config.count_no_ms2_as_detected,
        review_rows=review_rows,
    )

    ws_targets = wb.create_sheet("Targets")
    _build_targets_sheet(ws_targets, targets)

    ws_diagnostics = wb.create_sheet("Diagnostics")
    _build_diagnostics_sheet(ws_diagnostics, diagnostics)
    ws_diagnostics.sheet_state = "hidden"

    ws_metadata = wb.create_sheet("Run Metadata")
    _build_metadata_sheet(ws_metadata, config)

    if config.emit_score_breakdown and score_breakdown:
        ws_breakdown = wb.create_sheet("Score Breakdown")
        _build_score_breakdown_sheet(ws_breakdown, score_breakdown)

    wb.active = wb.index(ws_overview)
    _apply_sheet_role_styles(wb)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    wb.close()

    if config.emit_review_report:
        injection_order = (
            read_injection_order(config.injection_order_source)
            if config.injection_order_source is not None
            else None
        )
        report_writer(
            review_report_path_for_excel(output_path),
            rows,
            diagnostics=diagnostics,
            review_rows=review_rows,
            count_no_ms2_as_detected=config.count_no_ms2_as_detected,
            injection_order=injection_order,
        )
    return output_path


def run_from_config(
    config: ExtractionConfig,
    targets: list[Target],
    *,
    report_writer: ReviewReportWriter = write_review_report,
) -> Path:
    output_dir = config.output_csv.parent
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    excel_path = output_dir / f"xic_results_{timestamp}.xlsx"

    rows = _read_long_results(config, targets)
    if not rows:
        print("CSV is empty.")
        return excel_path

    diagnostics = _read_diagnostics(config.diagnostics_csv)
    score_breakdown = _read_score_breakdown(config)
    write_workbook_from_rows(
        config,
        targets,
        rows,
        diagnostics=diagnostics,
        score_breakdown=score_breakdown,
        output_path=excel_path,
        report_writer=report_writer,
    )
    _print_summary(excel_path, rows, config.count_no_ms2_as_detected)
    return excel_path

def _print_summary(
    excel_path: Path,
    rows: list[dict[str, str]],
    count_no_ms2_as_detected: bool,
) -> None:
    print(f"Saved : {excel_path}")
    sample_count = len({row.get("SampleName", "") for row in rows})
    print(f"Rows  : {sample_count}")
    for target in _target_summaries(rows):
        label = target["Target"]
        target_rows = [row for row in rows if row.get("Target") == label]
        detected = sum(
            1 for row in target_rows if _is_long_detected(row, count_no_ms2_as_detected)
        )
        note = (
            " (NL confirmed)" if any(row.get("NL", "") for row in target_rows) else ""
        )
        print(f"  {label} detected{note}: {detected}/{len(target_rows)}")
    for target in _target_summaries(rows):
        label = target["Target"]
        target_rows = [row for row in rows if row.get("Target") == label]
        if not any(row.get("NL", "") for row in target_rows):
            continue
        counts = _long_nl_counts(target_rows)
        print(
            f"  {label}_NL  OK:{counts['OK']}  WARN:{counts['WARN']}  "
            f"FAIL:{counts['NL_FAIL']}  NO_MS2:{counts['NO_MS2']}"
        )
    for target in _target_summaries(rows):
        if target.get("Role") != "ISTD":
            continue
        label = target["Target"]
        target_rows = [row for row in rows if row.get("Target") == label]
        detected = sum(
            1 for row in target_rows if _is_long_detected(row, count_no_ms2_as_detected)
        )
        if detected < len(target_rows):
            print(f"ISTD_ND: {label} {detected}/{len(target_rows)}")
