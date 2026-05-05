from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from scripts.csv_to_excel import (
    _apply_sheet_role_styles,
    _build_data_sheet,
    _build_diagnostics_sheet,
    _build_metadata_sheet,
    _build_overview_sheet,
    _build_review_queue_sheet,
    _build_score_breakdown_sheet,
    _build_summary_sheet,
    _build_targets_sheet,
    _review_queue_rows,
)
from xic_extractor.config import ExtractionConfig, Target
from xic_extractor.extractor import RunOutput
from xic_extractor.output import csv_writers
from xic_extractor.output.messages import DiagnosticRecord
from xic_extractor.output.review_report import (
    review_report_path_for_excel,
    write_review_report,
)


def write_excel_from_run_output(
    config: ExtractionConfig,
    targets: list[Target],
    run_output: RunOutput,
    *,
    output_path: Path,
) -> Path:
    """Write an xlsx workbook directly from in-memory extraction output."""
    rows = _run_output_to_long_rows(run_output, targets)
    diagnostics = _diagnostics_to_rows(run_output.diagnostics)
    score_breakdown = (
        _run_output_to_score_breakdown_rows(run_output)
        if config.emit_score_breakdown
        else []
    )
    review_rows = _review_queue_rows(rows, diagnostics)

    wb = Workbook()
    ws_overview = wb.active
    _build_overview_sheet(ws_overview, rows, diagnostics, review_rows)

    ws_review = wb.create_sheet("Review Queue")
    _build_review_queue_sheet(ws_review, review_rows)

    ws_data = wb.create_sheet("XIC Results")
    _build_data_sheet(ws_data, rows)

    ws_summary = wb.create_sheet("Summary")
    _build_summary_sheet(
        ws_summary,
        rows,
        count_no_ms2_as_detected=config.count_no_ms2_as_detected,
        review_rows=review_rows,
    )

    ws_targets = wb.create_sheet("Targets")
    _build_targets_sheet(ws_targets, targets)

    ws_diagnostics = wb.create_sheet("Diagnostics")
    _build_diagnostics_sheet(ws_diagnostics, diagnostics)
    ws_diagnostics.sheet_state = "hidden"

    ws_metadata = wb.create_sheet("Run Metadata")
    _build_metadata_sheet(ws_metadata, config)

    if config.emit_score_breakdown and score_breakdown:
        ws_breakdown = wb.create_sheet("Score Breakdown")
        _build_score_breakdown_sheet(ws_breakdown, score_breakdown)

    wb.active = wb.index(ws_overview)
    _apply_sheet_role_styles(wb)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    wb.close()
    if config.emit_review_report:
        write_review_report(
            review_report_path_for_excel(output_path),
            rows,
            diagnostics=diagnostics,
            review_rows=review_rows,
            count_no_ms2_as_detected=config.count_no_ms2_as_detected,
        )
    return output_path


def _run_output_to_long_rows(
    run_output: RunOutput, targets: list[Target]
) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for file_result in run_output.file_results:
        rows.extend(csv_writers._long_output_rows(file_result, targets))
    return rows


def _diagnostics_to_rows(
    diagnostics: list[DiagnosticRecord],
) -> list[dict[str, str]]:
    return [
        {
            "SampleName": record.sample_name,
            "Target": record.target_label,
            "Issue": record.issue,
            "Reason": record.reason,
        }
        for record in diagnostics
    ]


def _run_output_to_score_breakdown_rows(run_output: RunOutput) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for file_result in run_output.file_results:
        for result in file_result.extraction_results:
            severities = {label: severity for severity, label in result.severities}
            rows.append(
                {
                    "SampleName": file_result.sample_name,
                    "Target": result.target_label,
                    "symmetry": _format_optional_severity(severities.get("symmetry")),
                    "local_sn": _format_optional_severity(severities.get("local_sn")),
                    "nl_support": _format_optional_severity(
                        severities.get("nl_support")
                    ),
                    "rt_prior": _format_optional_severity(severities.get("rt_prior")),
                    "rt_centrality": _format_optional_severity(
                        severities.get("rt_centrality")
                    ),
                    "noise_shape": _format_optional_severity(
                        severities.get("noise_shape")
                    ),
                    "peak_width": _format_optional_severity(
                        severities.get("peak_width")
                    ),
                    "Quality Penalty": str(result.quality_penalty),
                    "Quality Flags": ",".join(result.quality_flags),
                    "Total Severity": str(result.total_severity),
                    "Confidence": result.confidence,
                    "Prior RT": _format_optional_number(result.prior_rt),
                    "Prior Source": result.prior_source,
                }
            )
    return rows


def _format_optional_severity(value: int | None) -> str:
    if value is None:
        return ""
    return str(value)


def _format_optional_number(value: float | None) -> str:
    if value is None:
        return "NA"
    return f"{value:g}"
