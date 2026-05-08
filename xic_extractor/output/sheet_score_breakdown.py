from __future__ import annotations

from openpyxl.utils import get_column_letter

from xic_extractor.output.workbook_styles import (
    _MS2_HEADER,
    BORDER,
    CENTER,
    GREY,
    WHITE,
    _apply,
    _excel_text,
    _fill,
    _header_style,
    _safe_float,
)


def _build_score_breakdown_sheet(ws, rows: list[dict[str, str]]) -> None:
    if not rows:
        return
    headers = list(rows[0].keys())
    for col_idx, header in enumerate(headers, start=1):
        _apply(
            ws.cell(row=1, column=col_idx, value=header),
            **_header_style(_MS2_HEADER),
        )
    for row_idx, row in enumerate(rows, start=2):
        fill_hex = GREY if row_idx % 2 == 0 else WHITE
        for col_idx, header in enumerate(headers, start=1):
            raw_value = row.get(header, "")
            value = (
                _safe_float(raw_value)
                if header
                in {
                    "symmetry",
                    "local_sn",
                    "nl_support",
                    "rt_prior",
                    "rt_centrality",
                    "noise_shape",
                    "peak_width",
                    "Quality Penalty",
                    "Total Severity",
                    "Prior RT",
                    "Base Score",
                    "Positive Points",
                    "Negative Points",
                    "Raw Score",
                }
                else _excel_text(raw_value)
            )
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            _apply(cell, fill=_fill(fill_hex), alignment=CENTER, border=BORDER)
            if isinstance(value, float) and header == "Prior RT":
                cell.number_format = "0.0000"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(1, len(rows) + 1)}"
    ws.freeze_panes = "A2"
