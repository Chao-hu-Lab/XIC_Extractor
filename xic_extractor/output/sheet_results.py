from __future__ import annotations

from openpyxl.utils import get_column_letter

from xic_extractor.output.schema import LONG_ADVANCED_HEADERS as _ADVANCED_HEADERS
from xic_extractor.output.schema import LONG_HEADERS as _LONG_HEADERS
from xic_extractor.output.workbook_styles import (
    _MS2_HEADER,
    _SAMPLE_HEADER,
    BORDER,
    CENTER,
    CENTER_WRAP,
    _apply,
    _header_style,
    _long_cell_fill,
    _long_cell_value,
    _long_column_width,
    _long_number_format,
)


def _build_data_sheet(ws, rows: list[dict[str, str]]) -> None:
    for col_idx, header in enumerate(_LONG_HEADERS, start=1):
        color = _MS2_HEADER if header in _ADVANCED_HEADERS else _SAMPLE_HEADER
        _apply(ws.cell(row=1, column=col_idx, value=header), **_header_style(color))
    ws.row_dimensions[1].height = 30

    for row_idx, row in enumerate(rows, start=2):
        alt = row_idx % 2 == 0
        for col_idx, header in enumerate(_LONG_HEADERS, start=1):
            raw_val = row.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx)
            value = _long_cell_value(header, raw_val)
            cell.value = value
            _apply(
                cell,
                fill=_long_cell_fill(header, raw_val, alt),
                alignment=CENTER,
                border=BORDER,
            )
            if isinstance(value, float):
                cell.number_format = _long_number_format(header)

    ws.auto_filter.ref = (
        f"A1:{get_column_letter(len(_LONG_HEADERS))}{max(1, len(rows) + 1)}"
    )
    ws.freeze_panes = "A2"
    for col_idx, header in enumerate(_LONG_HEADERS, start=1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = _long_column_width(header)
        if header in _ADVANCED_HEADERS:
            ws.column_dimensions[letter].hidden = True
            ws.column_dimensions[letter].outlineLevel = 1
    _merge_repeated_identity_cells(ws, rows)


def _merge_repeated_identity_cells(ws, rows: list[dict[str, str]]) -> None:
    if not rows:
        return
    block_start = 2
    previous = _identity_key(rows[0])
    for offset, row in enumerate(rows[1:], start=3):
        current = _identity_key(row)
        if current == previous:
            continue
        _merge_identity_block(ws, block_start, offset - 1)
        block_start = offset
        previous = current
    _merge_identity_block(ws, block_start, len(rows) + 1)


def _identity_key(row: dict[str, str]) -> tuple[str, str]:
    return row.get("SampleName", ""), row.get("Group", "")


def _merge_identity_block(ws, start_row: int, end_row: int) -> None:
    if end_row <= start_row:
        return
    for column in (1, 2):
        ws.merge_cells(
            start_row=start_row,
            start_column=column,
            end_row=end_row,
            end_column=column,
        )
        ws.cell(row=start_row, column=column).alignment = CENTER_WRAP
