from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from xic_extractor.output.schema import LONG_ADVANCED_HEADERS as _ADVANCED_HEADERS
from xic_extractor.output.workbook_values import (
    ND_ERROR,
    _excel_text,
    _nl_to_display,
    _safe_float,
)

_MS2_HEADER = "37474F"
_MS2_OK = "C8E6C9"
_MS2_WARN = "FFF9C4"
_MS2_FAIL = "FFCDD2"
_MS2_NO_MS2 = "E0E0E0"
_SAMPLE_HEADER = "2E4057"
_OVERVIEW_HEADER_FILL = "1F4E5F"
_DAILY_REVIEW_TAB = "1F4E5F"
_RESULT_TAB = "5B7C99"
_TECHNICAL_TAB = "B0BEC5"
WHITE = "FFFFFF"
GREY = "F5F5F5"
_THIN = Side(style="thin", color="BDBDBD")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
CENTER = Alignment(horizontal="center", vertical="center")
CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
_TARGETS_HEADERS = [
    "Label",
    "Role",
    "ISTD Pair",
    "m/z",
    "RT min",
    "RT max",
    "ppm tol",
    "NL (Da)",
    "Expected product m/z",
    "NL ppm warn",
    "NL ppm max",
]
_SUMMARY_HEADERS = [
    "Target",
    "Role",
    "ISTD Pair",
    "Detected",
    "Total",
    "Detection %",
    "Median Area (detected)",
    "Mean RT",
    "Area / ISTD ratio (paired detected)",
    "RT Delta vs ISTD",
    "NL OK",
    "NL WARN",
    "NL FAIL",
    "NO MS2",
    "Confidence HIGH",
    "Confidence MEDIUM",
    "Confidence LOW",
    "Confidence VERY_LOW",
    "Flagged Rows",
    "Flagged %",
    "MS2/NL Flags",
    "Low Confidence Rows",
]
_REVIEW_HEADERS = [
    "Priority",
    "Sample",
    "Target",
    "Role",
    "Status",
    "Why",
    "RT",
    "Area",
    "Action",
    "Issue Count",
    "Evidence",
]

def _fill(hex6: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex6)


def _apply(cell, **kw: object) -> None:
    for key, value in kw.items():
        setattr(cell, key, value)


def _header_style(hex6: str) -> dict[str, object]:
    return {
        "font": Font(bold=True, color="FFFFFF", size=11),
        "fill": _fill(hex6),
        "alignment": CENTER_WRAP,
        "border": BORDER,
    }


def _nl_cell_fill(value: str) -> PatternFill:
    if value == "OK":
        return _fill(_MS2_OK)
    if value.startswith("WARN_"):
        return _fill(_MS2_WARN)
    if value == "NO_MS2":
        return _fill(_MS2_NO_MS2)
    return _fill(_MS2_FAIL)


def _long_cell_value(header: str, raw_val: str) -> object:
    if header in {"SampleName", "Target", "ISTD Pair", "Confidence", "Reason"}:
        return _excel_text(raw_val)
    if header == "NL":
        return _nl_to_display(raw_val) if raw_val else ""
    if header in {"RT", "Area", "Int", "PeakStart", "PeakEnd", "PeakWidth"}:
        if raw_val in ND_ERROR:
            return raw_val
        parsed = _safe_float(raw_val)
        return parsed if parsed is not None else raw_val
    return raw_val


def _long_cell_fill(header: str, raw_val: str, alt: bool) -> PatternFill:
    if raw_val in ND_ERROR:
        return _fill("FFE0B2" if raw_val == "ND" else "FF7043")
    if header == "NL" and raw_val:
        return _nl_cell_fill(raw_val)
    if header in _ADVANCED_HEADERS:
        return _fill("ECEFF1")
    return _fill(GREY if alt else WHITE)


def _long_number_format(header: str) -> str:
    if header in {"RT", "PeakStart", "PeakEnd"}:
        return "0.0000"
    if header in {"Area", "Int"}:
        return "0.00E+00"
    if header == "PeakWidth":
        return "0.0000"
    return "#,##0"


def _long_column_width(header: str) -> int:
    return {
        "SampleName": 28,
        "Group": 14,
        "Target": 24,
        "Role": 12,
        "ISTD Pair": 24,
        "RT": 12,
        "Area": 16,
        "NL": 14,
        "Int": 14,
        "PeakStart": 14,
        "PeakEnd": 14,
        "PeakWidth": 14,
        "Confidence": 14,
        "Reason": 42,
    }[header]

def _review_cell_value(header: str, raw_value: str) -> object:
    if header in {"Priority", "Issue Count"}:
        parsed = _safe_float(raw_value)
        return int(parsed) if parsed is not None else raw_value
    if header in {"RT", "Area"}:
        if raw_value in ND_ERROR:
            return raw_value
        parsed = _safe_float(raw_value)
        return parsed if parsed is not None else raw_value
    if header in {
        "Sample",
        "Target",
        "Why",
        "Action",
        "Evidence",
    }:
        return _excel_text(raw_value)
    return raw_value


def _review_cell_fill(header: str, raw_value: str, row_fill_hex: str) -> PatternFill:
    return _fill(row_fill_hex)


def _review_priority_fill(priority: str) -> str:
    return {"1": _MS2_FAIL, "2": _MS2_WARN}.get(priority, GREY)


def _review_number_format(header: str) -> str:
    if header == "RT":
        return "0.0000"
    if header == "Area":
        return "0.00E+00"
    return "#,##0"



def _apply_sheet_role_styles(wb: Workbook) -> None:
    role_colors = {
        "Overview": _DAILY_REVIEW_TAB,
        "Review Queue": _DAILY_REVIEW_TAB,
        "XIC Results": _RESULT_TAB,
        "Summary": _RESULT_TAB,
        "Targets": _TECHNICAL_TAB,
        "Diagnostics": _TECHNICAL_TAB,
        "Run Metadata": _TECHNICAL_TAB,
        "Score Breakdown": _TECHNICAL_TAB,
    }
    for name, color in role_colors.items():
        if name in wb.sheetnames:
            wb[name].sheet_properties.tabColor = color
