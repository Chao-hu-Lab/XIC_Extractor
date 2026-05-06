from __future__ import annotations

import re

from openpyxl.utils import get_column_letter

from xic_extractor.output.schema import DIAGNOSTIC_HEADERS as _DIAGNOSTIC_HEADERS
from xic_extractor.output.workbook_styles import (
    _MS2_FAIL,
    _MS2_HEADER,
    _MS2_NO_MS2,
    _MS2_WARN,
    BORDER,
    CENTER,
    _apply,
    _excel_text,
    _fill,
    _header_style,
)


def _build_diagnostics_sheet(ws, rows: list[dict[str, str]]) -> None:
    for col_idx, header in enumerate(_DIAGNOSTIC_HEADERS, start=1):
        _apply(
            ws.cell(row=1, column=col_idx, value=header),
            **_header_style(_MS2_HEADER),
        )
    for row_idx, row in enumerate(rows, start=2):
        issue = row.get("Issue", "")
        for col_idx, header in enumerate(_DIAGNOSTIC_HEADERS, start=1):
            raw = (
                _diagnostic_reason_for_sheet(row)
                if header == "Reason"
                else row.get(header, "")
            )
            value = (
                _excel_text(raw)
                if header in {"SampleName", "Target", "Reason"}
                else raw
            )
            _apply(
                ws.cell(row=row_idx, column=col_idx, value=value),
                fill=_fill(_diagnostic_fill(issue)),
                alignment=CENTER,
                border=BORDER,
            )
    ws.auto_filter.ref = f"A1:D{max(1, len(rows) + 1)}"
    widths = [24, 20, 18, 60]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"


def _diagnostic_reason_for_sheet(row: dict[str, str]) -> str:
    issue = row.get("Issue", "")
    reason = row.get("Reason", "")
    if issue == "NL_FAIL":
        return "selected candidate lacks strict NL match"
    if issue == "NO_MS2":
        return "no aligned MS2 trigger scan"
    if issue == "NL_ANCHOR_FALLBACK":
        return "neutral-loss anchor fallback used"
    if issue == "ANCHOR_RT_MISMATCH":
        delta = _first_regex_group(row.get("Reason", ""), r"deviates ([0-9.]+) min")
        return f"anchor RT mismatch by {delta} min" if delta else "anchor RT mismatch"
    if issue == "MULTI_PEAK":
        count = _first_regex_group(row.get("Reason", ""), r"(\d+) prominent peaks")
        return (
            f"{count} prominent peaks in window"
            if count
            else "multiple peaks in window"
        )
    if issue in {"PEAK_NOT_FOUND", "NO_SIGNAL"}:
        return "peak not found"
    return reason or issue


def _diagnostic_fill(issue: str) -> str:
    if issue in {"NO_MS2", "WINDOW_TOO_SHORT"}:
        return _MS2_NO_MS2
    if issue in {"NL_FAIL", "PEAK_NOT_FOUND", "NO_SIGNAL", "FILE_ERROR"}:
        return _MS2_FAIL
    return _MS2_WARN



def _first_regex_group(text: str, pattern: str) -> str:
    match = re.search(pattern, text)
    return match.group(1) if match else ""
