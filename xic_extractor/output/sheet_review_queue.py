from __future__ import annotations

from openpyxl.utils import get_column_letter

from xic_extractor.output.workbook_styles import (
    _MS2_HEADER,
    _REVIEW_HEADERS,
    BORDER,
    CENTER,
    CENTER_WRAP,
    _apply,
    _header_style,
    _review_cell_fill,
    _review_cell_value,
    _review_number_format,
    _review_priority_fill,
)


def _build_review_queue_sheet(
    ws,
    queue_rows: list[dict[str, str]],
) -> None:
    for col_idx, header in enumerate(_REVIEW_HEADERS, start=1):
        _apply(
            ws.cell(row=1, column=col_idx, value=header),
            **_header_style(_MS2_HEADER),
        )
    ws.row_dimensions[1].height = 30

    for row_idx, row in enumerate(queue_rows, start=2):
        fill_hex = _review_priority_fill(row["Priority"])
        for col_idx, header in enumerate(_REVIEW_HEADERS, start=1):
            raw_value = row.get(header, "")
            value = _review_cell_value(header, raw_value)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            _apply(
                cell,
                fill=_review_cell_fill(header, raw_value, fill_hex),
                alignment=(CENTER_WRAP if header in {"Action", "Evidence"} else CENTER),
                border=BORDER,
            )
            if isinstance(value, float):
                cell.number_format = _review_number_format(header)

    widths = [10, 28, 24, 12, 12, 28, 12, 16, 38, 12, 72]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.auto_filter.ref = (
        f"A1:{get_column_letter(len(_REVIEW_HEADERS))}{max(1, len(queue_rows) + 1)}"
    )
    ws.freeze_panes = "A2"
