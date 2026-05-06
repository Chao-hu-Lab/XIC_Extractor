from __future__ import annotations

from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

from xic_extractor.config import Target
from xic_extractor.output.workbook_styles import (
    _SAMPLE_HEADER,
    _TARGETS_HEADERS,
    BORDER,
    CENTER,
    GREY,
    WHITE,
    _apply,
    _excel_text,
    _fill,
    _header_style,
)


def _build_targets_sheet(ws, targets: list[Target]) -> None:
    for col_idx, header in enumerate(_TARGETS_HEADERS, start=1):
        _apply(
            ws.cell(row=1, column=col_idx, value=header),
            **_header_style(_SAMPLE_HEADER),
        )
    ws["I1"].comment = Comment(
        "Nominal target product m/z = target m/z - neutral_loss_da. "
        "Strict observed-loss NL evidence is evaluated from each MS2 scan precursor "
        "and selected candidate alignment.",
        "XIC Extractor",
    )
    ws.row_dimensions[1].height = 30

    for row_idx, target in enumerate(targets, start=2):
        fill_hex = GREY if row_idx % 2 == 0 else WHITE
        expected_product = (
            target.mz - target.neutral_loss_da
            if target.neutral_loss_da is not None
            else None
        )
        values: list[object] = [
            _excel_text(target.label),
            "ISTD" if target.is_istd else "Analyte",
            _excel_text(target.istd_pair),
            target.mz,
            target.rt_min,
            target.rt_max,
            target.ppm_tol,
            target.neutral_loss_da if target.neutral_loss_da is not None else "—",
            expected_product if expected_product is not None else "—",
            target.nl_ppm_warn if target.nl_ppm_warn is not None else "—",
            target.nl_ppm_max if target.nl_ppm_max is not None else "—",
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            _apply(cell, fill=_fill(fill_hex), alignment=CENTER, border=BORDER)
            header = _TARGETS_HEADERS[col_idx - 1]
            if isinstance(value, float):
                if header in {"m/z", "Expected product m/z", "NL (Da)"}:
                    cell.number_format = "0.0000"
                elif header in {"RT min", "RT max"}:
                    cell.number_format = "0.00"

    widths = [24, 12, 24, 14, 10, 10, 10, 12, 22, 14, 14]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.auto_filter.ref = f"A1:K{max(1, len(targets) + 1)}"
    ws.freeze_panes = "A2"
