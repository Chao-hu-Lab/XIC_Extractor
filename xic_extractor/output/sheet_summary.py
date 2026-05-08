from __future__ import annotations

import statistics

from openpyxl.utils import get_column_letter

from xic_extractor.output.review_metrics import ReviewMetrics, build_review_metrics
from xic_extractor.output.workbook_styles import (
    _MS2_HEADER,
    _SUMMARY_HEADERS,
    BORDER,
    CENTER,
    GREY,
    WHITE,
    _apply,
    _excel_text,
    _fill,
    _header_style,
    _safe_float,
)
from xic_extractor.sample_groups import classify_sample_group


def _build_summary_sheet(
    ws,
    rows: list[dict[str, str]],
    count_no_ms2_as_detected: bool = False,
    review_rows: list[dict[str, str]] | None = None,
) -> None:
    for col_idx, header in enumerate(_SUMMARY_HEADERS, start=1):
        _apply(
            ws.cell(row=1, column=col_idx, value=header),
            **_header_style(_MS2_HEADER),
        )
    ws.row_dimensions[1].height = 30

    metrics = build_review_metrics(
        rows,
        diagnostics=[],
        review_rows=review_rows or [],
        count_no_ms2_as_detected=count_no_ms2_as_detected,
    )
    for row_idx, target in enumerate(_target_summaries(rows), start=2):
        fill_hex = GREY if row_idx % 2 == 0 else WHITE
        values = _summary_row_values(
            target,
            rows,
            count_no_ms2_as_detected,
            metrics,
        )
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            _apply(
                cell,
                fill=_fill(fill_hex),
                alignment=CENTER,
                border=BORDER,
            )
            if _SUMMARY_HEADERS[col_idx - 1] == "Median Area (detected)":
                cell.number_format = "0.00E+00"

    widths = [
        24,
        12,
        24,
        12,
        10,
        12,
        16,
        12,
        24,
        22,
        10,
        10,
        10,
        10,
        14,
        16,
        12,
        14,
        14,
        14,
        12,
        14,
    ]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.auto_filter.ref = (
        f"A1:{get_column_letter(len(_SUMMARY_HEADERS))}"
        f"{max(1, len(_target_summaries(rows)) + 1)}"
    )
    ws.freeze_panes = "A2"

def _target_summaries(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    seen: set[str] = set()
    summaries: list[dict[str, str]] = []
    for row in rows:
        target = row.get("Target", "")
        if not target or target in seen:
            continue
        seen.add(target)
        summaries.append(row)
    return summaries


def _summary_row_values(
    target_row: dict[str, str],
    rows: list[dict[str, str]],
    count_no_ms2_as_detected: bool,
    metrics: ReviewMetrics,
) -> list[object]:
    target = target_row["Target"]
    target_rows = [row for row in rows if row.get("Target") == target]
    detected_rows = [
        row for row in target_rows if _is_long_detected(row, count_no_ms2_as_detected)
    ]
    total = len(target_rows)
    detected = len(detected_rows)
    nl_counts = _long_nl_counts(target_rows)
    confidence_counts = _long_confidence_counts(target_rows)
    target_metrics = metrics.targets[target]
    return [
        _excel_text(target),
        target_row.get("Role", ""),
        _excel_text(target_row.get("ISTD Pair", "")),
        detected,
        total,
        f"{detected / total * 100:.0f}%" if total else "0%",
        _long_median_area(detected_rows),
        _long_mean_rt(detected_rows),
        _long_area_ratio(target_row, rows, count_no_ms2_as_detected),
        _long_rt_delta(target_row, rows, count_no_ms2_as_detected),
        nl_counts["OK"],
        nl_counts["WARN"],
        nl_counts["NL_FAIL"],
        nl_counts["NO_MS2"],
        confidence_counts["HIGH"],
        confidence_counts["MEDIUM"],
        confidence_counts["LOW"],
        confidence_counts["VERY_LOW"],
        target_metrics.flagged_rows,
        target_metrics.flagged_percent,
        target_metrics.ms2_nl_flags,
        target_metrics.low_confidence_rows,
    ]


def _long_confidence_counts(target_rows: list[dict[str, str]]) -> dict[str, int]:
    counts = {"HIGH": 0, "MEDIUM": 0, "LOW": 0, "VERY_LOW": 0}
    for row in target_rows:
        confidence = row.get("Confidence", "HIGH")
        if confidence in counts:
            counts[confidence] += 1
    return counts


def _is_long_detected(
    row: dict[str, str], count_no_ms2_as_detected: bool = False
) -> bool:
    if _safe_float(row.get("RT", "")) is None:
        return False
    if _safe_float(row.get("Area", "")) is None:
        return False
    if row.get("Confidence", "") == "VERY_LOW":
        return False
    nl = row.get("NL", "")
    if nl == "NO_MS2":
        return count_no_ms2_as_detected
    if nl == "NL_FAIL":
        return False
    return nl == "" or nl == "OK" or nl.startswith("WARN_")


def _long_mean_rt(rows: list[dict[str, str]]) -> str:
    values = [
        value for row in rows if (value := _safe_float(row.get("RT", ""))) is not None
    ]
    return f"{sum(values) / len(values):.4f}" if values else "—"


def _long_median_area(rows: list[dict[str, str]]) -> float | str:
    values = [
        value for row in rows if (value := _safe_float(row.get("Area", ""))) is not None
    ]
    return statistics.median(values) if values else "—"


def _long_nl_counts(rows: list[dict[str, str]]) -> dict[str, int]:
    values = [row.get("NL", "") for row in rows]
    ok = sum(1 for value in values if value == "OK")
    warn = sum(1 for value in values if value.startswith("WARN_"))
    no_ms2 = sum(1 for value in values if value == "NO_MS2")
    fail = sum(1 for value in values if value == "NL_FAIL")
    return {"OK": ok, "WARN": warn, "NL_FAIL": fail, "NO_MS2": no_ms2}


def _long_area_ratio(
    target_row: dict[str, str],
    rows: list[dict[str, str]],
    count_no_ms2_as_detected: bool,
) -> str:
    istd_pair = target_row.get("ISTD Pair", "")
    if not istd_pair:
        return "—"
    ratios: list[float] = []
    rows_by_sample = _rows_by_sample_and_target(rows)
    for row in rows:
        if row.get("Target") != target_row["Target"]:
            continue
        if classify_sample_group(row.get("SampleName", "")) != "QC":
            continue
        if not _is_long_detected(row, count_no_ms2_as_detected):
            continue
        istd = rows_by_sample.get((row.get("SampleName", ""), istd_pair))
        if istd is None or not _is_long_detected(istd, count_no_ms2_as_detected):
            continue
        analyte_area = _safe_float(row.get("Area", ""))
        istd_area = _safe_float(istd.get("Area", ""))
        if analyte_area is None or istd_area is None or istd_area == 0:
            continue
        ratios.append(analyte_area / istd_area)
    if not ratios:
        return "—"
    mean_ratio = sum(ratios) / len(ratios)
    sd_ratio = statistics.stdev(ratios) if len(ratios) > 1 else 0.0
    cv_text = (
        f"{sd_ratio / mean_ratio * 100:.1f}%"
        if len(ratios) > 1 and mean_ratio != 0
        else "—"
    )
    return f"{mean_ratio:.4f}±{sd_ratio:.4f}, CV={cv_text} (n={len(ratios)})"


def _long_rt_delta(
    target_row: dict[str, str],
    rows: list[dict[str, str]],
    count_no_ms2_as_detected: bool,
) -> str:
    istd_pair = target_row.get("ISTD Pair", "")
    if not istd_pair:
        return "—"
    deltas: list[float] = []
    rows_by_sample = _rows_by_sample_and_target(rows)
    for row in rows:
        if row.get("Target") != target_row["Target"]:
            continue
        if not _is_long_detected(row, count_no_ms2_as_detected):
            continue
        istd = rows_by_sample.get((row.get("SampleName", ""), istd_pair))
        if istd is None or not _is_long_detected(istd, count_no_ms2_as_detected):
            continue
        rt_analyte = _safe_float(row.get("RT", ""))
        rt_istd = _safe_float(istd.get("RT", ""))
        if rt_analyte is None or rt_istd is None:
            continue
        deltas.append(abs(rt_analyte - rt_istd))
    if not deltas:
        return "—"
    mean_delta = sum(deltas) / len(deltas)
    sd_delta = statistics.stdev(deltas) if len(deltas) > 1 else 0.0
    return f"{mean_delta:.4f}±{sd_delta:.4f} min (n={len(deltas)})"


def _rows_by_sample_and_target(
    rows: list[dict[str, str]],
) -> dict[tuple[str, str], dict[str, str]]:
    return {(row.get("SampleName", ""), row.get("Target", "")): row for row in rows}
