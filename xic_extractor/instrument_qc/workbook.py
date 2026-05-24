from __future__ import annotations

from dataclasses import asdict
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from xic_extractor.instrument_qc.models import (
    HCDAuditRow,
    InstrumentQCDiagnostic,
    SDOLEKTrendRow,
)
from xic_extractor.instrument_qc.workbook_manual_review import (
    format_counts,
    manual_review_rows,
)
from xic_extractor.instrument_qc.writers import (
    DIAGNOSTIC_TSV_COLUMNS,
    HCD_AUDIT_TSV_COLUMNS,
    TREND_TSV_COLUMNS,
)


def write_sdolek_workbook(
    path: Path,
    rows: Iterable[SDOLEKTrendRow],
    diagnostics: Iterable[InstrumentQCDiagnostic],
    *,
    metadata_source_status: dict[str, str] | None = None,
    mixstds_rows: Iterable[SDOLEKTrendRow] | None = None,
    hcd_rows: Iterable[HCDAuditRow] | None = None,
) -> Path:
    row_list = list(rows)
    mixstds_row_list = list(mixstds_rows) if mixstds_rows is not None else None
    hcd_row_list = list(hcd_rows) if hcd_rows is not None else None
    diagnostic_list = list(diagnostics)
    path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    overview = workbook.active
    overview.title = "Overview"
    _write_overview_sheet(
        overview,
        row_list,
        diagnostic_list,
        metadata_source_status or {},
        mixstds_row_list,
        hcd_row_list,
    )
    _write_trend_sheet(workbook.create_sheet("SDOLEK Trend"), row_list)
    if mixstds_row_list is not None:
        _write_trend_sheet(workbook.create_sheet("Mix STDs Trend"), mixstds_row_list)
    if hcd_row_list is not None:
        _write_hcd_sheet(workbook.create_sheet("HCD Audit"), hcd_row_list)
        _write_manual_review_sheet(
            workbook.create_sheet("Manual Review Queue"),
            row_list,
            mixstds_row_list or [],
            hcd_row_list,
        )
    _write_diagnostics_sheet(workbook.create_sheet("Diagnostics"), diagnostic_list)
    workbook.save(path)
    return path


def _write_overview_sheet(
    sheet: Any,
    rows: list[SDOLEKTrendRow],
    diagnostics: list[InstrumentQCDiagnostic],
    metadata_source_status: dict[str, str],
    mixstds_rows: list[SDOLEKTrendRow] | None,
    hcd_rows: list[HCDAuditRow] | None,
) -> None:
    sheet.append(["metric", "value"])
    _bold_header(sheet)
    status_counts = _counts(row.status for row in rows)
    compound_counts = _counts(row.compound for row in rows)
    diagnostic_counts = _counts(diag.issue for diag in diagnostics)
    hcd_status_counts = _counts(row.hcd_status for row in hcd_rows or [])
    overview_rows = [
        ("report_type", "Instrument QC SDOLEK Trend"),
        (
            "identity_evidence",
            (
                "MS1 trend plus audit-only CID/wHCD product-ion review; "
                "not a production identity gate"
            ),
        ),
        ("total_rows", len(rows)),
        ("detected_rows", status_counts.get("detected", 0)),
        ("sdo_rows", compound_counts.get("SDO", 0)),
        ("lek_rows", compound_counts.get("LEK", 0)),
        ("mixstds_rows", len(mixstds_rows or [])),
        ("hcd_rows", len(hcd_rows or [])),
        ("hcd_status_counts", format_counts(hcd_status_counts)),
        ("blank_stance", "blank_defer"),
        (
            "injection_order_status",
            metadata_source_status.get("injection_order_status", ""),
        ),
        (
            "injection_order_source",
            metadata_source_status.get("injection_order_source", ""),
        ),
        ("diagnostic_counts", format_counts(diagnostic_counts)),
        ("top_rt_delta_to_reference", _top_rt_delta(rows)),
        ("top_width_ratio_to_reference", _top_width_ratio(rows)),
    ]
    for key, value in overview_rows:
        sheet.append([key, _xlsx_value(value)])
    _autosize_columns(sheet)


def _write_trend_sheet(sheet: Any, rows: list[SDOLEKTrendRow]) -> None:
    _append_xlsx_row(sheet, TREND_TSV_COLUMNS)
    for row in rows:
        values = _trend_row_to_dict(row)
        _append_xlsx_row(sheet, [values[column] for column in TREND_TSV_COLUMNS])
    _finish_table_sheet(sheet)


def _write_hcd_sheet(sheet: Any, rows: list[HCDAuditRow]) -> None:
    _append_xlsx_row(sheet, HCD_AUDIT_TSV_COLUMNS)
    for row in rows:
        values = _hcd_row_to_dict(row)
        _append_xlsx_row(sheet, [values[column] for column in HCD_AUDIT_TSV_COLUMNS])
    _finish_table_sheet(sheet)


def _write_manual_review_sheet(
    sheet: Any,
    sdolek_rows: list[SDOLEKTrendRow],
    mixstds_rows: list[SDOLEKTrendRow],
    hcd_rows: list[HCDAuditRow],
) -> None:
    columns = [
        "priority",
        "queue_reason",
        "review_item",
        "compound",
        "precursor_mz",
        "row_count",
        "samples",
        "ms1_summary",
        "hcd_summary",
        "rt_drift_hint",
        "product_hint",
        "suggested_action",
        "manual_label",
        "manual_note",
    ]
    _append_xlsx_row(sheet, columns)
    for row in manual_review_rows(sdolek_rows, mixstds_rows, hcd_rows):
        _append_xlsx_row(sheet, [row.get(column, "") for column in columns])
    _finish_manual_review_sheet(sheet)


def _write_diagnostics_sheet(
    sheet: Any,
    diagnostics: list[InstrumentQCDiagnostic],
) -> None:
    _append_xlsx_row(sheet, DIAGNOSTIC_TSV_COLUMNS)
    for diagnostic in diagnostics:
        values = _diagnostic_to_dict(diagnostic)
        _append_xlsx_row(
            sheet,
            [values[column] for column in DIAGNOSTIC_TSV_COLUMNS],
        )
    _finish_table_sheet(sheet)


def _finish_table_sheet(sheet: Any) -> None:
    _bold_header(sheet)
    sheet.freeze_panes = "A2"
    if sheet.max_row > 1 and sheet.max_column > 1:
        sheet.auto_filter.ref = sheet.dimensions
    _autosize_columns(sheet)


def _finish_manual_review_sheet(sheet: Any) -> None:
    _finish_table_sheet(sheet)
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        priority = row[0].value
        fill = _priority_fill(str(priority or ""))
        row[0].fill = fill
        row[1].fill = fill
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _bold_header(sheet: Any) -> None:
    for cell in sheet[1]:
        cell.font = Font(name="Arial", bold=True)


def _append_xlsx_row(sheet: Any, values: Iterable[object]) -> None:
    sheet.append([_xlsx_value(value) for value in values])


def _xlsx_value(value: object) -> object:
    if value is None:
        return ""
    if isinstance(value, str):
        return _escape_excel_formula(value)
    if isinstance(value, Path):
        return _escape_excel_formula(str(value))
    return value


def _trend_row_to_dict(row: SDOLEKTrendRow) -> dict[str, object]:
    values = asdict(row)
    values["raw_path"] = str(row.raw_path)
    values["trend_flags"] = ";".join(row.trend_flags)
    return values


def _hcd_row_to_dict(row: HCDAuditRow) -> dict[str, object]:
    values = asdict(row)
    values["raw_path"] = str(row.raw_path)
    values["matched_products"] = ";".join(row.matched_products)
    values["review_flags"] = ";".join(row.review_flags)
    return values


def _diagnostic_to_dict(diagnostic: InstrumentQCDiagnostic) -> dict[str, object]:
    return {
        "sample_name": diagnostic.sample_name,
        "raw_path": str(diagnostic.raw_path),
        "issue": diagnostic.issue,
        "detail": diagnostic.detail,
    }


def _escape_excel_formula(value: str) -> str:
    if value.startswith(("=", "+", "-", "@")):
        return f"'{value}"
    return value


def _priority_fill(priority: str) -> PatternFill:
    colors = {
        "P1": "F4CCCC",
        "P2": "FFF2CC",
        "P3": "D9EAD3",
    }
    return PatternFill("solid", fgColor=colors.get(priority, "FFFFFF"))


def _counts(values: Iterable[str]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for value in values:
        counts[value] = counts.get(value, 0) + 1
    return counts


def _top_rt_delta(rows: list[SDOLEKTrendRow]) -> str:
    candidates = [
        row for row in rows if row.rt_delta_to_reference_min is not None
    ]
    if not candidates:
        return ""
    row = max(candidates, key=lambda item: abs(item.rt_delta_to_reference_min or 0.0))
    return (
        f"{row.sample_name} {row.compound}: "
        f"{row.rt_delta_to_reference_min:.3f} min"
    )


def _top_width_ratio(rows: list[SDOLEKTrendRow]) -> str:
    candidates = [
        row for row in rows if row.base_width_ratio_to_reference is not None
    ]
    if not candidates:
        return ""
    row = max(
        candidates,
        key=lambda item: abs((item.base_width_ratio_to_reference or 1.0) - 1.0),
    )
    return (
        f"{row.sample_name} {row.compound}: "
        f"{row.base_width_ratio_to_reference:.3f}"
    )


def _autosize_columns(sheet: Any) -> None:
    for column_cells in sheet.columns:
        header = column_cells[0]
        letter = header.column_letter
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[letter].width = min(max(max_length + 2, 10), 60)
