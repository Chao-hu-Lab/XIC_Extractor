from pathlib import Path

from xic_extractor.injection_rolling import read_injection_order, rolling_median_rt


def test_read_csv(tmp_path: Path) -> None:
    p = tmp_path / "info.csv"
    p.write_text("Sample_Name,Injection_Order\nS_A,1\nS_B,2\n", encoding="utf-8")
    assert read_injection_order(p) == {"S_A": 1, "S_B": 2}


def test_read_csv_handles_utf8_bom(tmp_path: Path) -> None:
    p = tmp_path / "info.csv"
    p.write_text(
        "Sample_Name,Injection_Order\nS_A,1\n",
        encoding="utf-8-sig",
    )
    assert read_injection_order(p) == {"S_A": 1}


def test_read_strips_whitespace(tmp_path: Path) -> None:
    p = tmp_path / "info.csv"
    p.write_text("Sample_Name,Injection_Order\n  S_A  ,5\n", encoding="utf-8")
    assert read_injection_order(p) == {"S_A": 5}


def test_rolling_median_uses_window() -> None:
    rts = {f"s{i}": float(i) for i in range(1, 11)}
    order = {f"s{i}": i for i in range(1, 11)}
    assert rolling_median_rt("istd", "s5", rts, order, window=2) == 5.0


def test_rolling_median_respects_gaps() -> None:
    rts = {"s1": 1.0, "s2": 2.0, "s10": 10.0}
    order = {"s1": 1, "s2": 2, "s10": 10}
    assert rolling_median_rt("istd", "s10", rts, order, window=1) is None


def test_rolling_median_missing_target_returns_none() -> None:
    assert rolling_median_rt("istd", "unknown", {}, {}, window=5) is None


def test_read_xlsx(tmp_path: Path) -> None:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(["Sample_Name", "Injection_Order"])
    ws.append(["S_X", 3])
    ws.append(["S_Y", 7])
    p = tmp_path / "info.xlsx"
    wb.save(p)
    assert read_injection_order(p) == {"S_X": 3, "S_Y": 7}


def test_read_xlsx_releases_file_handle(tmp_path: Path) -> None:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(["Sample_Name", "Injection_Order"])
    ws.append(["S_X", 3])
    p = tmp_path / "info.xlsx"
    wb.save(p)

    assert read_injection_order(p) == {"S_X": 3}
    p.unlink()
    assert not p.exists()


def test_read_xlsx_adds_canonical_aliases_for_tissue_sampleinfo(
    tmp_path: Path,
) -> None:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(["Sample_Name", "Injection_Order"])
    ws.append(["Tumor tissue BC2257_DNA ", 2])
    ws.append(["Normal tissue BC2257_DNA ", 36])
    ws.append(["Benign fat BC1055_DNA ", 76])
    ws.append(["Tumor tissue BC2286* DNA +RNA", 20])
    ws.append(["Breast Cancer Tissue_ pooled_QC_1 ", 1])
    ws.append(["Breast Cancer Tissue_pooled_QC_4", 49])
    path = tmp_path / "SampleInfo.xlsx"
    wb.save(path)

    order = read_injection_order(path)

    assert order["Tumor tissue BC2257_DNA"] == 2
    assert order["TumorBC2257_DNA"] == 2
    assert order["NormalBC2257_DNA"] == 36
    assert order["BenignfatBC1055_DNA"] == 76
    assert order["TumorBC2286_DNAandRNA"] == 20
    assert order["Breast_Cancer_Tissue_pooled_QC1"] == 1
    assert order["Breast_Cancer_Tissue_pooled_QC_4"] == 49


def test_canonical_alias_collision_raises(tmp_path: Path) -> None:
    p = tmp_path / "info.csv"
    p.write_text(
        "Sample_Name,Injection_Order\nTumor tissue BC2257_DNA,2\nTumorBC2257_DNA,3\n",
        encoding="utf-8",
    )

    try:
        read_injection_order(p)
    except ValueError as exc:
        assert "Conflicting injection order" in str(exc)
    else:
        raise AssertionError("expected conflicting alias to raise")
