from pathlib import Path

from openpyxl import load_workbook

from scripts import csv_to_excel
from xic_extractor import extractor as extractor_module
from xic_extractor.config import ExtractionConfig, Target
from xic_extractor.extractor import (
    ExtractionResult,
    FileResult,
    RunOutput,
)
from xic_extractor.output.csv_writers import (
    _long_output_rows,
    _output_row,
    write_all,
)
from xic_extractor.signal_processing import (
    PeakCandidate,
    PeakDetectionResult,
    PeakResult,
)


def _target(label: str, *, is_istd: bool = False) -> Target:
    return Target(
        label=label,
        mz=258.1085,
        rt_min=8.0,
        rt_max=10.0,
        ppm_tol=20.0,
        neutral_loss_da=116.0474,
        nl_ppm_warn=20.0,
        nl_ppm_max=50.0,
        is_istd=is_istd,
        istd_pair="",
    )


def _fabricate_run_output() -> RunOutput:
    peak = PeakResult(
        rt=9.03,
        intensity=500,
        intensity_smoothed=500,
        area=123.0,
        peak_start=8.9,
        peak_end=9.2,
    )
    candidate = PeakCandidate(
        peak=peak,
        smoothed_apex_rt=9.0,
        smoothed_apex_intensity=500.0,
        smoothed_apex_index=10,
        raw_apex_rt=9.03,
        raw_apex_intensity=500.0,
        raw_apex_index=10,
        prominence=100.0,
    )
    peak_result = PeakDetectionResult(
        status="OK",
        peak=peak,
        n_points=10,
        max_smoothed=500.0,
        n_prominent_peaks=1,
        candidates=(candidate,),
        confidence="HIGH",
        reason="all checks passed",
    )
    extraction_result = ExtractionResult(
        peak_result=peak_result,
        nl=None,
        target_label="d3-5-hmdC",
        role="ISTD",
        istd_pair="",
        confidence="VERY_LOW",
        reason="concerns: rt_prior (major); weak candidate: too_broad",
        severities=(
            (0, "symmetry"),
            (1, "local_sn"),
            (0, "nl_support"),
            (2, "rt_prior"),
            (0, "rt_centrality"),
            (1, "noise_shape"),
            (0, "peak_width"),
        ),
        prior_rt=9.01,
        prior_source="rolling_median",
        quality_penalty=1,
        quality_flags=("too_broad",),
    )
    file_result = FileResult(
        sample_name="S1",
        results={"d3-5-hmdC": extraction_result},
        group=None,
    )
    return RunOutput(file_results=[file_result], diagnostics=[])


def _config(tmp_path: Path, *, emit_score_breakdown: bool = False) -> ExtractionConfig:
    return ExtractionConfig(
        data_dir=tmp_path / "raw",
        dll_dir=tmp_path / "dll",
        output_csv=tmp_path / "output" / "xic_results.csv",
        diagnostics_csv=tmp_path / "output" / "xic_diagnostics.csv",
        smooth_window=7,
        smooth_polyorder=3,
        peak_rel_height=0.95,
        peak_min_prominence_ratio=0.1,
        ms2_precursor_tol_da=1.6,
        nl_min_intensity_ratio=0.01,
        emit_score_breakdown=emit_score_breakdown,
    )


def _write_workbook(tmp_path: Path, *, emit_score_breakdown: bool = False) -> Path:
    config = _config(tmp_path, emit_score_breakdown=emit_score_breakdown)
    target = _target("d3-5-hmdC", is_istd=True)
    run_output = _fabricate_run_output()
    write_all(
        config,
        [target],
        run_output.file_results,
        run_output.diagnostics,
        emit_score_breakdown=emit_score_breakdown,
    )
    return csv_to_excel.run(config, [target])


def test_extractor_no_longer_exposes_test_only_xlsx_writer() -> None:
    assert not hasattr(extractor_module, "_write_xlsx")


def test_main_sheet_has_confidence_and_reason(tmp_path: Path) -> None:
    out = _write_workbook(tmp_path)
    wb = load_workbook(out, read_only=True)
    ws = wb["XIC Results"]
    headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
    assert "Confidence" in headers
    assert "Reason" in headers


def test_summary_sheet_has_confidence_counts(tmp_path: Path) -> None:
    out = _write_workbook(tmp_path)
    wb = load_workbook(out, read_only=True)
    ws = wb["Summary"]
    headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
    for label in (
        "Confidence HIGH",
        "Confidence MEDIUM",
        "Confidence LOW",
        "Confidence VERY_LOW",
    ):
        assert label in headers


def test_score_breakdown_sheet_absent_by_default(tmp_path: Path) -> None:
    out = _write_workbook(tmp_path)
    wb = load_workbook(out, read_only=True)
    assert "Score Breakdown" not in wb.sheetnames


def test_score_breakdown_sheet_emitted_when_flag_on(tmp_path: Path) -> None:
    out = _write_workbook(tmp_path, emit_score_breakdown=True)
    wb = load_workbook(out, read_only=True)
    assert "Score Breakdown" in wb.sheetnames
    ws = wb["Score Breakdown"]
    rows = list(ws.iter_rows(values_only=True))
    headers = list(rows[0])
    for col in (
        "SampleName",
        "Target",
        "symmetry",
        "local_sn",
        "nl_support",
        "rt_prior",
        "rt_centrality",
        "noise_shape",
        "peak_width",
        "Quality Penalty",
        "Quality Flags",
        "Total Severity",
        "Confidence",
        "Prior RT",
        "Prior Source",
    ):
        assert col in headers
    assert len(rows) == 2
    row = dict(zip(headers, rows[1], strict=False))
    assert row["SampleName"] == "S1"
    assert row["Target"] == "d3-5-hmdC"
    assert row["local_sn"] == 1
    assert row["rt_prior"] == 2
    assert row["noise_shape"] == 1
    assert row["Quality Penalty"] == 1
    assert row["Quality Flags"] == "too_broad"
    assert row["Total Severity"] == 5
    assert row["Confidence"] == "VERY_LOW"
    assert row["Prior RT"] == 9.01
    assert row["Prior Source"] == "rolling_median"


def test_output_rows_use_smoothed_apex_rt_for_reporting() -> None:
    run_output = _fabricate_run_output()
    file_result = run_output.file_results[0]
    target = _target("d3-5-hmdC")

    wide_row = _output_row(file_result, [target])
    long_row = _long_output_rows(file_result, [target])[0]

    assert wide_row["d3-5-hmdC_RT"] == "9.0000"
    assert long_row["RT"] == "9.0000"
    assert wide_row["d3-5-hmdC_Area"] == "123.00"
    assert long_row["Area"] == "123.00"


def test_xlsx_results_sheet_uses_smoothed_apex_rt_for_reporting(
    tmp_path: Path,
) -> None:
    out = _write_workbook(tmp_path)
    wb = load_workbook(out, read_only=True)
    ws = wb["XIC Results"]
    rows = list(ws.iter_rows(values_only=True))
    headers = list(rows[0])
    row = dict(zip(headers, rows[1], strict=False))

    assert row["Target"] == "d3-5-hmdC"
    assert row["RT"] == 9
    assert row["Area"] == 123
