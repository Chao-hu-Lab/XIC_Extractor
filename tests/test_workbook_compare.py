from pathlib import Path

from openpyxl import Workbook

from scripts.compare_workbooks import compare_workbooks


def test_compare_workbooks_accepts_identical_workbooks(tmp_path: Path) -> None:
    left = tmp_path / "left.xlsx"
    right = tmp_path / "right.xlsx"
    _write_workbook(left)
    _write_workbook(right)

    result = compare_workbooks(left, right)

    assert result.matched
    assert result.differences == []


def test_compare_workbooks_rejects_changed_analytical_value(tmp_path: Path) -> None:
    left = tmp_path / "left.xlsx"
    right = tmp_path / "right.xlsx"
    _write_workbook(left, area=100.0)
    _write_workbook(right, area=110.0)

    result = compare_workbooks(left, right)

    assert not result.matched
    assert any(
        "XIC Results" in diff and "100" in diff and "110" in diff
        for diff in result.differences
    )


def test_compare_workbooks_ignores_generated_at_metadata(tmp_path: Path) -> None:
    left = tmp_path / "left.xlsx"
    right = tmp_path / "right.xlsx"
    _write_workbook(left, generated_at="2026-05-03T00:00:00Z")
    _write_workbook(right, generated_at="2026-05-04T00:00:00Z")

    result = compare_workbooks(left, right)

    assert result.matched
    assert result.differences == []


def test_compare_workbooks_ignores_runtime_and_output_path_metadata(
    tmp_path: Path,
) -> None:
    left = tmp_path / "left.xlsx"
    right = tmp_path / "right.xlsx"
    _write_workbook(
        left,
        elapsed_seconds="1.0",
        output_path="C:\\old\\xic_results.xlsx",
    )
    _write_workbook(
        right,
        elapsed_seconds="2.0",
        output_path="D:\\new\\xic_results.xlsx",
    )

    result = compare_workbooks(left, right)

    assert result.matched
    assert result.differences == []


def test_compare_workbooks_ignores_sheet_order(tmp_path: Path) -> None:
    left = tmp_path / "left.xlsx"
    right = tmp_path / "right.xlsx"
    _write_workbook(
        left,
        sheet_order=(
            "XIC Results",
            "Summary",
            "Review Queue",
            "Targets",
            "Diagnostics",
            "Run Metadata",
        ),
    )
    _write_workbook(
        right,
        sheet_order=(
            "Run Metadata",
            "Diagnostics",
            "Targets",
            "Review Queue",
            "Summary",
            "XIC Results",
        ),
    )

    result = compare_workbooks(left, right)

    assert result.matched
    assert result.differences == []


def test_compare_workbooks_compares_score_breakdown_when_present(
    tmp_path: Path,
) -> None:
    left = tmp_path / "left.xlsx"
    right = tmp_path / "right.xlsx"
    _write_workbook(left, include_score_breakdown=True, score=0.9)
    _write_workbook(right, include_score_breakdown=True, score=0.8)

    result = compare_workbooks(left, right)

    assert not result.matched
    assert any("Score Breakdown" in diff for diff in result.differences)


def _write_workbook(
    path: Path,
    *,
    area: float = 100.0,
    generated_at: str = "2026-05-03T00:00:00Z",
    elapsed_seconds: str = "1.0",
    output_path: str = "C:\\output\\xic_results.xlsx",
    include_score_breakdown: bool = False,
    score: float = 0.9,
    sheet_order: tuple[str, ...] = (
        "XIC Results",
        "Summary",
        "Review Queue",
        "Targets",
        "Diagnostics",
        "Run Metadata",
    ),
) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    for name in sheet_order:
        ws = wb.create_sheet(name)
        if name == "XIC Results":
            ws.append(["Sample", "Target", "Area"])
            ws.append(["SampleA", "Analyte", area])
        elif name == "Summary":
            ws.append(["Target", "Detected"])
            ws.append(["Analyte", 1])
        elif name == "Review Queue":
            ws.append(["Priority", "Sample", "Target", "Issue"])
        elif name == "Targets":
            ws.append(["label", "mz"])
            ws.append(["Analyte", 258.1085])
        elif name == "Diagnostics":
            ws.append(["Sample", "Issue"])
        elif name == "Run Metadata":
            ws.append(["key", "value"])
            ws.append(["generated_at", generated_at])
            ws.append(["elapsed_seconds", elapsed_seconds])
            ws.append(["output_path", output_path])
            ws.append(["resolver_mode", "legacy_savgol"])
    if include_score_breakdown:
        ws = wb.create_sheet("Score Breakdown")
        ws.append(["Sample", "Target", "Score"])
        ws.append(["SampleA", "Analyte", score])
    wb.save(path)
