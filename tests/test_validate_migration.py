import csv
from pathlib import Path

from openpyxl import load_workbook

from scripts.validate_migration import (
    ValidationCase,
    ValidationRow,
    _stage_cases,
    _update_settings_value,
    compare_validation_rows,
    parse_case_arg,
    strict_exit_code,
    write_validation_workbook,
)


def test_compare_validation_rows_passes_with_like_for_like_smoothed_match() -> None:
    rows = [
        ValidationRow(
            sample_name="Tumor_1",
            target="Analyte",
            rt_old=9.0000,
            int_old=1000.0,
            nl_old="OK",
            rt_new=9.0010,
            int_new_raw=1100.0,
            int_new_smoothed=1002.0,
            area_new=500.0,
            peak_start_new=8.9,
            peak_end_new=9.1,
            nl_new="OK",
        ),
        ValidationRow(
            sample_name="Tumor_2",
            target="Analyte",
            rt_old=9.1000,
            int_old=2000.0,
            nl_old="ND",
            rt_new=9.1020,
            int_new_raw=2100.0,
            int_new_smoothed=1998.0,
            area_new=900.0,
            peak_start_new=9.0,
            peak_end_new=9.2,
            nl_new="NO_MS2",
        ),
    ]

    report = compare_validation_rows(rows)

    assert report.failed is False
    assert report.targets[0].target == "Analyte"
    assert report.targets[0].status == "PASS"
    assert report.targets[0].median_rt_delta == 0.0015
    assert report.targets[0].smoothed_median_ratio == 1.0005
    assert report.targets[0].nl_agreement_pct == 100.0
    assert report.failures == []


def test_compare_validation_rows_groups_targets_and_preserves_report_order() -> None:
    rows = [
        _validation_row("Tumor_2", "Target-B", rt_old=10.0, rt_new=10.001),
        _validation_row("Tumor_1", "Target-A", rt_old=8.0, rt_new=8.001),
        _validation_row("Tumor_3", "Target-B", rt_old=10.1, rt_new=10.102),
    ]

    report = compare_validation_rows(rows)

    assert [target.target for target in report.targets] == ["Target-A", "Target-B"]
    assert [target.n_rows for target in report.targets] == [1, 2]
    assert report.targets[1].median_rt_delta == 0.0015


def test_compare_fails_before_thresholds_when_smoothed_missing() -> None:
    rows = [
        ValidationRow(
            sample_name="Tumor_1",
            target="Analyte",
            rt_old=9.0,
            int_old=1000.0,
            nl_old="OK",
            rt_new=9.0,
            int_new_raw=1100.0,
            int_new_smoothed=None,
            area_new=500.0,
            peak_start_new=8.9,
            peak_end_new=9.1,
            nl_new="OK",
        )
    ]

    report = compare_validation_rows(rows)

    assert report.failed is True
    assert report.targets[0].status == "FAIL"
    assert report.failures[0].issue == "MISSING_SMOOTHED_INTENSITY"


def test_compare_fails_when_old_peak_is_missing_from_new_pipeline() -> None:
    rows = [
        ValidationRow(
            sample_name="Tumor_1",
            target="Analyte",
            rt_old=9.0,
            int_old=1000.0,
            nl_old="OK",
            rt_new=None,
            int_new_raw=None,
            int_new_smoothed=None,
            area_new=None,
            peak_start_new=None,
            peak_end_new=None,
            nl_new="OK",
        )
    ]

    report = compare_validation_rows(rows)

    assert report.failed is True
    assert report.targets[0].status == "FAIL"
    assert report.failures[0].issue == "NEW_PEAK_MISSING"


def test_compare_validation_rows_flags_rt_smoothed_and_nl_failures() -> None:
    rows = [
        ValidationRow(
            sample_name="Tumor_1",
            target="Analyte",
            rt_old=9.0000,
            int_old=1000.0,
            nl_old="OK",
            rt_new=9.0200,
            int_new_raw=1500.0,
            int_new_smoothed=1300.0,
            area_new=500.0,
            peak_start_new=8.9,
            peak_end_new=9.1,
            nl_new="NL_FAIL",
        )
    ]

    report = compare_validation_rows(rows)

    issues = {failure.issue for failure in report.failures}
    assert report.failed is True
    assert {"RT_DRIFT", "SMOOTHED_INTENSITY_DRIFT", "NL_STATUS_MISMATCH"} <= issues
    assert strict_exit_code(report, strict=True, allow_overrides=False) == 1


def test_strict_exit_code_honors_explicit_override_mode() -> None:
    report = compare_validation_rows(
        [
            ValidationRow(
                sample_name="Tumor_1",
                target="Analyte",
                rt_old=9.0,
                int_old=1000.0,
                nl_old="OK",
                rt_new=9.0200,
                int_new_raw=1100.0,
                int_new_smoothed=1000.0,
                area_new=500.0,
                peak_start_new=8.9,
                peak_end_new=9.1,
                nl_new="OK",
            )
        ]
    )

    assert strict_exit_code(report, strict=False, allow_overrides=False) == 0
    assert strict_exit_code(report, strict=True, allow_overrides=False) == 1
    assert strict_exit_code(report, strict=True, allow_overrides=True) == 0


def test_merge_old_new_rows_preserves_old_only_rows_as_failures() -> None:
    old_rows = [
        ValidationRow(
            sample_name="Tumor_1",
            target="Analyte",
            rt_old=9.0,
            int_old=1000.0,
            nl_old="OK",
            rt_new=None,
            int_new_raw=None,
            int_new_smoothed=None,
            area_new=None,
            peak_start_new=None,
            peak_end_new=None,
            nl_new="",
            new_row_present=False,
        )
    ]

    from scripts.validate_migration import merge_old_new_rows

    merged = merge_old_new_rows(old_rows, [])
    report = compare_validation_rows(merged)

    assert len(merged) == 1
    assert report.failed is True
    assert report.failures[0].issue == "NEW_ROW_MISSING"


def test_write_validation_workbook_contains_summary_pertarget_and_fail_sheets(
    tmp_path: Path,
) -> None:
    rows = [
        ValidationRow(
            sample_name="Tumor_1",
            target="Analyte",
            rt_old=9.0,
            int_old=1000.0,
            nl_old="OK",
            rt_new=9.0200,
            int_new_raw=1100.0,
            int_new_smoothed=1000.0,
            area_new=500.0,
            peak_start_new=8.9,
            peak_end_new=9.1,
            nl_new="OK",
        )
    ]
    report = compare_validation_rows(rows)
    output = tmp_path / "validation.xlsx"

    write_validation_workbook(output, rows, report)

    wb = load_workbook(output)
    assert wb.sheetnames == ["Summary", "PerTarget", "FAIL"]
    assert wb["Summary"]["A1"].value == "Target"
    assert wb["PerTarget"]["A1"].value == "SampleName"
    fail_headers = [cell.value for cell in wb["FAIL"][1]]
    assert fail_headers[-4:] == [
        "OverrideDecision",
        "OverrideReason",
        "Reviewer",
        "ScreenshotPath",
    ]


def test_parse_case_arg_accepts_name_path_pairs(tmp_path: Path) -> None:
    case = parse_case_arg(f"Tumor={tmp_path}")

    assert case.name == "Tumor"
    assert case.path == tmp_path


def test_stage_cases_copies_raws_and_points_settings_at_validation_dir(
    tmp_path: Path,
) -> None:
    worktree = tmp_path / "worktree"
    config_dir = worktree / "config"
    config_dir.mkdir(parents=True)
    (config_dir / "settings.csv").write_text(
        "key,value,description\ndata_dir,C:\\old,raw folder\ndll_dir,C:\\dll,dll\n",
        encoding="utf-8-sig",
    )
    raw_file = tmp_path / "source.raw"
    raw_file.write_text("raw", encoding="utf-8")

    data_dir = _stage_cases(worktree, [ValidationCase("Tumor", raw_file)])

    assert (data_dir / "Tumor.raw").read_text(encoding="utf-8") == "raw"
    settings_text = (config_dir / "settings.csv").read_text(encoding="utf-8-sig")
    assert f"data_dir,{data_dir}" in settings_text


def test_stage_cases_removes_stale_validation_raw_files(tmp_path: Path) -> None:
    worktree = tmp_path / "worktree"
    config_dir = worktree / "config"
    config_dir.mkdir(parents=True)
    (config_dir / "settings.csv").write_text(
        "key,value,description\ndata_dir,C:\\old,raw folder\ndll_dir,C:\\dll,dll\n",
        encoding="utf-8-sig",
    )
    data_dir = worktree / "local_validation_raw"
    data_dir.mkdir(parents=True)
    (data_dir / "Stale.raw").write_text("stale", encoding="utf-8")
    raw_file = tmp_path / "source.raw"
    raw_file.write_text("raw", encoding="utf-8")

    _stage_cases(worktree, [ValidationCase("Tumor", raw_file)])

    assert not (data_dir / "Stale.raw").exists()
    assert (data_dir / "Tumor.raw").read_text(encoding="utf-8") == "raw"


def test_update_settings_value_preserves_settings_csv_contract(tmp_path: Path) -> None:
    settings_path = tmp_path / "settings.csv"
    settings_path.write_bytes(
        b"\xef\xbb\xbf"
        b"key,value,description,scope\r\n"
        b"data_dir,C:\\old,raw folder,local\r\n"
        b"dll_dir,C:\\dll,dll,global\r\n"
    )
    updated_data_dir = Path("C:/new data")
    appended_value = Path("D:/generated")

    _update_settings_value(settings_path, "data_dir", updated_data_dir)
    _update_settings_value(settings_path, "new_key", appended_value)

    raw = settings_path.read_bytes()
    assert raw.startswith(b"\xef\xbb\xbf")
    assert b"\r\n" in raw
    with settings_path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        assert reader.fieldnames == ["key", "value", "description", "scope"]
        rows = list(reader)
    assert rows[0] == {
        "key": "data_dir",
        "value": str(updated_data_dir),
        "description": "raw folder",
        "scope": "local",
    }
    assert rows[2] == {
        "key": "new_key",
        "value": str(appended_value),
        "description": "",
        "scope": "",
    }


def _validation_row(
    sample_name: str,
    target: str,
    *,
    rt_old: float,
    rt_new: float,
) -> ValidationRow:
    return ValidationRow(
        sample_name=sample_name,
        target=target,
        rt_old=rt_old,
        int_old=1000.0,
        nl_old="OK",
        rt_new=rt_new,
        int_new_raw=1000.0,
        int_new_smoothed=1000.0,
        area_new=500.0,
        peak_start_new=rt_new - 0.1,
        peak_end_new=rt_new + 0.1,
        nl_new="OK",
    )
