import importlib
import importlib.metadata
import re
from datetime import datetime, timezone
from pathlib import Path

import pytest
from openpyxl import Workbook

from xic_extractor.config import ExtractionConfig


def test_app_version_reads_installed_distribution_version(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    metadata = importlib.import_module("xic_extractor.output.metadata")

    def fake_version(name: str) -> str:
        assert name == "xic-extractor"
        return "1.2.3"

    monkeypatch.setattr(importlib.metadata, "version", fake_version)

    assert metadata.app_version() == "1.2.3"


def test_app_version_returns_unknown_when_package_is_not_installed(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    metadata = importlib.import_module("xic_extractor.output.metadata")

    def missing_version(name: str) -> str:
        raise importlib.metadata.PackageNotFoundError(name)

    monkeypatch.setattr(importlib.metadata, "version", missing_version)

    assert metadata.app_version() == "unknown"


def test_build_metadata_rows_returns_contract_keys_in_order() -> None:
    metadata = importlib.import_module("xic_extractor.output.metadata")
    config = _config(Path("output"))
    before = datetime.now(timezone.utc).replace(microsecond=0)

    rows = metadata.build_metadata_rows(config)

    assert [key for key, _value in rows] == [
        "config_hash",
        "app_version",
        "generated_at",
        "resolver_mode",
        "smooth_window",
        "smooth_polyorder",
        "peak_min_prominence_ratio",
        "nl_min_intensity_ratio",
        "ms2_precursor_tol_da",
    ]
    values = dict(rows)
    assert values["config_hash"] == "abc12345"
    assert values["resolver_mode"] == "local_minimum"
    assert values["smooth_window"] == 17
    assert values["smooth_polyorder"] == 3
    assert values["peak_min_prominence_ratio"] == 0.12
    assert values["nl_min_intensity_ratio"] == 0.02
    assert values["ms2_precursor_tol_da"] == 0.4
    assert re.fullmatch(r"\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}Z", values["generated_at"])
    generated_at = datetime.strptime(values["generated_at"], "%Y-%m-%dT%H:%M:%SZ")
    generated_at = generated_at.replace(tzinfo=timezone.utc)
    after = datetime.now(timezone.utc)
    assert before <= generated_at <= after


def test_build_metadata_sheet_writes_headers_values_layout_and_styles() -> None:
    csv_to_excel = importlib.import_module("scripts.csv_to_excel")
    wb = Workbook()
    ws = wb.active

    csv_to_excel._build_metadata_sheet(ws, _config(Path("output")))

    assert ws["A1"].value == "Key"
    assert ws["B1"].value == "Value"
    assert ws["A2"].value == "config_hash"
    assert ws["B2"].value == "abc12345"
    assert ws["A4"].value == "generated_at"
    assert ws["A1"].fill.fgColor.rgb.endswith("2E4057")
    assert ws["A1"].font.bold is True
    assert ws["B2"].border.left.style == "thin"
    assert ws.column_dimensions["A"].width == 28
    assert ws.column_dimensions["B"].width == 48
    assert ws.freeze_panes == "A2"


def _config(output_dir: Path) -> ExtractionConfig:
    return ExtractionConfig(
        data_dir=Path("raw"),
        dll_dir=Path("dll"),
        output_csv=output_dir / "xic_results.csv",
        diagnostics_csv=output_dir / "xic_diagnostics.csv",
        smooth_window=17,
        smooth_polyorder=3,
        peak_rel_height=0.95,
        peak_min_prominence_ratio=0.12,
        ms2_precursor_tol_da=0.4,
        nl_min_intensity_ratio=0.02,
        resolver_mode="local_minimum",
        config_hash="abc12345",
    )
