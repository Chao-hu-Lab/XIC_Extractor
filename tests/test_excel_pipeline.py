from pathlib import Path

from openpyxl import load_workbook

from xic_extractor.config import ExtractionConfig, Target
from xic_extractor.extractor import (
    DiagnosticRecord,
    ExtractionResult,
    FileResult,
    RunOutput,
)
from xic_extractor.neutral_loss import NLResult
from xic_extractor.signal_processing import (
    PeakCandidate,
    PeakDetectionResult,
    PeakResult,
)


def test_write_excel_from_run_output_uses_in_memory_rows_and_metadata(
    tmp_path: Path,
) -> None:
    from xic_extractor.output.excel_pipeline import write_excel_from_run_output

    config = _config(tmp_path)
    target = _target("WithNL")
    output = _run_output(with_diagnostics=True)
    output_path = tmp_path / "output" / "xic_results.xlsx"

    result_path = write_excel_from_run_output(
        config,
        [target],
        output,
        output_path=output_path,
    )

    assert result_path == output_path
    assert output_path.exists()
    assert not config.output_csv.exists()
    assert not config.output_csv.with_name("xic_results_long.csv").exists()
    assert not config.diagnostics_csv.exists()

    wb = load_workbook(output_path)
    assert wb.sheetnames == [
        "Overview",
        "Review Queue",
        "XIC Results",
        "Summary",
        "Targets",
        "Diagnostics",
        "Run Metadata",
    ]
    assert wb.active.title == "Overview"
    ws_overview = wb["Overview"]
    assert ws_overview["A1"].value == "XIC Review Overview"
    assert ws_overview["B3"].value == 1
    assert ws_overview["B5"].value == 1
    assert ws_overview["B6"].value == 1
    ws = wb["XIC Results"]
    assert ws["A2"].value == "SampleA"
    assert ws["C2"].value == "WithNL"
    assert ws["F2"].value == 9.0
    assert ws["G2"].value == 123.45
    assert ws["H2"].value == "⚠ 12.3ppm"
    assert ws["M2"].value == "LOW"

    ws_diagnostics = wb["Diagnostics"]
    assert ws_diagnostics.sheet_state == "hidden"
    assert ws_diagnostics["A2"].value == "SampleA"
    assert ws_diagnostics["C2"].value == "NL_FAIL"
    ws_review = wb["Review Queue"]
    assert ws_review["B2"].value == "SampleA"
    assert ws_review["C2"].value == "WithNL"
    assert ws_review["E2"].value == "Review"
    assert ws_review["F2"].value == "NL support failed"
    assert ws_review["I2"].value == "Check MS2 / NL evidence near selected RT"

    ws_metadata = wb["Run Metadata"]
    metadata_keys = {
        ws_metadata.cell(row=row, column=1).value
        for row in range(2, ws_metadata.max_row + 1)
    }
    assert {
        "config_hash",
        "app_version",
        "generated_at",
        "resolver_mode",
    } <= metadata_keys


def test_write_excel_from_run_output_adds_score_breakdown_when_enabled(
    tmp_path: Path,
) -> None:
    from xic_extractor.output.excel_pipeline import write_excel_from_run_output

    config = _config(tmp_path, emit_score_breakdown=True)
    output_path = tmp_path / "output" / "xic_results.xlsx"

    write_excel_from_run_output(
        config,
        [_target("WithNL")],
        _run_output(with_diagnostics=False),
        output_path=output_path,
    )

    wb = load_workbook(output_path)
    assert wb.sheetnames == [
        "Overview",
        "Review Queue",
        "XIC Results",
        "Summary",
        "Targets",
        "Diagnostics",
        "Run Metadata",
        "Score Breakdown",
    ]
    ws = wb["Score Breakdown"]
    headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
    values = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
    row = dict(zip(headers, values, strict=False))
    assert row["SampleName"] == "SampleA"
    assert row["Target"] == "WithNL"
    assert row["local_sn"] == 1
    assert row["Quality Penalty"] == 1
    assert row["Quality Flags"] == "too_broad"
    assert row["Total Severity"] == 2
    assert row["Confidence"] == "LOW"
    assert row["Prior RT"] is None


def test_write_excel_from_run_output_emits_review_report_when_enabled(
    tmp_path: Path,
) -> None:
    from xic_extractor.output.excel_pipeline import write_excel_from_run_output

    config = _config(tmp_path, emit_review_report=True)
    output_path = tmp_path / "output" / "xic_results_20260505_1200.xlsx"

    write_excel_from_run_output(
        config,
        [_target("WithNL")],
        _run_output(with_diagnostics=False),
        output_path=output_path,
    )

    report_path = output_path.with_name("review_report_20260505_1200.html")
    assert report_path.exists()
    assert "XIC Review Report" in report_path.read_text(encoding="utf-8")


def _config(
    tmp_path: Path,
    *,
    emit_score_breakdown: bool = False,
    emit_review_report: bool = False,
) -> ExtractionConfig:
    return ExtractionConfig(
        data_dir=tmp_path / "raw",
        dll_dir=tmp_path / "dll",
        output_csv=tmp_path / "output" / "xic_results.csv",
        diagnostics_csv=tmp_path / "output" / "xic_diagnostics.csv",
        smooth_window=15,
        smooth_polyorder=3,
        peak_rel_height=0.95,
        peak_min_prominence_ratio=0.10,
        ms2_precursor_tol_da=0.5,
        nl_min_intensity_ratio=0.01,
        emit_score_breakdown=emit_score_breakdown,
        emit_review_report=emit_review_report,
        config_hash="abc12345",
    )


def _target(label: str) -> Target:
    return Target(
        label=label,
        mz=258.1085,
        rt_min=8.0,
        rt_max=10.0,
        ppm_tol=20.0,
        neutral_loss_da=116.0474,
        nl_ppm_warn=20.0,
        nl_ppm_max=50.0,
        is_istd=False,
        istd_pair="",
    )


def _run_output(*, with_diagnostics: bool) -> RunOutput:
    peak = PeakResult(
        rt=9.03,
        intensity=500.0,
        intensity_smoothed=450.0,
        area=123.45,
        peak_start=8.9,
        peak_end=9.2,
    )
    candidate = PeakCandidate(
        peak=peak,
        selection_apex_rt=9.0,
        selection_apex_intensity=450.0,
        selection_apex_index=10,
        raw_apex_rt=9.03,
        raw_apex_intensity=500.0,
        raw_apex_index=11,
        prominence=100.0,
    )
    peak_result = PeakDetectionResult(
        status="OK",
        peak=peak,
        n_points=12,
        max_smoothed=450.0,
        n_prominent_peaks=1,
        candidates=(candidate,),
        confidence="LOW",
        reason="concerns: local_sn (minor)",
        severities=((1, "local_sn"),),
    )
    result = ExtractionResult(
        peak_result=peak_result,
        nl=NLResult("WARN", 12.34, None, 3, 0, 2),
        target_label="WithNL",
        role="Analyte",
        istd_pair="",
        confidence="LOW",
        reason="concerns: local_sn (minor)",
        severities=((1, "local_sn"),),
        quality_penalty=1,
        quality_flags=("too_broad",),
    )
    diagnostics = []
    if with_diagnostics:
        diagnostics.append(
            DiagnosticRecord(
                sample_name="SampleA",
                target_label="WithNL",
                issue="NL_FAIL",
                reason="best match 80 ppm",
            )
        )
    return RunOutput(
        file_results=[FileResult(sample_name="SampleA", results={"WithNL": result})],
        diagnostics=diagnostics,
    )
