from pathlib import Path
from typing import cast

from openpyxl import load_workbook

from xic_extractor.instrument_qc.models import (
    ActivationMethod,
    HCDAuditRow,
    HCDAuditStatus,
    InstrumentQCDiagnostic,
    SDOLEKTrendRow,
)
from xic_extractor.instrument_qc.workbook import write_sdolek_workbook


def test_write_sdolek_workbook_has_overview_trend_and_diagnostics(
    tmp_path: Path,
) -> None:
    path = write_sdolek_workbook(
        tmp_path / "instrument_qc_trend_sdolek.xlsx",
        [_row(tmp_path / "SDOLEK-pretest.raw")],
        [
            InstrumentQCDiagnostic(
                sample_name="SDOLEK-pretest",
                raw_path=tmp_path / "SDOLEK-pretest.raw",
                issue="INJECTION_ORDER_MISSING",
                detail="No injection-order file supplied.",
            )
        ],
        metadata_source_status={
            "injection_order_source": "",
            "injection_order_status": "missing",
        },
    )

    workbook = load_workbook(path, data_only=False)

    assert workbook.sheetnames == ["Overview", "SDOLEK Trend", "Diagnostics"]
    assert workbook["Overview"]["A1"].value == "metric"
    assert workbook["Overview"]["A2"].value == "report_type"
    assert workbook["Overview"]["B4"].value == 1
    assert workbook["SDOLEK Trend"]["A1"].value == "sample_name"
    assert workbook["SDOLEK Trend"]["D2"].value == "SDO"
    assert workbook["Diagnostics"]["C2"].value == "INJECTION_ORDER_MISSING"


def test_write_workbook_adds_mixstds_sheet_before_diagnostics(
    tmp_path: Path,
) -> None:
    path = write_sdolek_workbook(
        tmp_path / "instrument_qc_trend_sdolek.xlsx",
        [_row(tmp_path / "SDOLEK-pretest.raw")],
        [],
        mixstds_rows=[_row(tmp_path / "Mix_STDs_01.raw", sample_name="Mix_STDs_01")],
    )

    workbook = load_workbook(path, data_only=False)

    assert workbook.sheetnames == [
        "Overview",
        "SDOLEK Trend",
        "Mix STDs Trend",
        "Diagnostics",
    ]
    assert workbook["Mix STDs Trend"]["A1"].value == "sample_name"


def test_write_sdolek_workbook_escapes_formula_like_values(tmp_path: Path) -> None:
    path = write_sdolek_workbook(
        tmp_path / "instrument_qc_trend_sdolek.xlsx",
        [_row(Path("=raw.raw"), sample_name="=SDOLEK")],
        [],
    )

    workbook = load_workbook(path, data_only=False)

    assert workbook["SDOLEK Trend"]["A2"].value == "'=SDOLEK"
    assert workbook["SDOLEK Trend"]["B2"].value.endswith("'=raw.raw")


def test_write_workbook_adds_hcd_and_manual_review_sheets(tmp_path: Path) -> None:
    path = write_sdolek_workbook(
        tmp_path / "instrument_qc_trend_sdolek.xlsx",
        [_row(tmp_path / "SDOLEK-pretest.raw", compound="LEK", rt_delta=-0.8)],
        [],
        hcd_rows=[_hcd_row(tmp_path / "SDOLEK-pretest.raw")],
    )

    workbook = load_workbook(path, data_only=False)

    assert workbook.sheetnames == [
        "Overview",
        "SDOLEK Trend",
        "HCD Audit",
        "Manual Review Queue",
        "Diagnostics",
    ]
    assert workbook["HCD Audit"]["A1"].value == "sample_name"
    assert workbook["Manual Review Queue"]["A1"].value == "priority"
    assert workbook["Manual Review Queue"]["J1"].value == "rt_drift_hint"
    queue_reasons = [
        workbook["Manual Review Queue"].cell(row=row, column=2).value
        for row in range(2, workbook["Manual Review Queue"].max_row + 1)
    ]
    assert "hcd_no_product_match" in queue_reasons
    assert "lek_rt_shift" in queue_reasons


def test_manual_review_queue_skips_stable_lek_batch_prior_shift(
    tmp_path: Path,
) -> None:
    path = write_sdolek_workbook(
        tmp_path / "instrument_qc_trend_sdolek.xlsx",
        [
            _row(
                tmp_path / "SDOLEK-1.raw",
                sample_name="SDOLEK-1",
                compound="LEK",
                rt_delta=-0.90,
                apex_rt=5.50,
            ),
            _row(
                tmp_path / "SDOLEK-2.raw",
                sample_name="SDOLEK-2",
                compound="LEK",
                rt_delta=-0.80,
                apex_rt=5.60,
            ),
            _row(
                tmp_path / "SDOLEK-3.raw",
                sample_name="SDOLEK-3",
                compound="LEK",
                rt_delta=-0.70,
                apex_rt=5.70,
            ),
        ],
        [],
        hcd_rows=[],
    )

    workbook = load_workbook(path, data_only=False)

    assert workbook["Manual Review Queue"].max_row == 1


def test_manual_review_queue_skips_istd_supported_product_miss(
    tmp_path: Path,
) -> None:
    path = write_sdolek_workbook(
        tmp_path / "instrument_qc_trend_sdolek.xlsx",
        [],
        [],
        hcd_rows=[
            _hcd_row(
                tmp_path / "Mix_STDs.raw",
                sample_name="Mix_STDs",
                compound="5-hmdC",
                hcd_status="no_product_match",
                matched_products=(),
            ),
            _hcd_row(
                tmp_path / "Mix_STDs.raw",
                sample_name="Mix_STDs",
                compound="d3-5-hmdC",
                hcd_status="hcd_supported",
                matched_products=("CIDwHCD:NL-116.0474",),
            ),
        ],
    )

    workbook = load_workbook(path, data_only=False)

    assert workbook["Manual Review Queue"].max_row == 1


def test_manual_review_queue_skips_unmapped_product_miss(
    tmp_path: Path,
) -> None:
    path = write_sdolek_workbook(
        tmp_path / "instrument_qc_trend_sdolek.xlsx",
        [],
        [],
        hcd_rows=[
            _hcd_row(
                tmp_path / "Mix_STDs.raw",
                compound="Y",
                hcd_mapping_source="unmapped",
                hcd_product_group="",
                hcd_status="no_product_match",
            ),
        ],
    )

    workbook = load_workbook(path, data_only=False)

    assert workbook["Manual Review Queue"].max_row == 1


def test_manual_review_queue_reports_target_rt_window_mismatch(
    tmp_path: Path,
) -> None:
    path = write_sdolek_workbook(
        tmp_path / "instrument_qc_trend_sdolek.xlsx",
        [],
        [],
        mixstds_rows=[
            _row(
                tmp_path / "Mix_STDs.raw",
                sample_name="Mix_STDs",
                compound="t6A",
                rt_delta=0.0,
            )
        ],
        hcd_rows=[
            _hcd_row(
                tmp_path / "Mix_STDs.raw",
                sample_name="Mix_STDs",
                compound="t6A",
                hcd_status="hcd_supported",
                matched_products=("CIDwHCD:NL-132.0423",),
                review_flags=("target_rt_window_review",),
            ),
        ],
    )

    workbook = load_workbook(path, data_only=False)

    assert workbook["Manual Review Queue"]["B2"].value == "target_rt_window_mismatch"
    assert "outside" in workbook["Manual Review Queue"]["K2"].value


def test_manual_review_queue_skips_unmapped_target_rt_window_mismatch(
    tmp_path: Path,
) -> None:
    path = write_sdolek_workbook(
        tmp_path / "instrument_qc_trend_sdolek.xlsx",
        [],
        [],
        hcd_rows=[
            _hcd_row(
                tmp_path / "Mix_STDs.raw",
                compound="Y",
                hcd_mapping_source="unmapped",
                hcd_product_group="",
                hcd_status="hcd_supported",
                matched_products=("CIDwHCD:NL-116.0474",),
                review_flags=("target_rt_window_review",),
            ),
        ],
    )

    workbook = load_workbook(path, data_only=False)

    assert workbook["Manual Review Queue"].max_row == 1


def test_manual_review_queue_skips_isotope_supported_target_rt_window_mismatch(
    tmp_path: Path,
) -> None:
    path = write_sdolek_workbook(
        tmp_path / "instrument_qc_trend_sdolek.xlsx",
        [],
        [],
        hcd_rows=[
            _hcd_row(
                tmp_path / "Mix_STDs.raw",
                sample_name="Mix_STDs",
                compound="5-hmdC",
                hcd_mapping_source="heuristic",
                hcd_product_group="C",
                hcd_status="hcd_supported",
                matched_products=("CIDwHCD:NL-116.0474",),
                review_flags=("target_rt_window_review",),
            ),
            _hcd_row(
                tmp_path / "Mix_STDs.raw",
                sample_name="Mix_STDs",
                compound="d3-5-hmdC",
                hcd_mapping_source="heuristic",
                hcd_product_group="C",
                hcd_status="hcd_supported",
                matched_products=("CIDwHCD:NL-116.0474",),
            ),
        ],
    )

    workbook = load_workbook(path, data_only=False)

    assert workbook["Manual Review Queue"].max_row == 1


def _row(
    raw_path: Path,
    *,
    sample_name: str = "SDOLEK-pretest",
    compound: str = "SDO",
    rt_delta: float = 0.01,
    apex_rt: float = 6.27,
) -> SDOLEKTrendRow:
    return SDOLEKTrendRow(
        sample_name=sample_name,
        raw_path=raw_path,
        injection_order=4,
        compound=compound,
        precursor_mz=311.0814,
        identity_evidence="MS1_ONLY",
        reference_rt_min=6.26,
        rt_delta_to_reference_min=rt_delta,
        apex_rt_min=apex_rt,
        area=123.4,
        base_width_min=0.83,
        reference_base_width_min=0.83,
        base_width_ratio_to_reference=1.0,
        peak_start_rt_min=5.90,
        peak_end_rt_min=6.73,
        trend_confidence="clean",
        trend_flags=(),
        status="detected",
        reason="OK",
    )


def _hcd_row(
    raw_path: Path,
    *,
    sample_name: str = "SDOLEK-pretest",
    compound: str = "LEK",
    hcd_mapping_source: str = "sdolek_builtin",
    hcd_product_group: str = "",
    hcd_status: str = "no_product_match",
    matched_products: tuple[str, ...] = (),
    activation_method: str = "CIDwHCD",
    review_flags: tuple[str, ...] | None = None,
) -> HCDAuditRow:
    row_review_flags = (
        ("activation_unknown_review",)
        if activation_method == "unknown"
        else ()
    )
    if review_flags is not None:
        row_review_flags = review_flags
    return HCDAuditRow(
        sample_name=sample_name,
        raw_path=raw_path,
        injection_order=4,
        compound=compound,
        precursor_mz=556.2771,
        ms1_apex_rt_min=5.6,
        ms1_status="detected",
        instrument_method="20260105 SDOLEK",
        activation_method=cast(ActivationMethod, activation_method),
        hcd_mapping_source=hcd_mapping_source,
        hcd_product_group=hcd_product_group,
        hcd_status=cast(HCDAuditStatus, hcd_status),
        best_ms2_scan_rt_min=5.6,
        apex_ms2_delta_min=0.0,
        trigger_scan_count=1,
        expected_product_count=8,
        matched_product_count=len(matched_products),
        best_product_ppm=None,
        best_product_base_ratio=None,
        matched_products=matched_products,
        review_flags=row_review_flags,
        review_reason="MS2 trigger exists, but no expected product ion was observed.",
    )
