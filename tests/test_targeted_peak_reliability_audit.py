from __future__ import annotations

import csv
import json
from pathlib import Path

import pytest
from openpyxl import Workbook

from tools.diagnostics import targeted_peak_reliability_audit as audit


def test_audit_classifies_strong_review_and_negative_rows(tmp_path: Path) -> None:
    workbook = tmp_path / "targeted.xlsx"
    _write_targeted_workbook(
        workbook,
        targets=[
            _target("clean"),
            _target("low_conf"),
            _target("nl_fail"),
            _target("unicode_nl_fail"),
            _target("no_peak"),
        ],
        result_rows=[
            _result("S1", "clean", rt=10.0, area=1000.0, nl="OK", confidence="HIGH"),
            _result(
                "S1",
                "low_conf",
                rt=10.1,
                area=100.0,
                nl="OK",
                confidence="VERY_LOW",
            ),
            _result(
                "S1",
                "nl_fail",
                rt=10.2,
                area=120.0,
                nl="NL_FAIL",
                confidence="HIGH",
            ),
            _result(
                "S1",
                "unicode_nl_fail",
                rt=10.3,
                area=130.0,
                nl="✗ NL",
                confidence="HIGH",
            ),
            _result("S1", "no_peak", rt="ND", area="ND", nl="OK", confidence="HIGH"),
        ],
        score_rows=[
            _score("S1", "clean", confidence="HIGH"),
            _score("S1", "low_conf", confidence="VERY_LOW"),
            _score("S1", "nl_fail", confidence="HIGH", concerns="nl_fail"),
            _score("S1", "unicode_nl_fail", confidence="HIGH", concerns="nl_fail"),
            _score("S1", "no_peak", confidence="HIGH"),
        ],
    )

    outputs, result = audit.run_targeted_peak_reliability_audit(
        targeted_workbook=workbook,
        output_dir=tmp_path / "audit",
    )

    by_target = {row.target_label: row for row in result.rows}
    assert by_target["clean"].reliability_state == "benchmark_eligible"
    assert by_target["low_conf"].reliability_state == "targeted_review"
    assert by_target["low_conf"].risk_reasons == ("low_confidence",)
    assert by_target["nl_fail"].reliability_state == "targeted_review"
    assert by_target["nl_fail"].risk_reasons == ("hard_nl_conflict",)
    assert by_target["unicode_nl_fail"].reliability_state == "targeted_review"
    assert by_target["unicode_nl_fail"].risk_reasons == ("hard_nl_conflict",)
    assert by_target["no_peak"].reliability_state == "targeted_negative"
    assert by_target["no_peak"].risk_reasons == ("no_usable_peak",)

    rows = _read_tsv(outputs.rows_tsv)
    assert {row["target_label"]: row["reliability_state"] for row in rows} == {
        "clean": "benchmark_eligible",
        "low_conf": "targeted_review",
        "nl_fail": "targeted_review",
        "unicode_nl_fail": "targeted_review",
        "no_peak": "targeted_negative",
    }


def test_plausible_nl_dropout_is_review_positive_not_targeted_negative(
    tmp_path: Path,
) -> None:
    workbook = tmp_path / "targeted.xlsx"
    _write_targeted_workbook(
        workbook,
        targets=[
            _target("8-oxodG"),
            _target("hard_nl_conflict"),
        ],
        result_rows=[
            _result(
                "S1",
                "8-oxodG",
                rt=17.18,
                area=1_850_000.0,
                nl="✗ NL",
                confidence="VERY_LOW",
                reason=(
                    "decision: review only, not counted; "
                    "cap: VERY_LOW due to nl fail; "
                    "support: local S/N strong; shape clean; trace clean; "
                    "concerns: nl fail"
                ),
            ),
            _result(
                "S1",
                "hard_nl_conflict",
                rt=17.18,
                area=120_000.0,
                nl="NL_FAIL",
                confidence="VERY_LOW",
                reason=(
                    "decision: review only, not counted; "
                    "cap: VERY_LOW due to nl fail; "
                    "support: local S/N strong; "
                    "concerns: nl fail; shape poor"
                ),
            ),
        ],
        score_rows=[
            _score("S1", "8-oxodG", confidence="VERY_LOW", concerns="nl_fail"),
            _score(
                "S1",
                "hard_nl_conflict",
                confidence="VERY_LOW",
                concerns="nl_fail",
            ),
        ],
    )

    outputs, result = audit.run_targeted_peak_reliability_audit(
        targeted_workbook=workbook,
        output_dir=tmp_path / "audit",
    )

    by_target = {row.target_label: row for row in result.rows}
    assert by_target["8-oxodG"].reliability_state == "targeted_review_positive"
    assert by_target["8-oxodG"].risk_reasons == (
        "low_confidence",
        "plausible_nl_dropout",
    )
    assert by_target["hard_nl_conflict"].reliability_state == "targeted_review"
    assert by_target["hard_nl_conflict"].risk_reasons == (
        "low_confidence",
        "hard_nl_conflict",
    )
    summary = {row.target_label: row for row in result.summaries}
    assert summary["8-oxodG"].targeted_review_positive_count == 1
    assert summary["hard_nl_conflict"].targeted_review_count == 1

    rows = _read_tsv(outputs.rows_tsv)
    assert {row["target_label"]: row["reliability_state"] for row in rows} == {
        "8-oxodG": "targeted_review_positive",
        "hard_nl_conflict": "targeted_review",
    }
    payload = json.loads(outputs.json_path.read_text(encoding="utf-8"))
    assert payload["summary"]["targeted_review_positive_count"] == 1


def test_selected_candidate_evidence_can_mark_nl_fail_as_review_positive(
    tmp_path: Path,
) -> None:
    workbook = tmp_path / "targeted.xlsx"
    candidates = tmp_path / "peak_candidates.tsv"
    _write_targeted_workbook(
        workbook,
        targets=[_target("8-oxo-Guo")],
        result_rows=[
            _result(
                "TumorBC2275_DNA",
                "8-oxo-Guo",
                rt=13.9143,
                area=250_000.0,
                nl="✗ NL",
                confidence="VERY_LOW",
            ),
        ],
        score_rows=None,
    )
    _write_peak_candidates(
        candidates,
        [
            _peak_candidate(
                "TumorBC2275_DNA",
                "8-oxo-Guo",
                selected="TRUE",
                raw_score="15",
                support_labels="local_sn_strong;trace_clean",
                concern_labels="nl_fail;shape_borderline",
                proposal_sources="local_minimum;centwave_cwt",
                ms2_present="TRUE",
                nl_match="FALSE",
            )
        ],
    )

    outputs, result = audit.run_targeted_peak_reliability_audit(
        targeted_workbook=workbook,
        peak_candidates_tsv=candidates,
        output_dir=tmp_path / "audit",
    )

    row = result.rows[0]
    assert row.reliability_state == "targeted_review_positive"
    assert row.risk_reasons == (
        "low_confidence",
        "plausible_nl_dropout",
        "score_breakdown_unavailable",
    )
    assert result.benchmark_eligible_count == 0
    assert result.targeted_review_positive_count == 1


def test_selected_candidate_product_probe_reason_stays_review_positive_context(
    tmp_path: Path,
) -> None:
    workbook = tmp_path / "targeted.xlsx"
    candidates = tmp_path / "peak_candidates.tsv"
    _write_targeted_workbook(
        workbook,
        targets=[_target("8-oxo-Guo")],
        result_rows=[
            _result(
                "TumorBC2275_DNA",
                "8-oxo-Guo",
                rt=13.9143,
                area=250_000.0,
                nl="NL_FAIL",
                confidence="VERY_LOW",
            ),
        ],
        score_rows=None,
    )
    candidate = _peak_candidate(
        "TumorBC2275_DNA",
        "8-oxo-Guo",
        selected="TRUE",
        raw_score="15",
        support_labels="local_sn_strong;trace_clean",
        concern_labels="nl_fail;shape_borderline",
        proposal_sources="local_minimum;centwave_cwt",
        ms2_present="TRUE",
        nl_match="FALSE",
    )
    candidate.update(
        {
            "diagnostic_product_absence_reason": (
                "product_outside_diagnostic_window"
            ),
            "apex_ms2_delta_min": "0.02",
            "nearest_product_loss_ppm": "352.1",
            "nearest_product_base_ratio": "0.35",
            "nearest_product_mz": "171.0",
        }
    )
    _write_peak_candidates(candidates, [candidate])

    outputs, result = audit.run_targeted_peak_reliability_audit(
        targeted_workbook=workbook,
        peak_candidates_tsv=candidates,
        output_dir=tmp_path / "audit",
    )

    row = result.rows[0]
    assert row.reliability_state == "targeted_review_positive"
    assert row.risk_reasons == (
        "low_confidence",
        "plausible_nl_dropout",
        "score_breakdown_unavailable",
        "product_outside_diagnostic_window",
    )
    summary = result.summaries[0]
    assert "product_outside_diagnostic_window" in summary.top_risk_reasons


def test_selected_candidate_hard_conflict_does_not_mark_dropout_positive(
    tmp_path: Path,
) -> None:
    workbook = tmp_path / "targeted.xlsx"
    candidates = tmp_path / "peak_candidates.tsv"
    _write_targeted_workbook(
        workbook,
        targets=[_target("hard_nl_conflict")],
        result_rows=[
            _result(
                "S1",
                "hard_nl_conflict",
                rt=10.0,
                area=1000.0,
                nl="NL_FAIL",
                confidence="VERY_LOW",
            ),
        ],
        score_rows=None,
    )
    _write_peak_candidates(
        candidates,
        [
            _peak_candidate(
                "S1",
                "hard_nl_conflict",
                selected="TRUE",
                raw_score="15",
                support_labels="local_sn_strong;trace_clean",
                concern_labels="nl_fail;shape_poor",
                proposal_sources="legacy_savgol;centwave_cwt",
                ms2_present="TRUE",
                nl_match="FALSE",
            )
        ],
    )

    _outputs, result = audit.run_targeted_peak_reliability_audit(
        targeted_workbook=workbook,
        peak_candidates_tsv=candidates,
        output_dir=tmp_path / "audit",
    )

    row = result.rows[0]
    assert row.reliability_state == "targeted_review"
    assert row.risk_reasons == (
        "low_confidence",
        "hard_nl_conflict",
        "score_breakdown_unavailable",
    )


def test_peak_candidate_input_requires_expected_columns(tmp_path: Path) -> None:
    workbook = tmp_path / "targeted.xlsx"
    candidates = tmp_path / "peak_candidates.tsv"
    _write_targeted_workbook(
        workbook,
        targets=[_target("clean")],
        result_rows=[
            _result("S1", "clean", rt=10.0, area=1000.0, nl="OK", confidence="HIGH"),
        ],
        score_rows=None,
    )
    with candidates.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=["sample_name"], delimiter="\t")
        writer.writeheader()
        writer.writerow({"sample_name": "S1"})

    code = audit.main(
        [
            "--targeted-workbook",
            str(workbook),
            "--peak-candidates-tsv",
            str(candidates),
            "--output-dir",
            str(tmp_path / "audit"),
        ],
    )

    assert code == 2


def test_missing_score_breakdown_is_reported_without_demoting_strong_row(
    tmp_path: Path,
) -> None:
    workbook = tmp_path / "targeted.xlsx"
    _write_targeted_workbook(
        workbook,
        targets=[_target("clean")],
        result_rows=[
            _result("S1", "clean", rt=10.0, area=1000.0, nl="OK", confidence="HIGH"),
        ],
        score_rows=None,
    )

    _outputs, result = audit.run_targeted_peak_reliability_audit(
        targeted_workbook=workbook,
        output_dir=tmp_path / "audit",
    )

    row = result.rows[0]
    assert row.reliability_state == "benchmark_eligible"
    assert row.risk_reasons == ("score_breakdown_unavailable",)
    assert result.summaries[0].benchmark_eligible_count == 1


def test_accepted_low_istd_with_strict_nl_remains_benchmark_eligible(
    tmp_path: Path,
) -> None:
    workbook = tmp_path / "targeted.xlsx"
    _write_targeted_workbook(
        workbook,
        targets=[_target("ISTD-A")],
        result_rows=[
            _result(
                "S1",
                "ISTD-A",
                rt=10.0,
                area=1000.0,
                nl="OK",
                confidence="LOW",
                reason=(
                    "decision: accepted; support: strict NL OK; "
                    "concerns: MS2 trace weak"
                ),
            ),
        ],
        score_rows=None,
    )

    _outputs, result = audit.run_targeted_peak_reliability_audit(
        targeted_workbook=workbook,
        output_dir=tmp_path / "audit",
    )

    row = result.rows[0]
    assert row.reliability_state == "benchmark_eligible"
    assert row.risk_reasons == ("score_breakdown_unavailable",)
    assert result.summaries[0].benchmark_eligible_count == 1


def test_accepted_low_istd_with_hard_quality_stays_targeted_review(
    tmp_path: Path,
) -> None:
    workbook = tmp_path / "targeted.xlsx"
    _write_targeted_workbook(
        workbook,
        targets=[_target("ISTD-A")],
        result_rows=[
            _result(
                "S1",
                "ISTD-A",
                rt=10.0,
                area=1000.0,
                nl="OK",
                confidence="LOW",
                reason=(
                    "decision: accepted; support: strict NL OK; "
                    "concerns: hard quality flag; weak candidate: edge_clipped"
                ),
            ),
        ],
        score_rows=None,
    )

    _outputs, result = audit.run_targeted_peak_reliability_audit(
        targeted_workbook=workbook,
        output_dir=tmp_path / "audit",
    )

    row = result.rows[0]
    assert row.reliability_state == "targeted_review"
    assert "low_confidence" in row.risk_reasons


def test_weak_area_outlier_and_known_exception_are_reported(tmp_path: Path) -> None:
    workbook = tmp_path / "targeted.xlsx"
    _write_targeted_workbook(
        workbook,
        targets=[_target("d3-N6-medA")],
        result_rows=[
            _result(
                "S1",
                "d3-N6-medA",
                rt=10.0,
                area=1000.0,
                nl="OK",
                confidence="HIGH",
            ),
            _result(
                "S2",
                "d3-N6-medA",
                rt=10.1,
                area=900.0,
                nl="OK",
                confidence="HIGH",
            ),
            _result("S3", "d3-N6-medA", rt=10.2, area=5.0, nl="OK", confidence="HIGH"),
        ],
        score_rows=[
            _score("S1", "d3-N6-medA", confidence="HIGH"),
            _score("S2", "d3-N6-medA", confidence="HIGH"),
            _score("S3", "d3-N6-medA", confidence="HIGH"),
        ],
    )

    outputs, result = audit.run_targeted_peak_reliability_audit(
        targeted_workbook=workbook,
        output_dir=tmp_path / "audit",
        known_target_exceptions=("d3-N6-medA:AREA_MISMATCH",),
    )

    by_sample = {row.sample_name: row for row in result.rows}
    assert by_sample["S1"].reliability_state == "benchmark_eligible"
    assert by_sample["S1"].target_area_median == 900.0
    assert by_sample["S1"].area_to_target_median_ratio == pytest.approx(
        1000.0 / 900.0
    )
    assert by_sample["S1"].weak_area_threshold_ratio == 0.05
    assert by_sample["S3"].reliability_state == "targeted_review"
    assert "weak_area_rank" in by_sample["S3"].risk_reasons
    assert by_sample["S3"].target_area_median == 900.0
    assert by_sample["S3"].area_to_target_median_ratio == pytest.approx(5.0 / 900.0)
    assert by_sample["S3"].weak_area_threshold_ratio == 0.05
    assert by_sample["S3"].known_exception == "AREA_MISMATCH"
    summary = result.summaries[0]
    assert summary.benchmark_eligible_count == 2
    assert summary.targeted_review_count == 1
    assert summary.known_exception == "AREA_MISMATCH"
    assert "weak_area_rank" in summary.top_risk_reasons
    rows = _read_tsv(outputs.rows_tsv)
    by_sample_tsv = {row["sample_name"]: row for row in rows}
    assert by_sample_tsv["S3"]["target_area_median"] == "900"
    assert by_sample_tsv["S3"]["area_to_target_median_ratio"]
    assert by_sample_tsv["S3"]["weak_area_threshold_ratio"] == "0.05"


def test_main_writes_outputs_and_reports_missing_required_column(
    tmp_path: Path,
    capsys,
) -> None:
    workbook = tmp_path / "targeted.xlsx"
    _write_targeted_workbook(
        workbook,
        targets=[_target("clean")],
        result_rows=[
            _result("S1", "clean", rt=10.0, area=1000.0, nl="OK", confidence="HIGH"),
        ],
        score_rows=None,
        omit_result_columns={"Confidence"},
    )

    code = audit.main(
        [
            "--targeted-workbook",
            str(workbook),
            "--output-dir",
            str(tmp_path / "audit"),
        ],
    )

    assert code == 2
    assert "Confidence" in capsys.readouterr().err


def test_main_writes_json_markdown_and_tsv_outputs(tmp_path: Path) -> None:
    workbook = tmp_path / "targeted.xlsx"
    output_dir = tmp_path / "audit"
    _write_targeted_workbook(
        workbook,
        targets=[_target("clean")],
        result_rows=[
            _result("S1", "clean", rt=10.0, area=1000.0, nl="OK", confidence="HIGH"),
        ],
        score_rows=[
            _score("S1", "clean", confidence="HIGH"),
        ],
    )

    code = audit.main(
        [
            "--targeted-workbook",
            str(workbook),
            "--output-dir",
            str(output_dir),
        ],
    )

    assert code == 0
    assert (output_dir / "targeted_peak_reliability_summary.tsv").exists()
    assert (output_dir / "targeted_peak_reliability_rows.tsv").exists()
    assert (output_dir / "targeted_peak_reliability.md").exists()
    payload = json.loads(
        (output_dir / "targeted_peak_reliability.json").read_text(encoding="utf-8"),
    )
    assert payload["overall_status"] == "PASS"
    assert payload["summary"]["benchmark_eligible_count"] == 1


def _target(label: str) -> dict[str, object]:
    return {
        "Label": label,
        "Role": "ISTD",
        "ISTD Pair": "",
        "m/z": 100.0,
        "RT min": 9.5,
        "RT max": 10.5,
        "ppm tol": 20.0,
        "NL (Da)": 116.0474,
        "Expected product m/z": 50.0,
    }


def _result(
    sample: str,
    target: str,
    *,
    rt: object,
    area: object,
    nl: str,
    confidence: str,
    reason: str = "",
) -> dict[str, object]:
    return {
        "SampleName": sample,
        "Group": "QC",
        "Target": target,
        "Role": "ISTD",
        "ISTD Pair": "",
        "RT": rt,
        "Area": area,
        "NL": nl,
        "Int": 1000.0,
        "PeakStart": 9.9,
        "PeakEnd": 10.1,
        "PeakWidth": 0.2,
        "Confidence": confidence,
        "Reason": reason,
    }


def _score(
    sample: str,
    target: str,
    *,
    confidence: str,
    concerns: str = "",
) -> dict[str, object]:
    return {
        "SampleName": sample,
        "Target": target,
        "Final Confidence": confidence,
        "Detection Counted": "TRUE",
        "Caps": "",
        "Raw Score": "80",
        "Support": "strict_nl_ok",
        "Concerns": concerns,
        "Base Score": "80",
        "Positive Points": "80",
        "Negative Points": "0",
        "symmetry": "0",
        "local_sn": "0",
        "nl_support": "0",
        "rt_prior": "0",
        "rt_centrality": "0",
        "noise_shape": "0",
        "peak_width": "0",
        "Quality Penalty": "0",
        "Quality Flags": "",
        "Total Severity": "0",
        "Confidence": confidence,
        "Prior RT": "10",
        "Prior Source": "test",
    }


def _peak_candidate(
    sample: str,
    target: str,
    *,
    selected: str,
    raw_score: str,
    support_labels: str,
    concern_labels: str,
    proposal_sources: str,
    ms2_present: str,
    nl_match: str,
) -> dict[str, object]:
    return {
        "sample_name": sample,
        "target_label": target,
        "resolver_mode": "arbitrated",
        "candidate_id": f"{sample}:{target}",
        "proposal_sources": proposal_sources,
        "rt_apex_min": "10.0",
        "selected": selected,
        "confidence": "VERY_LOW",
        "raw_score": raw_score,
        "support_labels": support_labels,
        "concern_labels": concern_labels,
        "quality_flags": "",
        "ms2_present": ms2_present,
        "nl_match": nl_match,
    }


def _write_peak_candidates(
    path: Path,
    rows: list[dict[str, object]],
) -> None:
    columns = tuple(rows[0])
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns, delimiter="\t")
        writer.writeheader()
        writer.writerows(rows)


def _write_targeted_workbook(
    path: Path,
    *,
    targets: list[dict[str, object]],
    result_rows: list[dict[str, object]],
    score_rows: list[dict[str, object]] | None,
    omit_result_columns: set[str] | None = None,
) -> None:
    workbook = Workbook()
    targets_sheet = workbook.active
    targets_sheet.title = "Targets"
    target_columns = tuple(targets[0])
    targets_sheet.append(target_columns)
    for target in targets:
        targets_sheet.append([target[column] for column in target_columns])

    omit_result_columns = omit_result_columns or set()
    result_columns = tuple(
        column for column in result_rows[0] if column not in omit_result_columns
    )
    results = workbook.create_sheet("XIC Results")
    results.append(result_columns)
    for row in result_rows:
        results.append([row[column] for column in result_columns])

    if score_rows is not None:
        score_columns = tuple(score_rows[0])
        score_sheet = workbook.create_sheet("Score Breakdown")
        score_sheet.append(score_columns)
        for row in score_rows:
            score_sheet.append([row[column] for column in score_columns])

    workbook.save(path)


def _read_tsv(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8") as handle:
        return list(csv.DictReader(handle, delimiter="\t"))
