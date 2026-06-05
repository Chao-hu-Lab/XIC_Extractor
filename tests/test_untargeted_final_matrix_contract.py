import csv
from pathlib import Path
from types import SimpleNamespace

from openpyxl import load_workbook

from xic_extractor.alignment.matrix import AlignedCell, AlignmentMatrix
from xic_extractor.alignment.tsv_writer import (
    write_alignment_matrix_identity_tsv,
    write_alignment_matrix_tsv,
)
from xic_extractor.alignment.xlsx_writer import write_alignment_results_xlsx
from xic_extractor.peak_detection.hypotheses import IntegrationResult

FORBIDDEN_PRIMARY_STATUSES = {
    "detected",
    "rescued",
    "accepted_rescue",
    "review_rescue",
    "rejected_rescue",
    "duplicate_assigned",
    "ambiguous_ms1_owner",
    "absent",
    "unchecked",
}


def test_primary_outputs_hide_status_strings_and_keep_audit_reasons(
    tmp_path: Path,
):
    matrix = AlignmentMatrix(
        clusters=(
            _feature("FAM001", evidence="owner_complete_link;owner_count=2"),
            _feature("FAM_PROVISIONAL", evidence="single_sample_local_owner"),
            _feature("FAM002", evidence="", has_anchor=False),
            _feature("FAM003", evidence="owner_complete_link;owner_count=2"),
        ),
        sample_order=("s1", "s2", "s3"),
        cells=(
            _cell("s1", "FAM001", "detected", 100.0),
            _cell("s2", "FAM001", "rescued", 90.0),
            _cell(
                "s3",
                "FAM001",
                "detected",
                110.0,
                selected_integration=_integration(
                    raw_area=125.0,
                    asls_area=105.0,
                    morphology_area=115.0,
                ),
            ),
            _cell("s1", "FAM_PROVISIONAL", "detected", 85.0),
            _cell("s2", "FAM_PROVISIONAL", "rescued", 75.0),
            _cell("s1", "FAM002", "rescued", 80.0),
            _cell("s2", "FAM002", "absent", None),
            _cell("s1", "FAM003", "duplicate_assigned", 70.0),
            _cell("s2", "FAM003", "ambiguous_ms1_owner", None),
        ),
    )

    matrix_tsv = write_alignment_matrix_tsv(
        tmp_path / "alignment_matrix.tsv",
        matrix,
    )
    workbook_path = write_alignment_results_xlsx(
        tmp_path / "alignment_results.xlsx",
        matrix,
        metadata={"schema_version": "alignment-results-v3"},
    )

    tsv_rows = _read_tsv(matrix_tsv)
    assert list(tsv_rows[0]) == ["Mz", "RT", "s1", "s2", "s3"]
    assert [(row["Mz"], row["RT"]) for row in tsv_rows] == [("500.123", "8.49")]
    assert tsv_rows[0]["s1"] == "100"
    assert tsv_rows[0]["s2"] == "90"
    assert tsv_rows[0]["s3"] == "115"
    assert FORBIDDEN_PRIMARY_STATUSES.isdisjoint(
        value for row in tsv_rows for value in row.values()
    )

    workbook = load_workbook(workbook_path, data_only=True)
    matrix_rows = _worksheet_records(workbook["Matrix"])
    assert list(matrix_rows[0]) == ["Mz", "RT", "s1", "s2", "s3"]
    assert [(row["Mz"], row["RT"]) for row in matrix_rows] == [(500.123, 8.49)]
    assert matrix_rows[0]["s1"] == 100.0
    assert matrix_rows[0]["s2"] == 90.0
    assert matrix_rows[0]["s3"] == 115.0
    assert FORBIDDEN_PRIMARY_STATUSES.isdisjoint(
        str(value)
        for row in matrix_rows
        for value in row.values()
        if value is not None
    )

    audit_rows = _worksheet_records(workbook["Audit"])
    review_rows = _worksheet_records(workbook["Review"])
    audit_by_sample = {
        (row["feature_family_id"], row["sample_stem"]): row for row in audit_rows
    }
    assert audit_by_sample[("FAM001", "s3")]["area"] == 110.0
    assert audit_by_sample[("FAM001", "s3")]["primary_matrix_area"] == 115.0
    assert (
        audit_by_sample[("FAM001", "s3")]["primary_matrix_area_source"]
        == "gaussian15_positive_asls_residual"
    )
    review_decisions = {
        row["feature_family_id"]: row["identity_decision"] for row in review_rows
    }
    assert review_decisions["FAM_PROVISIONAL"] == "provisional_discovery"
    audit_decisions = {
        row["feature_family_id"]: row["identity_decision"] for row in audit_rows
    }
    assert audit_decisions["FAM_PROVISIONAL"] == "provisional_discovery"
    audit_blank_reasons = {row["blank_reason"] for row in audit_rows}
    assert "missing_row_identity_support" in audit_blank_reasons
    assert "duplicate_loser" in audit_blank_reasons
    assert "ambiguous_ms1_owner" in audit_blank_reasons

    identity_rows = _read_tsv(
        write_alignment_matrix_identity_tsv(
            tmp_path / "alignment_matrix_identity.tsv",
            matrix,
        )
    )
    assert len(identity_rows) == len(tsv_rows)
    assert identity_rows[0]["matrix_row_index"] == "1"
    assert identity_rows[0]["peak_hypothesis_id"] == "FAM001"
    assert identity_rows[0]["source_feature_family_ids"] == "FAM001"
    assert identity_rows[0]["row_identity_basis"] == "no_split_peak_hypothesis"
    assert (
        identity_rows[0]["split_evaluation_status"]
        == "complete_no_product_ready_split"
    )
    assert identity_rows[0]["projection_status"] == "not_projection"


def test_istd_fixture_records_explicit_matching_fields():
    fixture = Path(
        "tests/fixtures/untargeted_alignment/istd_false_missing_fixture.csv",
    )
    rows = _read_csv(fixture)

    assert len(rows) == 16
    assert {row["targeted_identity"] for row in rows} == {
        "d3-5-medC",
        "d3-5-hmdC",
    }
    assert all(row["targeted_confidence"] == "HIGH" for row in rows)
    assert all(row["targeted_nl"] == "✓" for row in rows)
    assert all(row["sample_mapping_rule"] for row in rows)
    assert {
        (row["old_row_coordinate"], row["targeted_identity"])
        for row in rows
    } == {
        ("245.1332/12.28", "d3-5-medC"),
        ("261.1283/8.97", "d3-5-hmdC"),
    }


def _feature(
    feature_family_id: str,
    *,
    evidence: str,
    has_anchor: bool = True,
) -> SimpleNamespace:
    return SimpleNamespace(
        feature_family_id=feature_family_id,
        neutral_loss_tag="DNA_dR",
        family_center_mz=500.123,
        family_center_rt=8.49,
        family_product_mz=384.076,
        family_observed_neutral_loss_da=116.047,
        has_anchor=has_anchor,
        event_cluster_ids=("OWN-s1-000001",),
        event_member_count=1,
        evidence=evidence,
        review_only=False,
    )


def _cell(
    sample_stem: str,
    cluster_id: str,
    status: str,
    area: float | None,
    *,
    selected_integration: IntegrationResult | None = None,
) -> AlignedCell:
    if (
        selected_integration is None
        and status in {"detected", "rescued"}
        and area is not None
        and area > 0
    ):
        selected_integration = _integration(
            raw_area=area,
            asls_area=area,
            morphology_area=area,
        )
    return AlignedCell(
        sample_stem=sample_stem,
        cluster_id=cluster_id,
        status=status,  # type: ignore[arg-type]
        area=area,
        apex_rt=8.49 if area is not None else None,
        height=100.0 if area is not None else None,
        peak_start_rt=8.4 if area is not None else None,
        peak_end_rt=8.6 if area is not None else None,
        rt_delta_sec=0.0 if area is not None else None,
        trace_quality="clean" if area is not None else status,
        scan_support_score=0.8 if area is not None else None,
        source_candidate_id=f"{sample_stem}#1" if status == "detected" else None,
        source_raw_file=Path(f"{sample_stem}.raw") if status == "detected" else None,
        reason=(
            "duplicate MS1 peak claim; winner=FAM001; original_status=detected"
            if status == "duplicate_assigned"
            else status
        ),
        selected_integration=selected_integration,
    )


def _integration(
    *,
    raw_area: float,
    asls_area: float | None,
    baseline_type: str = "asls",
    morphology_area: float | None = None,
) -> IntegrationResult:
    return IntegrationResult(
        rt_left_min=8.4,
        rt_apex_min=8.49,
        rt_right_min=8.6,
        raw_apex_rt_min=8.49,
        rt_width_min=0.2,
        height_raw=100.0,
        height_smoothed=100.0,
        area_raw_counts_seconds=raw_area,
        area_baseline_corrected=asls_area,
        baseline_type=baseline_type,
        boundary_sources=("test",),
        area_ms1_morphology=morphology_area,
        ms1_morphology_area_source=(
            "gaussian15_positive_asls_residual"
            if morphology_area is not None
            else ""
        ),
    )


def _read_tsv(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8") as handle:
        return list(csv.DictReader(handle, delimiter="\t"))


def _read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8") as handle:
        return list(csv.DictReader(handle))


def _worksheet_records(sheet) -> list[dict[str, object]]:
    rows = list(sheet.iter_rows(values_only=True))
    headers = [str(value) for value in rows[0]]
    return [
        dict(zip(headers, row))
        for row in rows[1:]
        if any(value is not None for value in row)
    ]
