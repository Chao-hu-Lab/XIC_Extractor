from pathlib import Path

import numpy as np
import pytest
from openpyxl import load_workbook

from xic_extractor import extractor
from xic_extractor.config import ExtractionConfig, Target
from xic_extractor.extractor import DiagnosticRecord, RunOutput
from xic_extractor.neutral_loss import NLResult
from xic_extractor.output.excel_pipeline import write_excel_from_run_output
from xic_extractor.signal_processing import PeakDetectionResult, PeakResult


@pytest.fixture(autouse=True)
def _disable_reader_preflight(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(
        "xic_extractor.extraction.pipeline.preflight_raw_reader",
        lambda _dll_dir: [],
        raising=False,
    )


def test_default_output_only_has_one_xlsx(tmp_path: Path, monkeypatch) -> None:
    xlsx_path = _run_pipeline(tmp_path, monkeypatch, keep_intermediate_csv=False)
    output_dir = tmp_path / "output"

    assert list(output_dir.glob("*.csv")) == []
    assert list(output_dir.glob("*.xlsx")) == [xlsx_path]


def test_keep_intermediate_csv_emits_csvs(tmp_path: Path, monkeypatch) -> None:
    _run_pipeline(tmp_path, monkeypatch, keep_intermediate_csv=True)

    output_dir = tmp_path / "output"
    actual = {path.name for path in output_dir.glob("*.csv")}
    assert {"xic_results.csv", "xic_results_long.csv", "xic_diagnostics.csv"} <= actual


def test_default_xlsx_has_overview_landing_sheet(tmp_path: Path, monkeypatch) -> None:
    xlsx_path = _run_pipeline(tmp_path, monkeypatch, emit_score_breakdown=False)

    wb = load_workbook(xlsx_path)
    assert wb.sheetnames == [
        "Overview",
        "Review Queue",
        "XIC Results",
        "Summary",
        "Targets",
        "Diagnostics",
        "Run Metadata",
    ]
    assert wb.active.title == "Overview"
    assert wb["Diagnostics"].sheet_state == "hidden"


def test_score_breakdown_appears_when_enabled(tmp_path: Path, monkeypatch) -> None:
    xlsx_path = _run_pipeline(tmp_path, monkeypatch, emit_score_breakdown=True)

    wb = load_workbook(xlsx_path)
    assert wb.sheetnames == [
        "Overview",
        "Review Queue",
        "XIC Results",
        "Summary",
        "Targets",
        "Diagnostics",
        "Run Metadata",
        "Score Breakdown",
    ]
    assert wb["Diagnostics"].sheet_state == "hidden"


def test_landing_sheet_when_diagnostics_empty(tmp_path: Path, monkeypatch) -> None:
    xlsx_path = _run_pipeline(tmp_path, monkeypatch, with_diagnostics=False)

    wb = load_workbook(xlsx_path)
    assert wb.active.title == "Overview"


def test_landing_sheet_when_diagnostics_present(tmp_path: Path, monkeypatch) -> None:
    xlsx_path = _run_pipeline(tmp_path, monkeypatch, with_diagnostics=True)

    wb = load_workbook(xlsx_path)
    assert wb.active.title == "Overview"


def test_run_metadata_sheet_has_required_keys(tmp_path: Path, monkeypatch) -> None:
    xlsx_path = _run_pipeline(tmp_path, monkeypatch)

    wb = load_workbook(xlsx_path)
    ws = wb["Run Metadata"]
    keys = {ws.cell(row=row, column=1).value for row in range(2, ws.max_row + 1)}
    assert {"config_hash", "app_version", "generated_at", "resolver_mode"} <= keys


def _run_pipeline(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    *,
    keep_intermediate_csv: bool = False,
    emit_score_breakdown: bool = False,
    with_diagnostics: bool = False,
) -> Path:
    config = _config(
        tmp_path,
        keep_intermediate_csv=keep_intermediate_csv,
        emit_score_breakdown=emit_score_breakdown,
    )
    config.data_dir.mkdir(parents=True)
    config.dll_dir.mkdir(parents=True)
    (config.data_dir / "SampleA.raw").write_text("", encoding="utf-8")
    target = _target("Analyte")

    monkeypatch.setattr("xic_extractor.extractor.open_raw", _open_raw_factory())
    monkeypatch.setattr(
        "xic_extractor.extractor.find_peak_and_area",
        lambda *_args, **_kwargs: _ok_peak(),
    )
    monkeypatch.setattr(
        "xic_extractor.extractor.check_nl",
        lambda *_args, **_kwargs: NLResult("OK", 1.0, None, 1, 0, 1),
    )

    run_output = extractor.run(config, [target])
    if with_diagnostics:
        run_output = RunOutput(
            file_results=run_output.file_results,
            diagnostics=[
                DiagnosticRecord(
                    sample_name="SampleA",
                    target_label="Analyte",
                    issue="NL_FAIL",
                    reason="forced diagnostic",
                )
            ],
        )
    xlsx_path = config.output_csv.parent / "xic_results_contract.xlsx"
    return write_excel_from_run_output(
        config,
        [target],
        run_output,
        output_path=xlsx_path,
    )


def _config(
    tmp_path: Path,
    *,
    keep_intermediate_csv: bool,
    emit_score_breakdown: bool,
) -> ExtractionConfig:
    return ExtractionConfig(
        data_dir=tmp_path / "raw",
        dll_dir=tmp_path / "dll",
        output_csv=tmp_path / "output" / "xic_results.csv",
        diagnostics_csv=tmp_path / "output" / "xic_diagnostics.csv",
        smooth_window=15,
        smooth_polyorder=3,
        peak_rel_height=0.95,
        peak_min_prominence_ratio=0.10,
        ms2_precursor_tol_da=0.5,
        nl_min_intensity_ratio=0.01,
        keep_intermediate_csv=keep_intermediate_csv,
        emit_score_breakdown=emit_score_breakdown,
        config_hash="abc12345",
    )


def _target(label: str) -> Target:
    return Target(
        label=label,
        mz=258.1085,
        rt_min=8.0,
        rt_max=10.0,
        ppm_tol=20.0,
        neutral_loss_da=116.0474,
        nl_ppm_warn=20.0,
        nl_ppm_max=50.0,
        is_istd=False,
        istd_pair="",
    )


def _ok_peak() -> PeakDetectionResult:
    peak = PeakResult(
        rt=9.0,
        intensity=1000.0,
        intensity_smoothed=900.0,
        area=1234.5,
        peak_start=8.8,
        peak_end=9.2,
    )
    return PeakDetectionResult(
        status="OK",
        peak=peak,
        n_points=12,
        max_smoothed=900.0,
        n_prominent_peaks=1,
        confidence="HIGH",
        reason="all checks passed",
        severities=((0, "local_sn"),),
    )


def _open_raw_factory():
    class _FakeRaw:
        def __enter__(self):
            return self

        def __exit__(self, *_args: object) -> None:
            return None

        def extract_xic(
            self, mz: float, rt_min: float, rt_max: float, ppm_tol: float
        ) -> tuple[np.ndarray, np.ndarray]:
            return np.asarray([8.8, 9.0, 9.2]), np.asarray([1.0, 3.0, 1.0])

        def iter_ms2_scans(self, rt_min: float, rt_max: float):
            return iter([])

    return lambda *_args, **_kwargs: _FakeRaw()
