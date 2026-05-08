import csv
from dataclasses import replace
from pathlib import Path

from openpyxl import Workbook, load_workbook

import scripts.csv_to_excel as csv_to_excel
from scripts.csv_to_excel import (
    _build_data_sheet,
    _build_diagnostics_sheet,
    _build_review_queue_sheet,
    _build_summary_sheet,
    _build_targets_sheet,
    _review_queue_rows,
    _wide_to_long_rows,
    run,
)
from xic_extractor.config import ExtractionConfig, Target


def test_build_data_sheet_uses_row_based_compact_view_with_hidden_advanced_columns(
    tmp_path: Path,
) -> None:
    rows = [
        _long_row("Tumor_1", "Analyte", "9.1234", "12345.6", "OK"),
        _long_row("Tumor_2", "Analyte", "9.2000", "22345.6", "WARN_12.3ppm"),
        _long_row("Tumor_3", "Analyte", "9.3000", "32345.6", "NL_FAIL"),
        _long_row("Tumor_4", "Analyte", "9.4000", "42345.6", "NO_MS2"),
        _long_row("Tumor_5", "Analyte", "ND", "ND", "ND"),
    ]
    wb = Workbook()
    ws = wb.active

    _build_data_sheet(ws, rows)

    assert [ws.cell(row=1, column=i).value for i in range(1, 15)] == [
        "SampleName",
        "Group",
        "Target",
        "Role",
        "ISTD Pair",
        "RT",
        "Area",
        "NL",
        "Int",
        "PeakStart",
        "PeakEnd",
        "PeakWidth",
        "Confidence",
        "Reason",
    ]
    assert ws["F2"].value == 9.1234
    assert ws["F2"].number_format == "0.0000"
    assert ws["G2"].value == 12345.6
    assert ws["G2"].number_format == "0.00E+00"
    assert ws["H2"].value == "✓"
    assert ws["H3"].value == "⚠ 12.3ppm"
    assert ws["H4"].value == "✗ NL"
    assert ws["H5"].value == "— MS2"
    assert ws["H2"].fill.fgColor.rgb.endswith("C8E6C9")
    assert ws["I2"].value == 1000.0
    assert ws["I2"].number_format == "0.00E+00"
    assert abs(ws["L2"].value - 0.4) < 1e-9
    assert ws["L2"].number_format == "0.0000"
    assert ws.column_dimensions["I"].hidden is True
    assert ws.column_dimensions["J"].hidden is True
    assert ws.column_dimensions["K"].hidden is True
    assert ws.column_dimensions["L"].hidden is True
    assert ws["M2"].value == "HIGH"
    assert ws["N2"].value == "all checks passed"
    assert ws.auto_filter.ref == "A1:N6"


def test_data_sheet_merges_repeated_sample_and_group_cells() -> None:
    rows = [
        _long_row("Tumor_1", "Analyte", "9.1", "10000", "OK"),
        _long_row("Tumor_1", "ISTD", "9.2", "20000", "OK", role="ISTD"),
        _long_row("Normal_1", "Analyte", "9.3", "30000", "OK"),
    ]
    wb = Workbook()
    ws = wb.active

    _build_data_sheet(ws, rows)

    merged_ranges = {str(cell_range) for cell_range in ws.merged_cells.ranges}
    assert "A2:A3" in merged_ranges
    assert "B2:B3" in merged_ranges
    assert "A4:A4" not in merged_ranges
    assert ws["A2"].value == "Tumor_1"
    assert ws["A3"].value is None
    assert ws["B2"].value == "Tumor"
    assert ws["B3"].value is None


def test_data_sheet_forces_sample_names_and_target_labels_to_literal_text() -> None:
    rows = [_long_row("+Sample", "=Bad", "9.1", "10000", "OK")]
    wb = Workbook()
    ws = wb.active

    _build_data_sheet(ws, rows)

    assert ws["A2"].value == "'+Sample"
    assert ws["A2"].data_type != "f"
    assert ws["C2"].value == "'=Bad"
    assert ws["C2"].data_type != "f"


def test_build_overview_sheet_summarizes_batch_review_health() -> None:
    rows = [
        _long_row("Tumor_1", "AnalyteA", "9.0", "10000", "OK"),
        _long_row(
            "Tumor_2",
            "AnalyteA",
            "ND",
            "ND",
            "NL_FAIL",
            confidence="VERY_LOW",
            reason="concerns: nl_support (major)",
        ),
        _long_row(
            "Tumor_2",
            "AnalyteB",
            "12.0",
            "20000",
            "NO_MS2",
            confidence="MEDIUM",
        ),
    ]
    diagnostics = [
        {
            "SampleName": "Tumor_2",
            "Target": "AnalyteA",
            "Issue": "NL_FAIL",
            "Reason": "strict observed neutral loss missing",
        }
    ]
    review_rows = _review_queue_rows(rows, diagnostics)
    wb = Workbook()
    ws = wb.active

    csv_to_excel._build_overview_sheet(ws, rows, diagnostics, review_rows)

    assert ws.title == "Overview"
    assert ws["A1"].value == "XIC Review Overview"
    assert ws["A3"].value == "Samples"
    assert ws["B3"].value == 2
    assert ws["A4"].value == "Targets"
    assert ws["B4"].value == 2
    assert ws["A5"].value == "Review Items"
    assert ws["B5"].value == 2
    assert ws["A6"].value == "Diagnostics"
    assert ws["B6"].value == 1
    section_labels = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
    assert "Top Targets" in section_labels
    assert "Top Samples" in section_labels


def test_build_overview_sheet_forces_top_labels_to_literal_text() -> None:
    rows = [
        _long_row("+Sample", "=Analyte", "ND", "ND", "NL_FAIL"),
    ]
    diagnostics = [
        {
            "SampleName": "+Sample",
            "Target": "=Analyte",
            "Issue": "NL_FAIL",
            "Reason": "strict observed neutral loss missing",
        }
    ]
    review_rows = _review_queue_rows(rows, diagnostics)
    wb = Workbook()
    ws = wb.active

    csv_to_excel._build_overview_sheet(ws, rows, diagnostics, review_rows)

    assert ws["A10"].value == "'=Analyte"
    assert ws["A10"].data_type != "f"
    assert ws["A14"].value == "'+Sample"
    assert ws["A14"].data_type != "f"


def test_overview_explains_detected_and_flagged_rates() -> None:
    rows = [_long_row("Tumor_1", "Analyte", "9.0", "10000", "OK")]
    wb = Workbook()
    ws = wb.active

    csv_to_excel._build_overview_sheet(ws, rows, diagnostics=[], review_rows=[])

    values = [
        ws.cell(row=row_idx, column=1).value for row_idx in range(1, ws.max_row + 1)
    ]
    joined = "\n".join(str(value) for value in values if value)
    assert "Start with Summary Detection %" in joined
    assert "Review Queue has one row per sample-target needing attention" in joined
    assert "Diagnostics is a hidden technical log" in joined
    assert "HTML Review Report is for visual batch QA" in joined
    assert "Flagged % is review workload" in joined
    assert "NL_FAIL rows are review evidence, not counted detections" in joined
    assert "detected-but-flagged" not in joined
    assert "Score Breakdown is a technical audit sheet" in joined


def test_build_summary_sheet_uses_row_based_target_metrics() -> None:
    rows = [
        _long_row(
            "Tumor_1",
            "Analyte",
            "9.0",
            "10000",
            "OK",
            istd_pair="ISTD",
            confidence="HIGH",
        ),
        _long_row("Tumor_1", "ISTD", "9.1", "20000", "OK", role="ISTD"),
        _long_row(
            "Tumor_2",
            "Analyte",
            "9.2",
            "30000",
            "WARN_5ppm",
            istd_pair="ISTD",
            confidence="MEDIUM",
        ),
        _long_row("Tumor_2", "ISTD", "9.1", "60000", "OK", role="ISTD"),
        _long_row(
            "Tumor_3",
            "Analyte",
            "9.3",
            "50000",
            "NO_MS2",
            istd_pair="ISTD",
            confidence="HIGH",
        ),
        _long_row("Tumor_3", "ISTD", "9.1", "100000", "NO_MS2", role="ISTD"),
        _long_row(
            "Tumor_4",
            "Analyte",
            "9.4",
            "70000",
            "OK",
            istd_pair="ISTD",
            confidence="LOW",
        ),
        _long_row("Tumor_4", "ISTD", "ND", "ND", "ND", role="ISTD"),
    ]
    wb = Workbook()
    ws = wb.active

    _build_summary_sheet(ws, rows, count_no_ms2_as_detected=True)
    data = _summary_rows(ws)

    assert "Mean Int" not in data["headers"]
    assert data["Analyte"]["Role"] == "Analyte"
    assert data["Analyte"]["Detected"] == 4
    assert data["Analyte"]["Total"] == 4
    assert data["Analyte"]["Detection %"] == "100%"
    assert "Median Area (detected)" in data["headers"]
    assert "Area / ISTD ratio (paired detected)" in data["headers"]
    assert data["Analyte"]["Median Area (detected)"] == 40000.0
    assert ws["G2"].number_format == "0.00E+00"
    assert data["Analyte"]["Area / ISTD ratio (paired detected)"] == "—"
    assert data["Analyte"]["NL OK"] == 2
    assert data["Analyte"]["NL WARN"] == 1
    assert data["Analyte"]["NL FAIL"] == 0
    assert data["Analyte"]["NO MS2"] == 1
    assert data["Analyte"]["RT Delta vs ISTD"] == "0.1333±0.0577 min (n=3)"
    assert data["Analyte"]["Confidence HIGH"] == 2
    assert data["Analyte"]["Confidence MEDIUM"] == 1
    assert data["Analyte"]["Confidence LOW"] == 1
    assert data["Analyte"]["Confidence VERY_LOW"] == 0
    assert data["ISTD"]["Area / ISTD ratio (paired detected)"] == "—"


def test_summary_sheet_includes_target_health_metrics() -> None:
    rows = [
        _long_row("Tumor_1", "Analyte", "9.0", "10000", "OK", confidence="HIGH"),
        _long_row("Tumor_2", "Analyte", "9.1", "11000", "NL_FAIL", confidence="LOW"),
        _long_row(
            "Tumor_3",
            "Analyte",
            "9.2",
            "12000",
            "NO_MS2",
            confidence="MEDIUM",
        ),
    ]
    diagnostics = [
        {
            "SampleName": "Tumor_2",
            "Target": "Analyte",
            "Issue": "NL_FAIL",
            "Reason": "strict observed neutral loss missing",
        }
    ]
    review_rows = _review_queue_rows(rows, diagnostics)
    wb = Workbook()
    ws = wb.active

    _build_summary_sheet(
        ws,
        rows,
        count_no_ms2_as_detected=False,
        review_rows=review_rows,
    )
    data = _summary_rows(ws)

    assert "Flagged Rows" in data["headers"]
    assert "Flagged %" in data["headers"]
    assert "MS2/NL Flags" in data["headers"]
    assert "Low Confidence Rows" in data["headers"]
    assert "Review Items" not in data["headers"]
    assert "Problem Rate" not in data["headers"]
    assert "NL Problems" not in data["headers"]
    assert "Low Confidence" not in data["headers"]
    assert data["Analyte"]["Flagged Rows"] == 2
    assert data["Analyte"]["Flagged %"] == "67%"
    assert data["Analyte"]["MS2/NL Flags"] == 2
    assert data["Analyte"]["Low Confidence Rows"] == 1


def test_summary_detection_excludes_very_low_and_non_positive_area_rows() -> None:
    rows = [
        _long_row("S1", "Analyte", "9.0", "100", "OK", confidence="VERY_LOW"),
        _long_row("S2", "Analyte", "9.1", "110", "OK", confidence="LOW"),
        _long_row("S3", "Analyte", "9.2", "0", "OK", confidence="LOW"),
        _long_row("S4", "Analyte", "9.3", "-5", "OK", confidence="LOW"),
    ]
    wb = Workbook()
    ws = wb.active

    _build_summary_sheet(ws, rows, count_no_ms2_as_detected=False, review_rows=[])
    data = _summary_rows(ws)

    assert data["Analyte"]["Detected"] == 1
    assert data["Analyte"]["Detection %"] == "25%"
    assert data["Analyte"]["Median Area (detected)"] == 110.0


def test_summary_detection_excludes_very_low_rows() -> None:
    rows = [
        _long_row("S1", "Analyte", "9.0", "100", "OK", confidence="VERY_LOW"),
        _long_row("S2", "Analyte", "9.1", "110", "OK", confidence="LOW"),
    ]
    wb = Workbook()
    ws = wb.active

    _build_summary_sheet(ws, rows, count_no_ms2_as_detected=False, review_rows=[])
    data = _summary_rows(ws)

    assert data["Analyte"]["Detected"] == 1
    assert data["Analyte"]["Detection %"] == "50%"


def test_summary_puts_detection_rate_before_review_workload() -> None:
    rows = [
        _long_row("S1", "Analyte", "9.0", "100", "OK", istd_pair="ISTD"),
        _long_row("S1", "ISTD", "8.9", "50", "OK", role="ISTD"),
        _long_row("S2", "Analyte", "ND", "ND", "ND", istd_pair="ISTD"),
        _long_row("S2", "ISTD", "8.8", "50", "OK", role="ISTD"),
    ]
    wb = Workbook()
    ws = wb.active

    _build_summary_sheet(ws, rows, count_no_ms2_as_detected=False, review_rows=[])

    headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
    assert headers[:10] == [
        "Target",
        "Role",
        "ISTD Pair",
        "Detected",
        "Total",
        "Detection %",
        "Median Area (detected)",
        "Mean RT",
        "Area / ISTD ratio (paired detected)",
        "RT Delta vs ISTD",
    ]
    assert headers.index("Flagged Rows") > headers.index("Confidence VERY_LOW")


def test_summary_analytical_aggregates_exclude_numeric_nl_fail_rows() -> None:
    rows = [
        _long_row(
            "QC_1",
            "Analyte",
            "9.0",
            "100",
            "OK",
            istd_pair="ISTD",
        ),
        _long_row("QC_1", "ISTD", "8.9", "50", "OK", role="ISTD"),
        _long_row(
            "QC_2",
            "Analyte",
            "9.5",
            "1000",
            "NL_FAIL",
            istd_pair="ISTD",
            confidence="LOW",
        ),
        _long_row("QC_2", "ISTD", "9.1", "50", "OK", role="ISTD"),
        _long_row(
            "QC_3",
            "Analyte",
            "9.7",
            "500",
            "NO_MS2",
            istd_pair="ISTD",
            confidence="MEDIUM",
        ),
        _long_row("QC_3", "ISTD", "9.6", "50", "OK", role="ISTD"),
    ]
    review_rows = _review_queue_rows(
        rows,
        [
            {
                "SampleName": "QC_2",
                "Target": "Analyte",
                "Issue": "NL_FAIL",
                "Reason": "strict observed neutral loss missing",
            }
        ],
    )
    wb = Workbook()
    ws = wb.active

    _build_summary_sheet(
        ws,
        rows,
        count_no_ms2_as_detected=False,
        review_rows=review_rows,
    )
    data = _summary_rows(ws)

    assert data["Analyte"]["Flagged Rows"] == 2
    assert data["Analyte"]["Detected"] == 1
    assert data["Analyte"]["Detection %"] == "33%"
    assert data["Analyte"]["Mean RT"] == "9.0000"
    assert data["Analyte"]["Median Area (detected)"] == 100.0
    assert data["Analyte"]["Area / ISTD ratio (paired detected)"] == (
        "2.0000±0.0000, CV=— (n=1)"
    )
    assert data["Analyte"]["RT Delta vs ISTD"] == "0.1000±0.0000 min (n=1)"


def test_build_review_queue_sheet_prioritizes_rows_that_need_manual_review() -> None:
    rows = [
        _long_row(
            "Tumor_1",
            "Analyte",
            "9.0",
            "10000",
            "OK",
            confidence="LOW",
            reason="concerns: peak_width (major); noise_shape (minor)",
        ),
        _long_row(
            "Tumor_2",
            "Analyte",
            "9.2",
            "30000",
            "NO_MS2",
            confidence="MEDIUM",
            reason="all checks passed",
        ),
        _long_row("Tumor_3", "Analyte", "9.3", "40000", "OK"),
    ]
    diagnostics = [
        {
            "SampleName": "Tumor_1",
            "Target": "Analyte",
            "Issue": "NL_FAIL",
            "Reason": "selected candidate has trigger-only MS2 evidence",
        }
    ]
    wb = Workbook()
    ws = wb.active

    _build_review_queue_sheet(ws, _review_queue_rows(rows, diagnostics))

    headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
    assert headers == [
        "Priority",
        "Sample",
        "Target",
        "Role",
        "Status",
        "Why",
        "RT",
        "Area",
        "Action",
        "Issue Count",
        "Evidence",
    ]
    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref == "A1:K3"
    assert ws["A2"].value == 1
    assert ws["E2"].value == "Review"
    assert ws["F2"].value == "NL support failed"
    assert ws["I2"].value == "Check MS2 / NL evidence near selected RT"
    assert ws["E3"].value == "Check"
    assert ws["F3"].value == "MS2 trigger missing"
    assert ws["I3"].value == "Check whether missing DDA trigger is acceptable"
    assert ws["A2"].fill.fgColor.rgb.endswith("FFCDD2")
    assert ws["A3"].fill.fgColor.rgb.endswith("FFF9C4")


def test_review_queue_aggregates_diagnostics_by_sample_target() -> None:
    rows = [
        _long_row(
            "Tumor_1",
            "Analyte",
            "9.0",
            "10000",
            "NL_FAIL",
            confidence="LOW",
            reason="concerns: nl_support (major); local_sn (minor)",
        ),
        _long_row(
            "Tumor_1",
            "Analyte",
            "9.1",
            "11000",
            "OK",
            confidence="MEDIUM",
            reason="all checks passed",
        ),
    ]
    diagnostics = [
        {
            "SampleName": "Tumor_1",
            "Target": "Analyte",
            "Issue": "NL_FAIL",
            "Reason": "strict observed neutral loss missing",
        },
        {
            "SampleName": "Tumor_1",
            "Target": "Analyte",
            "Issue": "ANCHOR_MISMATCH",
            "Reason": "selected RT is far from anchor",
        },
    ]

    review_rows = _review_queue_rows(rows, diagnostics)

    assert len(review_rows) == 1
    assert review_rows[0]["Priority"] == "1"
    assert review_rows[0]["Sample"] == "Tumor_1"
    assert review_rows[0]["Target"] == "Analyte"
    assert review_rows[0]["Status"] == "Review"
    assert review_rows[0]["Why"] == "NL support failed"
    assert review_rows[0]["Action"] == "Check MS2 / NL evidence near selected RT"
    assert review_rows[0]["Issue Count"] == "2"
    assert review_rows[0]["Evidence"] == "NL_FAIL; ANCHOR_MISMATCH"


def test_review_queue_sheet_uses_worklist_columns() -> None:
    rows = [
        {
            "Priority": "1",
            "Sample": "Tumor_1",
            "Target": "Analyte",
            "Role": "Analyte",
            "Status": "Review",
            "Why": "NL support failed",
            "RT": "9.0",
            "Area": "10000",
            "Action": "Check MS2 / NL evidence near selected RT",
            "Issue Count": "2",
            "Evidence": "strict observed neutral loss missing",
        }
    ]
    wb = Workbook()
    ws = wb.active

    _build_review_queue_sheet(ws, rows)

    assert [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)] == [
        "Priority",
        "Sample",
        "Target",
        "Role",
        "Status",
        "Why",
        "RT",
        "Area",
        "Action",
        "Issue Count",
        "Evidence",
    ]
    assert ws["E2"].value == "Review"
    assert ws["F2"].value == "NL support failed"


def test_build_review_queue_sheet_keeps_empty_queue_readable() -> None:
    wb = Workbook()
    ws = wb.active

    _build_review_queue_sheet(
        ws,
        _review_queue_rows(
            [_long_row("Tumor_1", "Analyte", "9.3", "40000", "OK")],
            [],
        ),
    )

    assert ws.max_row == 1
    assert ws.auto_filter.ref == "A1:K1"


def test_targets_sheet_marks_expected_product_as_nominal_reference() -> None:
    wb = Workbook()
    ws = wb.active

    _build_targets_sheet(ws, [_target("Analyte")])

    assert ws["I1"].value == "Expected product m/z"
    assert ws["I1"].comment is not None
    assert "Nominal target product" in ws["I1"].comment.text
    assert "strict observed-loss" in ws["I1"].comment.text.lower()


def test_summary_area_ratio_uses_only_explicit_qc_samples() -> None:
    rows = [
        _long_row("Tumor_1", "Analyte", "9.0", "10000", "OK", istd_pair="ISTD"),
        _long_row("Tumor_1", "ISTD", "9.0", "10000", "OK", role="ISTD"),
        _long_row("Tumor_2", "Analyte", "9.0", "90000", "OK", istd_pair="ISTD"),
        _long_row("Tumor_2", "ISTD", "9.0", "10000", "OK", role="ISTD"),
        _long_row(
            "Breast Cancer Tissue_pooled_QC1",
            "Analyte",
            "9.0",
            "20000",
            "OK",
            istd_pair="ISTD",
        ),
        _long_row(
            "Breast Cancer Tissue_pooled_QC1",
            "ISTD",
            "9.0",
            "10000",
            "OK",
            role="ISTD",
        ),
        _long_row(
            "Breast Cancer Tissue_pooled_QC_2",
            "Analyte",
            "9.0",
            "40000",
            "OK",
            istd_pair="ISTD",
        ),
        _long_row(
            "Breast Cancer Tissue_pooled_QC_2",
            "ISTD",
            "9.0",
            "10000",
            "OK",
            role="ISTD",
        ),
    ]
    wb = Workbook()
    ws = wb.active

    _build_summary_sheet(ws, rows)
    data = _summary_rows(ws)

    assert data["Analyte"]["Detected"] == 4
    assert data["Analyte"]["Area / ISTD ratio (paired detected)"] == (
        "3.0000±1.4142, CV=47.1% (n=2)"
    )


def test_wide_to_long_rows_classifies_qc_from_sample_name_token() -> None:
    rows = [
        {
            "SampleName": "Breast Cancer Tissue_pooled_QC_1",
            "Analyte_RT": "9.0",
            "Analyte_Area": "10000",
            "Analyte_NL": "OK",
        },
        {
            "SampleName": "SampleA",
            "Analyte_RT": "9.0",
            "Analyte_Area": "10000",
            "Analyte_NL": "OK",
        },
        {
            "SampleName": "Breast_Cancer_Tissue_pooled_QC1",
            "Analyte_RT": "9.0",
            "Analyte_Area": "10000",
            "Analyte_NL": "OK",
        },
    ]

    long_rows = _wide_to_long_rows(rows, [_target("Analyte")])

    assert long_rows[0]["Group"] == "QC"
    assert long_rows[1]["Group"] == "Other"
    assert long_rows[2]["Group"] == "QC"


def test_run_writes_row_based_results_sheet_and_makes_overview_active(
    tmp_path: Path,
) -> None:
    config = _config(tmp_path)
    targets = [_target("Analyte")]
    config.output_csv.parent.mkdir(parents=True)
    _write_csv(
        config.output_csv,
        [
            {
                "SampleName": "Tumor_1",
                "Analyte_RT": "ND",
                "Analyte_Int": "ND",
                "Analyte_Area": "ND",
                "Analyte_PeakStart": "ND",
                "Analyte_PeakEnd": "ND",
                "Analyte_PeakWidth": "ND",
                "Analyte_NL": "NL_FAIL",
            }
        ],
    )
    _write_csv(
        config.output_csv.with_name("xic_results_long.csv"),
        [_long_row("Tumor_1", "Analyte", "ND", "ND", "NL_FAIL")],
    )
    _write_csv(
        config.diagnostics_csv,
        [
            {
                "SampleName": "Tumor_1",
                "Target": "Analyte",
                "Issue": "NL_FAIL",
                "Reason": "=unsafe reason",
            }
        ],
    )

    excel_path = run(config, targets)

    wb = load_workbook(excel_path)
    assert wb.sheetnames == [
        "Overview",
        "Review Queue",
        "XIC Results",
        "Summary",
        "Targets",
        "Diagnostics",
        "Run Metadata",
    ]
    assert wb.active.title == "Overview"
    ws_overview = wb["Overview"]
    assert ws_overview["A1"].value == "XIC Review Overview"
    assert ws_overview["B3"].value == 1
    assert ws_overview["B5"].value == 1
    assert ws_overview["B6"].value == 1
    ws_results = wb["XIC Results"]
    assert ws_results["A1"].value == "SampleName"
    assert ws_results["C1"].value == "Target"
    assert ws_results["G1"].value == "Area"
    assert ws_results["I1"].value == "Int"
    assert ws_results["M1"].value == "Confidence"
    assert ws_results["N1"].value == "Reason"
    assert ws_results.column_dimensions["I"].hidden is True
    ws = wb["Diagnostics"]
    assert [ws.cell(row=1, column=i).value for i in range(1, 5)] == [
        "SampleName",
        "Target",
        "Issue",
        "Reason",
    ]
    assert ws.auto_filter.ref == "A1:D2"
    assert ws["C2"].value == "NL_FAIL"
    assert ws["D2"].value == "selected candidate lacks strict NL match"
    assert ws["C2"].fill.fgColor.rgb.endswith("FFCDD2")
    assert ws.sheet_state == "hidden"
    ws_metadata = wb["Run Metadata"]
    assert ws_metadata["A1"].value == "Key"
    assert ws_metadata["B1"].value == "Value"
    assert ws_metadata["A2"].value == "config_hash"
    assert config.output_csv.exists()
    assert config.output_csv.with_name("xic_results_long.csv").exists()
    assert config.diagnostics_csv.exists()
    ws_review = wb["Review Queue"]
    assert ws_review["E2"].value == "Review"
    assert ws_review["F2"].value == "NL support failed"
    assert ws_review["K2"].value == "NL_FAIL"


def test_run_hides_diagnostics_as_technical_log(tmp_path: Path) -> None:
    config = _config(tmp_path)
    config.output_csv.parent.mkdir(parents=True, exist_ok=True)
    _write_csv(
        config.output_csv,
        [_wide_row("S1", [_target("Analyte")])],
    )
    _write_empty_diagnostics_csv(config.diagnostics_csv)

    out = run(config, [_target("Analyte")])
    wb = load_workbook(out)

    assert wb["Diagnostics"].sheet_state == "hidden"
    assert wb["Review Queue"].sheet_state == "visible"


def test_run_can_build_long_results_from_legacy_wide_csv_when_needed(
    tmp_path: Path,
) -> None:
    config = _config(tmp_path)
    targets = [_target("Analyte")]
    config.output_csv.parent.mkdir(parents=True)
    _write_csv(
        config.output_csv,
        [
            {
                "SampleName": "Tumor_1",
                "Analyte_RT": "9.1",
                "Analyte_Int": "1000",
                "Analyte_Area": "10000",
                "Analyte_PeakStart": "8.9",
                "Analyte_PeakEnd": "9.3",
                "Analyte_PeakWidth": "0.4000",
                "Analyte_NL": "OK",
            }
        ],
    )
    _write_empty_diagnostics_csv(config.diagnostics_csv)

    excel_path = run(config, targets)

    wb = load_workbook(excel_path)
    assert wb.sheetnames == [
        "Overview",
        "Review Queue",
        "XIC Results",
        "Summary",
        "Targets",
        "Diagnostics",
        "Run Metadata",
    ]
    assert wb.active.title == "Overview"
    assert wb["XIC Results"]["C2"].value == "Analyte"
    assert wb["Diagnostics"].auto_filter.ref == "A1:D1"


def test_run_emits_score_breakdown_sheet_when_enabled(tmp_path: Path) -> None:
    config = _config(tmp_path, emit_score_breakdown=True)
    targets = [_target("Analyte")]
    config.output_csv.parent.mkdir(parents=True)
    _write_csv(
        config.output_csv.with_name("xic_results_long.csv"),
        [
            _long_row(
                "Tumor_1",
                "Analyte",
                "9.1",
                "10000",
                "OK",
                confidence="MEDIUM",
                reason="concerns: local_sn (minor)",
            )
        ],
    )
    _write_csv(
        config.output_csv.with_name("xic_score_breakdown.csv"),
        [
            {
                "SampleName": "Tumor_1",
                "Target": "Analyte",
                "Final Confidence": "HIGH",
                "Detection Counted": "TRUE",
                "Caps": "",
                "Raw Score": "90",
                "Support": "strict_nl_ok; local_sn_strong",
                "Concerns": "",
                "Base Score": "50",
                "Positive Points": "40",
                "Negative Points": "0",
                "symmetry": "0",
                "local_sn": "1",
                "nl_support": "0",
                "rt_prior": "0",
                "rt_centrality": "0",
                "noise_shape": "0",
                "peak_width": "0",
                "Quality Penalty": "1",
                "Quality Flags": "too_broad",
                "Total Severity": "2",
                "Confidence": "MEDIUM",
                "Prior RT": "9.05",
                "Prior Source": "rolling_median",
            }
        ],
    )
    _write_empty_diagnostics_csv(config.diagnostics_csv)

    excel_path = run(config, targets)

    wb = load_workbook(excel_path)
    assert wb.sheetnames == [
        "Overview",
        "Review Queue",
        "XIC Results",
        "Summary",
        "Targets",
        "Diagnostics",
        "Run Metadata",
        "Score Breakdown",
    ]
    ws = wb["Score Breakdown"]
    headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
    values = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
    row = dict(zip(headers, values, strict=False))
    assert row["SampleName"] == "Tumor_1"
    assert row["Base Score"] == 50
    assert row["Positive Points"] == 40
    assert row["Negative Points"] == 0
    assert row["Raw Score"] == 90
    assert row["Final Confidence"] == "HIGH"
    assert row["Detection Counted"] == "TRUE"
    assert row["Support"] == "strict_nl_ok; local_sn_strong"
    assert row["Concerns"] is None
    assert row["Quality Penalty"] == 1
    assert row["Quality Flags"] == "too_broad"
    assert row["Total Severity"] == 2
    assert row["Confidence"] == "MEDIUM"


def test_run_emits_review_report_when_enabled(tmp_path: Path) -> None:
    config = _config(tmp_path, emit_review_report=True)
    targets = [_target("Analyte")]
    config.output_csv.parent.mkdir(parents=True)
    _write_csv(
        config.output_csv.with_name("xic_results_long.csv"),
        [_long_row("Tumor_1", "Analyte", "9.1", "10000", "OK")],
    )
    _write_empty_diagnostics_csv(config.diagnostics_csv)

    excel_path = run(config, targets)

    report_path = excel_path.with_name(
        excel_path.name.replace("xic_results_", "review_report_")
    ).with_suffix(".html")
    assert report_path.exists()
    assert "XIC Review Report" in report_path.read_text(encoding="utf-8")


def test_run_passes_injection_order_to_review_report(
    tmp_path: Path, monkeypatch
) -> None:
    config = replace(
        _config(tmp_path, emit_review_report=True),
        injection_order_source=tmp_path / "SampleInfo.csv",
    )
    config.injection_order_source.write_text(
        "Sample_Name,Injection_Order\nS1,1\n",
        encoding="utf-8",
    )
    config.output_csv.parent.mkdir(parents=True, exist_ok=True)
    _write_csv(config.output_csv, [_wide_row("S1", [_target("Analyte")])])
    _write_empty_diagnostics_csv(config.diagnostics_csv)
    calls = {}

    def _fake_write_review_report(path, rows, **kwargs):
        calls["injection_order"] = kwargs["injection_order"]
        path.write_text("<html></html>", encoding="utf-8")
        return path

    monkeypatch.setattr(
        "scripts.csv_to_excel.write_review_report",
        _fake_write_review_report,
    )

    run(config, [_target("Analyte")])

    assert calls["injection_order"] == {"S1": 1}


def test_workbook_sheet_tabs_signal_review_and_technical_roles(tmp_path: Path) -> None:
    config = _config(tmp_path, emit_score_breakdown=True)
    targets = [_target("Analyte")]
    config.output_csv.parent.mkdir(parents=True)
    _write_csv(
        config.output_csv.with_name("xic_results_long.csv"),
        [_long_row("Tumor_1", "Analyte", "ND", "ND", "NL_FAIL")],
    )
    _write_csv(
        config.output_csv.with_name("xic_score_breakdown.csv"),
        [
            {
                "SampleName": "Tumor_1",
                "Target": "Analyte",
                "symmetry": "0",
                "local_sn": "1",
                "nl_support": "2",
                "rt_prior": "0",
                "rt_centrality": "0",
                "noise_shape": "0",
                "peak_width": "0",
                "Quality Penalty": "0",
                "Quality Flags": "",
                "Total Severity": "3",
                "Confidence": "LOW",
                "Prior RT": "NA",
                "Prior Source": "",
            }
        ],
    )
    _write_empty_diagnostics_csv(config.diagnostics_csv)

    excel_path = run(config, targets)

    wb = load_workbook(excel_path)
    assert wb["Overview"].sheet_properties.tabColor.rgb.endswith("1F4E5F")
    assert wb["Review Queue"].sheet_properties.tabColor.rgb.endswith("1F4E5F")
    assert wb["XIC Results"].sheet_properties.tabColor.rgb.endswith("5B7C99")
    assert wb["Summary"].sheet_properties.tabColor.rgb.endswith("5B7C99")
    assert wb["Diagnostics"].sheet_properties.tabColor.rgb.endswith("B0BEC5")
    assert wb["Score Breakdown"].sheet_properties.tabColor.rgb.endswith("B0BEC5")


def test_review_queue_evidence_uses_short_tags() -> None:
    rows = [
        _long_row(
            "BenignfatBC1055_DNA",
            "8-oxodG",
            "17.177",
            "1850221.22",
            "NL_FAIL",
            confidence="LOW",
            reason="concerns: rt_prior (major); weak candidate: too_broad",
        )
    ]
    diagnostics = [
        {
            "SampleName": "BenignfatBC1055_DNA",
            "Target": "8-oxodG",
            "Issue": "NL_FAIL",
            "Reason": (
                "selected candidate has 2 candidate-aligned MS2 trigger scans; "
                "strict observed neutral loss 116.047 Da not detected in any "
                "aligned scan; "
                "alignment=region"
            ),
        },
        {
            "SampleName": "BenignfatBC1055_DNA",
            "Target": "8-oxodG",
            "Issue": "ANCHOR_RT_MISMATCH",
            "Reason": (
                "Paired analyte peak RT 17.177 min deviates 0.71 min from "
                "ISTD anchor at 16.470 min (allowed +/-0.50 min)"
            ),
        },
        {
            "SampleName": "BenignfatBC1055_DNA",
            "Target": "8-oxodG",
            "Issue": "MULTI_PEAK",
            "Reason": "4 prominent peaks detected in window [16.0, 18.0]",
        },
    ]

    queue = _review_queue_rows(rows, diagnostics)

    assert queue[0]["Evidence"] == "NL_FAIL; anchor dRT=0.71 min; multi_peak=4"
    assert "selected candidate has" not in queue[0]["Evidence"]
    assert len(queue[0]["Evidence"]) < 80


def test_diagnostics_sheet_uses_short_reason_text() -> None:
    rows = [
        {
            "SampleName": "S1",
            "Target": "A",
            "Issue": "NL_FAIL",
            "Reason": (
                "selected candidate has 2 candidate-aligned MS2 trigger scans; "
                "strict observed neutral loss 116.047 Da not detected in any "
                "aligned scan; "
                "alignment=region"
            ),
        }
    ]
    wb = Workbook()
    ws = wb.active

    _build_diagnostics_sheet(ws, rows)

    assert ws["D2"].value == "selected candidate lacks strict NL match"


def test_diagnostics_sheet_falls_back_to_issue_when_reason_is_empty() -> None:
    rows = [
        {
            "SampleName": "S1",
            "Target": "A",
            "Issue": "ODD_TECHNICAL_WARNING",
            "Reason": "",
        }
    ]
    wb = Workbook()
    ws = wb.active

    _build_diagnostics_sheet(ws, rows)

    assert ws["D2"].value == "ODD_TECHNICAL_WARNING"


def _long_row(
    sample_name: str,
    target: str,
    rt: str,
    area: str,
    nl: str,
    *,
    role: str = "Analyte",
    istd_pair: str = "",
    confidence: str = "HIGH",
    reason: str = "all checks passed",
) -> dict[str, str]:
    return {
        "SampleName": sample_name,
        "Group": "Tumor" if sample_name.startswith("Tumor") else "QC",
        "Target": target,
        "Role": role,
        "ISTD Pair": istd_pair,
        "RT": rt,
        "Area": area,
        "NL": nl,
        "Int": "1000",
        "PeakStart": "8.9",
        "PeakEnd": "9.3",
        "PeakWidth": "0.4000",
        "Confidence": confidence,
        "Reason": reason,
    }


def _wide_row(sample_name: str, targets: list[Target]) -> dict[str, str]:
    row = {"SampleName": sample_name}
    for target in targets:
        row.update(
            {
                f"{target.label}_RT": "9.1",
                f"{target.label}_Int": "1000",
                f"{target.label}_Area": "10000",
                f"{target.label}_PeakStart": "8.9",
                f"{target.label}_PeakEnd": "9.3",
                f"{target.label}_PeakWidth": "0.4000",
                f"{target.label}_NL": "OK",
            }
        )
    return row


def _summary_rows(ws) -> dict[str, object]:
    headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
    data: dict[str, object] = {"headers": headers}
    for row_idx in range(2, ws.max_row + 1):
        target = ws.cell(row=row_idx, column=1).value
        data[target] = {
            headers[col_idx - 1]: ws.cell(row=row_idx, column=col_idx).value
            for col_idx in range(1, ws.max_column + 1)
        }
    return data


def _write_csv(path: Path, rows: list[dict[str, str]]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def _write_empty_diagnostics_csv(path: Path) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(
            handle, fieldnames=["SampleName", "Target", "Issue", "Reason"]
        )
        writer.writeheader()


def _config(
    tmp_path: Path,
    *,
    emit_score_breakdown: bool = False,
    emit_review_report: bool = False,
) -> ExtractionConfig:
    return ExtractionConfig(
        data_dir=tmp_path / "raw",
        dll_dir=tmp_path / "dll",
        output_csv=tmp_path / "output" / "xic_results.csv",
        diagnostics_csv=tmp_path / "output" / "xic_diagnostics.csv",
        smooth_window=15,
        smooth_polyorder=3,
        peak_rel_height=0.95,
        peak_min_prominence_ratio=0.10,
        ms2_precursor_tol_da=0.5,
        nl_min_intensity_ratio=0.01,
        count_no_ms2_as_detected=True,
        emit_score_breakdown=emit_score_breakdown,
        emit_review_report=emit_review_report,
    )


def _target(label: str) -> Target:
    return Target(
        label=label,
        mz=258.1085,
        rt_min=8.0,
        rt_max=10.0,
        ppm_tol=20.0,
        neutral_loss_da=116.0474,
        nl_ppm_warn=20.0,
        nl_ppm_max=50.0,
        is_istd=False,
        istd_pair="",
    )
