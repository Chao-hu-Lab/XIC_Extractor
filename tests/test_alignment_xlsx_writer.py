from pathlib import Path
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook

from xic_extractor.alignment.matrix import AlignedCell, AlignmentMatrix
from xic_extractor.alignment.owner_group_delivery import (
    CROSS_SAMPLE_GROUP_CELL_COLUMNS,
    GROUP_REVIEW_PROJECTION_COLUMNS,
)
from xic_extractor.alignment.xlsx_writer import write_alignment_results_xlsx
from xic_extractor.peak_detection.hypotheses import IntegrationResult


def test_alignment_results_xlsx_has_matrix_review_metadata_sheets(tmp_path: Path):
    matrix = sample_alignment_matrix()

    path = write_alignment_results_xlsx(
        tmp_path / "alignment_results.xlsx",
        matrix,
        metadata={
            "schema_version": "alignment-results-v3",
            "resolver_mode": "local_minimum",
        },
    )

    workbook = load_workbook(path, data_only=True)
    assert workbook.sheetnames == ["Matrix", "Review", "Audit", "Metadata"]
    assert [cell.value for cell in workbook["Matrix"][1]] == [
        "Mz",
        "RT",
        "s1",
        "s2",
        "s3",
    ]
    assert workbook["Matrix"]["C2"].value == 100.0
    assert workbook["Matrix"]["D2"].value is None
    assert workbook["Matrix"]["E2"].value == 150.0
    assert [cell.value for cell in workbook["Review"][1]] == [
        "feature_family_id",
        *GROUP_REVIEW_PROJECTION_COLUMNS,
        "neutral_loss_tag",
        "detected_count",
        "rescued_count",
        "identity_decision",
        "identity_confidence",
        "primary_evidence",
        "identity_reason",
        "quantifiable_detected_count",
        "quantifiable_rescue_count",
        "accepted_cell_count",
        "accepted_rescue_count",
        "review_rescue_count",
        "absent_count",
        "unchecked_count",
        "duplicate_assigned_count",
        "ambiguous_ms1_owner_count",
        "include_in_primary_matrix",
        "row_flags",
        "artificial_adduct_role",
        "artificial_adduct_name",
        "artificial_adduct_related_family_id",
        "artificial_adduct_mz_delta_error_ppm",
        "artificial_adduct_rt_delta_min",
    ]
    assert _sheet_value(workbook["Review"], "group_hypothesis_id") == "FAM000001"
    assert _sheet_value(workbook["Review"], "detected_count") == 2
    assert _sheet_value(workbook["Review"], "primary_evidence") == "owner_identity"
    assert _sheet_value(workbook["Review"], "ambiguous_ms1_owner_count") == 1
    assert [cell.value for cell in workbook["Audit"][1]] == [
        "feature_family_id",
        *CROSS_SAMPLE_GROUP_CELL_COLUMNS,
        "sample_stem",
        "neutral_loss_tag",
        "family_center_mz",
        "family_center_rt",
        "raw_status",
        "production_status",
        "rescue_tier",
        "write_matrix_value",
        "blank_reason",
        "area",
        "primary_matrix_area",
        "primary_matrix_area_source",
        "primary_matrix_area_reason",
        "apex_rt",
        "rt_delta_sec",
        "claim_state",
        "row_flags",
        "identity_decision",
        "identity_confidence",
        "primary_evidence",
        "identity_reason",
        "quantifiable_detected_count",
        "quantifiable_rescue_count",
        "accepted_cell_count",
        "accepted_rescue_count",
        "review_rescue_count",
        "duplicate_assigned_count",
        "ambiguous_ms1_owner_count",
        "artificial_adduct_role",
        "artificial_adduct_name",
        "artificial_adduct_related_family_id",
        "artificial_adduct_mz_delta_error_ppm",
        "artificial_adduct_rt_delta_min",
        "region_decision_status",
        "region_decision_class",
        "region_product_action",
        "region_promotion_reason",
        "region_baseline_method",
        "reason",
    ]
    assert workbook["Metadata"]["A1"].value == "key"
    assert workbook["Metadata"]["A3"].value == "schema_version"
    assert workbook["Metadata"]["B3"].value == "alignment-results-v3"


def test_alignment_results_xlsx_audit_projects_region_decision_fields(
    tmp_path: Path,
):
    matrix = AlignmentMatrix(
        clusters=(sample_feature("FAM000001", evidence="owner_identity"),),
        sample_order=("s1", "s2", "s3"),
        cells=(
            sample_cell(
                "s1",
                "FAM000001",
                "detected",
                100.0,
                region=True,
            ),
            sample_cell("s2", "FAM000001", "ambiguous_ms1_owner", None),
            sample_cell("s3", "FAM000001", "detected", 150.0),
        ),
    )

    path = write_alignment_results_xlsx(
        tmp_path / "alignment_results.xlsx",
        matrix,
        metadata={"schema_version": "alignment-results-v3"},
    )

    workbook = load_workbook(path, data_only=True)
    assert _sheet_value(workbook["Audit"], "region_decision_status") == "evaluated"
    assert _sheet_value(workbook["Audit"], "region_decision_class") == "merge_suggested"
    assert _sheet_value(workbook["Audit"], "region_product_action") == (
        "safe_merge_eligible"
    )
    assert _sheet_value(workbook["Audit"], "region_promotion_reason") == (
        "adjacent_wis_local_minimum_merge"
    )
    assert _sheet_value(workbook["Audit"], "region_baseline_method") == "asls"
    assert "region_decision_status" not in _sheet_headers(workbook["Review"])
    assert workbook["Matrix"]["C2"].value == 100.0


def test_alignment_results_xlsx_blanks_duplicate_assigned_matrix_area(
    tmp_path: Path,
):
    base = sample_alignment_matrix()
    matrix = AlignmentMatrix(
        clusters=base.clusters,
        sample_order=("s1", "s2", "s3"),
        cells=(
            sample_cell("s1", "FAM000001", "detected", 100.0),
            sample_cell("s2", "FAM000001", "duplicate_assigned", 200.0),
            sample_cell("s3", "FAM000001", "detected", 300.0),
        ),
    )

    path = write_alignment_results_xlsx(
        tmp_path / "alignment_results.xlsx",
        matrix,
        metadata={"schema_version": "alignment-results-v3"},
    )

    workbook = load_workbook(path, data_only=True)
    assert workbook["Matrix"]["C2"].value == 100.0
    assert workbook["Matrix"]["D2"].value is None
    assert workbook["Matrix"]["E2"].value == 300.0
    assert _sheet_value(workbook["Review"], "duplicate_assigned_count") == 1


def test_alignment_results_xlsx_prefers_ms1_morphology_area(
    tmp_path: Path,
):
    matrix = AlignmentMatrix(
        clusters=(
            sample_feature("FAM000001", evidence="owner_complete_link;owner_count=2"),
        ),
        sample_order=("s1", "s2"),
        cells=(
            sample_cell(
                "s1",
                "FAM000001",
                "detected",
                100.0,
                selected_integration=sample_integration(
                    raw_area=177.0,
                    asls_area=144.0,
                    morphology_area=155.0,
                ),
            ),
            sample_cell("s2", "FAM000001", "detected", 120.0),
        ),
    )

    path = write_alignment_results_xlsx(
        tmp_path / "alignment_results.xlsx",
        matrix,
        metadata={"schema_version": "alignment-results-v3"},
    )

    workbook = load_workbook(path, data_only=True)
    assert workbook.sheetnames == ["Matrix", "Review", "Audit", "Metadata"]
    assert workbook["Matrix"]["C2"].value == 155.0
    assert workbook["Matrix"]["D2"].value == 120.0
    assert _sheet_value(workbook["Audit"], "area") == 100.0


def test_alignment_results_xlsx_excludes_review_only_rows_from_matrix(
    tmp_path: Path,
):
    matrix = AlignmentMatrix(
        clusters=(
            sample_feature("FAM000001", evidence="owner_complete_link;owner_count=2"),
            sample_feature("FAM000002", evidence="", has_anchor=False),
        ),
        sample_order=("s1", "s2"),
        cells=(
            sample_cell("s1", "FAM000001", "detected", 100.0),
            sample_cell("s2", "FAM000001", "detected", 110.0),
            sample_cell("s1", "FAM000002", "rescued", 200.0),
        ),
    )

    path = write_alignment_results_xlsx(
        tmp_path / "alignment_results.xlsx",
        matrix,
        metadata={"schema_version": "alignment-results-v3"},
    )

    workbook = load_workbook(path, data_only=True)
    matrix_rows = _worksheet_records(workbook["Matrix"])
    assert len(matrix_rows) == 1
    assert matrix_rows[0]["Mz"] == 242.114
    assert matrix_rows[0]["RT"] == 12.593
    assert matrix_rows[0]["s1"] == 100.0
    assert matrix_rows[0]["s2"] == 110.0
    assert workbook["Audit"]["A2"].value == "FAM000001"
    assert workbook["Audit"]["A3"].value == "FAM000001"
    assert workbook["Audit"]["A4"].value == "FAM000002"
    assert _sheet_value(workbook["Audit"], "rescue_tier", row=4) == "review_rescue"
    assert _sheet_value(workbook["Audit"], "blank_reason", row=4) == (
        "missing_row_identity_support"
    )


def test_alignment_results_xlsx_audit_explains_duplicate_blank(
    tmp_path: Path,
):
    matrix = AlignmentMatrix(
        clusters=(
            sample_feature(
                "FAM000001",
                evidence="owner_complete_link;owner_count=2",
            ),
        ),
        sample_order=("s1",),
        cells=(sample_cell("s1", "FAM000001", "duplicate_assigned", 200.0),),
    )

    path = write_alignment_results_xlsx(
        tmp_path / "alignment_results.xlsx",
        matrix,
        metadata={"schema_version": "alignment-results-v3"},
    )

    workbook = load_workbook(path, data_only=True)
    assert workbook["Matrix"]["A2"].value is None
    assert workbook["Audit"]["A2"].value == "FAM000001"
    assert _sheet_value(workbook["Audit"], "raw_status") == "duplicate_assigned"
    assert _sheet_value(workbook["Audit"], "write_matrix_value") is False
    assert _sheet_value(workbook["Audit"], "blank_reason") == "duplicate_loser"


def test_alignment_results_xlsx_escapes_formula_like_external_strings(
    tmp_path: Path,
):
    matrix = AlignmentMatrix(
        clusters=(
            sample_feature(
                "=FAM000001",
                evidence="owner_identity",
                neutral_loss_tag="-DNA_dR",
            ),
        ),
        sample_order=("+Sample_A", "Sample_B"),
        cells=(
            sample_cell(
                "+Sample_A",
                "=FAM000001",
                "detected",
                100.0,
                reason="@audit reason",
            ),
            sample_cell(
                "Sample_B",
                "=FAM000001",
                "detected",
                120.0,
            ),
        ),
    )

    path = write_alignment_results_xlsx(
        tmp_path / "alignment_results.xlsx",
        matrix,
        metadata={"schema_version": "@metadata value"},
    )

    workbook = load_workbook(path, data_only=False)
    assert workbook["Matrix"]["C1"].value == "'+Sample_A"
    assert workbook["Matrix"]["C1"].data_type != "f"
    assert _sheet_value(workbook["Review"], "neutral_loss_tag") == "'-DNA_dR"
    assert _sheet_cell(workbook["Review"], "neutral_loss_tag").data_type != "f"
    assert _sheet_value(workbook["Audit"], "sample_stem") == "'+Sample_A"
    assert _sheet_value(workbook["Audit"], "reason") == "'@audit reason"
    assert _sheet_value(workbook["Audit"], "write_matrix_value") is True
    assert workbook["Metadata"]["B2"].value == "'@metadata value"
    assert workbook["Metadata"]["B2"].data_type != "f"


def test_alignment_results_xlsx_rejects_orphan_audit_cell(tmp_path: Path):
    matrix = AlignmentMatrix(
        clusters=(sample_feature("FAM000001", evidence="owner_identity"),),
        sample_order=("s1",),
        cells=(sample_cell("s1", "FAM999999", "detected", 100.0),),
    )

    with pytest.raises(ValueError, match="unknown cluster: FAM999999"):
        write_alignment_results_xlsx(
            tmp_path / "alignment_results.xlsx",
            matrix,
            metadata={"schema_version": "alignment-results-v3"},
        )


def _sheet_headers(sheet) -> list[str]:
    return [cell.value for cell in sheet[1]]


def _sheet_cell(sheet, column_name: str, *, row: int = 2):
    return sheet.cell(
        row=row,
        column=_sheet_headers(sheet).index(column_name) + 1,
    )


def _sheet_value(sheet, column_name: str, *, row: int = 2):
    return _sheet_cell(sheet, column_name, row=row).value


def _worksheet_records(sheet) -> list[dict[str, object]]:
    rows = list(sheet.iter_rows(values_only=True))
    headers = [str(value) for value in rows[0]]
    return [
        dict(zip(headers, row))
        for row in rows[1:]
        if any(value is not None for value in row)
    ]


def sample_alignment_matrix() -> AlignmentMatrix:
    cluster = sample_feature("FAM000001", evidence="owner_identity")
    return AlignmentMatrix(
        clusters=(cluster,),
        sample_order=("s1", "s2", "s3"),
        cells=(
            sample_cell("s1", "FAM000001", "detected", 100.0),
            sample_cell("s2", "FAM000001", "ambiguous_ms1_owner", None),
            sample_cell("s3", "FAM000001", "detected", 150.0),
        ),
    )


def sample_feature(
    feature_family_id: str,
    *,
    evidence: str,
    has_anchor: bool = True,
    neutral_loss_tag: str = "DNA_dR",
):
    return SimpleNamespace(
        feature_family_id=feature_family_id,
        neutral_loss_tag=neutral_loss_tag,
        family_center_mz=242.114,
        family_center_rt=12.593,
        family_product_mz=126.066,
        family_observed_neutral_loss_da=116.048,
        has_anchor=has_anchor,
        event_cluster_ids=("OWN-s1-000001",),
        event_member_count=1,
        evidence=evidence,
        review_only=False,
    )


def sample_cell(
    sample: str,
    cluster_id: str,
    status: str,
    area: float | None,
    *,
    reason: str | None = None,
    selected_integration: IntegrationResult | None = None,
    region: bool = False,
) -> AlignedCell:
    if (
        selected_integration is None
        and status in {"detected", "rescued"}
        and _positive_area(area)
    ):
        selected_integration = sample_integration(
            raw_area=area,
            asls_area=area,
            morphology_area=area,
        )
    return AlignedCell(
        sample_stem=sample,
        cluster_id=cluster_id,
        status=status,  # type: ignore[arg-type]
        area=area,
        apex_rt=12.593 if area else None,
        height=1000.0 if area else None,
        peak_start_rt=12.55 if area else None,
        peak_end_rt=12.64 if area else None,
        rt_delta_sec=0.0 if area else None,
        trace_quality=status,
        scan_support_score=None,
        source_candidate_id="s1#6095" if area else None,
        source_raw_file=None,
        reason=reason or status,
        region_decision_status="evaluated" if region else "",
        region_decision_class="merge_suggested" if region else "",
        region_product_action="safe_merge_eligible" if region else "",
        region_promotion_reason=(
            "adjacent_wis_local_minimum_merge" if region else ""
        ),
        region_baseline_method="asls" if region else "",
        selected_integration=selected_integration,
    )


def sample_integration(
    *,
    raw_area: float,
    asls_area: float | None,
    baseline_type: str = "asls",
    morphology_area: float | None = None,
) -> IntegrationResult:
    return IntegrationResult(
        rt_left_min=12.55,
        rt_apex_min=12.593,
        rt_right_min=12.64,
        raw_apex_rt_min=12.593,
        rt_width_min=0.09,
        height_raw=1000.0,
        height_smoothed=1000.0,
        area_raw_counts_seconds=raw_area,
        area_baseline_corrected=asls_area,
        baseline_type=baseline_type,
        boundary_sources=("test",),
        area_ms1_morphology=morphology_area,
        ms1_morphology_area_source=(
            "gaussian15_positive_asls_residual"
            if morphology_area is not None
            else ""
        ),
    )


def _positive_area(value: float | None) -> bool:
    return value is not None and value > 0
